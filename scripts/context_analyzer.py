"""
context_analyzer.py  ── SA-7 (Sub Agent 7: Context Analyzer)
================================================================
역할: 기사 맥락 기반 진행단계 판단 — 규칙기반(1단계) + Claude Haiku(2단계)

아키텍처 (2단계 분류):
  ┌─────────────────────────────────────────────────────────┐
  │  1단계 — 규칙기반 분류 (무료, 빠름, API 불필요)         │
  │    키워드 점수 집계 → 진행단계(PLANNING~OPERATION) 판정  │
  │    신뢰도 ≥ 0.65 → 확정 (Haiku 호출 생략)              │
  ├─────────────────────────────────────────────────────────┤
  │  2단계 — Claude Haiku 분석 (신뢰도 < 0.65인 기사만)     │
  │    knowledge_index Layer1 (사업개요·KPI) 주입            │
  │    → project_stage / milestone / next_watch 생성        │
  │    → 월 $0.06~0.10 수준 (HIGH 기사 기준 ~30건/주)       │
  └─────────────────────────────────────────────────────────┘

출력:
  data/agent_output/context_output.json   ← SA-8 MI 보고서가 읽음
  data/agent_output/stage_timeline.json   ← 대시보드 Timeline 뷰

실행:
  python3 scripts/context_analyzer.py              # 전체 실행
  python3 scripts/context_analyzer.py --rule-only  # 규칙기반만 (API 불필요)
  python3 scripts/context_analyzer.py --days 14    # 14일치 기사 분석

GitHub Actions 연동:
  daily_pipeline.yml — SA-6 품질검증 다음 단계로 실행

영구 제약:
  - Haiku 모델: claude-haiku-4-5-20251001 고정
  - date 키 fallback: article.get('date') or article.get('published_date')
  - Anthropic API: 번역 금지 — 맥락분석에만 사용
  - YAML inline python 금지: 이 스크립트를 직접 호출

버전: v1.0 (2026-04-24)
"""

import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import requests

# ── 경로 설정 ──────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent.parent
DATA_DIR      = BASE_DIR / 'data'
SHARED_DIR    = DATA_DIR / 'shared'
AGENT_OUT_DIR = DATA_DIR / 'agent_output'
EXCEL_PATH    = DATA_DIR / 'database' / 'Vietnam_Infra_News_Database_Final.xlsx'

# knowledge_index 탐색 경로 (v2.3 기준)
KI_PATHS = [
    BASE_DIR / 'docs'  / 'shared' / 'knowledge_index.json',   # 실제 경로 (Genspark 공유)
    DATA_DIR / 'shared' / 'knowledge_index.json',
    DATA_DIR / 'shared' / 'layer1_data.json',
    AGENT_OUT_DIR / 'knowledge_index.json',
]

CONTEXT_OUT  = AGENT_OUT_DIR / 'context_output.json'
TIMELINE_OUT = AGENT_OUT_DIR / 'stage_timeline.json'

# ── Anthropic API 설정 ────────────────────────────────────────────────────
ANTHROPIC_API_URL = 'https://api.anthropic.com/v1/messages'
HAIKU_MODEL       = 'claude-haiku-4-5-20251001'   # 영구 고정
MAX_TOKENS_HAIKU  = 600    # 진행단계 분석에 충분
HAIKU_TIMEOUT     = 40

# ── 로깅 ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='[SA-7 %(asctime)s] %(message)s',
    datefmt='%H:%M:%S'
)
log = logging.getLogger('SA-7')


# ══════════════════════════════════════════════════════════════════════════
#  1단계 — 규칙기반 분류 엔진
# ══════════════════════════════════════════════════════════════════════════

# ── 진행단계 신호어 사전 ──────────────────────────────────────────────────
# 각 단계별 한국어·영어·베트남어 키워드
# 값이 클수록 해당 단계임을 강하게 시사
STAGE_RULES: dict[str, dict] = {

    'PLANNING': {
        'weight': 1.0,
        'ko': [
            '계획','결정','승인','결의','발표','수립','검토','예정','허가','인허가',
            '법제화','고시','지정','확정','방향','목표','로드맵','전략','정책',
            '마스터플랜','기본계획','타당성','FS','예비타당성',
        ],
        'en': [
            'plan','planning','decision','approve','approval','announced',
            'resolution','policy','strategy','feasibility','roadmap',
            'target','milestone','framework','regulation','decree','law',
            'Decision','Resolution','Decree','Law',
        ],
        'vi': [
            'kế hoạch','quyết định','phê duyệt','công bố','chiến lược',
            'nghị quyết','mục tiêu','quy hoạch','chính sách','luật',
        ],
    },

    'TENDERING': {
        'weight': 1.2,
        'ko': [
            '입찰','조달','공고','RFP','선정','계약','협약','MOU','낙찰',
            '사업자 선정','파트너십','투자 유치','투자자 모집','PPP 계약',
            'EPC 계약','LOI','양해각서','투자 협약',
        ],
        'en': [
            'tender','bid','contract','MOU','LOI','selected','awarded',
            'procurement','RFP','partnership','invest','investor','EPC',
            'PPP','joint venture','consortium','agreement signed',
            'financial close','FID','FI decision',
        ],
        'vi': [
            'đấu thầu','hợp đồng','đầu tư','nhà đầu tư','MOU',
            'liên doanh','ký kết','chọn','lựa chọn',
        ],
    },

    'CONSTRUCTION': {
        'weight': 1.3,
        'ko': [
            '착공','건설','공사','시공','건축','설치','구축','투자 집행',
            '공정','착수','토목','인프라 건설','현장','공사 중','건설 중',
            '건설 착수','설비','배관','전선','파이프라인 건설',
        ],
        'en': [
            'construction','groundbreaking','build','building','install',
            'pipeline laid','under construction','work begins','EPC started',
            'civil work','infrastructure','commissioning','developing',
            'erected','assembled','drilling','excavation',
        ],
        'vi': [
            'xây dựng','khởi công','thi công','lắp đặt','công trình',
            'đang xây','hạ tầng','thiết kế','xây lắp',
        ],
    },

    'COMPLETION': {
        'weight': 1.5,
        'ko': [
            '준공','개통','개항','개장','완공','가동 시작','상업 운전 개시',
            '준공식','시운전','테스트 완료','인수','인도','개막','출범',
            '취항','운항 시작','정식 가동','운전 개시',
        ],
        'en': [
            'inaugurated','commissioned','opened','launched','completed',
            'operational','online','COD','commercial operation',
            'handover','delivered','first power','grid connected',
            'inaugural flight','ribbon cutting','ceremony',
        ],
        'vi': [
            'khánh thành','hoàn thành','đưa vào','vận hành','khai trương',
            'chính thức','hoạt động','bàn giao','mở cửa','ra mắt',
        ],
    },

    'OPERATION': {
        'weight': 1.0,
        'ko': [
            '운영 중','가동 중','처리량','발전량','수익','수율','유지보수',
            '확장','증설','효율화','성과','실적','연간','분기','생산량',
            '처리 실적','공급량','이용률','가동률',
        ],
        'en': [
            'operating','operational','running','generating','producing',
            'capacity factor','output','revenue','maintenance','expanding',
            'upgrade','performance','annual','quarterly','throughput',
            'MW generated','m3 treated','efficiency','utilization',
        ],
        'vi': [
            'đang vận hành','hoạt động','sản xuất','hiệu quả','mở rộng',
            'nâng cấp','doanh thu','công suất','sản lượng',
        ],
    },
}

# ── 고신뢰 완료 패턴 (정규식) ─────────────────────────────────────────────
# 이 패턴에 해당하면 단계를 즉시 COMPLETION으로 확정
COMPLETION_REGEX = [
    r'준공\s*(완료|식|됩니다|됐다|했다)',
    r'(개통|개항|개장)\s*(완료|식|됩니다)',
    r'commercial\s+operation\s+(started|commenced|began|date)',
    r'(inaugurated|commissioned|opened)\s+(?:on|by|the)',
    r'COD\s*:\s*\d{4}',
    r'khánh\s+thành',
    r'chính\s+thức\s+vận\s+hành',
]

# ── 진행단계 표시 레이블 ──────────────────────────────────────────────────
STAGE_META = {
    'PLANNING':     {'label_ko': '계획·승인 단계',    'color': '3B82F6', 'order': 1},
    'TENDERING':    {'label_ko': '입찰·계약 단계',    'color': 'F59E0B', 'order': 2},
    'CONSTRUCTION': {'label_ko': '건설·시공 단계',    'color': 'EF4444', 'order': 3},
    'COMPLETION':   {'label_ko': '준공·개통 완료',    'color': '10B981', 'order': 4},
    'OPERATION':    {'label_ko': '운영·확장 단계',    'color': '6366F1', 'order': 5},
    'UNKNOWN':      {'label_ko': '단계 미확정',       'color': '9CA3AF', 'order': 0},
}


def rule_based_classify(text: str) -> tuple[str, float, dict]:
    """
    1단계: 규칙기반 진행단계 분류.

    Args:
        text: 분석 대상 텍스트 (제목 + 요약)

    Returns:
        (stage, confidence, score_breakdown)
        confidence: 0.0~1.0 (≥0.65면 Haiku 호출 생략)
    """
    text_lower = text.lower()
    scores: dict[str, float] = {s: 0.0 for s in STAGE_RULES}

    # 정규식 완료 패턴 우선 검사
    for pattern in COMPLETION_REGEX:
        if re.search(pattern, text, re.IGNORECASE):
            return 'COMPLETION', 0.95, {'regex_match': pattern}

    # 키워드 점수 집계
    for stage, rules in STAGE_RULES.items():
        weight = rules['weight']
        for lang in ('ko', 'en', 'vi'):
            for kw in rules[lang]:
                if kw.lower() in text_lower:
                    # 제목에 있으면 2배 가중치 (첫 100자 기준)
                    bonus = 2.0 if kw.lower() in text_lower[:100] else 1.0
                    scores[stage] += weight * bonus

    if not any(scores.values()):
        return 'UNKNOWN', 0.0, scores

    total = sum(scores.values())
    best_stage = max(scores, key=scores.get)
    best_score = scores[best_stage]
    confidence = min(best_score / (total + 1e-6) + (best_score / 10), 1.0)

    return best_stage, confidence, scores


# ══════════════════════════════════════════════════════════════════════════
#  2단계 — Claude Haiku 분석
# ══════════════════════════════════════════════════════════════════════════

def call_haiku_stage_analysis(
    article:    dict,
    plan_data:  dict,
    rule_stage: str,
    api_key:    str,
) -> dict:
    """
    2단계: Haiku가 기사 맥락과 마스터플랜 사업개요를 비교해
    진행단계·마일스톤·다음 관찰 포인트를 생성.

    Returns:
        {
          'stage':       str,    # PLANNING / TENDERING / CONSTRUCTION / COMPLETION / OPERATION
          'confidence':  float,
          'milestone':   str,    # 이번 기사가 시사하는 핵심 이정표
          'next_watch':  str,    # 다음 주 추적해야 할 사항
          'insight':     str,    # 1~2문장 투자자 시사점
          'haiku_used':  bool,
        }
    """
    plan_id = plan_data.get('plan_id', '')
    desc    = plan_data.get('description_ko', '')
    kpi_txt = '\n'.join(
        f"  - {k.get('indicator', '')}: 목표 {k.get('target_2030', '')} / 현황 {k.get('current', '')}"
        for k in plan_data.get('kpi_targets', [])
    ) or '  (KPI 정보 없음)'

    title   = article.get('title_ko') or article.get('title_en', '')
    summary = article.get('summary_ko') or article.get('summary_en', '')
    date    = article.get('date') or article.get('published_date', '')
    source  = article.get('source', '')

    system_prompt = f"""당신은 베트남 인프라 사업 진행현황 전문 분석가입니다.
아래 마스터플랜의 사업 개요와 KPI를 숙지한 후, 제공된 기사를 분석하여
사업 진행단계와 핵심 마일스톤을 판단하세요.

【마스터플랜: {plan_id}】
사업 개요:
{desc}

KPI 목표:
{kpi_txt}

진행단계 정의:
  PLANNING    — 계획·승인·법령 단계 (결정문, Decision, Resolution)
  TENDERING   — 입찰·계약·투자 유치 단계 (MOU, EPC 계약, FID)
  CONSTRUCTION — 착공·건설·시공 단계 (공사 중, 공정률)
  COMPLETION  — 준공·개통·상업운전 개시 단계
  OPERATION   — 운영·확장·성과 관리 단계

분석 원칙:
1. 기사 날짜({date})와 사업 타임라인을 고려하여 현실적 단계를 판단
2. 수치·기관명·날짜를 근거로 사용
3. 한국어로 간결하게 작성
"""

    user_prompt = f"""기사를 분석하여 아래 JSON 형식으로만 답변하세요. 다른 텍스트는 금지.

【기사】
날짜: {date}
출처: {source}
제목: {title}
요약: {summary}

규칙기반 1단계 예측: {rule_stage}

JSON 형식 (마크다운 없이 순수 JSON만):
{{
  "stage": "PLANNING|TENDERING|CONSTRUCTION|COMPLETION|OPERATION",
  "confidence": 0.0~1.0,
  "milestone": "이번 기사가 시사하는 핵심 이정표 (1문장, 30자 이내)",
  "next_watch": "다음 주 추적해야 할 사항 (1문장, 40자 이내)",
  "insight": "투자자·사업개발자를 위한 핵심 시사점 (1~2문장, 80자 이내)",
  "stage_reason": "판단 근거 키워드 (20자 이내)"
}}"""

    headers = {
        'Content-Type':      'application/json',
        'x-api-key':         api_key,
        'anthropic-version': '2023-06-01',
    }
    payload = {
        'model':      HAIKU_MODEL,
        'max_tokens': MAX_TOKENS_HAIKU,
        'system':     system_prompt,
        'messages':   [{'role': 'user', 'content': user_prompt}],
    }

    try:
        resp = requests.post(ANTHROPIC_API_URL, headers=headers,
                             json=payload, timeout=HAIKU_TIMEOUT)
        resp.raise_for_status()
        data = resp.json()

        raw_text = ''
        for block in data.get('content', []):
            if block.get('type') == 'text':
                raw_text = block['text'].strip()
                break

        # JSON 파싱
        json_match = re.search(r'\{.*\}', raw_text, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
            result['haiku_used'] = True
            return result

    except requests.exceptions.Timeout:
        log.warning(f"  Haiku 타임아웃 — 규칙 결과 사용: {title[:40]}")
    except json.JSONDecodeError as e:
        log.warning(f"  Haiku JSON 파싱 실패: {e}")
    except Exception as e:
        log.warning(f"  Haiku 오류: {e}")

    # 실패 시 규칙 결과 반환
    return {
        'stage':        rule_stage,
        'confidence':   0.5,
        'milestone':    '',
        'next_watch':   '',
        'insight':      '',
        'stage_reason': '규칙기반 (Haiku 실패)',
        'haiku_used':   False,
    }


# ══════════════════════════════════════════════════════════════════════════
#  데이터 로더
# ══════════════════════════════════════════════════════════════════════════

def load_knowledge_index() -> dict:
    """knowledge_index.json v2.3 또는 layer1_data.json 로드."""
    for path in KI_PATHS:
        if path.exists():
            with open(path, 'r', encoding='utf-8') as f:
                ki = json.load(f)
            # masterplans dict 구조 (v2.1+)
            if 'masterplans' in ki:
                plans = ki['masterplans']
                # plan_id 주입
                for pid, plan in plans.items():
                    plan['plan_id'] = pid
                log.info(f"knowledge_index 로드: {len(plans)}개 플랜 [{path.name}]")
                return plans
            # layer1_data.json (flat dict)
            else:
                for pid, plan in ki.items():
                    plan['plan_id'] = pid
                log.info(f"layer1_data 로드: {len(ki)}개 플랜 [{path.name}]")
                return ki

    log.warning("knowledge_index를 찾을 수 없음 — 빈 플랜 사용")
    return {}


def load_articles_from_excel(days_back: int = 7) -> list[dict]:
    """
    Excel DB에서 최근 N일 기사 로드.
    ctx_grade = HIGH 또는 ctx_plans가 있는 기사 우선 반환.
    """
    if not EXCEL_PATH.exists():
        log.warning(f"Excel DB 없음: {EXCEL_PATH}")
        return []

    wb  = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if 'News Database' not in wb.sheetnames:
        wb.close()
        return []

    ws      = wb['News Database']
    headers = [str(c.value or '').strip().lower().replace(' ', '_')
               for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def ci(keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    COL = {
        'date':       ci(['date']),
        'sector':     ci(['business_sector', 'sector']),
        'area':       ci(['area']),
        'province':   ci(['province']),
        'title_ko':   ci(['title_ko']),
        'title_en':   ci(['title_en', 'title', 'news_title']),
        'summary_ko': ci(['summary_ko']),
        'summary_en': ci(['summary_en', 'short_summary']),
        'source':     ci(['source']),
        'url':        ci(['link', 'url']),
        'ctx_grade':  ci(['ctx_grade']),
        'ctx_plans':  ci(['ctx_plans']),
    }

    cutoff   = (datetime.now() - timedelta(days=days_back)).strftime('%Y-%m-%d')
    articles = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # date 키 fallback
        date_val = str(row[COL['date']] if COL['date'] is not None else '').strip()[:10]
        if not date_val or date_val < cutoff:
            continue

        title_ko = str(row[COL['title_ko']] if COL['title_ko'] is not None else '').strip()
        if not title_ko:
            continue

        articles.append({
            'date':       date_val,
            'sector':     str(row[COL['sector']]    or ''),
            'area':       str(row[COL['area']]      or ''),
            'province':   str(row[COL['province']]  or ''),
            'title_ko':   title_ko,
            'title_en':   str(row[COL['title_en']]  or ''),
            'summary_ko': str(row[COL['summary_ko']] or ''),
            'summary_en': str(row[COL['summary_en']] or ''),
            # source fallback — Python or 연산자 사용 (빈 문자열 방지)
            'source':     (row[COL['source']] if COL['source'] is not None else None) or 'Unknown',
            'url':        str(row[COL['url']]       or ''),
            'ctx_grade':  str(row[COL['ctx_grade']] or 'MEDIUM'),
            'ctx_plans':  str(row[COL['ctx_plans']] or ''),
        })

    wb.close()
    articles.sort(key=lambda x: x['date'], reverse=True)
    log.info(f"Excel DB: {len(articles)}건 로드 (최근 {days_back}일)")
    return articles


def group_articles_by_plan(articles: list[dict], plans: dict) -> dict:
    """
    ctx_plans 컬럼 기준으로 기사를 플랜별로 그룹핑.
    ctx_plans 없으면 sector 기반 fallback.
    """
    SECTOR_TO_PLAN = {
        'Waste Water':           ['VN-WW-2030'],
        'Water Supply/Drainage': ['VN-WAT-URBAN', 'VN-WAT-RESOURCES'],
        'Solid Waste':           ['VN-SWM-NATIONAL-2030'],
        'Power':                 ['VN-PWR-PDP8-RENEWABLE', 'VN-PWR-PDP8-LNG', 'VN-PWR-PDP8-NUCLEAR'],
        'Oil & Gas':             ['VN-OG-2030'],
        'Industrial Parks':      ['VN-IP-NORTH-2030'],
        'Smart City':            ['VN-URB-METRO-2030', 'VN-HAN-URBAN-2045'],
        'Transport':             ['VN-TRAN-2055'],
    }

    grouped: dict[str, list] = {pid: [] for pid in plans}

    for art in articles:
        plan_ids = [p.strip() for p in art['ctx_plans'].split(',')
                    if p.strip() and p.strip() in plans]

        if not plan_ids:
            plan_ids = [p for p in SECTOR_TO_PLAN.get(art['sector'], [])
                        if p in plans]

        for pid in plan_ids:
            grouped.setdefault(pid, []).append(art)

    return grouped


# ══════════════════════════════════════════════════════════════════════════
#  메인 분류 엔진
# ══════════════════════════════════════════════════════════════════════════

def analyze_articles(
    articles:  list[dict],
    plans:     dict,
    api_key:   str,
    rule_only: bool = False,
) -> list[dict]:
    """
    전체 기사에 대해 2단계 분류 수행.

    1단계: 규칙기반 → confidence ≥ 0.65 → 확정
    2단계: Haiku 분석 → confidence < 0.65 또는 HIGH 기사

    Returns:
        분석 결과 목록 (각 기사에 stage, confidence, milestone 등 추가)
    """
    results = []
    haiku_count = 0
    rule_count  = 0

    # 플랜별 그룹핑
    grouped = group_articles_by_plan(articles, plans)

    for plan_id, plan_arts in grouped.items():
        if not plan_arts:
            continue

        plan_data = plans.get(plan_id, {})
        plan_data['plan_id'] = plan_id

        # HIGH 등급 기사 우선 + 최신 8건 제한
        plan_arts_sorted = sorted(
            plan_arts,
            key=lambda x: (0 if x['ctx_grade'] == 'HIGH' else 1, x['date']),
        )[:8]

        for art in plan_arts_sorted:
            # 분석 텍스트: 제목 + 요약 (ko 우선, 없으면 en)
            text = ' '.join(filter(None, [
                art.get('title_ko'),
                art.get('title_en'),
                art.get('summary_ko'),
                art.get('summary_en'),
            ]))

            # 1단계: 규칙기반
            rule_stage, confidence, score_breakdown = rule_based_classify(text)

            # Haiku 호출 조건:
            #   - rule_only=False
            #   - api_key 있음
            #   - (신뢰도 < 0.65) OR (HIGH 등급 기사)
            use_haiku = (
                not rule_only
                and bool(api_key)
                and (confidence < 0.65 or art['ctx_grade'] == 'HIGH')
            )

            if use_haiku:
                result = call_haiku_stage_analysis(art, plan_data, rule_stage, api_key)
                haiku_count += 1
                # Haiku rate limit 방지
                time.sleep(0.3)
            else:
                result = {
                    'stage':        rule_stage,
                    'confidence':   confidence,
                    'milestone':    _auto_milestone(rule_stage, text),
                    'next_watch':   _auto_next_watch(rule_stage, plan_id),
                    'insight':      '',
                    'stage_reason': f"규칙기반 (신뢰도 {confidence:.2f})",
                    'haiku_used':   False,
                }
                rule_count += 1

            results.append({
                **art,
                'plan_id':     plan_id,
                'rule_stage':  rule_stage,
                'rule_conf':   round(confidence, 3),
                'stage':       result.get('stage', rule_stage),
                'confidence':  round(result.get('confidence', confidence), 3),
                'milestone':   result.get('milestone', ''),
                'next_watch':  result.get('next_watch', ''),
                'insight':     result.get('insight', ''),
                'stage_reason': result.get('stage_reason', ''),
                'haiku_used':  result.get('haiku_used', False),
                'analyzed_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
            })

    log.info(f"분류 완료: 규칙기반={rule_count}건, Haiku={haiku_count}건")
    return results


def _auto_milestone(stage: str, text: str) -> str:
    """규칙기반 결과에서 간단한 마일스톤 텍스트 자동 생성."""
    MILESTONE_TEMPLATES = {
        'PLANNING':     '정책·법령 발표',
        'TENDERING':    '입찰·계약 진행',
        'CONSTRUCTION': '착공·건설 진행 중',
        'COMPLETION':   '준공·운영 개시',
        'OPERATION':    '운영 성과 업데이트',
        'UNKNOWN':      '단계 미확정',
    }
    return MILESTONE_TEMPLATES.get(stage, '')


def _auto_next_watch(stage: str, plan_id: str) -> str:
    """단계별 다음 추적 포인트 자동 생성."""
    NEXT_WATCH_TEMPLATES = {
        'PLANNING':     '입찰 공고 또는 EPC 계약 발표 여부 확인',
        'TENDERING':    '낙찰자 선정 및 계약 서명 일정 확인',
        'CONSTRUCTION': '공정률 업데이트 및 준공 예정일 변경 여부',
        'COMPLETION':   '상업운전 성과 지표 (처리량·발전량) 첫 보고',
        'OPERATION':    '분기 성과 및 확장 투자 계획 발표 여부',
        'UNKNOWN':      '추가 기사 수집 후 재분류 필요',
    }
    return NEXT_WATCH_TEMPLATES.get(stage, '')


# ══════════════════════════════════════════════════════════════════════════
#  Timeline 생성 (SA-8 MI 보고서 + 대시보드 활용)
# ══════════════════════════════════════════════════════════════════════════

def build_stage_timeline(results: list[dict], plans: dict) -> dict:
    """
    플랜별 진행단계 타임라인 생성.

    Returns:
        {
          plan_id: {
            'current_stage':   str,
            'stage_history':   [{date, stage, milestone, source}],
            'next_watch':      str,
            'latest_insight':  str,
            'stage_color':     str (hex),
            'kpi_targets':     list,
          }
        }
    """
    timeline = {}

    # 플랜별 그룹
    plan_results: dict[str, list] = {}
    for r in results:
        pid = r.get('plan_id', '')
        if pid:
            plan_results.setdefault(pid, []).append(r)

    for pid, arts in plan_results.items():
        if not arts:
            continue

        # 최신 기사 기준 현재 단계
        arts_sorted  = sorted(arts, key=lambda x: x['date'], reverse=True)
        latest       = arts_sorted[0]
        current_stage = latest.get('stage', 'UNKNOWN')
        stage_meta   = STAGE_META.get(current_stage, STAGE_META['UNKNOWN'])

        # 단계 히스토리 (중복 제거)
        history = []
        seen_milestones = set()
        for art in arts_sorted:
            m = art.get('milestone', '')
            if m and m not in seen_milestones:
                seen_milestones.add(m)
                history.append({
                    'date':      art['date'],
                    'stage':     art.get('stage', 'UNKNOWN'),
                    'milestone': m,
                    'source':    art.get('source', ''),
                    'title_ko':  art.get('title_ko', ''),
                })

        plan_data = plans.get(pid, {})
        timeline[pid] = {
            'plan_id':        pid,
            'title_ko':       plan_data.get('title_ko', pid),
            'sector':         plan_data.get('sector', ''),
            'area':           plan_data.get('area', ''),
            'current_stage':  current_stage,
            'stage_label_ko': stage_meta['label_ko'],
            'stage_color':    stage_meta['color'],
            'stage_order':    stage_meta['order'],
            'stage_history':  history[:10],         # 최근 10개 이정표
            'next_watch':     latest.get('next_watch', ''),
            'latest_insight': latest.get('insight', ''),
            'latest_article_date': latest['date'],
            'article_count':  len(arts),
            'haiku_count':    sum(1 for a in arts if a.get('haiku_used')),
            'kpi_targets':    plan_data.get('kpi_targets', []),
        }

    return timeline


# ══════════════════════════════════════════════════════════════════════════
#  출력 저장
# ══════════════════════════════════════════════════════════════════════════

def save_outputs(results: list[dict], timeline: dict) -> None:
    """context_output.json + stage_timeline.json 저장."""
    AGENT_OUT_DIR.mkdir(parents=True, exist_ok=True)

    # context_output.json
    output = {
        'generated_at':  datetime.now().strftime('%Y-%m-%d %H:%M'),
        'model':         HAIKU_MODEL,
        'total_articles': len(results),
        'haiku_used':    sum(1 for r in results if r.get('haiku_used')),
        'stage_summary': {
            stage: sum(1 for r in results if r.get('stage') == stage)
            for stage in STAGE_META
        },
        'articles': results,
    }
    with open(CONTEXT_OUT, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    log.info(f"저장: {CONTEXT_OUT} ({len(results)}건)")

    # stage_timeline.json
    timeline_out = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'plans':        timeline,
    }
    with open(TIMELINE_OUT, 'w', encoding='utf-8') as f:
        json.dump(timeline_out, f, ensure_ascii=False, indent=2)
    log.info(f"저장: {TIMELINE_OUT} ({len(timeline)}개 플랜)")


# ══════════════════════════════════════════════════════════════════════════
#  main()
# ══════════════════════════════════════════════════════════════════════════

def main():
    rule_only = '--rule-only' in sys.argv

    # days 파라미터 파싱
    days_back = 7
    for arg in sys.argv:
        if arg.startswith('--days='):
            try:
                days_back = int(arg.split('=')[1])
            except ValueError:
                pass

    api_key = os.getenv('ANTHROPIC_API_KEY', '')
    if not api_key and not rule_only:
        log.warning("ANTHROPIC_API_KEY 없음 → rule-only 모드로 전환")
        rule_only = True

    log.info("=" * 65)
    log.info(f"SA-7 Context Analyzer v1.0 — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    log.info(f"모드: {'규칙기반만' if rule_only else 'Haiku 2단계'} | 기간: {days_back}일")
    log.info("=" * 65)

    # Step 1: 데이터 로드
    plans    = load_knowledge_index()
    articles = load_articles_from_excel(days_back)

    if not articles:
        log.warning("분석할 기사 없음 — 종료")
        return

    # Step 2: 2단계 분류 실행
    results = analyze_articles(articles, plans, api_key, rule_only)

    # Step 3: 타임라인 생성
    timeline = build_stage_timeline(results, plans)

    # Step 4: 저장
    save_outputs(results, timeline)

    # Step 5: 요약 출력
    stage_counts = {}
    for r in results:
        s = r.get('stage', 'UNKNOWN')
        stage_counts[s] = stage_counts.get(s, 0) + 1

    log.info("")
    log.info("━" * 65)
    log.info(f"✅ SA-7 완료 — {len(results)}건 분석")
    for stage, count in sorted(stage_counts.items(),
                                key=lambda x: STAGE_META.get(x[0], STAGE_META['UNKNOWN'])['order']):
        label = STAGE_META.get(stage, STAGE_META['UNKNOWN'])['label_ko']
        log.info(f"   {label:<15}: {count:3d}건")
    log.info(f"   Haiku 호출: {sum(1 for r in results if r.get('haiku_used'))}건")
    log.info("━" * 65)


if __name__ == '__main__':
    main()
