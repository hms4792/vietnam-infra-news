"""
cleanup_warnings.py
===================
Excel DB에서 MYMEMORY WARNING 오염 데이터를 정리하는 1회성 스크립트.

처리 방식:
  케이스 1: News Title(원문)은 정상, title_ko/vi/summary_ko/vi만 WARNING
            → WARNING 값을 빈 문자열로 교체 (batch_translate가 나중에 재번역)
  케이스 2: News Title 자체가 WARNING
            → 해당 행 전체 삭제

실행 방법 (GitHub Actions):
  python3 scripts/cleanup_warnings.py

또는 로컬 (PowerShell):
  cd C:\\Users\\hms47\\vietnam-infra-news
  python scripts/cleanup_warnings.py
"""

import os
from pathlib import Path

EXCEL_PATH = os.environ.get(
    'EXCEL_PATH',
    'data/database/Vietnam_Infra_News_Database_Final.xlsx'
)


def is_warning(val) -> bool:
    """MYMEMORY 경고 메시지 여부 확인"""
    if not val:
        return False
    v = str(val).upper().strip()
    return (
        v.startswith('MYMEMORY WARNING') or
        v.startswith('PLEASE SELECT') or
        v.startswith('YOU USED ALL AVAILABLE') or
        'MYMEMORY WARNING' in v
    )


def main():
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl 미설치")
        return

    p = Path(EXCEL_PATH)
    if not p.exists():
        print(f"[ERROR] 파일 없음: {p}")
        return

    print(f"[정리] Excel 로드: {p}")
    wb = openpyxl.load_workbook(p)
    ws = wb['News Database']

    # 컬럼 위치 확인
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            col_map[h] = c

    # 필수 컬럼 확인
    required = ['News Title', 'title_ko', 'title_en', 'title_vi',
                'summary_ko', 'summary_en', 'summary_vi']
    for r in required:
        if r not in col_map:
            print(f"[WARN] 컬럼 없음: {r} — 헤더: {list(col_map.keys())}")

    c_title    = col_map.get('News Title', 4)
    c_tko      = col_map.get('title_ko', 9)
    c_ten      = col_map.get('title_en', 10)
    c_tvi      = col_map.get('title_vi', 11)
    c_sko      = col_map.get('summary_ko', 12)
    c_sen      = col_map.get('summary_en', 13)
    c_svi      = col_map.get('summary_vi', 14)

    # 1단계: 삭제할 행 찾기 (News Title 자체가 WARNING)
    rows_to_delete = []
    cleaned = 0

    for r in range(2, ws.max_row + 1):
        orig_title = ws.cell(r, c_title).value
        if is_warning(orig_title):
            rows_to_delete.append(r)

    # 역순으로 삭제 (행 번호 밀림 방지)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
        print(f"  [삭제] 행{r}: News Title 자체가 WARNING")

    print(f"[삭제 완료] {len(rows_to_delete)}행 삭제")

    # 2단계: title_ko/vi/summary 의 WARNING 값만 빈 문자열로 교체
    # (News Title 원문은 정상이므로 title_ko만 지움 → batch_translate가 재번역)
    warn_cols = [
        (c_tko,  'title_ko'),
        (c_tvi,  'title_vi'),
        (c_sko,  'summary_ko'),
        (c_svi,  'summary_vi'),
    ]
    # title_en, summary_en도 WARNING이면 원문(title/summary)으로 복구
    for r in range(2, ws.max_row + 1):
        orig_title   = str(ws.cell(r, c_title).value or '')
        orig_summary = str(ws.cell(r, col_map.get('Short Summary', 8)).value or '')

        # title_en 복구
        if is_warning(ws.cell(r, c_ten).value):
            ws.cell(r, c_ten).value = orig_title if not is_warning(orig_title) else ''
            cleaned += 1

        # summary_en 복구
        if is_warning(ws.cell(r, c_sen).value):
            ws.cell(r, c_sen).value = orig_summary if not is_warning(orig_summary) else ''
            cleaned += 1

        # 번역 컬럼 (ko/vi) — 빈 문자열로만 교체 (batch_translate가 재번역)
        for c_col, col_name in warn_cols:
            if is_warning(ws.cell(r, c_col).value):
                ws.cell(r, c_col).value = ''
                cleaned += 1

    print(f"[정리 완료] WARNING 값 {cleaned}개 → 빈 문자열로 교체")
    print(f"  → 교체된 title_ko/vi/summary_ko/vi는 batch_translate가 자동 재번역 예정")

    wb.save(p)
    print(f"[저장 완료] {p}")
    print()
    print(f"최종 결과:")
    print(f"  삭제된 행:    {len(rows_to_delete)}건")
    print(f"  정리된 셀:    {cleaned}개")
    print(f"  남은 기사:    {ws.max_row - 1}건")


if __name__ == '__main__':
    main()
