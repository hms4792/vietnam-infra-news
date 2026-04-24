"""
build_timeline.py  ── Stage Timeline Builder
=============================================
역할: SA-7 context_output.json → Excel "Stage Timeline" 시트 자동 누적 기록

핵심 개념:
  매번 실행할 때마다 새 기사의 진행단계 이정표를 Excel에 추가(누적)합니다.
  삭제 없이 계속 쌓이므로, 마스터플랜별 진행현황을 시간순으로 추적할 수 있습니다.

Excel 시트 구조 (Stage Timeline):
  A열: Plan ID        — VN-WW-2030
  B열: 날짜 (Date)    — 2026-04-24
  C열: 진행단계       — COMPLETION
  D열: 단계(한국어)   — 준공·개통 완료
  E열: 마일스톤       — 옌짜 WWTP 준공 완료
  F열: 다음 관찰 포인트 — 처리량 첫 성과 보고
  G열: Expert Insight — 북부 최대 WWTP 가동으로 하노이 처리율 50% 달성
  H열: 기사 제목      — 옌짜 WWTP 공식 준공 — 하노이 폐수처리율 향상
  I열: 출처 (Source)  — VietnamPlus
  J열: 신뢰도         — 0.95
  K열: AI 분석 여부   — True
  L열: 기록 시각      — 2026-04-24 20:05

실행:
  python3 scripts/build_timeline.py          # 표준 실행
  python3 scripts/build_timeline.py --reset  # 시트 초기화 후 재생성 (주의)

GitHub Actions 연동:
  daily_pipeline.yml — SA-7 다음 단계로 실행
  collect_weekly.yml — 주간 전체 기록

영구 제약:
  - Excel 파일: data/news_database.xlsx
  - 시트명: Stage Timeline (고정)
  - 헤더 행: 1행 (수정 금지)
  - 신규 행: 항상 마지막 행에 추가 (insert_rows 사용 안 함)
  - date 키: article.get('date') or article.get('published_date')

버전: v1.0 (2026-04-24)
"""

import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 설정 ──────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent.parent
DATA_DIR      = BASE_DIR / 'data'
AGENT_OUT_DIR = DATA_DIR / 'agent_output'

EXCEL_PATH    = DATA_DIR / 'news_database.xlsx'
CONTEXT_OUT   = AGENT_OUT_DIR / 'context_output.json'
TIMELINE_OUT  = AGENT_OUT_DIR / 'stage_timeline.json'

SHEET_NAME    = 'Stage Timeline'   # 고정 시트명

# ── 로깅 ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='[Timeline %(asctime)s] %(message)s',
    datefmt='%H:%M:%S'
)
log = logging.getLogger('Timeline')

# ── 진행단계 색상 (행 배경) ──────────────────────────────────────────────
STAGE_COLORS = {
    'PLANNING':     'E6F1FB',   # 파란색 계열
    'TENDERING':    'FFF9C4',   # 노란색 계열
    'CONSTRUCTION': 'FFE0B2',   # 주황색 계열
    'COMPLETION':   'E8F5E9',   # 초록색 계열
    'OPERATION':    'EDE7F6',   # 보라색 계열
    'UNKNOWN':      'F5F5F5',   # 회색
}

STAGE_LABELS = {
    'PLANNING':     '계획·승인',
    'TENDERING':    '입찰·계약',
    'CONSTRUCTION': '건설·시공',
    'COMPLETION':   '준공·개통',
    'OPERATION':    '운영·확장',
    'UNKNOWN':      '미확정',
}

# ── 헤더 정의 ─────────────────────────────────────────────────────────────
HEADERS = [
    ('Plan ID',         16),
    ('날짜',            12),
    ('진행단계',        12),
    ('단계(한국어)',    14),
    ('마일스톤',        30),
    ('다음 관찰 포인트', 28),
    ('Expert Insight',  40),
    ('기사 제목',       40),
    ('출처',            16),
    ('신뢰도',          10),
    ('AI 분석',         10),
    ('기록 시각',       18),
]


# ══════════════════════════════════════════════════════════════════════════
#  Excel 시트 초기화 / 로드
# ══════════════════════════════════════════════════════════════════════════
def get_or_create_sheet(wb: openpyxl.Workbook, reset: bool = False):
    """
    'Stage Timeline' 시트를 가져오거나 새로 만듭니다.

    reset=True 이면 기존 시트를 삭제하고 새로 만듭니다.
    주의: reset은 모든 누적 기록을 지웁니다.
    """
    if SHEET_NAME in wb.sheetnames:
        if reset:
            log.warning(f"  [주의] '{SHEET_NAME}' 시트 초기화 — 기존 데이터 삭제됨")
            del wb[SHEET_NAME]
        else:
            return wb[SHEET_NAME], False   # (시트, 신규여부)

    # 신규 시트 생성
    ws = wb.create_sheet(SHEET_NAME)
    _write_header(ws)
    log.info(f"  '{SHEET_NAME}' 시트 신규 생성")
    return ws, True


def _write_header(ws):
    """헤더 행 작성 (1행, 고정)."""
    hdr_fill   = PatternFill('solid', fgColor='0C2340')
    hdr_font   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side  = Side(style='thin', color='CCCCCC')
    hdr_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side, bottom=thin_side,
    )

    for col, (label, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_align
        cell.border = hdr_border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'   # 헤더 고정


# ══════════════════════════════════════════════════════════════════════════
#  중복 체크 — 같은 날짜 + 같은 플랜 + 같은 단계는 재기록하지 않음
# ══════════════════════════════════════════════════════════════════════════
def load_existing_keys(ws) -> set:
    """
    이미 기록된 (plan_id, date, stage) 조합을 읽어서 중복 방지.
    A열=plan_id, B열=date, C열=stage
    """
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        plan_id = str(row[0] or '').strip()
        date    = str(row[1] or '').strip()
        stage   = str(row[2] or '').strip()
        if plan_id and date and stage:
            keys.add((plan_id, date, stage))
    return keys


# ══════════════════════════════════════════════════════════════════════════
#  데이터 로드
# ══════════════════════════════════════════════════════════════════════════
def load_sa7_data() -> tuple[list[dict], dict]:
    """
    SA-7 출력 파일 로드.

    Returns:
        (articles, timeline_plans)
        articles: context_output.json의 기사 목록
        timeline_plans: stage_timeline.json의 플랜별 요약
    """
    articles = []
    timeline_plans = {}

    # context_output.json
    if CONTEXT_OUT.exists():
        with open(CONTEXT_OUT, 'r', encoding='utf-8') as f:
            ctx = json.load(f)
        articles = ctx.get('articles', [])
        log.info(f"  context_output.json: {len(articles)}건")
    else:
        log.warning(f"  context_output.json 없음: {CONTEXT_OUT}")

    # stage_timeline.json
    if TIMELINE_OUT.exists():
        with open(TIMELINE_OUT, 'r', encoding='utf-8') as f:
            tl = json.load(f)
        timeline_plans = tl.get('plans', {})
        log.info(f"  stage_timeline.json: {len(timeline_plans)}개 플랜")
    else:
        log.warning(f"  stage_timeline.json 없음: {TIMELINE_OUT}")

    return articles, timeline_plans


# ══════════════════════════════════════════════════════════════════════════
#  행 데이터 조립
# ══════════════════════════════════════════════════════════════════════════
def build_row_data(article: dict, timeline_plans: dict) -> dict | None:
    """
    기사 1건의 행 데이터를 조립합니다.

    stage_timeline.json의 플랜 요약에서 next_watch / insight를 보완합니다.
    """
    plan_id = article.get('plan_id', '').strip()
    if not plan_id:
        return None

    # date 키 fallback (영구 제약)
    date_val = (article.get('date') or article.get('published_date', '')).strip()[:10]
    if not date_val:
        date_val = datetime.now().strftime('%Y-%m-%d')

    stage      = article.get('stage', 'UNKNOWN').strip()
    milestone  = article.get('milestone', '').strip()
    next_watch = article.get('next_watch', '').strip()
    insight    = article.get('insight', '').strip()
    title_ko   = (article.get('title_ko') or article.get('title_en', '')).strip()
    source     = (article.get('source') or 'Unknown').strip()   # or 연산자
    confidence = round(float(article.get('confidence', 0.0)), 3)
    haiku_used = bool(article.get('haiku_used', False))

    # stage_timeline.json에서 보완 (Haiku가 생성 못한 경우)
    tl_plan = timeline_plans.get(plan_id, {})
    if not next_watch:
        next_watch = tl_plan.get('next_watch', '')
    if not insight:
        insight = tl_plan.get('latest_insight', '')

    # milestone 없으면 기사 제목 첫 40자로 대체
    if not milestone and title_ko:
        milestone = title_ko[:40]

    return {
        'plan_id':    plan_id,
        'date':       date_val,
        'stage':      stage,
        'stage_ko':   STAGE_LABELS.get(stage, '미확정'),
        'milestone':  milestone,
        'next_watch': next_watch,
        'insight':    insight,
        'title_ko':   title_ko[:80],
        'source':     source,
        'confidence': confidence,
        'haiku_used': '예' if haiku_used else '규칙',
        'recorded_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
    }


# ══════════════════════════════════════════════════════════════════════════
#  행 스타일링
# ══════════════════════════════════════════════════════════════════════════
def _style_row(ws, row_num: int, stage: str, is_new: bool):
    """행 배경색 + 폰트 설정."""
    bg_color   = STAGE_COLORS.get(stage, 'FFFFFF')
    fill       = PatternFill('solid', fgColor=bg_color)
    base_font  = Font(name='Arial', size=9)
    bold_font  = Font(name='Arial', size=9, bold=True)
    thin_side  = Side(style='thin', color='DDDDDD')
    border     = Border(left=thin_side, right=thin_side,
                        top=thin_side, bottom=thin_side)
    wrap_align = Alignment(wrap_text=True, vertical='top')

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = wrap_align
        # Plan ID(A)와 단계(C)는 굵게
        cell.font = bold_font if col in (1, 3) else base_font

    ws.row_dimensions[row_num].height = 40


# ══════════════════════════════════════════════════════════════════════════
#  플랜별 요약 행 (월간/주간 요약용)
# ══════════════════════════════════════════════════════════════════════════
def append_plan_summary_rows(ws, timeline_plans: dict, existing_keys: set) -> int:
    """
    stage_timeline.json의 stage_history에서
    아직 기록되지 않은 이정표를 시트에 추가합니다.

    이 방식으로 SA-7이 기사별로 생성하지 못한 플랜 수준의 요약도 기록됩니다.
    """
    added = 0
    for plan_id, plan in timeline_plans.items():
        for hist in plan.get('stage_history', []):
            date_val  = str(hist.get('date', '')).strip()[:10]
            stage     = hist.get('stage', 'UNKNOWN').strip()
            milestone = hist.get('milestone', '').strip()
            source    = (hist.get('source') or 'Timeline').strip()
            title_ko  = hist.get('title_ko', '')[:80]

            if not date_val:
                continue

            key = (plan_id, date_val, stage)
            if key in existing_keys:
                continue

            row_data = {
                'plan_id':    plan_id,
                'date':       date_val,
                'stage':      stage,
                'stage_ko':   STAGE_LABELS.get(stage, '미확정'),
                'milestone':  milestone or title_ko[:40],
                'next_watch': plan.get('next_watch', ''),
                'insight':    plan.get('latest_insight', ''),
                'title_ko':   title_ko,
                'source':     source,
                'confidence': 0.0,
                'haiku_used': '요약',
                'recorded_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
            }

            next_row = ws.max_row + 1
            _write_row(ws, next_row, row_data, stage)
            existing_keys.add(key)
            added += 1

    return added


def _write_row(ws, row_num: int, data: dict, stage: str):
    """딕셔너리 데이터를 시트 1행에 씁니다."""
    values = [
        data['plan_id'],
        data['date'],
        data['stage'],
        data['stage_ko'],
        data['milestone'],
        data['next_watch'],
        data['insight'],
        data['title_ko'],
        data['source'],
        data['confidence'],
        data['haiku_used'],
        data['recorded_at'],
    ]
    for col, val in enumerate(values, start=1):
        ws.cell(row=row_num, column=col, value=val)

    _style_row(ws, row_num, stage, is_new=True)


# ══════════════════════════════════════════════════════════════════════════
#  최신화 요약 시트 (Summary) 업데이트
# ══════════════════════════════════════════════════════════════════════════
def update_timeline_summary(wb: openpyxl.Workbook, timeline_plans: dict):
    """
    'Timeline Summary' 시트에 플랜별 현재 진행단계를 한눈에 볼 수 있게 정리.
    이 시트는 매번 덮어씁니다 (누적 X).
    """
    SUMMARY_SHEET = 'Timeline Summary'
    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]

    ws = wb.create_sheet(SUMMARY_SHEET)

    # 헤더
    summary_headers = [
        ('Plan ID', 18), ('플랜명', 36), ('섹터', 20),
        ('현재 진행단계', 16), ('단계(한국어)', 14),
        ('마지막 기사 날짜', 16), ('기사 수', 10),
        ('다음 관찰 포인트', 36), ('최근 인사이트', 40),
    ]
    hdr_fill  = PatternFill('solid', fgColor='0C2340')
    hdr_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    thin_side = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin_side, right=thin_side,
                       top=thin_side, bottom=thin_side)

    for col, (label, width) in enumerate(summary_headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'

    # 플랜 데이터 기록 (영역 순 정렬)
    AREA_ORDER = {'Environment': 0, 'Energy Develop.': 1, 'Urban Develop.': 2}
    sorted_plans = sorted(
        timeline_plans.items(),
        key=lambda x: (AREA_ORDER.get(x[1].get('area', ''), 9), x[0])
    )

    for row_num, (plan_id, plan) in enumerate(sorted_plans, start=2):
        stage    = plan.get('current_stage', 'UNKNOWN')
        bg_color = STAGE_COLORS.get(stage, 'FFFFFF')

        values = [
            plan_id,
            plan.get('title_ko', plan_id),
            plan.get('sector', ''),
            stage,
            STAGE_LABELS.get(stage, '미확정'),
            plan.get('latest_article_date', ''),
            plan.get('article_count', 0),
            plan.get('next_watch', ''),
            plan.get('latest_insight', ''),
        ]
        fill = PatternFill('solid', fgColor=bg_color)
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font      = Font(name='Arial', size=9,
                                  bold=(col in (1, 4)))

        ws.row_dimensions[row_num].height = 36

    log.info(f"  'Timeline Summary' 시트 업데이트: {len(sorted_plans)}개 플랜")


# ══════════════════════════════════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════════════════════════════════
def main():
    reset = '--reset' in sys.argv

    log.info("=" * 60)
    log.info(f"Stage Timeline Builder v1.0 — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    if reset:
        log.warning("  --reset 모드: 기존 Timeline 데이터 초기화")
    log.info("=" * 60)

    # Step 1: SA-7 데이터 로드
    log.info("Step 1: SA-7 데이터 로드")
    articles, timeline_plans = load_sa7_data()

    if not articles and not timeline_plans:
        log.warning("SA-7 데이터 없음 — build_timeline.py를 SA-7 이후에 실행하세요")
        return

    # Step 2: Excel 파일 로드
    log.info("Step 2: Excel DB 로드")
    if not EXCEL_PATH.exists():
        log.error(f"Excel DB 없음: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Step 3: Stage Timeline 시트 준비
    log.info(f"Step 3: '{SHEET_NAME}' 시트 준비")
    ws, is_new = get_or_create_sheet(wb, reset=reset)

    # 기존 키 로드 (중복 방지)
    existing_keys = set() if is_new else load_existing_keys(ws)
    log.info(f"  기존 기록: {len(existing_keys)}건")

    # Step 4: 기사별 진행단계 기록
    log.info("Step 4: 기사별 진행단계 기록")
    added_articles = 0
    for art in articles:
        row_data = build_row_data(art, timeline_plans)
        if not row_data:
            continue

        key = (row_data['plan_id'], row_data['date'], row_data['stage'])
        if key in existing_keys:
            continue   # 중복 건너뜀

        next_row = ws.max_row + 1
        _write_row(ws, next_row, row_data, row_data['stage'])
        existing_keys.add(key)
        added_articles += 1

    log.info(f"  기사 기록: {added_articles}건 추가")

    # Step 5: 플랜 요약 이정표 기록 (stage_history)
    log.info("Step 5: 플랜 이정표 기록 (stage_history)")
    added_summary = append_plan_summary_rows(ws, timeline_plans, existing_keys)
    log.info(f"  이정표 기록: {added_summary}건 추가")

    # Step 6: Timeline Summary 시트 업데이트
    log.info("Step 6: Timeline Summary 시트 업데이트")
    update_timeline_summary(wb, timeline_plans)

    # Step 7: 날짜 기준 정렬 (B열)
    log.info("Step 7: 날짜 기준 정렬")
    if ws.max_row > 2:
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # 플랜ID(A) 1차, 날짜(B) 2차 정렬
        data_rows.sort(key=lambda r: (str(r[0] or ''), str(r[1] or '')))

        # 정렬된 데이터 재기록
        for row_idx, row_vals in enumerate(data_rows, start=2):
            for col_idx, val in enumerate(row_vals, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            stage = str(row_vals[2] or 'UNKNOWN')
            _style_row(ws, row_idx, stage, is_new=False)

    # Step 8: 저장
    log.info("Step 8: Excel 저장")
    wb.save(EXCEL_PATH)

    total_added = added_articles + added_summary
    log.info("")
    log.info("━" * 60)
    log.info(f"✅ Timeline 빌드 완료")
    log.info(f"   신규 기록: {total_added}건 (기사 {added_articles} + 이정표 {added_summary})")
    log.info(f"   전체 Timeline 행: {ws.max_row - 1}건")
    log.info(f"   Excel: {EXCEL_PATH}")
    log.info("━" * 60)


if __name__ == '__main__':
    main()
