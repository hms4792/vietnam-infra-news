# -*- coding: utf-8 -*-
"""
specialist_crawler.py v5.0
전문미디어 직접 크롤러 — RSS 차단 사이트 우회 수집

변경사항 (v5.0):
  - Jina.ai Reader API fallback 추가 (403/429/차단 시 자동 전환)
    사용법: r.jina.ai/{url} → 무료, API키 불필요
  - fetch_page() 3단계 fallback 구조:
      1단계: requests + Custom User-Agent (기존 방식)
      2단계: User-Agent 교체 후 재시도 (3초 sleep)
      3단계: Jina.ai Reader API 우회 (신규)
  - 폐쇄 소스 목록 명시 (절대 재추가 금지)

대상 사이트:
  1. theinvestor.vn — 카테고리 HTML 크롤링
  2. vir.com.vn    — 카테고리 HTML 크롤링
  (hanoitimes.vn은 RSS 정상 작동 → news_collector.py 담당)

실행:
  python3 scripts/specialist_crawler.py
  python3 scripts/specialist_crawler.py --from-date 2025-01-01
  python3 scripts/specialist_crawler.py --dry-run
"""

import os
import re
import sys
import time
import sqlite3
import hashlib
import logging
import argparse
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# BeautifulSoup (bs4)
try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    import requests as req_lib
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ── 경로 설정 ─────────────────────────────────────────────
SCRIPTS_DIR = Path(__file__).parent
ROOT_DIR    = SCRIPTS_DIR.parent
DB_PATH     = str(ROOT_DIR / 'data' / 'vietnam_infra_news.db')
EXCEL_PATH  = str(ROOT_DIR / 'data' / 'database' / 'Vietnam_Infra_News_Database_Final.xlsx')

# ── 로깅 ──────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
)
logger = logging.getLogger(__name__)

def log(msg): logger.info(msg)

# ── 크롤링 설정 ───────────────────────────────────────────
DELAY     = 2.0   # 요청 간 대기(초)
MAX_PAGES = 5     # 카테고리당 최대 페이지
TIMEOUT   = 20    # 요청 타임아웃(초)
MIN_TITLE = 15    # 제목 최소 글자수

# ── 영구 폐쇄 소스 (절대 재추가 금지) ─────────────────────
PERMANENTLY_CLOSED = [
    'theinvestor.vn/feed',       # 404 영구폐쇄
    'vir.com.vn/rss',            # 410 Gone
    'constructionvietnam.net',   # 폐쇄
    'monre.gov.vn/rss',          # 봇 차단
    'vea.gov.vn',                # 봇 차단
    'mic.gov.vn/rss',            # 봇 차단
    'smartcity.mobi',            # 폐쇄
    'ictvietnam.vn/feed',        # 봇 차단
    'baotintuc.vn',              # 봇 차단
    'kinhtemoitruong.vn',        # 봇 차단
    'hanoimoi.vn',               # 봇 차단
    'moitruong.com.vn/feed',     # 봇 차단
]

# ── 공통 헤더 ────────────────────────────────────────────
HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    ),
    'Accept'         : 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9,vi;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection'     : 'keep-alive',
    'Cache-Control'  : 'no-cache',
}

# The Investor 전용 헤더
INVESTOR_HEADERS = {
    **HEADERS,
    'Referer': 'https://theinvestor.vn/',
    'Origin' : 'https://theinvestor.vn',
}

# VIR 전용 헤더
VIR_HEADERS = {
    **HEADERS,
    'Referer': 'https://vir.com.vn/',
    'Origin' : 'https://vir.com.vn',
}

# Mac User-Agent (재시도용)
UA_MAC = (
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
    'AppleWebKit/605.1.15 (KHTML, like Gecko) '
    'Version/17.0 Safari/605.1.15'
)

# ── theinvestor.vn 카테고리 ───────────────────────────────
INVESTOR_CATEGORIES = [
    'https://theinvestor.vn/infrastructure/',
    'https://theinvestor.vn/energy/',
    'https://theinvestor.vn/industrial-real-estate/',
    'https://theinvestor.vn/environment/',
    'https://theinvestor.vn/smart-city/',
    'https://theinvestor.vn/transport/',
]

# theinvestor.vn 기사 URL 패턴 (-d숫자.html)
INVESTOR_ARTICLE_RE = re.compile(
    r'https://theinvestor\.vn/[a-z0-9\-]+-d\d+\.html'
)

# ── vir.com.vn 카테고리 ──────────────────────────────────
VIR_CATEGORIES = [
    'https://vir.com.vn/infrastructure.html',
    'https://vir.com.vn/energy.html',
    'https://vir.com.vn/industrial-zones.html',
    'https://vir.com.vn/environment.html',
    'https://vir.com.vn/smart-cities.html',
]

# VIR 기사 URL 패턴 (-숫자.html)
VIR_ARTICLE_RE = re.compile(
    r'https://vir\.com\.vn/[a-z0-9\-]+-\d+\.html'
)


# ══════════════════════════════════════════════════════════
# 핵심: fetch_page — 3단계 fallback (requests → retry → Jina.ai)
# ══════════════════════════════════════════════════════════
def fetch_page(url: str, headers: dict = None,
               use_jina_fallback: bool = True) -> Optional[str]:
    """
    웹 페이지 HTML 가져오기 — 3단계 fallback 구조

    1단계: requests + Custom User-Agent (기존 방식)
    2단계: User-Agent를 Mac Safari로 교체 후 3초 sleep 재시도
    3단계: Jina.ai Reader API 우회 (r.jina.ai/{url})
            → 봇 차단/403/429 사이트 우회 가능
            → 무료, API키 불필요
            → 단, 동적 JS 렌더링은 제한적

    Returns:
        HTML 문자열 또는 None (3단계 모두 실패 시)
    """
    hdrs = headers or HEADERS

    # ── 1단계: 기본 requests ──────────────────────────────
    if HAS_REQUESTS:
        try:
            session = req_lib.Session()
            r = session.get(url, headers=hdrs, timeout=TIMEOUT, allow_redirects=True)
            if r.status_code == 200 and len(r.text) > 500:
                log(f"  [fetch] 1단계 성공: {url[:60]}")
                return r.text
            log(f"  [fetch] 1단계 HTTP {r.status_code}: {url[:60]}")
        except Exception as e:
            log(f"  [fetch] 1단계 오류: {e}")

        # ── 2단계: User-Agent 교체 후 재시도 ──────────────
        time.sleep(3)
        try:
            hdrs2 = {**hdrs, 'User-Agent': UA_MAC}
            r2 = req_lib.get(url, headers=hdrs2, timeout=TIMEOUT, allow_redirects=True)
            if r2.status_code == 200 and len(r2.text) > 500:
                log(f"  [fetch] 2단계 성공 (Mac UA): {url[:60]}")
                return r2.text
            log(f"  [fetch] 2단계 HTTP {r2.status_code}: {url[:60]}")
        except Exception as e:
            log(f"  [fetch] 2단계 오류: {e}")
    else:
        # requests 없으면 urllib 사용
        try:
            req = urllib.request.Request(url, headers=hdrs)
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                if resp.status == 200:
                    return resp.read().decode('utf-8', errors='ignore')
        except Exception as e:
            log(f"  [fetch] urllib 오류: {e}")

    # ── 3단계: Jina.ai Reader API fallback ────────────────
    # r.jina.ai/{url} 형식으로 요청 → 봇 차단 우회
    # 무료 플랜: 분당 20회 / API키 없이 사용 가능
    # 유료 플랜(Jina.ai 계정): 더 빠른 속도, 더 많은 요청
    if use_jina_fallback:
        time.sleep(2)
        try:
            jina_url = f"https://r.jina.ai/{url}"
            jina_headers = {
                'User-Agent' : 'Mozilla/5.0',
                'Accept'     : 'text/html,text/plain',
                # API키 있으면 아래 줄 활성화:
                # 'Authorization': f'Bearer {os.environ.get("JINA_API_KEY","")}',
                'X-Return-Format': 'html',    # HTML 형식 요청
                'X-No-Cache' : 'true',
            }
            if HAS_REQUESTS:
                rj = req_lib.get(jina_url, headers=jina_headers, timeout=30)
                if rj.status_code == 200 and len(rj.text) > 200:
                    log(f"  [fetch] 3단계 Jina.ai 성공: {url[:60]}")
                    return rj.text
                log(f"  [fetch] 3단계 Jina.ai HTTP {rj.status_code}")
            else:
                req_j = urllib.request.Request(jina_url, headers=jina_headers)
                with urllib.request.urlopen(req_j, timeout=30) as rj:
                    if rj.status == 200:
                        content = rj.read().decode('utf-8', errors='ignore')
                        if len(content) > 200:
                            log(f"  [fetch] 3단계 Jina.ai 성공: {url[:60]}")
                            return content
        except Exception as e:
            log(f"  [fetch] 3단계 Jina.ai 오류: {e}")

    log(f"  [fetch] 3단계 모두 실패: {url[:60]}")
    return None


def jina_fetch_summary(url: str, max_chars: int = 600) -> Optional[str]:
    """
    Jina.ai로 페이지 본문 텍스트 추출 (요약용)
    X-Return-Format: text → 마크다운 텍스트로 반환

    기사 상세 페이지의 본문 요약 추출에 사용
    """
    try:
        jina_url = f"https://r.jina.ai/{url}"
        headers = {
            'User-Agent'     : 'Mozilla/5.0',
            'Accept'         : 'text/plain',
            'X-Return-Format': 'text',    # 텍스트 형식 (마크다운)
            'X-No-Cache'     : 'true',
        }
        # JINA_API_KEY 환경변수 있으면 사용 (무료 플랜보다 빠름)
        jina_key = os.environ.get('JINA_API_KEY', '')
        if jina_key:
            headers['Authorization'] = f'Bearer {jina_key}'

        if HAS_REQUESTS:
            r = req_lib.get(jina_url, headers=headers, timeout=25)
            if r.status_code == 200:
                text = r.text[:max_chars].strip()
                # 불필요한 헤더/URL 라인 제거
                lines = [l for l in text.splitlines()
                         if l.strip() and not l.startswith('http') and len(l) > 10]
                return ' '.join(lines)[:max_chars]
        else:
            req_obj = urllib.request.Request(jina_url, headers=headers)
            with urllib.request.urlopen(req_obj, timeout=25) as r:
                content = r.read().decode('utf-8', errors='ignore')[:max_chars]
                return content.strip()
    except Exception as e:
        log(f"  [jina_summary] 오류: {e}")
    return None


# ══════════════════════════════════════════════════════════
# 유틸리티
# ══════════════════════════════════════════════════════════
def generate_url_hash(url: str) -> str:
    return hashlib.md5(url.encode()).hexdigest()


def parse_date_str(date_str: str) -> Optional[datetime]:
    """다양한 날짜 형식 파싱"""
    if not date_str:
        return None
    fmts = [
        '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d', '%d/%m/%Y', '%B %d, %Y',
        '%d %B %Y', '%b %d, %Y',
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(date_str.strip()[:len(fmt)+2], fmt)
        except ValueError:
            continue
    return None


def load_existing_urls() -> set:
    """SQLite + Excel에서 기존 URL 해시 로드"""
    hashes = set()
    # SQLite
    if os.path.exists(DB_PATH):
        try:
            conn = sqlite3.connect(DB_PATH)
            rows = conn.execute('SELECT url_hash FROM articles').fetchall()
            hashes.update(r[0] for r in rows)
            conn.close()
            log(f"  기존 URL {len(hashes)}개 로드 (SQLite)")
        except Exception as e:
            log(f"  SQLite 로드 오류: {e}")
    # Excel
    if os.path.exists(EXCEL_PATH):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(EXCEL_PATH, read_only=True)
            if 'News Database' in wb.sheetnames:
                ws = wb['News Database']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    url = str(row[9] or '')   # J열: URL
                    if url.startswith('http'):
                        hashes.add(generate_url_hash(url))
            wb.close()
        except Exception as e:
            log(f"  Excel 로드 오류: {e}")
    return hashes


# ══════════════════════════════════════════════════════════
# theinvestor.vn 크롤러
# ══════════════════════════════════════════════════════════
def _get_investor_article_links(cat_url: str, page: int = 1) -> list[str]:
    """
    theinvestor.vn 카테고리 페이지에서 기사 URL 목록 추출
    URL 패턴: /기사제목-d숫자.html
    """
    # 페이지네이션: ?page=N 또는 /page/N
    if page > 1:
        url = f"{cat_url.rstrip('/')}?page={page}"
    else:
        url = cat_url

    html = fetch_page(url, INVESTOR_HEADERS)
    if not html:
        return []

    if not HAS_BS4:
        # 정규식 fallback
        return list(set(INVESTOR_ARTICLE_RE.findall(html)))

    soup = BeautifulSoup(html, 'html.parser')
    links = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        if not href.startswith('http'):
            href = 'https://theinvestor.vn' + href
        if INVESTOR_ARTICLE_RE.match(href):
            links.append(href)
    return list(set(links))


def _get_investor_article_info(article_url: str) -> tuple[str, str, str]:
    """
    theinvestor.vn 기사 상세 페이지에서 (제목, 날짜, 요약) 추출
    Jina.ai로 본문 요약도 추출 가능
    """
    html = fetch_page(article_url, INVESTOR_HEADERS)
    if not html:
        return '', '', ''

    title = pub_date = summary = ''

    if HAS_BS4:
        soup = BeautifulSoup(html, 'html.parser')
        # 제목
        for sel in ['h1.article-title', 'h1.post-title', 'h1', 'title']:
            el = soup.select_one(sel)
            if el and len(el.get_text(strip=True)) > MIN_TITLE:
                title = el.get_text(strip=True)
                break
        # 날짜
        for sel in ['time[datetime]', '.article-date', '.post-date', 'time']:
            el = soup.select_one(sel)
            if el:
                pub_date = el.get('datetime', '') or el.get_text(strip=True)
                break
        # 요약 (리드 문단)
        for sel in ['.article-lead', '.article-sapo', '.lead', 'p.intro']:
            el = soup.select_one(sel)
            if el and len(el.get_text(strip=True)) > 30:
                summary = el.get_text(strip=True)[:500]
                break
        if not summary:
            # 첫 번째 p 태그
            p = soup.find('p')
            if p:
                summary = p.get_text(strip=True)[:500]
    else:
        # 정규식 fallback
        m_title = re.search(r'<h1[^>]*>([^<]+)</h1>', html)
        if m_title:
            title = re.sub(r'<[^>]+>', '', m_title.group(1)).strip()
        m_date = re.search(r'datetime="([^"]+)"', html)
        if m_date:
            pub_date = m_date.group(1)

    # 요약이 짧으면 Jina.ai로 보완 (옵션)
    if len(summary) < 50:
        jina_text = jina_fetch_summary(article_url, max_chars=500)
        if jina_text and len(jina_text) > len(summary):
            summary = jina_text

    return title, pub_date, summary[:500]


def scrape_theinvestor(from_date_dt: datetime,
                       existing_urls: set,
                       dry_run: bool = False) -> list[dict]:
    """
    theinvestor.vn 크롤러 메인
    fetch_page 3단계 fallback(Jina.ai 포함) 사용
    """
    log("[1/2] theinvestor.vn 크롤링 시작")
    articles = []
    seen_urls = set(existing_urls)

    for cat_url in INVESTOR_CATEGORIES:
        log(f"  카테고리: {cat_url}")
        cat_arts = 0
        stop = False

        for page in range(1, MAX_PAGES + 1):
            if stop:
                break
            links = _get_investor_article_links(cat_url, page)
            if not links:
                log(f"    page {page}: 링크 없음 → 중단")
                break

            page_arts = 0
            for art_url in links:
                url_hash = generate_url_hash(art_url)
                if url_hash in seen_urls:
                    continue

                time.sleep(DELAY)
                title, pub_date_raw, summary = _get_investor_article_info(art_url)

                if not title or len(title) < MIN_TITLE:
                    continue

                # 날짜 파싱
                pub_dt = parse_date_str(pub_date_raw)
                if pub_dt:
                    if pub_dt < from_date_dt:
                        stop = True
                        continue
                    date_str = pub_dt.strftime('%Y-%m-%d')
                else:
                    date_str = datetime.now().strftime('%Y-%m-%d')

                seen_urls.add(url_hash)
                articles.append({
                    'title'    : title,
                    'url'      : art_url,
                    'date'     : date_str,
                    'source'   : 'The Investor',
                    'src_type' : 'Specialist Crawler',
                    'summary'  : summary,
                    'province' : '',
                    'sector'   : '',
                })
                page_arts += 1
                if not dry_run:
                    log(f"    [{date_str}] {title[:55]}")

            log(f"    page {page}: {len(links)}링크 → {page_arts}건")
            if page_arts == 0:
                break
            cat_arts += page_arts
            time.sleep(DELAY)

        log(f"  → {cat_url.split('/')[-2]}: {cat_arts}건")

    log(f"theinvestor.vn 총 {len(articles)}건")
    return articles


# ══════════════════════════════════════════════════════════
# vir.com.vn 크롤러
# ══════════════════════════════════════════════════════════
def _get_vir_article_links(cat_url: str, page: int = 1) -> list[str]:
    """
    vir.com.vn 카테고리 페이지에서 기사 URL 목록 추출
    URL 패턴: /기사제목-숫자.html
    """
    if page > 1:
        base = cat_url.replace('.html', '')
        url  = f"{base}-p{page}.html"
    else:
        url = cat_url

    html = fetch_page(url, VIR_HEADERS)
    if not html:
        return []

    if not HAS_BS4:
        return list(set(VIR_ARTICLE_RE.findall(html)))

    soup = BeautifulSoup(html, 'html.parser')
    links = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        if not href.startswith('http'):
            href = 'https://vir.com.vn' + href
        if VIR_ARTICLE_RE.match(href) and href not in VIR_CATEGORIES:
            links.append(href)
    return list(set(links))


def _get_vir_article_info(article_url: str) -> tuple[str, str, str]:
    """vir.com.vn 기사 상세 페이지 정보 추출"""
    html = fetch_page(article_url, VIR_HEADERS)
    if not html:
        # VIR는 특히 봇 차단이 강함 → Jina.ai 우선 시도
        log(f"  VIR 직접 접근 실패 → Jina.ai 시도: {article_url[:60]}")
        html = fetch_page(article_url, VIR_HEADERS, use_jina_fallback=True)
    if not html:
        return '', '', ''

    title = pub_date = summary = ''

    if HAS_BS4:
        soup = BeautifulSoup(html, 'html.parser')
        for sel in ['h1.article__title', 'h1.title', 'h1']:
            el = soup.select_one(sel)
            if el and len(el.get_text(strip=True)) > MIN_TITLE:
                title = el.get_text(strip=True)
                break
        for sel in ['time[datetime]', '.article__date', '.date', 'time']:
            el = soup.select_one(sel)
            if el:
                pub_date = el.get('datetime', '') or el.get_text(strip=True)
                break
        for sel in ['.article__sapo', '.sapo', '.article__lead', 'p.lead']:
            el = soup.select_one(sel)
            if el and len(el.get_text(strip=True)) > 30:
                summary = el.get_text(strip=True)[:500]
                break
        if not summary:
            p = soup.find('p')
            if p:
                summary = p.get_text(strip=True)[:500]
    else:
        m_title = re.search(r'<h1[^>]*>([^<]+)</h1>', html)
        if m_title:
            title = re.sub(r'<[^>]+>', '', m_title.group(1)).strip()
        m_date = re.search(r'datetime="([^"]+)"', html)
        if m_date:
            pub_date = m_date.group(1)

    if len(summary) < 50:
        jina_text = jina_fetch_summary(article_url, max_chars=500)
        if jina_text and len(jina_text) > len(summary):
            summary = jina_text

    return title, pub_date, summary[:500]


def scrape_vir(from_date_dt: datetime,
               existing_urls: set,
               dry_run: bool = False) -> list[dict]:
    """vir.com.vn 크롤러 메인"""
    log("[2/2] vir.com.vn 크롤링 시작")
    articles = []
    seen_urls = set(existing_urls)

    for cat_url in VIR_CATEGORIES:
        log(f"  카테고리: {cat_url}")
        cat_arts = 0
        stop = False

        for page in range(1, MAX_PAGES + 1):
            if stop:
                break
            links = _get_vir_article_links(cat_url, page)
            if not links:
                log(f"    page {page}: 링크 없음 → 중단")
                break

            page_arts = 0
            for art_url in links:
                url_hash = generate_url_hash(art_url)
                if url_hash in seen_urls:
                    continue

                time.sleep(DELAY)
                title, pub_date_raw, summary = _get_vir_article_info(art_url)

                if not title or len(title) < MIN_TITLE:
                    continue

                pub_dt = parse_date_str(pub_date_raw)
                if pub_dt:
                    if pub_dt < from_date_dt:
                        stop = True
                        continue
                    date_str = pub_dt.strftime('%Y-%m-%d')
                else:
                    date_str = datetime.now().strftime('%Y-%m-%d')

                seen_urls.add(url_hash)
                articles.append({
                    'title'    : title,
                    'url'      : art_url,
                    'date'     : date_str,
                    'source'   : 'Vietnam Investment Review',
                    'src_type' : 'Specialist Crawler',
                    'summary'  : summary,
                    'province' : '',
                    'sector'   : '',
                })
                page_arts += 1
                if not dry_run:
                    log(f"    [{date_str}] {title[:55]}")

            log(f"    page {page}: {len(links)}링크 → {page_arts}건")
            if page_arts == 0:
                break
            cat_arts += page_arts
            time.sleep(DELAY)

        log(f"  → {cat_url.split('/')[-1].replace('.html','')}: {cat_arts}건")

    log(f"vir.com.vn 총 {len(articles)}건")
    return articles


# ══════════════════════════════════════════════════════════
# 저장 — SQLite + Excel
# ══════════════════════════════════════════════════════════
def classify_and_save(articles: list[dict], dry_run: bool = False) -> int:
    """
    수집된 기사를 섹터 분류 후 SQLite + Excel에 저장

    섹터 분류는 news_collector.py의 classify_sector 로직과 동일하게
    키워드 기반으로 처리
    """
    if not articles:
        return 0

    # 섹터 키워드 (news_collector.py와 동일 기준)
    SECTOR_KW = {
        'Waste Water'           : ['wastewater','sewage','wwtp','nước thải','thoát nước'],
        'Water Supply/Drainage' : ['water supply','clean water','cấp nước','nước sạch','drainage'],
        'Solid Waste'           : ['solid waste','waste management','rác thải','landfill','recycling'],
        'Power'                 : ['electricity','power','renewable','solar','wind','pdp8','evn','điện'],
        'Oil & Gas'             : ['lng','petroleum','petrovietnam','oil','gas','dầu khí'],
        'Industrial Parks'      : ['industrial park','industrial zone','khu công nghiệp','vsip','fdi'],
        'Smart City'            : ['smart city','metro','digital','thành phố thông minh','đường sắt'],
        'Transport'             : ['expressway','highway','airport','port','cao tốc','cảng'],
        'Environment'           : ['environment','emission','carbon','môi trường','phát thải'],
    }

    def classify(title: str, summary: str = '') -> str:
        text = (title + ' ' + summary).lower()
        for sector, kws in SECTOR_KW.items():
            if any(kw in text for kw in kws):
                return sector
        return 'Environment'

    saved = 0

    if dry_run:
        for a in articles:
            sector = classify(a['title'], a.get('summary', ''))
            log(f"  [DRY-RUN] [{sector}] {a['title'][:60]}")
        return len(articles)

    # SQLite 저장
    if os.path.exists(DB_PATH):
        try:
            conn = sqlite3.connect(DB_PATH)
            for a in articles:
                sector = classify(a['title'], a.get('summary', ''))
                a['sector'] = sector
                url_hash = generate_url_hash(a['url'])
                try:
                    conn.execute(
                        '''INSERT OR IGNORE INTO articles
                           (url_hash, title, url, date, source, sector, summary)
                           VALUES (?,?,?,?,?,?,?)''',
                        (url_hash, a['title'], a['url'], a['date'],
                         a['source'], sector, a.get('summary', ''))
                    )
                    saved += 1
                except sqlite3.Error as e:
                    log(f"  DB 저장 오류: {e}")
            conn.commit()
            conn.close()
            log(f"SQLite 저장: {saved}건")
        except Exception as e:
            log(f"SQLite 오류: {e}")

    # Excel 저장 (ExcelUpdater 사용)
    if os.path.exists(EXCEL_PATH) and articles:
        try:
            sys.path.insert(0, str(SCRIPTS_DIR))
            from excel_updater import ExcelUpdater

            for a in articles:
                if not a.get('sector'):
                    a['sector'] = classify(a['title'], a.get('summary', ''))
                a['src_type']       = 'Specialist Crawler'
                a['title_en']       = a.get('title', '')
                a['title_ko']       = ''
                a['title_vi']       = ''
                a['sum_ko']         = ''
                a['sum_en']         = a.get('summary', '')
                a['sum_vi']         = ''
                a['plan']           = ''
                a['published_date'] = a.get('date', '')

            updater = ExcelUpdater(EXCEL_PATH)
            stats   = updater.update_all(articles)
            log(f"Excel 저장 완료: {stats}")
        except ImportError:
            log("  ExcelUpdater 없음 → Excel 저장 건너뜀")
        except Exception as e:
            log(f"  Excel 저장 오류: {e}")

    return saved


# ══════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description='전문미디어 크롤러 v5.0 (Jina.ai fallback)')
    parser.add_argument('--from-date', default='2025-01-01',
                        help='수집 시작 날짜 (기본: 2025-01-01, 형식: YYYY-MM-DD)')
    parser.add_argument('--dry-run', action='store_true',
                        help='실제 저장 없이 수집만 테스트')
    parser.add_argument('--no-jina', action='store_true',
                        help='Jina.ai fallback 비활성화')
    args = parser.parse_args()

    # Jina.ai fallback 전역 설정
    if args.no_jina:
        # fetch_page의 use_jina_fallback 기본값을 False로 재정의
        import functools
        global fetch_page
        _orig = fetch_page
        fetch_page = functools.partial(_orig, use_jina_fallback=False)
        log("Jina.ai fallback 비활성화")

    from_date_str = args.from_date
    try:
        from_date_dt = datetime.strptime(from_date_str, '%Y-%m-%d')
    except ValueError:
        log(f"날짜 형식 오류: {from_date_str} → YYYY-MM-DD")
        sys.exit(1)

    log('=' * 60)
    log(f"전문미디어 크롤러 v5.0 | from_date={from_date_str} | dry_run={args.dry_run}")
    log(f"Jina.ai fallback: {'비활성' if args.no_jina else '활성 (403/429 자동 우회)'}")
    log('=' * 60)

    # 기존 URL 로드
    existing = load_existing_urls()
    log(f"기존 URL {len(existing)}개 로드 완료")

    all_articles = []

    # theinvestor.vn
    try:
        inv_arts = scrape_theinvestor(from_date_dt, existing, args.dry_run)
        all_articles.extend(inv_arts)
        existing.update(generate_url_hash(a['url']) for a in inv_arts)
    except Exception as e:
        log(f"theinvestor.vn 오류: {e}")

    # vir.com.vn
    try:
        vir_arts = scrape_vir(from_date_dt, existing, args.dry_run)
        all_articles.extend(vir_arts)
    except Exception as e:
        log(f"vir.com.vn 오류: {e}")

    log(f"\n크롤링 원시 수집: {len(all_articles)}건")

    # 저장
    if all_articles:
        saved = classify_and_save(all_articles, args.dry_run)
        log(f"최종: {saved}건 처리")
    else:
        log("수집 기사 없음 — 종료")

    log('=' * 60)


if __name__ == '__main__':
    main()
