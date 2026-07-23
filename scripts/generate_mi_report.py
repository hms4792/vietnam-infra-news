#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_mi_report.py  ── SA-8 보고서 생성기 v3.3
===================================================
역할: knowledge_index.json + Excel DB → PPT + Word 보고서 생성 + 이메일 발송

v3.1 버그픽스 (2026-05-10):
  [Fix 1] ★★★ 날짜 필터 버그 수정
          - collector_output.json의 'published_date' 키 인식 추가
          - datetime 파싱 실패 시 경고 출력 (조용한 스킵 제거)
          - 파싱 실패 기사도 날짜 없음으로 포함 처리

  [Fix 2] ★★ Excel DB 직접 읽기로 기사 소스 변경
          - SA-6/SA-7과 동일하게 Excel Matched_Plan 시트 읽기
          - collector_output.json(12건)이 아닌 DB(1097건) 사용

  [Fix 3] ★★ Word docx → docs/reports/ 복사 로직 추가
          - yml 의존 없이 Python에서 직접 복사

  [Fix 4] ★ 이메일 발송 조건 변경
          - 기존: KPI변동 OR 신규기사 있을 때만 발송
          - 변경: 주간 보고서 항상 발송 (토요일 기준)

  [Fix 5] ★ kpi_dashboard Excel DB에서 읽어 채움

영구 제약 (변경 불가):
  - Anthropic API: 번역 금지 -- 분석(Layer2/Executive Summary)에만 사용
  - EMAIL_USERNAME / EMAIL_PASSWORD 시크릿
  - ExcelUpdater.update_all() 메서드명
  - docs/index.html 은 Claude 전용, docs/genspark/ 는 Genspark 전용
  - context-based collection (Haiku) 제거 금지
"""

import json
import logging
import os
import shutil
import time

# Anthropic Haiku 설정 (SA-6/SA-7과 동일 패턴, 번역 금지)
ANTHROPIC_API_URL = 'https://api.anthropic.com/v1/messages'
HAIKU_MODEL       = 'claude-haiku-4-5-20251001'  # 절대 변경 금지
HAIKU_TIMEOUT     = 45

# 기존 ANTHROPIC_API_KEY 아래에 추가
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "").strip()
GEMINI_MODEL = "gemini-3.6-flash"
GEMINI_API_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

# Gemini API 설정 (SA-8 v3.3 추가 -- 번역 금지, Layer2 분석 전용)
GEMINI_API_URL = 'https://generativelanguage.googleapis.com/v1beta/models'
GEMINI_MODEL   = 'gemini-3.6-flash-lite'   # 무료 티어 — 2.0 Flash 2026-06-01 종료로 교체
GEMINI_TIMEOUT = 60

import subprocess
import sys
import re
import smtplib
from datetime import datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

log = logging.getLogger('generate_mi_report')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [SA-8] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# ── 경로 설정 ────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).parent.parent
SCRIPTS_DIR  = BASE_DIR / 'scripts'
DATA_DIR     = BASE_DIR / 'data'
AGENT_OUT    = DATA_DIR / 'agent_output'
SHARED_DOCS  = BASE_DIR / 'docs' / 'shared'
DOCS_DIR     = BASE_DIR / 'docs'
REPORTS_DIR  = DOCS_DIR / 'reports'

# knowledge_index 탐색 순서 (공유 경로 우선)
KNOWLEDGE_INDEX_PATHS = [
    SHARED_DOCS / 'knowledge_index.json',
    DATA_DIR / 'shared' / 'knowledge_index.json',
    AGENT_OUT / 'knowledge_index.json',
    BASE_DIR / 'knowledge_index.json',
]

# Excel DB 탐색
EXCEL_PATHS = [
    DATA_DIR / 'database' / 'Vietnam_Infra_News_Database_Final.xlsx',
    DATA_DIR / 'Vietnam_Infra_News_Database_Final.xlsx',
    DATA_DIR / 'news_database.xlsx',
]

COLLECTOR_OUT  = AGENT_OUT / 'collector_output.json'
CONTEXT_OUT    = AGENT_OUT / 'context_output.json'   # SA-7 출력
PAYLOAD_FILE   = AGENT_OUT / 'sa8_report_payload.json'
PREV_PAYLOAD   = AGENT_OUT / 'sa8_report_payload_prev.json'

PPT_BUILDER    = SCRIPTS_DIR / 'build_mi_ppt.js'
DOCX_BUILDER   = SCRIPTS_DIR / 'build_mi_report_sa8.js'

# 영역 정의
DEFAULT_AREAS = [
    {
        'name_ko': '환경 인프라',
        'name_en': 'Environment Infrastructure',
        'sector_keywords': ['Waste Water', 'Wastewater', 'Water Supply', 'Drainage',
                            'Solid Waste', 'Environment', 'Water Resources'],
        'area_keywords':   ['Environment', '환경'],
    },
    {
        'name_ko': '에너지·전력',
        'name_en': 'Energy & Power',
        'sector_keywords': ['Power', 'Oil', 'Gas', 'Energy', 'LNG', 'Nuclear',
                            'Renewable', 'Hydrogen', 'Solar', 'Wind'],
        'area_keywords':   ['Energy', '에너지'],
    },
    {
        'name_ko': '도시·교통·산업',
        'name_en': 'Urban & Transport',
        'sector_keywords': ['Smart City', 'Industrial', 'Transport', 'Urban',
                            'Metro', 'Road', 'Airport', 'Infrastructure'],
        'area_keywords':   ['Urban', '도시', '교통'],
    },
]

# ══════════════════════════════════════════════════════════════════════════
# 1. knowledge_index.json 로드
# ══════════════════════════════════════════════════════════════════════════
def load_knowledge_index():
    for kpath in KNOWLEDGE_INDEX_PATHS:
        if kpath.exists():
            log.info(f'knowledge_index 로드: {kpath}')
            try:
                with open(kpath, encoding='utf-8') as f:
                    ki = json.load(f)
                plans = ki.get('masterplans', {})
                # v2.0 이전 list 구조 대응
                if isinstance(plans, list):
                    plans = {p.get('id', p.get('plan_id', f'PLAN_{i}')): p
                             for i, p in enumerate(plans)}
                log.info(f'마스터플랜: {len(plans)}개')
                return ki, plans
            except Exception as e:
                log.warning(f'knowledge_index 파싱 오류: {e}')
    log.error('knowledge_index.json 없음!')
    return {}, {}

# ══════════════════════════════════════════════════════════════════════════
# 2. [Fix 2] Excel DB에서 기사 직접 읽기 (SA-6/SA-7 방식 동일)
# ══════════════════════════════════════════════════════════════════════════
def load_articles_from_excel(days_back=14):
    """
    Excel DB의 News Database 시트(또는 기본 시트)에서 최근 N일 기사 읽기.
    SA-6/SA-7과 동일한 Excel 헤더 구조 사용:
    ['area','sector','no','date','title_en','title_vi','tit_ko','source',
     'src_type','province','plan_id','grade','url','sum_ko','sum_en','sum_vi','qc']
    """
    articles = []
    cutoff = datetime.now() - timedelta(days=days_back)

    excel_path = None
    for ep in EXCEL_PATHS:
        if ep.exists():
            excel_path = ep
            break

    if not excel_path:
        log.warning('Excel DB 없음 — collector_output.json으로 폴백')
        return load_articles_from_collector(days_back)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        # 시트 탐색 (News Database > 기본 시트 순서)
        ws = None
        for sname in ['News Database', 'Database', '기사', wb.sheetnames[0]]:
            if sname in wb.sheetnames:
                ws = wb[sname]
                break
        if ws is None:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return []

        # 헤더 매핑 (SA-6 방식과 동일)
        raw_headers = [str(c or '').strip().lower() for c in rows[0]]
        col_map = {}
        for i, h in enumerate(raw_headers):
            # 정규화: tit_ko → title_ko 등
            norm = h.replace('tit_ko', 'title_ko').replace('sum_ko', 'summary_ko') \
                    .replace('sum_en', 'summary_en').replace('sum_vi', 'summary_vi')
            col_map[norm] = i
        log.info(f'Excel 헤더: {raw_headers}')

        # 날짜 컬럼 찾기
        date_col = (col_map.get('date') or col_map.get('published_date') or
                    col_map.get('날짜') or 3)  # 기본 4번째 컬럼

        # 컬럼 인덱스 확정
        ci = {
            'date':       date_col,
            'title_ko':   col_map.get('title_ko', col_map.get('tit_ko', 6)),
            'title_en':   col_map.get('title_en', 4),
            'summary_ko': col_map.get('summary_ko', col_map.get('sum_ko', 13)),
            'summary_en': col_map.get('summary_en', col_map.get('sum_en', 14)),
            'source':     col_map.get('source', 7),
            'province':   col_map.get('province', 9),
            'plan_id':    col_map.get('plan_id', 10),
            'grade':      col_map.get('grade', col_map.get('ctx_grade', 11)),
            'url':        col_map.get('url', 12),
            'sector':     col_map.get('sector', 1),
            'area':       col_map.get('area', 0),
        }

        parsed_total, skipped_date = 0, 0
        for row in rows[1:]:
            if not row or not any(row):
                continue
            parsed_total += 1

            # 날짜 파싱 [Fix 1]
            raw_date = row[ci['date']] if len(row) > ci['date'] else None
            art_date = _parse_date_flexible(raw_date)

            # 날짜 필터 (파싱 실패도 일단 포함 — 필터링으로 제거하지 않음)
            if art_date and art_date < cutoff:
                skipped_date += 1
                continue

            def gc(key):
                idx = ci.get(key, -1)
                if idx < 0 or idx >= len(row):
                    return ''
                return str(row[idx] or '').strip()

            articles.append({
                'date':         gc('date'),
                'published_date': gc('date'),   # 양쪽 키 모두 세팅
                'title_ko':     gc('title_ko'),
                'title_en':     gc('title_en'),
                'summary_ko':   gc('summary_ko'),
                'summary_en':   gc('summary_en'),
                'source':       gc('source'),
                'province':     gc('province'),
                'plan_id':      gc('plan_id'),   # 기존 매핑 보존
                'grade':        gc('grade'),
                'url':          gc('url'),
                'sector':       gc('sector'),
                'area':         gc('area'),
                'isNew':        _is_new_article(gc('url')),
            })

        wb.close()
        log.info(f'Excel DB → {len(articles)}건 (최근 {days_back}일) / 전체 {parsed_total}건 / 이전기사 스킵 {skipped_date}건')
        return articles

    except Exception as e:
        log.error(f'Excel 읽기 오류: {e}')
        return load_articles_from_collector(days_back)


_prev_urls_cache = None  # 이전 payload URL 캐시

def _build_prev_urls():
    """이전 payload에서 URL 집합 로드"""
    global _prev_urls_cache
    if _prev_urls_cache is not None:
        return _prev_urls_cache
    _prev_urls_cache = set()
    if PREV_PAYLOAD.exists():
        try:
            with open(PREV_PAYLOAD, encoding='utf-8') as f:
                prev = json.load(f)
            for pdata in prev.get('plans', {}).values():
                for art in pdata.get('articles', []):
                    if art.get('url'):
                        _prev_urls_cache.add(art['url'])
        except Exception:
            pass
    return _prev_urls_cache


def _is_new_article(url):
    if not url:
        return False
    prev = _build_prev_urls()
    if not prev:
        return True   # 이전 payload 없으면 모두 신규
    return url not in prev


def _parse_date_flexible(raw):
    """
    [Fix 1] 날짜 파싱 — 다양한 포맷 지원, 실패 시 None (경고 출력)
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.replace(tzinfo=None)
    s = str(raw).strip()
    if not s or s == 'None':
        return None

    fmts = [
        '%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%SZ',
        '%Y-%m-%dT%H:%M:%S.%f', '%Y/%m/%d',
        '%d/%m/%Y', '%m/%d/%Y',
        '%a, %d %b %Y %H:%M:%S %z', '%a, %d %b %Y %H:%M:%S %Z',
        '%d %b %Y', '%b %d, %Y', '%B %d, %Y',
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s[:25], fmt).replace(tzinfo=None)
        except Exception:
            pass
    # ISO 포맷 부분 추출 시도
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', s)
    if m:
        try:
            return datetime.strptime(m.group(1), '%Y-%m-%d')
        except Exception:
            pass
    log.debug(f'날짜 파싱 실패: "{s[:30]}"')
    return None


# ══════════════════════════════════════════════════════════════════════════
# 2b. collector_output.json 폴백 (Excel 없을 때)
# ══════════════════════════════════════════════════════════════════════════
def load_articles_from_collector(days_back=14):
    """
    [Fix 1] collector_output.json 날짜 필터 버그 수정판
    published_date / date 양쪽 키 모두 인식
    """
    if not COLLECTOR_OUT.exists():
        log.warning('collector_output.json 없음')
        return []

    try:
        with open(COLLECTOR_OUT, encoding='utf-8') as f:
            raw = json.load(f)
        arts = raw if isinstance(raw, list) else raw.get('articles', raw.get('items', []))
    except Exception as e:
        log.error(f'collector_output.json 읽기 실패: {e}')
        return []

    cutoff = datetime.now() - timedelta(days=days_back)
    filtered, total, skipped_old, skipped_parse = [], len(arts), 0, 0

    for art in arts:
        # [Fix 1] published_date 우선, date 폴백
        raw_date = (art.get('published_date') or art.get('date') or
                    art.get('Published_Date') or art.get('Date') or '')
        art_date = _parse_date_flexible(raw_date)

        if art_date is None:
            # 날짜 파싱 실패 → 포함 (기존: 조용히 스킵 → 버그)
            skipped_parse += 1
            art['isNew'] = _is_new_article(art.get('url', ''))
            filtered.append(art)
        elif art_date >= cutoff:
            art['isNew'] = _is_new_article(art.get('url', ''))
            filtered.append(art)
        else:
            skipped_old += 1

    log.info(f'collector_output.json → {len(filtered)}건 (파싱실패={skipped_parse} 기간외={skipped_old} 전체={total})')
    return filtered

# ══════════════════════════════════════════════════════════════════════════
# 3. 기사 ↔ 플랜 매핑
# ══════════════════════════════════════════════════════════════════════════
def match_articles_to_plans(articles, plans):
    """
    1순위: plan_id 컬럼이 이미 있으면 직접 사용 (Excel DB)
    2순위: knowledge_index keywords 매칭
    기사 없는 플랜도 빈 리스트로 반드시 포함 (Layer1 보존)
    """
    grouped = {pid: [] for pid in plans}

    for art in articles:
        # 1순위: plan_id 직접 매핑 (Excel Matched_Plan or plan_id 컬럼)
        direct_pid = (art.get('plan_id') or art.get('Matched_Plan') or
                      art.get('matched_plan') or '').strip()
        if direct_pid and direct_pid in plans:
            grouped[direct_pid].append(art)
            continue

        # 2순위: 키워드 매핑
        text = ' '.join([
            str(art.get('title_ko', '')),
            str(art.get('title_en', '') or art.get('title', '')),
            str(art.get('summary_ko', '')),
            str(art.get('summary_en', '')),
        ]).lower()

        best_pid, best_score = None, 0
        for pid, pdata in plans.items():
            kws = (pdata.get('keywords_en', []) or
                   pdata.get('keywords', []) or
                   pdata.get('keywords_vi', []))
            if isinstance(kws, str):
                kws = [kws]
            score = sum(1 for kw in kws if kw and kw.lower() in text)
            if score > best_score:
                best_score, best_pid = score, pid

        if best_pid and best_score >= 1:
            grouped[best_pid].append(art)

    matched_total = sum(len(v) for v in grouped.values())
    log.info(f'기사 매핑: {matched_total}건 매핑 / {len(articles)}건 전체')
    return grouped

# ══════════════════════════════════════════════════════════════════════════
# 4. KPI 변동 감지
# ══════════════════════════════════════════════════════════════════════════
def detect_kpi_changes(plans):
    kpi_changes = []
    if not PREV_PAYLOAD.exists():
        return kpi_changes
    try:
        with open(PREV_PAYLOAD, encoding='utf-8') as f:
            prev = json.load(f)
        prev_plans = prev.get('plans', {})
        for pid, pdata in plans.items():
            prev_pdata = prev_plans.get(pid, {})
            curr_kpis = {k.get('label', k.get('indicator', '')): k
                         for k in pdata.get('kpi_targets', [])}
            prev_kpis = {k.get('label', k.get('indicator', '')): k
                         for k in prev_pdata.get('kpi_targets', [])}
            for label, curr in curr_kpis.items():
                prev_k = prev_kpis.get(label)
                if not prev_k:
                    kpi_changes.append({
                        'plan_id': pid, 'indicator': label,
                        'from': '미포함', 'to': str(curr.get('target', '')),
                        'reason': f'{pid} 신규 KPI',
                    })
                    curr['changed'] = True
                elif str(curr.get('target', '')) != str(prev_k.get('target', '')):
                    kpi_changes.append({
                        'plan_id': pid, 'indicator': label,
                        'from': str(prev_k.get('target', '')),
                        'to':   str(curr.get('target', '')),
                        'reason': f'{pid} — {label} 목표값 변경',
                    })
                    curr['changed'] = True
    except Exception as e:
        log.warning(f'KPI 변동 감지 실패: {e}')
    log.info(f'KPI 변동: {len(kpi_changes)}건')
    return kpi_changes

# ══════════════════════════════════════════════════════════════════════════
# 5. 페이로드 조립
# ══════════════════════════════════════════════════════════════════════════

# ============================================================
# Layer2 AI 분석 생성 함수 (SA-8 v3.2 신규)
# ============================================================

def _call_haiku_sa8(system_prompt, user_prompt, api_key):
    """SA-6/SA-7과 동일한 Haiku 호출 패턴 -- 번역 금지, 분석 전용"""
    import json as _j
    import urllib.request
    try:
        headers = {
            'Content-Type':      'application/json',
            'x-api-key':         api_key,
            'anthropic-version': '2023-06-01',
        }
        body = _j.dumps({
            'model':    HAIKU_MODEL,
            'max_tokens': 600,
            'system':   system_prompt,
            'messages': [{'role': 'user', 'content': user_prompt}],
        }).encode('utf-8')
        req = urllib.request.Request(
            ANTHROPIC_API_URL, data=body, headers=headers, method='POST')
        with urllib.request.urlopen(req, timeout=HAIKU_TIMEOUT) as resp:
            data = _j.loads(resp.read().decode('utf-8'))
            return data['content'][0]['text'].strip()
    except Exception as e:
        log.warning(f'  Haiku SA8 호출 오류: {e}')
        logger.warning("Anthropic API 실패 -> Gemini API로 대체 호출합니다.")
        return _call_gemini_fallback(prompt)
    
def _call_gemini_fallback(prompt: str) -> str:
    if not GEMINI_API_KEY:
        logger.error("GEMINI_API_KEY가 존재하지 않습니다.")
        return ""
    headers = {"Content-Type": "application/json"}
    params = {"key": GEMINI_API_KEY}
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 2048}
    }
    try:
        import requests
        res = requests.post(GEMINI_API_URL, headers=headers, params=params, json=payload, timeout=60)
        res.raise_for_status()
        candidates = res.json().get("candidates", [])
        if candidates:
            parts = candidates[0].get("content", {}).get("parts", [])
            if parts:
                return parts[0].get("text", "").strip()
        return ""
    except Exception as e:
        logger.error(f"Gemini API 호출 실패: {e}")
        return ""

def _call_gemini_sa8(system_prompt, user_prompt, gemini_key, use_search=True):
    """Gemini Flash 호출 -- 번역 금지, Layer2 분석 전용
    use_search=True: Google Search Grounding (정부발표·ADB 등 실시간 보완)
    """
    import json as _j
    import urllib.request
    try:
        url = (f'{GEMINI_API_URL}/{GEMINI_MODEL}:generateContent'
               f'?key={gemini_key}')
        tools = [{'google_search': {}}] if use_search else []
        payload = {
            'system_instruction': {'parts': [{'text': system_prompt}]},
            'contents': [{'parts': [{'text': user_prompt}], 'role': 'user'}],
            'generationConfig': {'maxOutputTokens': 800, 'temperature': 0.3},
        }
        if tools:
            payload['tools'] = tools
        body = _j.dumps(payload).encode('utf-8')
        req = urllib.request.Request(
            url, data=body,
            headers={'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=GEMINI_TIMEOUT) as resp:
            data = _j.loads(resp.read().decode('utf-8'))
            return data['candidates'][0]['content']['parts'][0]['text'].strip()
    except Exception as e:
        log.warning(f'  Gemini SA8 호출 오류: {e}')
        return ''


def generate_layer2_analysis(plans_payload, api_key):
    """플랜별 Layer2 AI 분석 -- 기사 있는 플랜에만 Haiku 호출"""
    if not api_key:
        log.warning('[Layer2] ANTHROPIC_API_KEY 없음 -- AI 분석 건너뜀')
        return
    system = (
        '당신은 베트남 인프라 개발 전문 시장정보(MI) 애널리스트입니다. '
        '수집된 뉴스 기사를 바탕으로 경영진 보고용 인사이트를 한국어로 작성합니다. '
        '반드시 다음 3개 항목을 포함하세요: '
        '1. 이번 주 핵심 진행사항 (사업 추진 현황) '
        '2. 투자 및 사업 기회 시그널 (한국 기업 관점) '
        '3. 리스크 또는 지연 징후 (없으면 특이사항 없음) '
        '각 항목은 1~2문장, 전체 250자 이내로 작성하세요.'
    )
    targets = {pid: p for pid, p in plans_payload.items() if p.get('articles')}
    log.info(f'[Layer2] AI 분석 대상: {len(targets)}개 플랜')
    for pid, pdata in targets.items():
        arts = pdata.get('articles', [])
        plan_name = pdata.get('plan_name_ko') or pid
        art_lines = []
        for a in arts[:8]:
            title   = (a.get('title_ko') or '')[:60]
            summary = (a.get('summary_ko') or '')[:100]
            date    = (a.get('date') or '')[:10]
            art_lines.append(f'- [{date}] {title}: {summary}')
        user = (
            f'마스터플랜: {plan_name}\n'
            f'수집 기사 ({len(arts)}건):\n' + '\n'.join(art_lines) +
            '\n\n위 기사를 분석하여 경영진 보고용 인사이트 3개 항목을 작성하세요.'
        )
        # Gemini 우선 -> 실패 시 Haiku fallback
        gemini_key = os.environ.get('GEMINI_API_KEY', '').strip()
        if gemini_key:
            result = _call_gemini_sa8(system, user, gemini_key, use_search=True)
            if result:
                pdata['analysis_ko'] = result
                log.info(f'  [Layer2/Gemini] {pid}: {len(result)}자')
                time.sleep(4)  # Rate Limit 방지
                continue
        result = _call_haiku_sa8(system, user, api_key)
        if result:
            pdata['analysis_ko'] = result
            log.info(f'  [Layer2/Haiku] {pid}: {len(result)}자')
        else:
            log.warning(f'  [Layer2] {pid}: 생성 실패')
        # Gemini Rate Limit 방지 — 플랜 간 4초 대기 (15 req/min 초과 방지)
        time.sleep(4)


def generate_executive_summary(plans_payload, new_articles, api_key):
    """Executive Summary -- 신규 기사 기반 Haiku 1회 호출"""
    if not api_key or not new_articles:
        return ''
    system = (
        '당신은 베트남 인프라 MI 애널리스트입니다. '
        '이번 주 베트남 인프라 동향을 경영진에게 보고하는 요약문을 한국어로 작성합니다. '
        '형식: [이번 주 3대 주요 동향] 1. 2. 3. '
        '[섹터별 주요 움직임] Power/에너지: / 환경인프라: / 교통산업단지: '
        '전체 400자 이내, 사실 기반으로만 작성하세요.'
    )
    art_lines = []
    for a in new_articles[:15]:
        title   = (a.get('title_ko') or a.get('tit_ko') or '')[:60]
        summary = (a.get('summary_ko') or a.get('sum_ko') or '')[:80]
        sector  = (a.get('sector') or '')
        art_lines.append(f'- [{sector}] {title}: {summary}')
    user = (
        f'이번 주 신규 수집 기사 ({len(new_articles)}건):\n' +
        '\n'.join(art_lines) +
        '\n\n위 기사를 종합하여 경영진 보고용 Executive Summary를 작성하세요.'
    )
    # Gemini 우선 (Search Grounding으로 최신 동향 반영) -> Haiku fallback
    gemini_key = os.environ.get('GEMINI_API_KEY', '').strip()
    if gemini_key:
        result = _call_gemini_sa8(system, user, gemini_key, use_search=True)
        if result:
            log.info(f'[Exec Summary/Gemini] {len(result)}자')
            return result
    result = _call_haiku_sa8(system, user, api_key)
    log.info(f'[Exec Summary/Haiku] {len(result)}자' if result else '[Exec Summary] 생성 실패')
    return result


def assemble_payload(ki, plans, grouped_arts, all_articles, kpi_changes):
    today_str     = datetime.now().strftime('%Y-%m-%d')
    period_start  = (datetime.now() - timedelta(days=13)).strftime('%Y-%m-%d')

    # [Fix 5] kpi_dashboard: knowledge_index에서 주요 KPI 집계
    kpi_dashboard, seen_labels = [], set()
    for pdata in plans.values():
        for kpi in pdata.get('kpi_targets', []):
            label = (kpi.get('label') or kpi.get('indicator') or
                     kpi.get('indicator_ko') or '').strip()
            if label and label not in seen_labels:
                seen_labels.add(label)
                kpi_dashboard.append({
                    'label':   label,
                    'target':  str(kpi.get('target', '')),
                    'current': str(kpi.get('current') or kpi.get('current_value') or kpi.get('baseline') or ''),
                    'changed': bool(kpi.get('changed', False)),
                })
            if len(kpi_dashboard) >= 12:
                break
        if len(kpi_dashboard) >= 12:
            break

    # KPI 달성률 (프로그레스 바용)
    kpi_achievement = []
    for kpi in kpi_dashboard[:8]:
        curr_nums   = re.findall(r'[\d.]+', str(kpi.get('current', '')))
        target_nums = re.findall(r'[\d.]+', str(kpi.get('target', '')))
        try:
            cn = float(curr_nums[0]) if curr_nums else 0
            tn = float(target_nums[0]) if target_nums else 100
            pct = min(int(cn / tn * 100), 100) if tn else 0
        except Exception:
            pct = 0
        kpi_achievement.append({'label': kpi['label'], 'current_pct': pct})

    # 영역별 플랜 분류
    areas = _classify_areas(plans)

    # 플랜별 데이터 조립 (Layer1 완전 보존)
    plans_payload = {}
    for pid, pdata in plans.items():
        arts = grouped_arts.get(pid, [])

        # Layer1 KPI 정규화 (절대 삭제 금지)
        # knowledge_index 실제 필드: kpi_targets[].indicator / kpis[] / key_targets[] 모두 대응
        # ★ key_targets는 문자열 배열 → dict로 변환하여 통일
        kpi_source = pdata.get('kpi_targets') or pdata.get('kpis') or []
        if not kpi_source:
            raw_kt = pdata.get('key_targets') or []
            kpi_source = [
                {'indicator': str(t), 'target': '', 'current': ''}
                for t in raw_kt if t
            ]
        norm_kpis = []
        for k in kpi_source:
            if isinstance(k, dict):
                norm_kpis.append({
                    'label':   (k.get('label') or k.get('indicator') or
                                k.get('indicator_ko') or '').strip(),
                    'target':  str(k.get('target', '')),
                    'current': str(k.get('current') or k.get('current_value') or
                                   k.get('baseline') or ''),
                    'changed': bool(k.get('changed', False)),
                })

        # Layer1 프로젝트 정규화 (절대 삭제 금지)
        # knowledge_index 실제 필드: projects[].name / key_projects[].name_ko 양쪽 대응
        proj_source = pdata.get('key_projects') or pdata.get('projects') or []
        norm_projs = []
        for p in proj_source:
            if isinstance(p, dict):
                norm_projs.append({
                    'name_ko':  (p.get('name_ko') or p.get('name') or '').strip(),
                    'location': (p.get('location') or p.get('province') or '').strip(),
                    'capacity': (p.get('capacity') or p.get('size') or '').strip(),
                    'note':     (p.get('note') or p.get('description') or '').strip(),
                    'status':   (p.get('status') or '').strip(),
                })

        # 기사 목록 (Layer2)
        arts_payload = []
        for a in arts:
            arts_payload.append({
                'title_ko':   str(a.get('title_ko') or a.get('tit_ko') or a.get('title') or ''),
                'summary_ko': str(a.get('summary_ko') or a.get('sum_ko') or '')[:200],
                'source':     str(a.get('source') or ''),
                'date':       str(a.get('date') or a.get('published_date') or '')[:10],
                'url':        str(a.get('url') or ''),
                'isNew':      bool(a.get('isNew', False)),
            })

        # ★ description_ko fallback: 없으면 title_ko + decision + key_targets로 구성
        _desc = (pdata.get('description_ko') or pdata.get('overview') or
                 pdata.get('description') or '')
        if not _desc:
            _parts = []
            if pdata.get('title_ko'):
                _parts.append(pdata['title_ko'])
            if pdata.get('decision'):
                _parts.append(f"근거: {pdata['decision']}")
            if pdata.get('period'):
                _parts.append(f"기간: {pdata['period']}")
            raw_kt = pdata.get('key_targets') or []
            if raw_kt:
                _parts.append('주요 목표: ' + ' / '.join(str(t) for t in raw_kt[:4]))
            _desc = '\n'.join(_parts)

        plans_payload[pid] = {
            'plan_name_ko': (pdata.get('name_ko') or pdata.get('plan_name_ko') or pid),
            'sector':       (pdata.get('sector') or
                             (pdata.get('sectors', [''])[0] if pdata.get('sectors') else '')),
            'area':         (pdata.get('area') or ''),
            # knowledge_index: 'legal' 필드, 페이로드: 'decision' 필드
            'decision':     (pdata.get('decision') or pdata.get('legal') or
                             pdata.get('legal_basis') or ''),
            # ★ Layer1 필수 필드 — 절대 삭제/변경 금지
            'description_ko': _desc,
            'kpi_targets':    norm_kpis,
            'key_projects':   norm_projs,
            # Layer2
            'analysis_ko':  pdata.get('analysis_ko') or '',
            'kpi_changes':  [c for c in kpi_changes if c.get('plan_id') == pid],
            'articles':     arts_payload,
        }

    new_count = sum(1 for a in all_articles if a.get('isNew', False))

    return {
        'report_date':         today_str,
        'report_period':       f'{period_start} ~ {today_str}',
        'knowledge_version':   ki.get('version', 'v2.x'),
        'total_articles':      len(all_articles),
        'new_articles_count':  new_count,
        'executive_summary_ko': '',
        'kpi_dashboard':       kpi_dashboard,
        'kpi_changes':         kpi_changes,
        'kpi_achievement':     kpi_achievement,
        'areas':               areas,
        'plans':               plans_payload,
    }


def _classify_areas(plans):
    areas = []
    for area_def in DEFAULT_AREAS:
        matched_ids = []
        for pid, pdata in plans.items():
            sector = (pdata.get('sector') or
                      (pdata.get('sectors', [''])[0] if pdata.get('sectors') else ''))
            area   = pdata.get('area', '')
            text   = f'{sector} {area}'.lower()
            for kw in area_def['sector_keywords'] + area_def.get('area_keywords', []):
                if kw.lower() in text:
                    matched_ids.append(pid)
                    break
        if matched_ids:
            areas.append({
                'name_ko':  area_def['name_ko'],
                'name_en':  area_def['name_en'],
                'plan_ids': matched_ids,
            })
    return areas

# ══════════════════════════════════════════════════════════════════════════
# 6. PPT / Word 빌더 실행
# ══════════════════════════════════════════════════════════════════════════
def run_ppt_builder(payload_path, output_path):
    if not PPT_BUILDER.exists():
        log.warning(f'PPT 빌더 없음: {PPT_BUILDER}')
        return False
    env = os.environ.copy()
    env['SA8_DATA_FILE']   = str(payload_path)
    env['SA8_OUTPUT_PATH'] = str(output_path)
    r = subprocess.run(['node', str(PPT_BUILDER)], env=env,
                       capture_output=True, text=True)
    if r.returncode != 0:
        log.error(f'PPT 빌더 실패:\n{r.stderr[-500:]}')
        return False
    log.info(f'PPT 생성: {output_path}')
    return True


def run_docx_builder(payload_path, output_path):
    if not DOCX_BUILDER.exists():
        log.warning(f'Word 빌더 없음: {DOCX_BUILDER}')
        return False
    r = subprocess.run(
        ['node', str(DOCX_BUILDER), str(payload_path), str(output_path)],
        capture_output=True, text=True
    )
    if r.returncode != 0:
        log.error(f'Word 빌더 실패:\n{r.stderr[-500:]}')
        return False
    log.info(f'Word 생성: {output_path}')
    return True

# ══════════════════════════════════════════════════════════════════════════
# 7. [Fix 3] docs/reports/ 복사
# ══════════════════════════════════════════════════════════════════════════
def copy_to_reports(pptx_path, docx_path):
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    for src in [pptx_path, docx_path]:
        if src and Path(src).exists():
            dst = REPORTS_DIR / Path(src).name
            if Path(src).resolve() == dst.resolve():
                log.info(f'보고서 복사 스킵 (동일 경로): {dst}')
                continue
            shutil.copy2(src, dst)
            log.info(f'보고서 복사: {dst}')

# ══════════════════════════════════════════════════════════════════════════
# 8. [Fix 4] 이메일 발송 — 주간 보고서 항상 발송
# ══════════════════════════════════════════════════════════════════════════
def send_email(pptx_path, docx_path, payload, kpi_changes):
    """
    [Fix 4] 주간 보고서는 항상 발송.
    (기존: KPI변동 OR 신규기사 조건 → 이번 실행처럼 기사 0건이면 미발송)
    """
    username = os.environ.get('EMAIL_USERNAME')
    password = os.environ.get('EMAIL_PASSWORD')
    if not username or not password:
        log.warning('EMAIL_USERNAME / EMAIL_PASSWORD 미설정 — 이메일 건너뜀')
        return False

    today_str   = payload.get('report_date', datetime.now().strftime('%Y-%m-%d'))
    plan_count  = len(payload.get('plans', {}))
    art_count   = payload.get('total_articles', 0)
    new_count   = payload.get('new_articles_count', 0)

    subject = f'[베트남 인프라 MI] 주간 보고서 — {today_str}'
    if kpi_changes:
        subject += f' ★ KPI 변동 {len(kpi_changes)}건'

    body_lines = [
        f'안녕하세요,\n\n베트남 인프라 MI 주간 보고서({today_str})를 첨부합니다.\n',
        f'■ 수집 기간: {payload.get("report_period", "")}',
        f'■ 전체 기사: {art_count}건 (신규 {new_count}건)',
        f'■ 마스터플랜: {plan_count}개 전체 포함',
    ]
    if kpi_changes:
        body_lines.append(f'\n★ KPI 변동사항 ({len(kpi_changes)}건):')
        for ch in kpi_changes:
            body_lines.append(
                f'  - {ch.get("plan_id")}: {ch.get("indicator")} '
                f'{ch.get("from")} → {ch.get("to")}'
            )
    body_lines += [
        '\n■ 첨부 파일:',
        '  • PPT: 경영진 보고용 슬라이드 (21개 플랜 전체)',
        '  • Word: 상세 분석 보고서 (Layer1 사업개요 + Layer2 AI분석)',
        '\n대시보드: https://hms4792.github.io/vietnam-infra-news/',
        '\n본 메일은 Claude SA-8이 자동 생성하였습니다.',
    ]
    body = '\n'.join(body_lines)

    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From']    = username
    msg['To']      = username
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    attached = []
    for fpath in [pptx_path, docx_path]:
        if fpath and Path(fpath).exists():
            with open(fpath, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Path(fpath).name)
            msg.attach(part)
            attached.append(Path(fpath).name)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(username, password)
            smtp.send_message(msg)
        log.info(f'이메일 발송 완료 → {username} | 첨부: {attached}')
        return True
    except Exception as e:
        log.error(f'이메일 발송 실패: {e}')
        return False

# ══════════════════════════════════════════════════════════════════════════
# 9. 메인
# ══════════════════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description='SA-8 MI 보고서 생성기 v3.3')
    parser.add_argument('--days-back',  type=int, default=14,
                        help='기사 수집 기간(일) [기본: 14]')
    parser.add_argument('--send-email', action='store_true',
                        help='이메일 발송')
    parser.add_argument('--daily-only', action='store_true',
                        help='Daily pipeline mode (ignored, for yml compatibility)')
    parser.add_argument('--dry-run',    action='store_true',
                        help='페이로드만 생성 (빌더 미실행)')
    parser.add_argument('--output-dir', default=str(DOCS_DIR),
                        help='출력 디렉토리')
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    AGENT_OUT.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    today_tag  = datetime.now().strftime('%Y%m%d')
    pptx_path  = output_dir / f'VN_Infra_MI_Weekly_Report_{today_tag}.pptx'
    # ★ Fix: 파일명을 PPT와 동일하게 Weekly_Report로 통일
    #   이전: VN_Infra_MI_Report_*.docx → scan_reports() 패턴 불일치
    #   수정: VN_Infra_MI_Weekly_Report_*.docx → 정상 인식
    docx_path  = REPORTS_DIR / f'VN_Infra_MI_Weekly_Report_{today_tag}.docx'

    log.info('=' * 60)
    log.info('SA-8 MI 보고서 생성기 v3.3 시작')
    log.info('=' * 60)

    # Step 1: knowledge_index 로드
    ki, plans = load_knowledge_index()
    if not plans:
        log.error('마스터플랜 데이터 없음. 종료.')
        sys.exit(1)

    # Step 2: [Fix 2] Excel DB에서 기사 로드
    all_articles = load_articles_from_excel(days_back=args.days_back)

    # Step 3: 기사 ↔ 플랜 매핑
    grouped_arts = match_articles_to_plans(all_articles, plans)

    # Step 4: KPI 변동 감지
    if PREV_PAYLOAD.exists():
        shutil.copy(PAYLOAD_FILE, PREV_PAYLOAD) if PAYLOAD_FILE.exists() else None
    kpi_changes = detect_kpi_changes(plans)

    # Step 5: 페이로드 조립
    payload = assemble_payload(ki, plans, grouped_arts, all_articles, kpi_changes)

    # Step 5-a: Layer2 AI 분석 + Executive Summary (Haiku) -- SA-8 v3.2
    _api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if _api_key:
        generate_layer2_analysis(payload['plans'], _api_key)
        _new_arts = [a for a in all_articles if a.get('isNew', False)]
        payload['executive_summary_ko'] = generate_executive_summary(
            payload['plans'], _new_arts, _api_key)
    else:
        log.warning('[SA-8] ANTHROPIC_API_KEY 없음 -- Layer2 분석 건너뜀')

    # 이전 payload 백업
    if PAYLOAD_FILE.exists():
        shutil.copy(PAYLOAD_FILE, PREV_PAYLOAD)

    # 페이로드 저장
    with open(PAYLOAD_FILE, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    log.info(f'페이로드 저장: {PAYLOAD_FILE}')
    log.info(f'  플랜: {len(payload["plans"])}개 | 기사: {payload["total_articles"]}건 | 신규: {payload["new_articles_count"]}건')
    log.info(f'  kpi_dashboard: {len(payload["kpi_dashboard"])}개 | kpi_changes: {len(kpi_changes)}건')

    if args.dry_run:
        log.info('DRY-RUN 모드 — 빌더 실행 생략')
        return

    # Step 6: PPT 생성
    ppt_ok = run_ppt_builder(PAYLOAD_FILE, pptx_path)

    # Step 7: Word 생성
    docx_ok = run_docx_builder(PAYLOAD_FILE, docx_path)

    # Step 8: [Fix 3] docs/reports/ 복사
    if ppt_ok or docx_ok:
        copy_to_reports(
            pptx_path if ppt_ok else None,
            docx_path if docx_ok else None,
        )

    # Step 9: [Fix 4] 이메일 발송 (주간 보고서 항상)
    if args.send_email:
        send_email(
            pptx_path if ppt_ok else None,
            docx_path if docx_ok else None,
            payload,
            kpi_changes,
        )

    log.info('=' * 60)
    log.info(f'SA-8 완료: PPT={ppt_ok} | Word={docx_ok}')
    log.info('=' * 60)


if __name__ == '__main__':
    main()
