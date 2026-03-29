#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
reset_sqlite_db.py
SQLite DB를 Excel DB 기준으로 재동기화

목적: 잘못된 EXCEL_PATH로 실행되어 SQLite에만 저장된 url_hash 제거
     → 다음 수집 실행 시 중복 판단 없이 신규 기사 정상 수집
용도: GitHub Actions 워크플로에서 main.py 실행 전 매번 호출

[메모리항목12] 별도 .py 파일로 분리 — yml 인라인 코드 금지
"""
import sqlite3, hashlib, os
from pathlib import Path

EXCEL_PATH = os.environ.get('EXCEL_PATH', 'data/database/Vietnam_Infra_News_Database_Final.xlsx')
DB_PATH    = os.environ.get('DB_PATH', 'data/vietnam_infrastructure_news.db')

print(f"Excel : {EXCEL_PATH}")
print(f"SQLite: {DB_PATH}")

# ── 1. Excel DB에서 기존 URL 수집 ────────────────────────────
excel_hashes = set()
try:
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    # Link 컬럼 위치 찾기 (헤더에서 'link' 검색)
    link_col = 7  # 기본값 (G열 = Link)
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or '').lower()
        if h in ('link', 'url'):
            link_col = c
            break
    # URL → MD5 hash 변환
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[link_col - 1] if link_col - 1 < len(row) else None
        if url and str(url).startswith('http'):
            excel_hashes.add(hashlib.md5(str(url).encode()).hexdigest())
    wb.close()
    print(f"Excel hash 수: {len(excel_hashes)}")
except Exception as e:
    print(f"Excel 읽기 오류 (무시하고 계속): {e}")

# ── 2. SQLite DB 연결 및 테이블 생성 ────────────────────────
Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)
conn = sqlite3.connect(DB_PATH)
conn.execute("""
    CREATE TABLE IF NOT EXISTS articles (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        url_hash       TEXT UNIQUE,
        url            TEXT,
        title          TEXT,
        title_vi       TEXT,
        title_ko       TEXT,
        summary        TEXT,
        summary_vi     TEXT,
        summary_ko     TEXT,
        source         TEXT,
        sector         TEXT,
        area           TEXT,
        province       TEXT,
        confidence     INTEGER DEFAULT 0,
        published_date TEXT,
        collected_date TEXT,
        processed      INTEGER DEFAULT 0
    )
""")
conn.commit()

cur = conn.execute("SELECT COUNT(*) FROM articles")
before = cur.fetchone()[0]
print(f"SQLite 현재: {before}건")

if not excel_hashes:
    print("Excel hash 없음 — SQLite 전체 초기화")
    conn.execute("DELETE FROM articles")
    conn.commit()
    print("SQLite 초기화 완료")
    conn.close()
    print("완료")
    exit(0)

# ── 3. Excel에 없는 hash 삭제 ────────────────────────────────
# SQLite IN절 최대 999개 제한 → 청크 단위로 분할 처리
excel_list = list(excel_hashes)
CHUNK = 900  # 999 미만으로 안전하게

# Excel에 있는 hash를 임시 테이블로 저장 후 NOT IN 대신 LEFT JOIN 방식
conn.execute("CREATE TEMP TABLE IF NOT EXISTS keep_hashes (h TEXT PRIMARY KEY)")
conn.execute("DELETE FROM keep_hashes")

# 청크 단위로 INSERT
for i in range(0, len(excel_list), CHUNK):
    chunk = excel_list[i:i+CHUNK]
    conn.executemany(
        "INSERT OR IGNORE INTO keep_hashes (h) VALUES (?)",
        [(h,) for h in chunk]
    )
conn.commit()

# Excel에 없는 행 삭제
conn.execute("""
    DELETE FROM articles
    WHERE url_hash NOT IN (SELECT h FROM keep_hashes)
""")
conn.commit()

cur = conn.execute("SELECT COUNT(*) FROM articles")
after = cur.fetchone()[0]
print(f"SQLite 정리 후: {after}건 (삭제: {before - after}건)")

conn.close()
print("완료 — 다음 main.py 실행 시 신규 기사 정상 수집됩니다.")
