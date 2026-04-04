"""
archive_unclassified.py
Business Sector = "Unclassified" 행을 Unclassified_Archive 시트로 이동.
News Database에서는 제거 (완전 삭제 아님).
"""

import os

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "database", "Vietnam_Infra_News_Database_Final.xlsx")

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl 미설치")

COL_SECTOR = 2   # B열 Business Sector

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["News Database"]

# ── 헤더 읽기 ────────────────────────────────────────────────────────────────
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
n_cols  = len(headers)

# ── Unclassified 행 인덱스 수집 ──────────────────────────────────────────────
unclassified_indices = []   # 1-based row index

for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row_idx, COL_SECTOR).value
    if val == "Unclassified":
        unclassified_indices.append(row_idx)

print(f"Unclassified 행: {len(unclassified_indices)}건")

# ── Unclassified_Archive 시트 생성 (또는 초기화) ─────────────────────────────
if "Unclassified_Archive" in wb.sheetnames:
    del wb["Unclassified_Archive"]
archive_ws = wb.create_sheet("Unclassified_Archive")

# 헤더 복사
for c_idx, h in enumerate(headers, 1):
    archive_ws.cell(1, c_idx).value = h

# 데이터 복사 (값만)
for archive_row, row_idx in enumerate(unclassified_indices, start=2):
    for c in range(1, n_cols + 1):
        archive_ws.cell(archive_row, c).value = ws.cell(row_idx, c).value

print(f"Unclassified_Archive 시트: {len(unclassified_indices)}행 복사 완료")

# ── News Database에서 행 삭제 (아래→위 순서로) ──────────────────────────────
for row_idx in reversed(unclassified_indices):
    ws.delete_rows(row_idx)

remaining = ws.max_row - 1   # 헤더 제외
print(f"News Database 남은 행: {remaining}건")

# ── 저장 ─────────────────────────────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f"[OK] Excel 저장 완료")
print(f"\n결과: News Database {remaining}건 / Unclassified_Archive {len(unclassified_indices)}건")
