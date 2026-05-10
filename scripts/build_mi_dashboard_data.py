"""
build_mi_dashboard_data.py
===========================
역할: MI Dashboard에 필요한 모든 데이터를 JSON으로 생성

입력:
  - data/database/Vietnam_Infra_News_Database_Final.xlsx  (News Database + Matched_Plan 시트)
  - data/agent_output/context_output.json                 (SA-7 맥락분석 결과)
  - data/agent_output/stage_timeline.json                 (SA-7 진행단계 타임라인)
  - data/agent_output/mi_daily_*.json                     (SA-8 일일 분석 결과)
  - docs/shared/knowledge_index.json                      (마스터플랜 메타데이터)

출력:
  - docs/shared/mi_dashboard_data.json  ← mi_dashboard.html이 fetch로 로드

실행:
  python3 scripts/build_mi_dashboard_data.py

GitHub Actions 연동:
  daily_pipeline.yml  — SA-8 직후 실행 (매일 KST 20:00)
  collect_weekly.yml  — SA-8 주간 보고서 생성 후 실행 (토 KST 22:00)

영구 제약:
  - EXCEL_PATH: data/database/Vietnam_Infra_News_Database_Final.xlsx
  - KI_PATH: docs/shared/knowledge_index.json 우선 탐색
  - date 키: article.get('date') or article.get('published_date')

버전: v1.2 (2026-05-10)
변경:
  ★ v1.1 BUG FIX: load_matched_articles() — Matched_Plan 헤더 오프셋 수정
    원인: Matched_Plan 시트 구조
      Row1 = 메타 행 ("★ SA-7 맥락 확정 기사 271건...")  ← merged cell
      Row2 = 실제 헤더 (No, ctx_tag, ctx_grade, Plan_ID ...)
      Row3~ = 데이터
    버그: min_row=1로 헤더를 읽으면 Row1 메타가 헤더로 인식됨
          → plan_id 컬럼을 못 찾음 (None) → 전체 기사 skip → 0건
    수정: 헤더 읽기 min_row=1 → min_row=2
          데이터 읽기 min_row=2 → min_row=3
    효과: 0건 → 271건 정상 로드

  ★ v1.2 BUG FIX: scan_reports() — 파일명 패턴 불일치 수정
    원인:
      SA-8이 생성하는 Word 파일명: VN_Infra_MI_Report_YYYYMMDD.docx
      SA-8이 생성하는 PPT 파일명:  VN_Infra_MI_Weekly_Report_YYYYMMDD.pptx
      scan_reports()가 찾는 패턴:  VN_Infra_MI_Weekly_Report_*.docx
      → Word 파일이 패턴에 안 맞아 word_exists=False → 대시보드 날짜 미갱신
    수정1: generate_mi_report.py에서 docx 파일명도 Weekly_Report 통일
           (VN_Infra_MI_Report → VN_Infra_MI_Weekly_Report)
    수정2: scan_reports()에서 두 패턴 모두 스캔하도록 보강
           (VN_Infra_MI_Weekly_Report_*.docx + VN_Infra_MI_Report_*.docx)
    수정3: 대시보드 최종갱신일을 stats.updated_at에 명시적으로 포함
    효과: 대시보드 "주간 MI 보고서 날짜"가 최신 실행일로 자동 갱신됨
"""

import json
import logging
import os
import glob
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# ── 경로 ──────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent.parent
DATA_DIR      = BASE_DIR / 'data'
AGENT_OUT_DIR = DATA_DIR / 'agent_output'
DOCS_SHARED   = BASE_DIR / 'docs' / 'shared'
REPORTS_DIR   = BASE_DIR / 'docs' / 'reports'

EXCEL_PATH    = DATA_DIR / 'database' / 'Vietnam_Infra_News_Database_Final.xlsx'
CONTEXT_OUT   = AGENT_OUT_DIR / 'context_output.json'
TIMELINE_OUT  = AGENT_OUT_DIR / 'stage_timeline.json'
OUTPUT_PATH   = DOCS_SHARED / 'mi_dashboard_data.json'

KI_PATHS = [
    DOCS_SHARED / 'knowledge_index.json',
    DATA_DIR / 'shared' / 'knowledge_index.json',
    DATA_DIR / 'shared' / 'layer1_data.json',
]

# ── 로깅 ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='[DashData %(asctime)s] %(message)s',
    datefmt='%H:%M:%S'
)
log = logging.getLogger('DashData')

# ── 진행단계 한국어 매핑 ──────────────────────────────────────────────────
STAGE_LABEL = {
    'PLANNING':     {'ko': '계획·승인',  'en': 'Planning',     'vi': 'Lập kế hoạch'},
    'TENDERING':    {'ko': '입찰·계약',  'en': 'Tendering',    'vi': 'Đấu thầu'},
    'CONSTRUCTION': {'ko': '건설·시공',  'en': 'Construction', 'vi': 'Xây dựng'},
    'COMPLETION':   {'ko': '준공·개통',  'en': 'Completion',   'vi': 'Hoàn thành'},
    'OPERATION':    {'ko': '운영·확장',  'en': 'Operation',    'vi': 'Vận hành'},
    'UNKNOWN':      {'ko': '미확정',     'en': 'Unknown',      'vi': 'Chưa xác định'},
}


# ══════════════════════════════════════════════════════════════════════════
#  Step 1: knowledge_index 로드
# ══════════════════════════════════════════════════════════════════════════
def load_knowledge_index() -> dict:
    for path in KI_PATHS:
        if path.exists():
            with open(path, 'r', encoding='utf-8') as f:
                ki = json.load(f)
            plans = ki.get('masterplans', ki)
            for pid, p in plans.items():
                p['plan_id'] = pid
            log.info(f"knowledge_index 로드: {len(plans)}개 플랜 [{path.name}]")
            return plans
    log.warning("knowledge_index 없음 — 빈 플랜 사용")
    return {}


# ══════════════════════════════════════════════════════════════════════════
#  Step 2: Excel에서 Matched_Plan 기사 추출
# ══════════════════════════════════════════════════════════════════════════
def load_matched_articles() -> tuple[list[dict], dict]:
    """
    Excel Matched_Plan 시트 또는 News Database에서 플랜 매핑 기사 추출.

    ★ v1.1 수정: Matched_Plan 시트 행 구조
      Row1 = 메타 행 (merged cell: "★ SA-7 맥락 확정 기사 271건...")
      Row2 = 실제 헤더 (No, ctx_tag, ctx_grade, Plan_ID, Title_EN ...)
      Row3~ = 데이터
      → 헤더: min_row=2, 데이터: min_row=3 으로 수정

    Returns:
        (matched_articles, stats)
    """
    if not EXCEL_PATH.exists():
        log.warning(f"Excel DB 없음: {EXCEL_PATH}")
        return [], {'total_articles': 0, 'matched_count': 0, 'updated_at': ''}

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # 전체 기사 수
    total = 0
    if 'News Database' in wb.sheetnames:
        ws_nd = wb['News Database']
        total = ws_nd.max_row - 1  # 헤더 제외

    matched_articles = []
    matched_count    = 0

    # ── Matched_Plan 시트 우선 시도 ──────────────────────────────────────
    if 'Matched_Plan' in wb.sheetnames:
        ws = wb['Matched_Plan']

        # ★ v1.1 수정: min_row=2 (Row2가 실제 헤더, Row1은 메타 행)
        headers = [str(c or '').strip().lower().replace(' ', '_')
                   for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]

        def ci(keys):
            for k in keys:
                for i, h in enumerate(headers):
                    if k in h:
                        return i
            return None

        COL = {
            'plan_id':    ci(['plan_id', 'matched_plan']),
            'date':       ci(['date']),
            'title_ko':   ci(['title_ko']),
            'title_en':   ci(['title_en', 'title_en_orig', 'title_(en/vi)', 'title', 'news_title']),
            'title_vi':   ci(['title_vi']),
            'summary_ko': ci(['summary_ko']),
            'summary_en': ci(['summary_en', 'short_sum', 'short_summary']),
            'source':     ci(['source']),
            'url':        ci(['link', 'url']),
            'ctx_grade':  ci(['ctx_grade']),
            'stage':      ci(['stage', 'ctx_stage']),
            'milestone':  ci(['milestone']),
        }

        log.info(f"  Matched_Plan 헤더: {headers[:8]}")
        log.info(f"  컬럼 매핑: plan_id={COL['plan_id']} date={COL['date']} "
                 f"title_ko={COL['title_ko']} ctx_grade={COL['ctx_grade']}")

        # ★ v1.1 수정: min_row=3 (Row3부터 데이터, Row1=메타 Row2=헤더)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue

            plan_id  = str(row[COL['plan_id']] if COL['plan_id'] is not None and len(row) > COL['plan_id'] else '').strip()
            date_val = str(row[COL['date']]    if COL['date']    is not None and len(row) > COL['date']    else '').strip()[:10]

            if not plan_id or not date_val:
                continue

            def rv(key):
                idx = COL.get(key)
                if idx is not None and len(row) > idx:
                    return str(row[idx] or '').strip()
                return ''

            title_ko  = rv('title_ko')
            title_en  = rv('title_en')
            grade_raw = rv('ctx_grade')
            grade     = grade_raw if grade_raw in ('HIGH','MEDIUM','POLICY','LOW') else 'MEDIUM'

            matched_articles.append({
                'plan_id':    plan_id,
                'date':       date_val,
                'title_ko':   title_ko or title_en,
                'title_en':   title_en or title_ko,
                'title_vi':   rv('title_vi'),
                'summary_ko': rv('summary_ko')[:200],
                'summary_en': rv('summary_en')[:200],
                'source':     rv('source') or 'Unknown',
                'url':        rv('url'),
                'grade':      grade,
                'stage':      rv('stage') or 'UNKNOWN',
                'milestone':  rv('milestone'),
            })
            matched_count += 1

        log.info(f"Matched_Plan 시트: {matched_count}건")

    else:
        # ── Matched_Plan 없음 → News Database에서 ctx_plans 있는 것만 추출 ──
        log.warning("Matched_Plan 시트 없음 — News Database에서 추출")
        if 'News Database' in wb.sheetnames:
            ws = wb['News Database']
            # News Database는 Row1이 헤더 (메타 행 없음)
            headers = [str(c or '').strip().lower().replace(' ', '_')
                       for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            def ci2(keys):
                for k in keys:
                    for i, h in enumerate(headers):
                        if k in h:
                            return i
                return None

            COL2 = {
                'ctx_plans':  ci2(['ctx_plans', 'plan_id', 'matched_plan']),
                'date':       ci2(['date']),
                'title_ko':   ci2(['tit_ko', 'title_ko']),
                'title_en':   ci2(['title_en', 'title_(en/vi)', 'news_title']),
                'summary_ko': ci2(['sum_ko', 'summary_ko']),
                'summary_en': ci2(['sum_en', 'summary_en', 'short_summary']),
                'source':     ci2(['source']),
                'url':        ci2(['url', 'link']),
                'ctx_grade':  ci2(['grade', 'ctx_grade']),
            }

            cutoff = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                ctx_plans = str(row[COL2['ctx_plans']] if COL2['ctx_plans'] is not None else '').strip()
                if not ctx_plans:
                    continue
                date_val = str(row[COL2['date']] if COL2['date'] is not None else '').strip()[:10]
                if date_val < cutoff:
                    continue

                for pid in [p.strip() for p in ctx_plans.split(',') if p.strip()]:
                    title_ko = str(row[COL2['title_ko']] if COL2['title_ko'] is not None else '').strip()
                    title_en = str(row[COL2['title_en']] if COL2['title_en'] is not None else '').strip()
                    grade_r  = str(row[COL2['ctx_grade']] if COL2['ctx_grade'] is not None else '').strip()
                    matched_articles.append({
                        'plan_id':    pid,
                        'date':       date_val,
                        'title_ko':   title_ko or title_en,
                        'title_en':   title_en or title_ko,
                        'title_vi':   '',
                        'summary_ko': str(row[COL2['summary_ko']] if COL2['summary_ko'] is not None else '').strip()[:200],
                        'summary_en': str(row[COL2['summary_en']] if COL2['summary_en'] is not None else '').strip()[:200],
                        'source':     str(row[COL2['source']] if COL2['source'] is not None else '') or 'Unknown',
                        'url':        str(row[COL2['url']] if COL2['url'] is not None else '').strip(),
                        'grade':      grade_r if grade_r in ('HIGH','MEDIUM','POLICY','LOW') else 'MEDIUM',
                        'stage':      'UNKNOWN',
                        'milestone':  '',
                    })
                    matched_count += 1

    wb.close()

    # 날짜 기준 정렬
    matched_articles.sort(key=lambda x: x.get('date', ''), reverse=True)

    stats = {
        'total_articles': total,
        'matched_count':  matched_count,
        'updated_at':     datetime.now().strftime('%Y-%m-%d %H:%M'),
    }
    return matched_articles, stats


# ══════════════════════════════════════════════════════════════════════════
#  Step 3: SA-7 컨텍스트 + 타임라인 로드
# ══════════════════════════════════════════════════════════════════════════
def load_sa7_data() -> tuple[dict, dict]:
    ctx_by_plan: dict[str, dict] = {}
    timeline:    dict[str, dict] = {}

    if CONTEXT_OUT.exists():
        with open(CONTEXT_OUT, 'r', encoding='utf-8') as f:
            ctx = json.load(f)
        for art in ctx.get('articles', []):
            pid = art.get('plan_id', '')
            if not pid:
                continue
            existing = ctx_by_plan.get(pid)
            if not existing or art.get('confidence', 0) > existing.get('confidence', 0):
                ctx_by_plan[pid] = {
                    'stage':      art.get('stage', 'UNKNOWN'),
                    'confidence': art.get('confidence', 0.0),
                    'milestone':  art.get('milestone', ''),
                    'next_watch': art.get('next_watch', ''),
                    'insight':    art.get('insight', ''),
                    'haiku_used': art.get('haiku_used', False),
                }
        log.info(f"SA-7 context: {len(ctx_by_plan)}개 플랜")

    if TIMELINE_OUT.exists():
        with open(TIMELINE_OUT, 'r', encoding='utf-8') as f:
            tl = json.load(f)
        timeline = tl.get('plans', {})
        log.info(f"SA-7 timeline: {len(timeline)}개 플랜")

    return ctx_by_plan, timeline


# ══════════════════════════════════════════════════════════════════════════
#  Step 4: SA-8 일일 분석 JSON 로드 (최신 파일)
# ══════════════════════════════════════════════════════════════════════════
def load_sa8_daily() -> dict:
    pattern = str(AGENT_OUT_DIR / 'mi_daily_*.json')
    files   = sorted(glob.glob(pattern), reverse=True)
    if not files:
        log.info("SA-8 daily JSON 없음")
        return {}
    latest = files[0]
    with open(latest, 'r', encoding='utf-8') as f:
        data = json.load(f)
    log.info(f"SA-8 daily 로드: {Path(latest).name}")
    return data


# ══════════════════════════════════════════════════════════════════════════
#  Step 5: 보고서 파일 목록 스캔
#  ★ v1.2 수정: 두 가지 docx 파일명 패턴 모두 인식
#    - VN_Infra_MI_Weekly_Report_*.docx  (신규 통일 명칭)
#    - VN_Infra_MI_Report_*.docx         (구버전 호환)
# ══════════════════════════════════════════════════════════════════════════
def scan_reports() -> list[dict]:
    import re
    report_map = {}

    if REPORTS_DIR.exists():
        # ── Word 파일 스캔 (두 패턴 모두 인식) ──────────────────────────
        # 패턴1: VN_Infra_MI_Weekly_Report_*.docx (신규 통일 명칭)
        for f in sorted(REPORTS_DIR.glob('VN_Infra_MI_Weekly_Report_*.docx'), reverse=True):
            key = f.stem.replace('_Full', '')
            report_map.setdefault(key, {})['word_url']    = f'reports/{f.name}'
            report_map.setdefault(key, {})['word_exists'] = True

        # 패턴2: VN_Infra_MI_Report_*.docx (구버전 호환 — Weekly 없는 이름)
        for f in sorted(REPORTS_DIR.glob('VN_Infra_MI_Report_*.docx'), reverse=True):
            # 날짜 태그 추출해서 Weekly_Report 키와 동일한 key로 매핑
            m = re.search(r'(\d{8})', f.stem)
            if m:
                key = f'VN_Infra_MI_Weekly_Report_{m.group(1)}'
            else:
                key = f.stem
            # 이미 Weekly_Report 패턴에서 찾은 게 없을 때만 추가
            if not report_map.get(key, {}).get('word_exists'):
                report_map.setdefault(key, {})['word_url']    = f'reports/{f.name}'
                report_map.setdefault(key, {})['word_exists'] = True

        # ── PPT 파일 스캔 ────────────────────────────────────────────────
        for f in sorted(REPORTS_DIR.glob('VN_Infra_MI_Weekly_Report_*.pptx'), reverse=True):
            key = f.stem
            report_map.setdefault(key, {})['pptx_url']    = f'reports/{f.name}'
            report_map.setdefault(key, {})['pptx_exists'] = True

    reports = []
    for key, info in sorted(report_map.items(), reverse=True):
        m = re.search(r'(\d{8})', key)
        date_str = week_str = ''
        if m:
            d = m.group(1)
            try:
                dt = datetime.strptime(d, '%Y%m%d')
                date_str = dt.strftime('%Y-%m-%d')
                iso      = dt.isocalendar()
                week_str = f'{iso[0]}-W{iso[1]:02d}'
            except ValueError:
                date_str = d

        reports.append({
            'week':        week_str or key,
            'date':        date_str,
            'word_url':    info.get('word_url', ''),
            'pptx_url':    info.get('pptx_url', ''),
            'word_exists': info.get('word_exists', False),
            'pptx_exists': info.get('pptx_exists', False),
        })

    log.info(f"보고서 파일 스캔: {len(reports)}개")
    if reports:
        latest = reports[0]
        log.info(f"  최신 보고서: {latest['date']} | Word={latest['word_exists']} PPT={latest['pptx_exists']}")
    return reports


# ══════════════════════════════════════════════════════════════════════════
#  Step 6: 플랜별 데이터 조립
# ══════════════════════════════════════════════════════════════════════════
def assemble_plan_data(
    plans:            dict,
    matched_articles: list[dict],
    ctx_by_plan:      dict,
    timeline:         dict,
    sa8_daily:        dict,
) -> dict:
    # 기사를 플랜별로 그룹핑
    art_by_plan: dict[str, list] = {}
    for art in matched_articles:
        pid = art.get('plan_id', '')
        if pid:
            art_by_plan.setdefault(pid, []).append(art)

    result = {}
    all_pids = set(plans.keys()) | set(art_by_plan.keys())

    for pid in all_pids:
        plan_meta = plans.get(pid, {})
        arts      = art_by_plan.get(pid, [])
        ctx       = ctx_by_plan.get(pid, {})
        tl        = timeline.get(pid, {})

        stage = (
            ctx.get('stage')
            or tl.get('current_stage')
            or plan_meta.get('stage')
            or 'UNKNOWN'
        )

        sa8_plan = {}
        if sa8_daily:
            for sec in sa8_daily.get('plan_sections', []):
                if sec.get('plan_id') == pid:
                    sa8_plan = sec
                    break

        tl_history = tl.get('stage_history', [])

        result[pid] = {
            'plan_id':    pid,
            'stage':      stage,
            'stage_label': STAGE_LABEL.get(stage, STAGE_LABEL['UNKNOWN']),
            'confidence': ctx.get('confidence', 0.0),
            'haiku_used': ctx.get('haiku_used', False),

            'title_ko':       plan_meta.get('title_ko', pid),
            'description_ko': plan_meta.get('description_ko', ''),
            'decision':       plan_meta.get('decision', ''),
            'sector':         plan_meta.get('sector', ''),
            'area':           plan_meta.get('area', ''),
            'kpi_targets':    plan_meta.get('kpi_targets', []),
            'key_projects':   plan_meta.get('key_projects', []),

            'articles': [
                {
                    'date':       a.get('date', ''),
                    'title_ko':   a.get('title_ko', ''),
                    'title_en':   a.get('title_en', ''),
                    'title_vi':   a.get('title_vi', ''),
                    'summary_ko': a.get('summary_ko', ''),
                    'summary_en': a.get('summary_en', ''),
                    'grade':      a.get('grade', 'MEDIUM'),
                    'stage':      a.get('stage', 'UNKNOWN'),
                    'source':     a.get('source', ''),
                    'url':        a.get('url', ''),
                    'milestone':  a.get('milestone', ''),
                }
                for a in arts[:8]
            ],
            'article_count': len(arts),

            'insight':    ctx.get('insight', '') or sa8_plan.get('insight_ko', ''),
            'next_watch': (
                ctx.get('next_watch', '')
                or tl.get('next_watch', '')
                or sa8_plan.get('next_watch_ko', '')
            ),
            'milestone': ctx.get('milestone', ''),

            'timeline': [
                {
                    'date':         item.get('date', ''),
                    'stage':        item.get('stage', 'UNKNOWN'),
                    'milestone_ko': item.get('milestone', ''),
                    'source':       item.get('source', ''),
                }
                for item in tl_history[:10]
            ],
        }

    log.info(f"플랜 데이터 조립: {len(result)}개")
    return result


# ══════════════════════════════════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════
#  Step 7: mi_dashboard.html BACKEND_DATA 자동 교체
#  - 화면 로직(JS) 완전 보존, /*__BACKEND_DATA__*/const KI={...}; 만 교체
#  - 교체 항목: articles, stats, gap_issues, ai_analysis, totals
#  - 보존 항목: overview, kpis, projects, name, legal 등 knowledge_index 필드
# ══════════════════════════════════════════════════════════════════════════
def update_mi_dashboard_html(ki_plans: dict, plan_data: dict,
                              matched_articles: list, stats: dict,
                              reports: list):
    """
    mi_dashboard.html의 /*__BACKEND_DATA__*/ 섹션만 교체.
    기존 화면 JS 로직, CSS, HTML 구조 완전 보존.

    교체 필드 (자동 업데이트):
      plans[].articles   → Excel Matched_Plan 최신 기사
      plans[].stats      → Excel DB 집계
      plans[].gap_issues → SA-7 분석
      plans[].ai_analysis→ SA-7 분석
      totals             → 전체 집계 + today 날짜

    보존 필드 (knowledge_index 수동 관리):
      plans[].overview, kpis, projects, name, legal 등
    """
    html_path = BASE_DIR / 'docs' / 'mi_dashboard.html'
    if not html_path.exists():
        log.warning(f"mi_dashboard.html 없음: {html_path}")
        return

    html = html_path.read_text(encoding='utf-8')

    # ── 기사를 플랜별로 그룹핑 ──────────────────────────────────────────
    art_by_plan: dict[str, list] = {}
    for art in matched_articles:
        pid = art.get('plan_id', '')
        if pid:
            art_by_plan.setdefault(pid, []).append(art)

    # ── 기존 KI 추출 (knowledge_index 필드 보존용) ──────────────────────
    m = re.search(r'/\*__BACKEND_DATA__\*/const KI=(\{.*?\});\s*\n',
                  html, re.DOTALL)
    if not m:
        log.warning("BACKEND_DATA 패턴 미발견 — 교체 건너뜀")
        return

    import json
    try:
        old_ki = json.loads(m.group(1))
    except Exception as e:
        log.error(f"기존 KI JSON 파싱 실패: {e}")
        return

    # ── 각 플랜 데이터 조립 ─────────────────────────────────────────────
    new_plans = {}
    for pid, old_plan in old_ki.get('plans', {}).items():
        arts = art_by_plan.get(pid, [])

        # article → HTML 형식 변환 (d, te, tk, sk, src, pv, g, url)
        new_articles = []
        for a in arts[:30]:
            new_articles.append({
                'd':   a.get('date', ''),
                'te':  a.get('title_en', '') or a.get('title_ko', ''),
                'tk':  a.get('title_ko', '') or a.get('title_en', ''),
                'sk':  (a.get('summary_ko', '') or a.get('summary_en', ''))[:300],
                'src': a.get('source', ''),
                'pv':  a.get('province', ''),
                'g':   a.get('grade', 'MEDIUM'),
                'url': a.get('url', ''),
            })

        # stats 계산
        year_dist: dict[str, int] = {}
        for a in arts:
            y = str(a.get('date', ''))[:4]
            if y.isdigit():
                year_dist[y] = year_dist.get(y, 0) + 1

        cutoff_week = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        this_week = sum(1 for a in arts if a.get('date', '') >= cutoff_week)
        latest_date = max((a.get('date', '') for a in arts), default='')

        # gap_issues
        from_plan_data = plan_data.get(pid, {})
        new_gap = from_plan_data.get('gap_issues',
                   old_plan.get('gap_issues', []))

        # ai_analysis
        new_ai = from_plan_data.get('ai_analysis',
                  old_plan.get('ai_analysis', []))

        # 기존 플랜 기반 + 자동 필드만 교체
        new_plan = dict(old_plan)   # ← overview, kpis, projects 등 보존
        new_plan['articles']    = new_articles
        new_plan['stats']       = {
            'total':       len(arts),
            'this_week':   this_week,
            'year_dist':   year_dist,
            'latest_date': latest_date,
        }
        new_plan['gap_issues']  = new_gap
        new_plan['ai_analysis'] = new_ai
        new_plans[pid] = new_plan

    # ── totals 재계산 ────────────────────────────────────────────────────
    total_news   = stats.get('total_articles', old_ki['totals'].get('total_news', 0))
    total_mapped = stats.get('matched_count',  old_ki['totals'].get('total_mapped', 0))
    this_week_total = sum(
        p.get('stats', {}).get('this_week', 0) for p in new_plans.values()
    )
    high = sum(1 for a in matched_articles if a.get('grade') == 'HIGH')
    med  = sum(1 for a in matched_articles if a.get('grade') == 'MEDIUM')
    pol  = sum(1 for a in matched_articles if a.get('grade') == 'POLICY')

    new_totals = dict(old_ki['totals'])
    new_totals.update({
        'total_news':   total_news,
        'total_mapped': total_mapped,
        'this_week':    this_week_total,
        'high':         high,
        'med':          med,
        'pol':          pol,
        'today':        datetime.now().strftime('%Y-%m-%d'),
    })

    # ── 보고서 링크 교체 ─────────────────────────────────────────────────
    latest_report_date = reports[0]['date'] if reports else ''
    latest_word = next((r for r in reports if r.get('word_exists')), None)
    latest_ppt  = next((r for r in reports if r.get('pptx_exists')), None)

    # ── 새 KI JSON 생성 ──────────────────────────────────────────────────
    new_ki = {
        'plans':       new_plans,
        'totals':      new_totals,
        'area_groups': old_ki.get('area_groups', {}),
    }
    new_ki_json = json.dumps(new_ki, ensure_ascii=False, separators=(',', ':'))

    new_backend = f'/*__BACKEND_DATA__*/const KI={new_ki_json};'
    new_html = html[:m.start()] + new_backend + html[m.end()-1:]

    # ── 보고서 링크 패치 ─────────────────────────────────────────────────
    if latest_word:
        new_html = re.sub(
            r'href="reports/VN_Infra_MI_[^"]*\.docx"',
            f'href="{latest_word["word_url"]}"', new_html)
    if latest_ppt:
        new_html = re.sub(
            r'href="reports/VN_Infra_MI_[^"]*\.pptx"',
            f'href="{latest_ppt["pptx_url"]}"', new_html)

    html_path.write_text(new_html, encoding='utf-8')
    log.info(f"mi_dashboard.html BACKEND_DATA 교체 완료")
    log.info(f"  플랜: {len(new_plans)}개 | 기사: {total_mapped}건 | today: {new_totals['today']}")

def main():
    log.info("=" * 58)
    log.info(f"MI Dashboard Data Builder v1.2 — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    log.info("=" * 58)

    plans            = load_knowledge_index()
    matched_articles, stats = load_matched_articles()
    ctx_by_plan, timeline   = load_sa7_data()
    sa8_daily        = load_sa8_daily()
    reports          = scan_reports()
    plan_data        = assemble_plan_data(
        plans, matched_articles, ctx_by_plan, timeline, sa8_daily
    )

    # ★ v1.2: 최신 보고서 날짜를 stats에 포함 → 대시보드 갱신일 업데이트
    latest_report_date = reports[0]['date'] if reports else ''

    output = {
        'generated_at':        datetime.now().strftime('%Y-%m-%d %H:%M'),
        'generated_date':      datetime.now().strftime('%Y-%m-%d'),
        'latest_report_date':  latest_report_date,   # ★ 신규 필드
        'stats': {
            'total_articles':      stats.get('total_articles', 0),
            'matched_count':       stats.get('matched_count', 0),
            'plan_count':          len(plan_data),
            'ki_plan_count':       len(plans),
            'report_count':        len(reports),
            'updated_at':          stats.get('updated_at', ''),        # Excel 갱신시각
            'latest_report_date':  latest_report_date,                 # 최신 보고서 날짜
        },
        'reports': reports,
        'plans':   plan_data,
    }

    DOCS_SHARED.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    size_kb = OUTPUT_PATH.stat().st_size // 1024
    # ★ mi_dashboard.html BACKEND_DATA 자동 교체
    update_mi_dashboard_html(
        ki_plans=plans,
        plan_data=plan_data,
        matched_articles=matched_articles,
        stats=stats,
        reports=reports,
    )

    log.info("")
    log.info("━" * 58)
    log.info("✅ MI Dashboard 데이터 빌드 완료")
    log.info(f"   출력: {OUTPUT_PATH}")
    log.info(f"   크기: {size_kb} KB")
    log.info(f"   플랜: {len(plan_data)}개")
    log.info(f"   기사: {stats.get('matched_count', 0)}건 (매핑)")
    log.info(f"   보고서: {len(reports)}개")
    log.info(f"   최신 보고서 날짜: {latest_report_date}")   # ★ 신규 로그
    log.info("━" * 58)


if __name__ == '__main__':
    main()
