"""
dashboard_updater.py
====================
Gemini 진단 수정사항 #3 완전 반영:

  문제: 대시보드 HTML 기사 카드에 data-ko / data-en / data-vi 속성이 없음
       → 3개국어 전환 JavaScript가 속성을 찾지 못해 버튼 동작 불가

  수정: 모든 기사 카드 <h3>, <p> 태그에 data-ko / data-en / data-vi 속성 삽입
       → 버튼 클릭 시 JavaScript가 해당 속성값으로 텍스트 교체 가능

JavaScript 언어 전환 동작 원리:
  1. 버튼 클릭 → setLanguage('ko') 호출
  2. document.querySelectorAll('[data-ko]') 로 모든 다국어 요소 탐색
  3. element.textContent = element.dataset.ko 로 텍스트 교체
  ↑ 이 흐름이 작동하려면 data-ko/en/vi 속성이 HTML에 반드시 있어야 함!
"""

import html
import json
import logging
import os
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

# ── 경로 설정 ──────────────────────────────────────────────
TEMPLATE_PATH = Path("templates/dashboard_template.html")
OUTPUT_PATH   = Path("docs/index.html")   # GitHub Pages 배포 경로
EXCEL_PATH    = Path("data/database/Vietnam_Infra_News_Database_Final.xlsx")


# ════════════════════════════════════════════════════════════
# 핵심: 기사 카드 HTML 생성 함수 (다국어 속성 완전 포함)
# ════════════════════════════════════════════════════════════

def _build_article_card(article: dict, idx: int) -> str:
    """
    단일 기사 카드 HTML 생성
    
    Gemini 수정사항 #3:
    - <h3> 제목 태그에 data-en / data-ko / data-vi 속성 추가
    - <p> 요약 태그에 data-en / data-ko / data-vi 속성 추가
    - 이 속성들이 없으면 3개국어 버튼이 동작하지 않음!
    
    Args:
        article: 번역 완료 기사 dict
                 (title_en, title_ko, title_vi, summary_en, summary_ko, summary_vi 포함)
        idx: 카드 인덱스 (짝수/홀수 스타일 구분용)
    """
    # ── 다국어 제목 추출 ─────────────────────────────────────
    # html.escape(): HTML 특수문자(<, >, &, ", ')를 안전하게 변환
    #   → XSS 방지 및 HTML 속성 내 특수문자 오류 방지
    title_en = html.escape(article.get('title_en') or article.get('title', '') or '')
    title_ko = html.escape(article.get('title_ko') or article.get('title', '') or '')
    title_vi = html.escape(article.get('title_vi') or article.get('title', '') or '')

    # ── 다국어 요약 추출 ─────────────────────────────────────
    summary_en = html.escape(article.get('summary_en', '') or '')
    summary_ko = html.escape(article.get('summary_ko', '') or '')
    summary_vi = html.escape(article.get('summary_vi', '') or '')

    # 요약이 없을 경우 기본값
    if not summary_en:
        summary_en = html.escape(f"{article.get('sector', 'Infrastructure')} project in Vietnam.")
    if not summary_ko:
        summary_ko = html.escape(f"베트남 {article.get('sector', '인프라')} 관련 사업.")
    if not summary_vi:
        summary_vi = html.escape(f"Dự án {article.get('sector', 'hạ tầng')} tại Việt Nam.")

    # ── 메타 정보 ─────────────────────────────────────────────
    sector  = html.escape(article.get('sector', 'General'))
    source  = html.escape(article.get('source', ''))
    date    = html.escape(str(article.get('date', '')))
    url     = article.get('url', '#')
    area    = html.escape(article.get('area', ''))

    # 섹터별 색상 클래스 매핑
    sector_class = _get_sector_class(article.get('sector', ''))

    # ── HTML 카드 생성 ────────────────────────────────────────
    # ★★★ 핵심 수정: data-en / data-ko / data-vi 속성 반드시 포함 ★★★
    card_html = f"""
    <div class="article-card" data-sector="{sector}" data-area="{area}" data-idx="{idx}">

        <!-- 섹터 배지 -->
        <div class="card-header">
            <span class="sector-badge {sector_class}">{sector}</span>
            <span class="article-date">{date}</span>
        </div>

        <!-- 기사 제목: data-en/ko/vi 속성으로 3개국어 저장 -->
        <!-- JavaScript: element.textContent = element.dataset[lang] 으로 교체 -->
        <h3 class="article-title"
            data-en="{title_en}"
            data-ko="{title_ko}"
            data-vi="{title_vi}">
            {title_en}
        </h3>

        <!-- 기사 요약: data-en/ko/vi 속성으로 3개국어 저장 -->
        <p class="article-summary"
           data-en="{summary_en}"
           data-ko="{summary_ko}"
           data-vi="{summary_vi}">
           {summary_en}
        </p>

        <!-- 하단 메타 정보 -->
        <div class="card-footer">
            <span class="source-tag">📰 {source}</span>
            {f'<span class="area-tag">📍 {area}</span>' if area else ''}
            <a href="{url}" target="_blank" rel="noopener" class="read-more">
                <span data-en="Read original"
                      data-ko="원문 보기"
                      data-vi="Đọc bản gốc">Read original</span>
                →
            </a>
        </div>

    </div>
    """
    return card_html


def _get_sector_class(sector: str) -> str:
    """섹터명에 따른 CSS 클래스 반환"""
    mapping = {
        'Waste Water':    'sector-ww',
        'Water Supply':   'sector-ws',
        'Drainage':       'sector-ws',
        'Solid Waste':    'sector-sw',
        'Power':          'sector-pw',
        'Oil & Gas':      'sector-og',
        'Industrial':     'sector-ip',
        'Smart City':     'sector-sc',
        'Transport':      'sector-tr',
        'Construction':   'sector-cn',
    }
    for key, cls in mapping.items():
        if key.lower() in sector.lower():
            return cls
    return 'sector-general'


# ════════════════════════════════════════════════════════════
# 언어 전환 JavaScript (3개국어 버튼 동작 로직)
# ════════════════════════════════════════════════════════════

LANGUAGE_SWITCH_JS = """
<script>
// ═══════════════════════════════════════════════════════════
// 3개국어 전환 JavaScript
// ═══════════════════════════════════════════════════════════
// 동작 원리:
//   1. 언어 버튼 클릭 → setLanguage('en'|'ko'|'vi') 호출
//   2. data-en / data-ko / data-vi 속성을 가진 모든 요소 탐색
//   3. 해당 속성값으로 요소의 텍스트 교체
//
// 필수 조건:
//   HTML 요소에 data-en, data-ko, data-vi 속성이 있어야 함!
//   (dashboard_updater.py에서 카드 생성 시 반드시 포함)
// ═══════════════════════════════════════════════════════════

let currentLang = 'en';  // 기본 언어: 영어

function setLanguage(lang) {
    currentLang = lang;
    
    // ① 모든 다국어 요소 일괄 업데이트
    // 'data-en' 속성이 있는 모든 요소를 대상으로 함
    document.querySelectorAll('[data-en]').forEach(function(el) {
        const text = el.dataset[lang];  // el.dataset.en / el.dataset.ko / el.dataset.vi
        if (text && text.trim() !== '') {
            el.textContent = text;
        }
    });

    // ② 버튼 active 상태 업데이트
    document.querySelectorAll('.lang-btn').forEach(function(btn) {
        btn.classList.remove('active');
    });
    const activeBtn = document.getElementById('btn-' + lang);
    if (activeBtn) activeBtn.classList.add('active');

    // ③ html lang 속성 업데이트 (접근성)
    document.documentElement.lang = lang === 'ko' ? 'ko' : lang === 'vi' ? 'vi' : 'en';

    // ④ localStorage에 선택 언어 저장 (페이지 재방문 시 유지)
    try { localStorage.setItem('preferred_lang', lang); } catch(e) {}
}

// 페이지 로드 시 저장된 언어 복원
document.addEventListener('DOMContentLoaded', function() {
    try {
        const saved = localStorage.getItem('preferred_lang') || 'en';
        setLanguage(saved);
    } catch(e) {
        setLanguage('en');
    }
    
    // 필터 초기화
    initFilters();
});

// ── 섹터/지역 필터 기능 ──────────────────────────────────
function initFilters() {
    // 섹터 필터
    document.querySelectorAll('.sector-filter-btn').forEach(function(btn) {
        btn.addEventListener('click', function() {
            const sector = this.dataset.sector || 'all';
            filterBySector(sector);
            document.querySelectorAll('.sector-filter-btn').forEach(b => b.classList.remove('active'));
            this.classList.add('active');
        });
    });
}

function filterBySector(sector) {
    document.querySelectorAll('.article-card').forEach(function(card) {
        if (sector === 'all' || card.dataset.sector === sector) {
            card.style.display = '';
        } else {
            card.style.display = 'none';
        }
    });
}

// 검색 기능
function searchArticles(query) {
    const q = query.toLowerCase().trim();
    document.querySelectorAll('.article-card').forEach(function(card) {
        const title   = (card.querySelector('.article-title')   || {}).textContent || '';
        const summary = (card.querySelector('.article-summary') || {}).textContent || '';
        const match   = title.toLowerCase().includes(q) || summary.toLowerCase().includes(q);
        card.style.display = (q === '' || match) ? '' : 'none';
    });
}
</script>
"""

# 언어 버튼 HTML
LANGUAGE_BUTTONS_HTML = """
<div class="language-switcher" role="navigation" aria-label="Language selector">
    <button id="btn-en" class="lang-btn active"
            onclick="setLanguage('en')"
            title="Switch to English">
        🇺🇸 EN
    </button>
    <button id="btn-ko" class="lang-btn"
            onclick="setLanguage('ko')"
            title="한국어로 전환">
        🇰🇷 KO
    </button>
    <button id="btn-vi" class="lang-btn"
            onclick="setLanguage('vi')"
            title="Chuyển sang Tiếng Việt">
        🇻🇳 VI
    </button>
</div>
"""


# ════════════════════════════════════════════════════════════
# 전체 대시보드 HTML 생성
# ════════════════════════════════════════════════════════════

def generate_dashboard_html(articles: list) -> str:
    """
    전체 대시보드 HTML 생성
    
    Args:
        articles: processed_articles (번역 완료, title_en/ko/vi, summary_en/ko/vi 포함)
    Returns:
        완성된 HTML 문자열
    """
    if not articles:
        logger.warning("[Dashboard] 기사 없음 - 빈 대시보드 생성")

    # ── 통계 계산 ─────────────────────────────────────────
    today = datetime.now().strftime('%Y-%m-%d')
    today_articles = [a for a in articles if str(a.get('date', '')).startswith(today)]
    sectors = list(set(a.get('sector', 'General') for a in articles if a.get('sector')))
    sectors.sort()

    # ── 기사 카드 HTML 생성 ─────────────────────────────────
    cards_html = '\n'.join(
        _build_article_card(article, idx)
        for idx, article in enumerate(articles[:100])  # 최대 100건 표시
    )

    # ── 섹터 필터 버튼 ─────────────────────────────────────
    sector_filter_html = '<button class="sector-filter-btn active" data-sector="all">All</button>\n'
    for s in sectors:
        sector_filter_html += f'<button class="sector-filter-btn" data-sector="{html.escape(s)}">{html.escape(s)}</button>\n'

    # ── BACKEND_DATA: JavaScript에서 사용할 전체 데이터 ──────
    # 대시보드 JS에서 차트나 추가 처리에 사용
    backend_data = []
    for a in articles:
        backend_data.append({
            'title': {
                'en': a.get('title_en') or a.get('title', ''),
                'ko': a.get('title_ko') or a.get('title', ''),
                'vi': a.get('title_vi') or a.get('title', ''),
            },
            'summary': {
                'en': a.get('summary_en', ''),
                'ko': a.get('summary_ko', ''),
                'vi': a.get('summary_vi', ''),
            },
            'sector':  a.get('sector', ''),
            'source':  a.get('source', ''),
            'date':    str(a.get('date', '')),
            'url':     a.get('url', ''),
            'area':    a.get('area', ''),
        })

    backend_data_json = json.dumps(backend_data, ensure_ascii=False, indent=2)

    # ── 완성 HTML ──────────────────────────────────────────
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Vietnam Infrastructure News Dashboard</title>
    <style>
        /* ── CSS 변수 (테마 색상) ── */
        :root {{
            --primary:   #1a5276;
            --secondary: #2e86c1;
            --accent:    #f39c12;
            --bg:        #f4f6f9;
            --card-bg:   #ffffff;
            --text:      #2c3e50;
            --text-light:#7f8c8d;
            --border:    #dce1e7;
            --shadow:    0 2px 8px rgba(0,0,0,0.08);
        }}

        * {{ box-sizing: border-box; margin: 0; padding: 0; }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 
                         'Noto Sans KR', Arial, sans-serif;
            background: var(--bg);
            color: var(--text);
            line-height: 1.6;
        }}

        /* ── 헤더 ── */
        .dashboard-header {{
            background: linear-gradient(135deg, var(--primary), var(--secondary));
            color: white;
            padding: 24px 32px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 16px;
        }}

        .header-title h1 {{
            font-size: 1.8rem;
            font-weight: 700;
            margin-bottom: 4px;
        }}

        .header-title p {{
            font-size: 0.9rem;
            opacity: 0.85;
        }}

        /* ── 언어 전환 버튼 ── */
        .language-switcher {{
            display: flex;
            gap: 8px;
        }}

        .lang-btn {{
            background: rgba(255,255,255,0.15);
            color: white;
            border: 1px solid rgba(255,255,255,0.4);
            padding: 8px 18px;
            border-radius: 20px;
            cursor: pointer;
            font-size: 0.9rem;
            font-weight: 600;
            transition: all 0.2s;
        }}

        .lang-btn:hover,
        .lang-btn.active {{
            background: white;
            color: var(--primary);
            border-color: white;
        }}

        /* ── 통계 바 ── */
        .stats-bar {{
            background: white;
            padding: 16px 32px;
            display: flex;
            gap: 32px;
            border-bottom: 1px solid var(--border);
            flex-wrap: wrap;
        }}

        .stat-item {{
            text-align: center;
        }}

        .stat-value {{
            font-size: 1.6rem;
            font-weight: 700;
            color: var(--primary);
        }}

        .stat-label {{
            font-size: 0.8rem;
            color: var(--text-light);
        }}

        /* ── 필터 바 ── */
        .filter-bar {{
            padding: 16px 32px;
            background: white;
            border-bottom: 1px solid var(--border);
            display: flex;
            gap: 8px;
            flex-wrap: wrap;
            align-items: center;
        }}

        .filter-bar input {{
            flex: 1;
            min-width: 200px;
            padding: 8px 16px;
            border: 1px solid var(--border);
            border-radius: 20px;
            font-size: 0.9rem;
            outline: none;
        }}

        .filter-bar input:focus {{
            border-color: var(--secondary);
            box-shadow: 0 0 0 3px rgba(46,134,193,0.15);
        }}

        .sector-filter-btn {{
            padding: 6px 14px;
            border-radius: 16px;
            border: 1px solid var(--border);
            background: white;
            cursor: pointer;
            font-size: 0.8rem;
            transition: all 0.2s;
        }}

        .sector-filter-btn:hover,
        .sector-filter-btn.active {{
            background: var(--primary);
            color: white;
            border-color: var(--primary);
        }}

        /* ── 기사 카드 그리드 ── */
        .articles-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(360px, 1fr));
            gap: 20px;
            padding: 24px 32px;
            max-width: 1600px;
            margin: 0 auto;
        }}

        .article-card {{
            background: var(--card-bg);
            border-radius: 12px;
            box-shadow: var(--shadow);
            padding: 20px;
            border-left: 4px solid var(--secondary);
            transition: transform 0.2s, box-shadow 0.2s;
        }}

        .article-card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(0,0,0,0.12);
        }}

        .card-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 12px;
        }}

        .sector-badge {{
            font-size: 0.75rem;
            font-weight: 600;
            padding: 3px 10px;
            border-radius: 12px;
            background: var(--secondary);
            color: white;
        }}

        /* 섹터별 색상 */
        .sector-ww {{ background: #1abc9c; }}
        .sector-ws {{ background: #3498db; }}
        .sector-sw {{ background: #e67e22; }}
        .sector-pw {{ background: #e74c3c; }}
        .sector-og {{ background: #8e44ad; }}
        .sector-ip {{ background: #2ecc71; }}
        .sector-sc {{ background: #f39c12; }}
        .sector-tr {{ background: #16a085; }}
        .sector-cn {{ background: #7f8c8d; }}

        .article-date {{
            font-size: 0.8rem;
            color: var(--text-light);
        }}

        .article-title {{
            font-size: 1rem;
            font-weight: 600;
            margin-bottom: 10px;
            line-height: 1.5;
            color: var(--primary);
        }}

        .article-summary {{
            font-size: 0.88rem;
            color: #555;
            line-height: 1.6;
            margin-bottom: 14px;
            display: -webkit-box;
            -webkit-line-clamp: 4;
            -webkit-box-orient: vertical;
            overflow: hidden;
        }}

        .card-footer {{
            display: flex;
            align-items: center;
            gap: 12px;
            flex-wrap: wrap;
            font-size: 0.8rem;
            color: var(--text-light);
        }}

        .read-more {{
            margin-left: auto;
            color: var(--secondary);
            font-weight: 600;
            text-decoration: none;
        }}

        .read-more:hover {{ text-decoration: underline; }}

        .no-articles {{
            text-align: center;
            padding: 60px;
            color: var(--text-light);
            font-size: 1.1rem;
        }}

        /* ── 반응형 ── */
        @media (max-width: 768px) {{
            .dashboard-header {{ padding: 16px; }}
            .articles-grid {{ padding: 16px; grid-template-columns: 1fr; }}
            .stats-bar {{ padding: 12px 16px; gap: 16px; }}
            .filter-bar {{ padding: 12px 16px; }}
        }}
    </style>
</head>
<body>

<!-- ════ 헤더 ════════════════════════════════════════════ -->
<header class="dashboard-header">
    <div class="header-title">
        <h1>🇻🇳 Vietnam Infrastructure News</h1>
        <p>
            <span data-en="Automated Infrastructure Intelligence Dashboard"
                  data-ko="자동화 인프라 인텔리전스 대시보드"
                  data-vi="Bảng tin hạ tầng tự động">
                Automated Infrastructure Intelligence Dashboard
            </span>
            &nbsp;|&nbsp; Updated: {today}
        </p>
    </div>

    <!-- ★ 3개국어 전환 버튼 -->
    {LANGUAGE_BUTTONS_HTML}
</header>

<!-- ════ 통계 바 ════════════════════════════════════════ -->
<div class="stats-bar">
    <div class="stat-item">
        <div class="stat-value">{len(articles)}</div>
        <div class="stat-label"
             data-en="Total Articles"
             data-ko="전체 기사"
             data-vi="Tổng bài viết">Total Articles</div>
    </div>
    <div class="stat-item">
        <div class="stat-value">{len(today_articles)}</div>
        <div class="stat-label"
             data-en="Today"
             data-ko="오늘"
             data-vi="Hôm nay">Today</div>
    </div>
    <div class="stat-item">
        <div class="stat-value">{len(sectors)}</div>
        <div class="stat-label"
             data-en="Sectors"
             data-ko="섹터"
             data-vi="Lĩnh vực">Sectors</div>
    </div>
</div>

<!-- ════ 필터 바 ══════════════════════════════════════════ -->
<div class="filter-bar">
    <input type="text"
           id="search-input"
           placeholder="Search articles..."
           oninput="searchArticles(this.value)">
    {sector_filter_html}
</div>

<!-- ════ 기사 카드 그리드 ═══════════════════════════════ -->
<main class="articles-grid" id="articles-container">
    {cards_html if cards_html.strip() else '<div class="no-articles">No articles available today.</div>'}
</main>

<!-- ════ 백엔드 데이터 (JavaScript용) ══════════════════════ -->
<!-- 
    BACKEND_DATA: 전체 기사 데이터를 JavaScript 객체로 저장
    title/summary는 각 언어별 객체로 구성되어 언어 전환 시 사용
-->
<script>
const BACKEND_DATA = {backend_data_json};
</script>

<!-- ════ 3개국어 전환 JavaScript ════════════════════════ -->
{LANGUAGE_SWITCH_JS}

<!-- ════ 페이지 로드 확인 ═════════════════════════════════ -->
<script>
console.log('[Dashboard] 로드 완료');
console.log('[Dashboard] 기사 수:', BACKEND_DATA.length);
console.log('[Dashboard] 다국어 요소 수:', document.querySelectorAll('[data-en]').length);
</script>

</body>
</html>"""

    return html_content


# ════════════════════════════════════════════════════════════
# DashboardUpdater 클래스 (main.py에서 호출)
# ════════════════════════════════════════════════════════════

class DashboardUpdater:
    """대시보드 생성 및 저장 관리 클래스"""

    def generate(self, articles: list) -> str:
        """
        대시보드 HTML 생성 및 저장
        
        Args:
            articles: processed_articles (번역 완료본)
                      반드시 title_en/ko/vi, summary_en/ko/vi 필드 포함!
        Returns:
            생성된 파일 경로 문자열
        """
        logger.info(f"[DashboardUpdater] 기사 {len(articles)}건으로 대시보드 생성 시작")

        # 번역 필드 검증
        sample = articles[0] if articles else {}
        has_multilang = all(
            k in sample
            for k in ['title_en', 'title_ko', 'title_vi', 'summary_en', 'summary_ko', 'summary_vi']
        )
        if not has_multilang:
            logger.warning(
                "[DashboardUpdater] 다국어 필드 누락! "
                "main.py에서 step2_translate 후 이 함수를 호출했는지 확인하세요."
            )

        # HTML 생성
        html_str = generate_dashboard_html(articles)

        # 출력 디렉토리 생성
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

        # HTML 파일 저장
        with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
            f.write(html_str)

        logger.info(f"[DashboardUpdater] 저장 완료: {OUTPUT_PATH}")
        return str(OUTPUT_PATH)

    def generate_from_excel(self) -> str:
        """Excel 데이터에서 직접 대시보드 생성 (dashboard-only 모드)"""
        articles = self._load_excel_articles()
        return self.generate(articles)

    def _load_excel_articles(self) -> list:
        """Excel 파일에서 기사 데이터 로드"""
        try:
            import openpyxl
            if not EXCEL_PATH.exists():
                logger.error(f"Excel 파일 없음: {EXCEL_PATH}")
                return []

            wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            articles = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                article = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        article[headers[i]] = val
                articles.append(article)

            wb.close()

            # 컬럼명 정규화 (Excel 헤더 → 내부 키 변환)
            return [self._normalize_article(a) for a in articles]

        except Exception as e:
            logger.error(f"Excel 로드 실패: {e}", exc_info=True)
            return []

    def _normalize_article(self, raw: dict) -> dict:
        """Excel 헤더명 → 내부 표준 필드명으로 변환"""
        return {
            'title':      raw.get('Title') or raw.get('title') or '',
            'title_en':   raw.get('title_en') or raw.get('Title') or '',
            'title_ko':   raw.get('title_ko') or '',
            'title_vi':   raw.get('title_vi') or raw.get('Title') or '',
            'summary_en': raw.get('summary_en') or raw.get('Short summary') or raw.get('Summary') or '',
            'summary_ko': raw.get('summary_ko') or '',
            'summary_vi': raw.get('summary_vi') or '',
            'sector':     raw.get('Sector') or raw.get('sector') or 'General',
            'source':     raw.get('Source') or raw.get('source') or '',
            'date':       str(raw.get('Date') or raw.get('date') or ''),
            'url':        raw.get('Link') or raw.get('URL') or raw.get('url') or '#',
            'area':       raw.get('Area') or raw.get('area') or '',
        }
