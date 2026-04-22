#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
batch_translate.py  v2.0
Vietnam Infra News — 배치 번역 스크립트

변경사항 (v2.0):
  - 번역 우선순위: 2025년 이후 → 2024년 → 이하 순
  - Jina.ai Reader API를 통한 차단 소스 본문 fetch 시도
  - NewsData.io domain 직접 쿼리 보완 로직
  - MyMemory + deep-translator 이중 fallback 유지
  - 배치 크기 20건 → 40건 (속도 향상)

[영구 제약]
  - Google Translate only (MyMemory primary + deep-translator secondary)
  - Anthropic API 금지 in GitHub Actions
"""

import os, sys, time, json, logging
from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    from openpyxl import load_workbook

try:
    from deep_translator import GoogleTranslator
    HAS_DEEP = True
except ImportError:
    HAS_DEEP = False

import urllib.request, urllib.parse

# ── 경로 설정 ────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent
DATA_DIR   = BASE_DIR / 'data'
EXCEL_PATH = Path(os.environ.get('EXCEL_PATH', '')) or \
             DATA_DIR / 'database' / 'Vietnam_Infra_News_Database_Final.xlsx'

BATCH_SIZE  = 40   # 배치당 번역 건수 (v2.0: 20→40)
SLEEP_SEC   = 0.25 # 요청 간격

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
log = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════
# 번역 함수 (MyMemory primary → deep-translator fallback)
# ════════════════════════════════════════════════════════
def detect_lang(text: str) -> str:
    """VI 문자 포함 여부로 언어 감지"""
    vi_chars = 'àáâãèéêìíòóôõùúýăđơưạảấầẩẫậắằẳẵặẹẻẽếềểễệỉịọỏốồổỗộớờởỡợụủứừửữựỳỷỹ'
    return 'vi' if any(c in str(text) for c in vi_chars) else 'en'


def mymemory_translate(text: str, src: str = 'en', tgt: str = 'ko',
                       max_len: int = 450) -> Optional[str]:
    """MyMemory API (무료, 50만 단어/월)"""
    if not text or len(str(text).strip()) < 3:
        return None
    try:
        t = str(text)[:max_len]
        params = urllib.parse.urlencode({
            'q': t, 'langpair': f'{src}|{tgt}', 'de': 'vietnam_infra@mi.vn'
        })
        url = f'https://api.mymemory.translated.net/get?{params}'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=8) as r:
            data = json.loads(r.read())
            if data.get('responseStatus') == 200:
                result = data['responseData']['translatedText']
                # MyMemory 실패 문자열 필터
                if result and 'MYMEMORY WARNING' not in result:
                    return result.strip()
    except Exception:
        pass
    return None


def deep_translate(text: str, src: str = 'auto', tgt: str = 'ko',
                   max_len: int = 450) -> Optional[str]:
    """deep-translator (Google Translate 백엔드)"""
    if not HAS_DEEP or not text or len(str(text).strip()) < 3:
        return None
    try:
        t = str(text)[:max_len]
        result = GoogleTranslator(source=src, target=tgt).translate(t)
        return result.strip() if result else None
    except Exception:
        return None


def translate(text: str, src: str = 'auto', tgt: str = 'ko') -> str:
    """번역 메인 함수: MyMemory → deep-translator fallback"""
    if not text or len(str(text).strip()) < 3:
        return str(text)

    src_detected = detect_lang(text) if src == 'auto' else src

    # 1차: MyMemory
    result = mymemory_translate(text, src=src_detected, tgt=tgt)
    if result and result != str(text):
        return result

    time.sleep(SLEEP_SEC)

    # 2차: deep-translator (Google)
    result = deep_translate(text, src=src_detected, tgt=tgt)
    if result and result != str(text):
        return result

    return str(text)  # 번역 실패 시 원문 반환


# ════════════════════════════════════════════════════════
# Jina.ai Reader — 차단 소스 본문 fetch 우회
# ════════════════════════════════════════════════════════
def jina_fetch(url: str, max_chars: int = 800) -> Optional[str]:
    """
    Jina.ai Reader API로 차단된 페이지 본문 가져오기
    사용법: r.jina.ai/https://theinvestor.vn/article-url
    """
    try:
        jina_url = f"https://r.jina.ai/{url}"
        req = urllib.request.Request(jina_url, headers={
            'User-Agent': 'Mozilla/5.0',
            'Accept': 'text/plain',
        })
        with urllib.request.urlopen(req, timeout=12) as r:
            content = r.read().decode('utf-8', errors='ignore')
            # 첫 800자만 요약용으로 사용
            return content[:max_chars].strip()
    except Exception:
        return None


def newsdata_domain_query(domain: str, keyword: str,
                          api_key: str = '') -> list:
    """
    NewsData.io domain 직접 쿼리
    차단된 전문미디어 기사 보완 수집
    주의: /api/1/latest 엔드포인트만 사용 (from_date 금지)
    """
    if not api_key:
        api_key = os.environ.get('NEWSDATA_API_KEY', '')
    if not api_key:
        return []

    params = urllib.parse.urlencode({
        'apikey': api_key,
        'q': keyword,
        'country': 'vn',
        'language': 'en',
        # domain 파라미터는 422 오류 유발 가능 → q에 포함
    })
    url = f"https://newsdata.io/api/1/latest?{params}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            return data.get('results', [])
    except Exception as e:
        log.warning("NewsData.io 쿼리 실패 (%s): %s", keyword, e)
        return []


# ════════════════════════════════════════════════════════
# 메인 배치 번역 실행
# ════════════════════════════════════════════════════════
def run_batch_translate(excel_path: Optional[str] = None,
                        batch_size: int = BATCH_SIZE,
                        priority_year: int = 2024) -> dict:
    """
    Excel DB에서 미번역 기사를 우선순위별로 배치 번역

    우선순위:
      1. 2025년 이후 기사 (최신)
      2. 2024년 기사
      3. 2023년 이하 기사

    Returns: 번역 결과 통계
    """
    path = Path(excel_path) if excel_path else EXCEL_PATH
    if not path.exists():
        log.error("Excel 파일 없음: %s", path)
        return {}

    log.info("Excel 로드: %s", path)
    wb = load_workbook(str(path))

    # News Database 시트 탐색
    sheet_name = next(
        (s for s in ('News Database', 'All_Articles') if s in wb.sheetnames),
        None
    )
    if not sheet_name:
        log.error("News Database 시트 없음")
        return {}

    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col_of(name):
        return headers.index(name) + 1 if name in headers else None

    date_col      = col_of('Date')
    title_col     = col_of('News Title')
    title_ko_col  = col_of('title_ko')
    title_en_col  = col_of('title_en')
    title_vi_col  = col_of('title_vi')
    summ_col      = col_of('Short Summary')
    summ_ko_col   = col_of('summary_ko')
    summ_en_col   = col_of('summary_en')

    if not all([date_col, title_col, title_ko_col]):
        log.error("필수 컬럼 없음")
        return {}

    # 미번역 행 수집 (우선순위 정렬)
    pending = []
    for ri in range(2, ws.max_row + 1):
        ko_val = ws.cell(ri, title_ko_col).value
        if ko_val:  # 이미 번역됨
            continue
        date_raw = ws.cell(ri, date_col).value
        try:
            year = int(str(date_raw)[:4])
        except Exception:
            year = 0
        priority = (
            0 if year >= 2025 else
            1 if year == 2024 else
            2
        )
        pending.append((priority, year, ri))

    # 우선순위 정렬 후 배치만큼 처리
    pending.sort()
    to_process = pending[:batch_size]

    log.info("미번역 총 %d건 | 이번 배치: %d건 (우선순위: %d년 이후)",
             len(pending), len(to_process), priority_year)

    stats = {
        'total_pending': len(pending),
        'processed': 0,
        'success': 0,
        'fail': 0,
        'today': datetime.now().strftime('%Y-%m-%d'),
    }

    YELLOW = PatternFill("solid", fgColor="FFF9C4")

    for idx, (priority, year, ri) in enumerate(to_process):
        title_raw  = str(ws.cell(ri, title_col).value or '').strip()
        summ_raw   = str(ws.cell(ri, summ_col).value or '').strip() if summ_col else ''

        if not title_raw:
            stats['fail'] += 1
            continue

        src = detect_lang(title_raw)

        # 제목 번역 (KO + VI)
        title_ko = translate(title_raw, src=src, tgt='ko')
        time.sleep(SLEEP_SEC)
        title_vi = translate(title_raw, src=src, tgt='vi')
        time.sleep(SLEEP_SEC)

        # 요약 번역
        summ_ko = translate(summ_raw, src=src, tgt='ko') if summ_raw else ''
        time.sleep(SLEEP_SEC)
        summ_vi = translate(summ_raw, src=src, tgt='vi') if summ_raw else ''
        time.sleep(SLEEP_SEC)

        # 셀에 기입
        ws.cell(ri, title_ko_col, title_ko).fill = YELLOW
        if title_en_col:
            ws.cell(ri, title_en_col, title_raw)
        if title_vi_col:
            ws.cell(ri, title_vi_col, title_vi).fill = YELLOW
        if summ_ko_col:
            ws.cell(ri, summ_ko_col, summ_ko).fill = YELLOW
        if summ_en_col:
            ws.cell(ri, summ_en_col, summ_raw)

        ok = title_ko != title_raw
        stats['processed'] += 1
        stats['success' if ok else 'fail'] += 1

        log.info("[%d/%d] %s %s → %s",
                 idx + 1, len(to_process),
                 "✅" if ok else "❌",
                 title_raw[:40],
                 title_ko[:40] if ok else "(번역 실패)")

    # 저장
    wb.save(str(path))
    log.info("저장 완료: %s", path)
    log.info("결과: 처리 %d건 / 성공 %d건 / 실패 %d건",
             stats['processed'], stats['success'], stats['fail'])
    log.info("잔여 미번역: %d건", stats['total_pending'] - stats['processed'])

    # Collection_Log 업데이트
    _update_collection_log(wb, stats, sheet_name)
    wb.save(str(path))

    return stats


def _update_collection_log(wb, stats: dict, ref_sheet: str) -> None:
    """Collection_Log 시트에 번역 결과 기록"""
    if 'Collection_Log' not in wb.sheetnames:
        return
    ws_log = wb['Collection_Log']
    last_row = ws_log.max_row
    ws_log.cell(last_row + 1, 1, stats.get('today', ''))
    ws_log.cell(last_row + 1, 10, stats.get('success', 0))   # Translation OK
    ws_log.cell(last_row + 1, 11, stats.get('fail', 0))      # Translation Fail


# ════════════════════════════════════════════════════════
# HN 전용 NewsData.io 쿼리 (④ 전문기사 우회)
# ════════════════════════════════════════════════════════
HN_NEWSDATA_QUERIES = [
    # HN-URBAN-NORTH: 동아인·BRG
    "Dong Anh smart city Hanoi BRG Sumitomo",
    "Co Loa urban development Hanoi 2025",
    "Noi Bai airport terminal 3 expansion",
    "Me Linh urban development Hanoi FDI",
    # HN-URBAN-WEST: 호아락
    "Hoa Lac hi-tech park expansion 2025 2026",
    "Hoa Lac silicon valley Vietnam innovation",
    "Xuan Mai son Tay Hanoi western development",
    # HN-URBAN-INFRA: 링로드4·메트로
    "Hanoi Ring Road 4 capital region 2026",
    "Hanoi metro line 2 3 4 construction 2025",
    "To Lich river cleanup Yen Xa wastewater Hanoi",
    # VN-WW-2030
    "wastewater treatment plant Vietnam ODA JICA 2025",
    # VN-SWM-NATIONAL-2030
    "waste-to-energy Vietnam incineration plant 2025",
    # VN-PWR-PDP8
    "Vietnam PDP8 offshore wind solar decree 2025",
    "Vietnam LNG power plant Nhon Trach decree 100",
]

def collect_hn_articles(api_key: str = '') -> list:
    """HN 도시개발 전용 NewsData.io 수집"""
    results = []
    for query in HN_NEWSDATA_QUERIES:
        arts = newsdata_domain_query('', query, api_key)
        for a in arts:
            results.append({
                'title': a.get('title', ''),
                'source': a.get('source_name', ''),
                'date': (a.get('pubDate', '') or '')[:10],
                'url': a.get('link', ''),
                'summary': a.get('description', '') or a.get('content', ''),
                'query': query,
            })
        time.sleep(0.5)  # Rate limit
    log.info("HN NewsData.io 수집: %d건", len(results))
    return results


# ════════════════════════════════════════════════════════
# CLI
# ════════════════════════════════════════════════════════
if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Batch Translate v2.0')
    parser.add_argument('--batch', type=int, default=BATCH_SIZE,
                        help=f'배치 크기 (기본: {BATCH_SIZE})')
    parser.add_argument('--excel', type=str, help='Excel 파일 경로')
    parser.add_argument('--collect-hn', action='store_true',
                        help='HN 전용 NewsData.io 수집 실행')
    parser.add_argument('--test-jina', type=str,
                        help='Jina.ai fetch 테스트 URL')
    args = parser.parse_args()

    if args.test_jina:
        log.info("Jina.ai fetch 테스트: %s", args.test_jina)
        content = jina_fetch(args.test_jina)
        print(content[:500] if content else "fetch 실패")
        sys.exit(0)

    if args.collect_hn:
        api_key = os.environ.get('NEWSDATA_API_KEY', '')
        arts = collect_hn_articles(api_key)
        print(f"HN 수집: {len(arts)}건")
        for a in arts[:5]:
            print(f"  {a['date']} | {a['title'][:60]}")
        sys.exit(0)

    excel_path = args.excel or os.environ.get('EXCEL_PATH') or str(EXCEL_PATH)
    run_batch_translate(excel_path=excel_path, batch_size=args.batch)
