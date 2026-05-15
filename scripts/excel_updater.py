"""
excel_updater.py — v4.0 (2026-05-15)
====================================================
[v4.0 변경사항 2가지]

  ★ FIX 1: Power 섹터 과분류 수정 (54% → 예상 18%)
    원인: _filter_and_enrich()에서 sector 미판별 시 기본값 'Power'
    수정A: _sector_from_text() 반환값 '' → 'General' (or 'Power' 오류 차단)
    수정B: _filter_and_enrich() 기본값 or 'Power' → or 'General'
    수정C: _sector_from_text() 키워드 강화
           - Power: 'power'/'solar'/'wind' 단독 → 복합어만 허용
           - Oil&Gas: 'oil'/'gas' 단독 → 복합어만 허용
           - Power/Oil&Gas: Vietnam 기사일 때만 분류 (해외 에너지 차단)

  ★ FIX 2: Matched_Plan 최신 기사 상단 표시
    원인: 신규 기사를 시트 마지막 행에 append → 구기사 상단, 신기사 하단
    수정: 신규 기사를 날짜 역순 정렬 후 헤더(Row2) 바로 아래 insert_rows(3)
          → News Database와 동일한 최신-우선 구조

[v3.9 핵심 수정] Matched_Plan 과팽창 버그 수정 (URL 3중 복합키)
[v3.8 핵심 수정] UnboundLocalError + MergedCell 버그 수정
[영구 제약] ExcelUpdater / update_all(articles) / insert_rows(2) — 변경 금지
"""

import os
import re
import tempfile
import shutil
from datetime import datetime, date, timedelta
from collections import defaultdict
from pathlib import Path
from itertools import groupby

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def _safe_set_merged_cell(ws, row, col, value):
    cell = ws.cell(row, col)
    merged_ranges = [mc for mc in ws.merged_cells.ranges if cell.coordinate in mc]
    if merged_ranges:
        for mc in merged_ranges:
            ws.unmerge_cells(str(mc))
        cell.value = value
        for mc in merged_ranges:
            ws.merge_cells(str(mc))
    else:
        cell.value = value


_VI_CHARS = set('ăâđêôơưĂÂĐÊÔƠƯáàảãạắằẳẵặấầẩẫậéèẻẽẹếềểễệíìỉĩịóòỏõọốồổỗộớờởỡợúùủũụứừửữựýỳỷỹỵ')
def _is_vietnamese(text: str) -> bool:
    if not text: return False
    return sum(1 for c in str(text) if c in _VI_CHARS) > 2

_SCRIPTS_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
_ROOT_DIR    = _SCRIPTS_DIR.parent
EXCEL_PATH   = Path(os.environ.get(
    "EXCEL_PATH",
    str(_ROOT_DIR / "data" / "database" / "Vietnam_Infra_News_Database_Final.xlsx")
))

SECTOR_ORDER = ["Waste Water","Water Supply/Drainage","Solid Waste",
                "Power","Oil & Gas","Transport","Industrial Parks",
                "Smart City","Construction"]

PLAN_PREFIX = {
    "VN-WW":"Waste Water","VN-WAT":"Water Supply/Drainage","VN-SWM":"Solid Waste",
    "VN-PWR":"Power","VN-PDP":"Power","VN-OG":"Oil & Gas",
    "VN-TRAN":"Transport","VN-URB":"Transport","VN-METRO":"Transport",
    "VN-IP":"Industrial Parks","VN-MEKONG":"Transport","HN":"Transport",
    "VN-EV":"Transport","VN-ENV":"Environment",
}
AREA_MAP = {
    "Waste Water":"Environment","Water Supply/Drainage":"Environment","Solid Waste":"Environment",
    "Power":"Energy Develop.","Oil & Gas":"Energy Develop.",
    "Transport":"Urban Develop.","Industrial Parks":"Urban Develop.",
    "Smart City":"Urban Develop.","Construction":"Urban Develop.",
    "Environment":"Environment","General":"Urban Develop.",
}

INFRA_KEYWORDS = {
    'en': ['wastewater','sewage','water supply','drainage','waste management',
           'power','electricity','energy','solar','wind','lng','hydropower',
           'oil','gas','petroleum','pipeline','refinery',
           'transport','highway','airport','railway','metro','bridge','port','seaport',
           'industrial park','industrial zone','smart city','infrastructure',
           'master plan','development plan','construction','urban','wwtp',
           'ring road','expressway','renewable','grid','transmission'],
    'vi': ['nước thải','xử lý nước','cấp nước','thoát nước','chất thải','môi trường',
           'điện','năng lượng','dầu khí','khí đốt','giao thông','đường','cầu',
           'sân bay','metro','khu công nghiệp','khu kinh tế','đô thị thông minh',
           'cơ sở hạ tầng','quy hoạch','xây dựng','cảng','đường sắt'],
    'ko': ['폐수','상수도','하수도','전력','에너지','태양광','풍력','석유','가스',
           '교통','도로','공항','철도','지하철','산업단지','스마트시티','인프라']
}

def _f(h): return PatternFill("solid", fgColor=h)
FILL = {
    "HIGH":   _f("FFF9C4"), "MEDIUM": _f("E8F0FE"), "POLICY": _f("E8F5E9"),
    "BOTH":   _f("FFF3E0"), "WHITE":  _f("FFFFFF"), "EVEN":   _f("F5F9FF"),
    "GRAY":   _f("F2F2F2"), "RED":    _f("FFEBEE"), "HDR_B":  _f("1F4E78"),
    "HDR_G":  _f("375623"), "HDR_N":  _f("17375E"), "KH_NEW": _f("FFF9C4"),
}

FONT_HDR        = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
FONT_DATA       = Font(name="맑은 고딕", size=10, color="000000")
FONT_TITLE_BOLD = Font(name="맑은 고딕", bold=True, size=11)
FONT_META       = Font(name="맑은 고딕", bold=True, size=11, color="000000")

NEWS_HEADERS = ["Area","Sector","No","Date","Title_EN","Title_VI","Tit_ko",
                "Source","Src_Type","Province","Plan_ID","Grade","URL",
                "sum_ko","sum_en","sum_vi","QC"]
NEWS_WIDTHS  = [16,18,5,12,45,40,30,18,10,15,20,8,40,30,30,30,8]

MP_HEADERS = ["No","ctx_tag","ctx_grade","Plan_ID","Title_EN","Date","Source",
              "Province","Sector","title_ko","title_en_orig","title_vi",
              "summary_ko","summary_en","summary_vi","short_sum","Link"]
MP_WIDTHS  = [5,12,10,22,40,12,18,15,15,30,30,30,40,40,40,40,30]

KH_HEADERS = ["Sector","Province","Date","Title (En)","Title (Ko)","Source","Grade","Plan_ID"]
KH_WIDTHS  = [22,18,12,55,40,18,8,22]

SECT_PRI = {s:i for i,s in enumerate(SECTOR_ORDER)}


def _is_infra_article(title_en: str, title_ko: str, plan_id: str) -> bool:
    if str(plan_id or '').strip():
        return True
    text = (str(title_en or '') + ' ' + str(title_ko or '')).lower()
    if any(kw in text for kw in INFRA_KEYWORDS['en']): return True
    if any(kw in text for kw in INFRA_KEYWORDS['vi']): return True
    if any(kw in text for kw in INFRA_KEYWORDS['ko']): return True
    return False

def _sector_from_plan(plan_id: str) -> str:
    p = str(plan_id or '').upper().strip()
    for pf, s in PLAN_PREFIX.items():
        if p.startswith(pf): return s
    return ''

# ★ v4.0: _sector_from_text() 전면 강화
# 핵심 변경:
#   1) 'power'/'solar'/'wind'/'oil'/'gas' 단독 키워드 제거 → 복합어로 대체
#   2) Power/Oil&Gas: Vietnam 관련 기사일 때만 분류 (해외 에너지 기사 차단)
#   3) 반환값 '' → 'General' (or 'Power' 오류 원천 차단)
def _sector_from_text(title_en: str, title_ko: str, plan_id: str) -> str:
    s = _sector_from_plan(plan_id)
    if s: return s

    txt = (str(title_en or '') + ' ' + str(title_ko or '')).lower()

    # 베트남 관련 기사인지 확인 (Power/Oil&Gas 해외 기사 차단용)
    is_vietnam = any(kw in txt for kw in [
        'vietnam', 'viet nam', 'việt nam', 'hanoi', 'hà nội', 'ha noi',
        'ho chi minh', 'hcmc', 'saigon', 'haiphong', 'mekong', 'đà nẵng',
        'da nang', 'can tho', 'cần thơ', 'quang ninh', 'bình dương',
    ])

    rules = [
        ("Waste Water", [
            "wastewater", "sewage", "wwtp", "nước thải", "effluent",
            "sewerage", "xử lý nước thải", "thoát nước",
        ]),
        ("Water Supply/Drainage", [
            "water supply", "water treatment plant", "cấp nước", "nước sạch",
            "waterworks", "water pipe", "water network",
        ]),
        ("Solid Waste", [
            "solid waste", "waste management", "landfill", "waste-to-energy",
            "wte plant", "incineration plant", "rác thải", "chất thải rắn",
            "đốt rác", "epr recycling", "thu gom rác",
        ]),
        # ★ Power: 단독 'power'/'solar'/'wind' 제거, 복합어 위주
        ("Power", [
            "power plant", "power grid", "power development", "power capacity",
            "electricity grid", "electricity capacity", "electricity sector",
            "rooftop solar", "solar farm", "solar power", "solar capacity",
            "floating solar", "wind farm", "wind power", "wind turbine",
            "offshore wind", "renewable energy", "pdp8", "lng power",
            "lng power plant", "nuclear power", "battery storage",
            "energy storage", "bess", "dppa", "power purchase agreement",
            "điện mặt trời", "điện gió", "nhà máy điện", "lưới điện",
            "năng lượng tái tạo", "quy hoạch điện", "quyết định 768",
        ]),
        # ★ Oil&Gas: 단독 'oil'/'gas' 제거, 복합어 위주
        ("Oil & Gas", [
            "oil field", "oil refinery", "oil production", "crude oil vietnam",
            "petroleum", "lng terminal", "lng import", "gas pipeline",
            "petrovietnam", "pvn", "pv gas",
            "dầu khí", "khí đốt", "lọc dầu",
        ]),
        ("Industrial Parks", [
            "industrial park", "industrial zone", "economic zone",
            "export processing zone", "vsip", "deep c",
            "khu công nghiệp", "khu kinh tế", "khu chế xuất",
        ]),
        ("Smart City", [
            "smart city", "urban rail", "metro line", "digital city",
            "đô thị thông minh", "tàu điện ngầm",
        ]),
        ("Transport", [
            "expressway", "highway", "airport", "seaport", "logistics hub",
            "metro", "ring road", "north-south", "long thanh",
            "cao tốc", "sân bay", "cảng biển", "đường cao tốc", "đường sắt",
        ]),
    ]

    for sec, kws in rules:
        if any(kw in txt for kw in kws):
            # ★ Power/Oil&Gas: 베트남 기사일 때만 분류
            if sec in ('Power', 'Oil & Gas') and not is_vietnam:
                return 'General'
            return sec

    return 'General'   # ★ '' 대신 'General' — or 'Power' 오류 원천 차단


def _grade_fill(grade: str, qc: str = '') -> PatternFill:
    g = str(grade or '').upper().strip()
    q = str(qc    or '').upper().strip()
    p = ('POLICY' in q) or (g == 'POLICY')
    if g == 'HIGH' and p: return FILL['BOTH']
    if g == 'HIGH':       return FILL['HIGH']
    if g == 'MEDIUM':     return FILL['MEDIUM']
    if g == 'POLICY' or p:return FILL['POLICY']
    return FILL['WHITE']

def _hdr(ws, row, col, val, fill_key="HDR_B"):
    c = ws.cell(row, col)
    c.value = val; c.font = FONT_HDR; c.fill = FILL[fill_key]
    c.alignment = Alignment(horizontal="center", vertical="center")


class ExcelUpdater:
    def __init__(self, excel_path=EXCEL_PATH):
        self.path   = Path(excel_path)
        self.today  = date.today().strftime("%Y-%m-%d")
        self.now    = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.cutoff = (date.today() - timedelta(days=7)).strftime("%Y-%m-%d")

    def update_all(self, articles: list) -> None:
        print(f"[ExcelUpdater v4.0] update_all: 신규 {len(articles)}건")

        if not self.path.exists():
            self._create_empty_workbook()

        import warnings
        warnings.filterwarnings('ignore', message=".*MergedCell.*")

        try:
            wb = openpyxl.load_workbook(str(self.path))
        except Exception as e:
            if 'MergedCell' in str(e) or 'read-only' in str(e):
                print(f"[WARNING] 워크북 로드 MergedCell 오류 무시: {str(e)[:80]}")
                wb = openpyxl.load_workbook(str(self.path), data_only=False)
            else:
                raise

        articles = self._filter_and_enrich(articles)
        print(f"  인프라 관련 필터링 후: {len(articles)}건")

        new_arts = self._deduplicate(wb, articles)
        print(f"  중복 제거 후 신규: {len(new_arts)}건")

        if new_arts:
            self._insert_news(wb, new_arts)

        self._rebuild_matched_plan(wb)
        self._rebuild_keywords_history(wb)
        self._update_summary(wb, len(new_arts))
        self._update_collection_log(wb, len(new_arts))
        self._update_source(wb)
        self._update_stats(wb)
        self._update_context_stats(wb)
        self._update_timeline(wb)
        self._update_province_keywords(wb)

        try:
            if 'Matched_Plan' in wb.sheetnames:
                matched_plan_sheet = wb['Matched_Plan']
                merged_ranges_backup = []
                for merged_range in list(matched_plan_sheet.merged_cells.ranges):
                    merged_ranges_backup.append(str(merged_range))
                    try:
                        matched_plan_sheet.unmerge_cells(str(merged_range))
                    except Exception:
                        pass
                print(f"  [MergedCell] {len(merged_ranges_backup)}개 제거 후 저장")
            else:
                merged_ranges_backup = []

            wb.save(str(self.path))
            print(f"  [완료] Excel 저장: {self.path}")

            if merged_ranges_backup:
                wb_reload = openpyxl.load_workbook(str(self.path), data_only=False)
                ws_reload = wb_reload['Matched_Plan']
                for merged_range_str in merged_ranges_backup:
                    try:
                        ws_reload.merge_cells(merged_range_str)
                    except Exception:
                        pass
                wb_reload.save(str(self.path))
                print(f"  [완료] MergedCell 복구 완료: {self.path}")

        except Exception as e:
            print(f"[WARNING] Excel 저장 오류 발생했으나 계속 진행: {str(e)[:120]}")

    def _filter_and_enrich(self, articles: list) -> list:
        result = []
        for a in articles:
            a = dict(a)
            title    = a.get("title") or a.get("title_en", "")
            title_ko = a.get("title_ko", "")
            plan_id  = a.get("plan_id", "") or a.get("ctx_plans", "")

            if not _is_infra_article(title, title_ko, plan_id):
                continue

            if not a.get("sector"):
                # ★ v4.0: or 'Power' → or 'General' (과분류 원천 차단)
                a["sector"] = _sector_from_text(title, title_ko, plan_id) or 'General'
            if not a.get("area"):
                a["area"] = AREA_MAP.get(a["sector"], "Urban Develop.")

            qc = str(a.get("qc","") or "").upper()
            cg = str(a.get("grade","") or a.get("ctx_grade","")).upper()
            if cg in ('HIGH','MEDIUM','POLICY','LOW'):
                a['grade'] = cg
            elif 'SA7+POLICY' in qc: a['grade'] = 'HIGH'
            elif 'SA7_MATCH' in qc:  a['grade'] = 'MEDIUM'
            elif 'POLICY_MATCH' in qc:a['grade'] = 'POLICY'
            else:                     a['grade'] = ''

            result.append(a)
        return result

    def _deduplicate(self, wb, articles):
        if "News Database" not in wb.sheetnames: return articles
        ws = wb["News Database"]
        eu, et = set(), set()
        for r in range(2, ws.max_row+1):
            u  = ws.cell(r,13).value
            t  = ws.cell(r,5).value
            t2 = ws.cell(r,6).value
            if u: eu.add(str(u).strip())
            if t: et.add(str(t)[:80].strip())
            if t2: et.add(str(t2)[:80].strip())
        return [a for a in articles
                if str(a.get("url","") or "").strip() not in eu
                and str(a.get("title") or a.get("title_en",""))[:80].strip() not in et]

    def _insert_news(self, wb, articles):
        if "News Database" not in wb.sheetnames:
            ws = wb.create_sheet("News Database", 0)
            self._write_news_header(ws)
        else:
            ws = wb["News Database"]
            cur_headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
            if cur_headers != NEWS_HEADERS:
                for ci, h in enumerate(NEWS_HEADERS, 1):
                    _hdr(ws, 1, ci, h)

        for a in sorted(articles, key=lambda x: str(x.get("date","") or ""), reverse=True):
            ws.insert_rows(2)
            self._write_news_row(ws, 2, a)

        self._renumber(ws)
        print(f"    [News DB] {len(articles)}건 삽입 → 총 {ws.max_row-1}건")

    def _write_news_header(self, ws):
        for ci, (h, w) in enumerate(zip(NEWS_HEADERS, NEWS_WIDTHS), 1):
            _hdr(ws, 1, ci, h)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def _write_news_row(self, ws, row, a):
        title    = a.get("title") or a.get("title_en", "")
        title_ko = a.get("title_ko", "")
        grade    = str(a.get("grade","") or "")
        qc       = str(a.get("qc","") or "")
        plan_id  = str(a.get("plan_id","") or a.get("ctx_plans","") or "")
        d        = str(a.get("date","") or "")
        sector   = a.get("sector", "")
        area     = a.get("area", "")

        title_en = a.get("title_en", "")
        title_vi = a.get("title_vi", "")
        if not title_en and not title_vi and title:
            if _is_vietnamese(title): title_vi = title
            else:                     title_en = title

        fill = _grade_fill(grade, qc)
        vals = [
            area, sector, "", d, title_en, title_vi, title_ko,
            a.get("source",""), a.get("src_type","News"), a.get("province",""),
            plan_id, grade, a.get("url",""),
            a.get("sum_ko","") or a.get("summary_ko",""),
            a.get("sum_en","") or a.get("summary_en",""),
            a.get("sum_vi","") or a.get("summary_vi",""),
            qc,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row, ci); c.value = v; c.fill = fill
            c.font = FONT_TITLE_BOLD if ci in (5,6,7,14,15,16) else FONT_DATA
            c.alignment = Alignment(vertical="top", wrap_text=False)

    def _renumber(self, ws):
        for r in range(2, ws.max_row+1):
            try: ws.cell(r,3).value = r-1
            except Exception: pass

    def _rebuild_matched_plan(self, wb):
        """
        ★ v4.0: 신규 기사를 날짜 역순 정렬 후 Row3(헤더 바로 아래)에 insert
                 → Matched_Plan이 항상 최신 기사 상단 구조 유지
        ★ v3.9: 3중 복합키 중복 체크 유지
        """
        ws_news = wb["News Database"]
        existing_links       = set()
        existing_plan_titles = set()
        existing_plan_dates  = set()
        mp_max_row = 2

        if "Matched_Plan" not in wb.sheetnames:
            nidx  = wb.sheetnames.index("News Database")
            ws_mp = wb.create_sheet("Matched_Plan", nidx+1)
            meta  = "★ SA-7 맥락 확정 기사 0건 | HIGH(노란)=0 MEDIUM(연파랑)=0 POLICY(연녹)=0"
            ws_mp.cell(1,1).value = meta
            ws_mp.cell(1,1).font  = FONT_META
            ws_mp.cell(1,1).fill  = FILL['HIGH']
            ws_mp.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
            for ci, (h, w) in enumerate(zip(MP_HEADERS, MP_WIDTHS), 1):
                _hdr(ws_mp, 2, ci, h)
                ws_mp.column_dimensions[get_column_letter(ci)].width = w
            ws_mp.row_dimensions[2].height = 20
            ws_mp.freeze_panes = "A3"
        else:
            ws_mp = wb["Matched_Plan"]
            mp_max_row = ws_mp.max_row
            # v3.9 3중 키 수집
            for r in range(3, mp_max_row+1):
                link     = str(ws_mp.cell(r,17).value or '').strip()
                plan_id  = str(ws_mp.cell(r, 4).value or '').strip()
                title_en = str(ws_mp.cell(r, 5).value or '').strip()[:60]
                date_val = str(ws_mp.cell(r, 6).value or '')[:10].strip()
                source   = str(ws_mp.cell(r, 7).value or '').strip()
                if link:                          existing_links.add(link)
                if plan_id and title_en:          existing_plan_titles.add(f"{plan_id}|{title_en}")
                if plan_id and date_val and source:existing_plan_dates.add(f"{plan_id}|{date_val}|{source}")

        # News DB에서 신규 기사 수집
        new_articles = []
        for r in range(2, ws_news.max_row+1):
            plan_id = str(ws_news.cell(r,11).value or '').strip()
            if not plan_id: continue

            link      = str(ws_news.cell(r,13).value or '').strip()
            title_en_raw = str(ws_news.cell(r,5).value or ws_news.cell(r,6).value or '').strip()
            title_key = title_en_raw[:60]
            date_val  = str(ws_news.cell(r,4).value  or '')[:10].strip()
            source    = str(ws_news.cell(r,8).value  or '').strip()

            if link and link in existing_links:                                     continue
            if plan_id and title_key and f"{plan_id}|{title_key}" in existing_plan_titles: continue
            if plan_id and date_val and source and f"{plan_id}|{date_val}|{source}" in existing_plan_dates: continue

            grade    = str(ws_news.cell(r,12).value or '').upper()
            qc       = str(ws_news.cell(r,17).value or '')
            title_en = ws_news.cell(r,5).value
            title_vi = ws_news.cell(r,6).value
            title    = title_en or title_vi
            sec      = ws_news.cell(r,2).value or _sector_from_plan(plan_id) or _sector_from_text(title, ws_news.cell(r,7).value, plan_id)

            ctx_grade = grade
            if not ctx_grade:
                if 'SA7+POLICY' in qc.upper(): ctx_grade = 'HIGH'
                elif 'SA7' in qc.upper():       ctx_grade = 'MEDIUM'
                elif 'POLICY' in qc.upper():    ctx_grade = 'POLICY'
                else:                           ctx_grade = 'MEDIUM'

            new_articles.append({
                'ctx_tag':    'SA7_MATCH' if 'SA7' in qc.upper() else ('POLICY_MATCH' if 'POLICY' in qc.upper() else 'SA7_MATCH'),
                'ctx_grade':  ctx_grade,
                'plan_id':    plan_id,
                'title_en':   title_en or title_vi,
                'title_vi':   title_vi,
                'date':       date_val,
                'source':     ws_news.cell(r,8).value,
                'province':   ws_news.cell(r,10).value,
                'sector':     sec,
                'title_ko':   ws_news.cell(r,7).value,
                'summary_ko': ws_news.cell(r,14).value,
                'summary_en': ws_news.cell(r,15).value,
                'summary_vi': ws_news.cell(r,16).value,
                'link':       link,
            })

        # ★ v4.0: 날짜 역순 정렬 후 Row3에 insert_rows로 삽입 (최신 상단)
        new_articles_sorted = sorted(
            new_articles,
            key=lambda x: str(x.get('date', '')),
            reverse=True
        )

        for a in new_articles_sorted:
            ws_mp.insert_rows(3)          # 헤더(Row2) 바로 아래에 삽입
            ri = 3
            ctx_grade = a.get('ctx_grade','')
            fill = _grade_fill(ctx_grade)
            vals = [None, a.get('ctx_tag',''), ctx_grade, a.get('plan_id',''),
                    a.get('title_en',''), a.get('date',''), a.get('source',''),
                    a.get('province',''), a.get('sector',''), a.get('title_ko',''),
                    a.get('title_en',''), a.get('title_vi',''),
                    a.get('summary_ko',''), a.get('summary_en',''), a.get('summary_vi',''),
                    '', a.get('link','')]
            for ci, v in enumerate(vals, 1):
                c = ws_mp.cell(ri, ci)
                c.value = v; c.fill = fill
                c.font = FONT_TITLE_BOLD if ci in (5,10,13,14) else FONT_DATA
                c.alignment = Alignment(vertical="top", wrap_text=False)

        # Row1 메타 업데이트
        total_mp = ws_mp.max_row - 2
        high_c   = sum(1 for r in range(3, ws_mp.max_row+1) if str(ws_mp.cell(r,3).value or '').upper()=='HIGH')
        med_c    = sum(1 for r in range(3, ws_mp.max_row+1) if str(ws_mp.cell(r,3).value or '').upper()=='MEDIUM')
        pol_c    = sum(1 for r in range(3, ws_mp.max_row+1) if str(ws_mp.cell(r,3).value or '').upper()=='POLICY')
        meta = (f"★ SA-7 맥락 확정 기사 {total_mp}건 | "
                f"HIGH(노란)={high_c} MEDIUM(연파랑)={med_c} POLICY(연녹)={pol_c}")

        if ws_mp.merged_cells:
            for mc in list(ws_mp.merged_cells.ranges):
                if 'A1' in str(mc):
                    ws_mp.unmerge_cells(str(mc))
                    ws_mp.cell(1,1).value = meta
                    ws_mp.merge_cells(str(mc))
                    break
        else:
            ws_mp.cell(1,1).value = meta
        ws_mp.cell(1,1).font = FONT_META
        ws_mp.cell(1,1).fill = FILL['HIGH']

        existing_mp = mp_max_row - 2
        print(f"    [Matched_Plan] {total_mp}건 (기존 {existing_mp}건 보존 + 신규 {len(new_articles)}건 상단 삽입 ★날짜역순)")

    def _rebuild_keywords_history(self, wb):
        ws_news = wb["News Database"]
        kh = []
        for r in range(2, ws_news.max_row+1):
            pid      = str(ws_news.cell(r,11).value or '').strip()
            title_en = ws_news.cell(r,5).value
            title_vi = ws_news.cell(r,6).value
            title    = title_en or title_vi
            sec      = ws_news.cell(r,2).value or _sector_from_plan(pid) or _sector_from_text(title, ws_news.cell(r,7).value, pid)
            if not sec: continue
            kh.append({
                'sector':   sec, 'province': str(ws_news.cell(r,10).value or ''),
                'date':     str(ws_news.cell(r,4).value or '')[:10],
                'title_en': title, 'title_ko': ws_news.cell(r,7).value,
                'source':   ws_news.cell(r,8).value,
                'grade':    str(ws_news.cell(r,12).value or ''),
                'plan_id':  pid,
            })

        pre    = sorted(kh, key=lambda x: (SECT_PRI.get(x['sector'], 99), str(x.get('province','') or '')))
        result = []
        for (s, p), grp in groupby(pre, key=lambda x: (x['sector'], x['province'])):
            result.extend(sorted(grp, key=lambda x: str(x.get('date','')), reverse=True))

        if "Keywords History" in wb.sheetnames: del wb["Keywords History"]
        mp_idx = wb.sheetnames.index("Matched_Plan")
        ws_kh  = wb.create_sheet("Keywords History", mp_idx+1)
        for ci, (h, w) in enumerate(zip(KH_HEADERS, KH_WIDTHS), 1):
            _hdr(ws_kh, 1, ci, h, "HDR_G")
            ws_kh.column_dimensions[get_column_letter(ci)].width = w
        ws_kh.row_dimensions[1].height = 22
        ws_kh.freeze_panes = "A2"

        for ri, a in enumerate(result, 2):
            d      = str(a.get('date',''))
            is_new = d >= self.cutoff
            fill   = FILL['KH_NEW'] if is_new else FILL['WHITE']
            vals   = [a.get('sector',''), a.get('province',''), d,
                      str(a.get('title_en','') or '')[:100],
                      str(a.get('title_ko','') or '')[:80],
                      a.get('source',''), a.get('grade',''), a.get('plan_id','')]
            for ci, v in enumerate(vals, 1):
                c = ws_kh.cell(ri, ci); c.value = v; c.fill = fill
                c.font = FONT_TITLE_BOLD if ci in (4,5) else FONT_DATA
                c.alignment = Alignment(vertical="top")

        new_c = sum(1 for a in result if str(a.get('date','')) >= self.cutoff)
        print(f"    [Keywords History] {len(result)}건 (신규 {new_c}건 노란표시)")

    def _update_summary(self, wb, new_count):
        if "Summary" not in wb.sheetnames: return
        ws_news = wb["News Database"]
        total   = ws_news.max_row - 1
        matched = 0; high_c = 0
        if "Matched_Plan" in wb.sheetnames:
            ws_mp   = wb["Matched_Plan"]
            matched = ws_mp.max_row - 2
            high_c  = sum(1 for r in range(3, ws_mp.max_row+1)
                          if str(ws_mp.cell(r,3).value or '').upper()=='HIGH')
        ws_sum    = wb["Summary"]
        meta_text = (f"Updated: {self.now} | Total News: {total}건 | SA-7: {matched}건 | HIGH: {high_c}건")
        for r in range(1, min(5, ws_sum.max_row+1)):
            v = str(ws_sum.cell(r,1).value or '')
            if 'Updated' in v or 'Total' in v:
                _safe_set_merged_cell(ws_sum, r, 1, meta_text); break
        print(f"    [Summary] Total={total} SA-7={matched} HIGH={high_c}")

    def _update_collection_log(self, wb, new_count):
        if "Collection_Log" not in wb.sheetnames: return
        ws    = wb["Collection_Log"]
        total = wb["News Database"].max_row - 1
        ws.insert_rows(2)
        ws.cell(2,1).value = f"{self.now} KST"
        ws.cell(2,2).value = new_count
        ws.cell(2,3).value = "Daily Automated"
        ws.cell(2,4).value = "✅"
        ws.cell(2,5).value = f"신규 {new_count}건 추가 | 전체 {total}건"
        for c in range(1, 6): ws.cell(2,c).font = FONT_DATA

    def _update_source(self, wb):
        if "Source" not in wb.sheetnames: return
        ws_news = wb["News Database"]
        src_cnt = defaultdict(int)
        for r in range(2, ws_news.max_row+1):
            s = str(ws_news.cell(r,8).value or '').strip()
            if s: src_cnt[s] += 1
        print(f"    [Source] {len(src_cnt)}개 출처 감지")

    def _update_stats(self, wb):
        if "Stats" not in wb.sheetnames: return
        ws_mp  = wb["Matched_Plan"]
        total_mp = ws_mp.max_row - 2
        ws_st  = wb["Stats"]
        ws_st.cell(1,1).value = f"VIETNAM INFRASTRUCTURE NEWS — 플랜별 SA-7 매칭 현황 ({self.now})"
        ws_st.cell(2,1).value = f"전체 SA-7: {total_mp}건"

    def _update_context_stats(self, wb):
        if "Context_Stats" not in wb.sheetnames: return
        total_mp = wb["Matched_Plan"].max_row - 2
        total    = wb["News Database"].max_row - 1
        ws_cs    = wb["Context_Stats"]
        ws_cs.cell(1,1).value = 'SA-7 + Policy Sentinel Total'
        ws_cs.cell(1,2).value = total_mp
        ws_cs.cell(1,3).value = f'{round(total_mp/total*100,1)}%' if total else '0%'

    def _update_timeline(self, wb):
        if "Timeline" not in wb.sheetnames: return
        ws_mp = wb["Matched_Plan"]
        plan_grade = defaultdict(lambda: defaultdict(int))
        for r in range(3, ws_mp.max_row+1):
            pid = str(ws_mp.cell(r,4).value or '').strip()
            g   = str(ws_mp.cell(r,3).value or '').upper()
            if pid: plan_grade[pid][g] += 1
        print(f"    [Timeline] {len(plan_grade)}개 Plan")

    def _update_province_keywords(self, wb):
        if "Province_Keywords" not in wb.sheetnames: return
        ws_news    = wb["News Database"]
        ws_mp      = wb["Matched_Plan"]
        prov_total = defaultdict(int); prov_unmap = defaultdict(int)
        prov_sec   = defaultdict(list); prov_src   = defaultdict(list)
        for r in range(2, ws_news.max_row+1):
            pv = str(ws_news.cell(r,10).value or '').strip()
            if not pv: continue
            prov_total[pv] += 1
            pid = str(ws_news.cell(r,11).value or '').strip()
            src = str(ws_news.cell(r, 8).value or '').strip()
            if pid: prov_sec[pv].append(_sector_from_plan(pid) or '')
            else:   prov_unmap[pv] += 1
            if src: prov_src[pv].append(src)
        prov_sa7 = defaultdict(int)
        for r in range(3, ws_mp.max_row+1):
            pv = str(ws_mp.cell(r,8).value or '').strip()
            if pv: prov_sa7[pv] += 1
        ws_pk = wb["Province_Keywords"]
        for r in range(2, ws_pk.max_row+1):
            pv  = str(ws_pk.cell(r,1).value or '').strip()
            if not pv: continue
            tot = prov_total.get(pv,0); sa7 = prov_sa7.get(pv,0)
            unm = prov_unmap.get(pv,0)
            rate= round(unm/tot*100,1) if tot else 0
            secs= [s for s in prov_sec.get(pv,[]) if s]
            ms  = max(set(secs), key=secs.count) if secs else ''
            srcs= prov_src.get(pv,[])
            mss = max(set(srcs), key=srcs.count) if srcs else ''
            if tot==0:       q='⚪ 데이터없음'
            elif rate<20:    q='🟢 양호'
            elif rate<40:    q='🟡 보통'
            elif rate<60:    q='🟠 주의'
            else:            q='🔴 요개선'
            ws_pk.cell(r,2).value = tot;  ws_pk.cell(r,3).value = sa7
            ws_pk.cell(r,4).value = ms;   ws_pk.cell(r,5).value = mss[:30] if mss else ''
            ws_pk.cell(r,6).value = f'{rate}%'; ws_pk.cell(r,7).value = q

    def _create_empty_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "News Database"
        self._write_news_header(ws)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(self.path))


if __name__ == "__main__":
    print("ExcelUpdater v4.0 단독 실행")
    updater = ExcelUpdater()
    updater.update_all([])
