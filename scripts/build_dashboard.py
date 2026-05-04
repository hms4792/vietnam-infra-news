"""
build_dashboard.py — v7.4
==========================
Excel → docs/index.html 재생성

[v7.4 핵심 수정 — Matched 버튼 기준 교정]

  ★ Matched 버튼 기준: Matched_Plan 시트 271건 기준
    - v7.3: Plan_ID OR Grade → 409건 (과다, 부정확)
    - v7.4: Matched_Plan 시트의 Link URL + Title_EN/VI 교차 매칭 → 최대 269건
    
    매칭 우선순위:
      1순위: Matched_Plan.Link ↔ News DB.URL 직접 매칭 (185건)
      2순위: Matched_Plan.Title_EN ↔ News DB.Title_EN/VI 앞60자 매칭 (추가 84건)
      → 총 269건 (나머지 2건은 News DB에 없는 기사)

[v7.3 핵심 수정 — 열 매핑 전면 교정]

  ★ 근본 원인: v7.2.1의 열 매핑이 실제 Excel v3.x 17컬럼 구조와 전혀 달랐음
  
  v7.2.1 잘못된 주석 (구 8컬럼 기준):
    row[0]=Area  row[1]=Date  row[2]=Title(En/Vi)  row[3]=Tit_ko
    row[4]=Source  row[5]=Src_Type  row[6]=Province  row[7]=Plan_ID
    row[8]=Grade  row[9]=URL  row[10]=sum_ko  row[11]=sum_en  row[12]=sum_vi
  
  실제 Excel v3.x 17컬럼 구조 (0-based):
    row[0]=Area   row[1]=Sector   row[2]=No      row[3]=Date
    row[4]=Title_EN  row[5]=Title_VI  row[6]=Tit_ko  row[7]=Source
    row[8]=Src_Type  row[9]=Province  row[10]=Plan_ID  row[11]=Grade
    row[12]=URL   row[13]=sum_ko  row[14]=sum_en  row[15]=sum_vi  row[16]=QC

열 매핑 (v3.x 17컬럼, 0-based):
  row[0] =Area      row[1] =Sector    row[2] =No        row[3] =Date
  row[4] =Title_EN  row[5] =Title_VI  row[6] =Tit_ko    row[7] =Source
  row[8] =Src_Type  row[9] =Province  row[10]=Plan_ID   row[11]=Grade
  row[12]=URL       row[13]=sum_ko    row[14]=sum_en    row[15]=sum_vi
  row[16]=QC

특징:
  ✓ 기존 템플릿 파일 사용 (templates/dashboard_template.html)
  ✓ BACKEND_DATA 구조 100% 동일
  ✓ Sector는 Excel의 Sector 컬럼(row[1]) 직접 사용 (가장 정확)
  ✓ Area는 Excel의 Area 컬럼(row[0]) 직접 사용
  ✓ Matched 버튼: Matched_Plan 시트 271건 기준 (URL + Title 교차 매칭)
"""

import os
import sys
import json
import re
import logging
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s'
)
logger = logging.getLogger('build_dashboard')

SCRIPTS_DIR   = os.path.dirname(os.path.abspath(__file__))
BASE_DIR      = os.path.dirname(SCRIPTS_DIR)
EXCEL_PATH    = os.path.join(BASE_DIR, 'data', 'database',
                             'Vietnam_Infra_News_Database_Final.xlsx')
TEMPLATE_PATH = os.path.join(BASE_DIR, 'templates', 'dashboard_template.html')
OUTPUT_PATH   = os.path.join(BASE_DIR, 'docs', 'index.html')

# ── Sector / Area 상수 ────────────────────────────────────────────────────
ENV_SECTORS   = {'Waste Water', 'Water Supply/Drainage', 'Solid Waste', 'Environment'}
ENERGY_SECTORS = {'Power', 'Oil & Gas'}
VALID_SECTORS = {
    'Waste Water', 'Water Supply/Drainage', 'Solid Waste',
    'Power', 'Oil & Gas', 'Transport', 'Industrial Parks',
    'Smart City', 'Construction', 'Environment'
}
VALID_AREAS   = {'Environment', 'Energy Develop.', 'Urban Develop.'}
VALID_GRADES  = {'HIGH', 'MEDIUM', 'POLICY', 'LOW'}

# Area ↔ Sector 매핑 (Area 컬럼 오염 시 복구용)
AREA_MAP = {
    'Waste Water':           'Environment',
    'Water Supply/Drainage': 'Environment',
    'Solid Waste':           'Environment',
    'Environment':           'Environment',
    'Power':                 'Energy Develop.',
    'Oil & Gas':             'Energy Develop.',
    'Transport':             'Urban Develop.',
    'Industrial Parks':      'Urban Develop.',
    'Smart City':            'Urban Develop.',
    'Construction':          'Urban Develop.',
    'General':               'Urban Develop.',
    'Bilateral':             'Urban Develop.',
}


def _plan_to_sector(plan: str) -> str:
    """Plan ID → Sector 추론 (fallback용)"""
    p = plan.upper()
    if 'WW' in p:                              return 'Waste Water'
    if 'SWM' in p or 'SOLID' in p:            return 'Solid Waste'
    if 'WAT' in p:                             return 'Water Supply/Drainage'
    if any(x in p for x in ('PDP8','RENEW','LNG','NUCLEAR','PWR')):
                                               return 'Power'
    if 'OG' in p or 'OIL' in p:              return 'Oil & Gas'
    if 'IP' in p or 'INDUST' in p or 'ENV' in p:
                                               return 'Industrial Parks'
    if any(x in p for x in ('SC','SMART','METRO','TRAN','URB','EV','MEKONG','HN')):
                                               return 'Transport'
    return 'Environment'


def _safe_area(area_val: str, sector_val: str) -> str:
    """
    Area 컬럼값 안전 처리.
    Area가 올바른 값이면 그대로, Sector명이 오염된 경우 AREA_MAP으로 교정.
    """
    if area_val in VALID_AREAS:
        return area_val
    # Area에 Sector명이 들어온 경우 (오염) → AREA_MAP으로 교정
    if area_val in AREA_MAP:
        return AREA_MAP[area_val]
    # Sector 기반으로 유추
    if sector_val in AREA_MAP:
        return AREA_MAP[sector_val]
    return 'Urban Develop.'


def _clean(val) -> str:
    """문자열 정제 (제어문자 제거)"""
    if val is None:
        return ''
    s = re.sub(r'[\x00-\x1f\x7f]', ' ', str(val))
    return re.sub(r'\s+', ' ', s).strip()


def _load_matched_plan_keys(excel_path: str) -> tuple[set, set]:
    """
    Matched_Plan 시트에서 매칭 키 두 세트를 로드.

    Returns:
        url_set   : Matched_Plan.Link URL 세트 (185건) — 1순위 매칭
        title_set : Link 없는 행의 Title_EN/VI 앞 60자 세트 (86건) — 2순위 매칭

    ★ title_set은 반드시 Link 없는 행만 수집
       Link 있는 행의 Title까지 넣으면 News DB 비매칭 기사와 제목이 겹쳐 과다 매칭 발생
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb['Matched_Plan']
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        logger.warning(f"Matched_Plan 로드 실패: {e}")
        return set(), set()

    if len(rows) < 3:
        return set(), set()

    # Row1=메타, Row2=헤더, Row3~=데이터
    hdr = [str(c or '').strip() for c in rows[1]]
    def _ci(name):
        return hdr.index(name) if name in hdr else -1

    link_ci  = _ci('Link')
    title_ci = _ci('Title_EN')

    url_set   = set()
    title_set = set()  # ★ Link 없는 행의 Title만 수집

    for r in rows[2:]:
        link = ''
        if link_ci >= 0 and len(r) > link_ci:
            link = str(r[link_ci] or '').strip()
            if link and link not in ('nan', 'None', ''):
                url_set.add(link)
            else:
                link = ''  # 무효 처리

        # ★ Link 없는 행만 Title 세트에 추가 (Link 있는 행은 URL로 매칭되므로 불필요)
        if not link and title_ci >= 0 and len(r) > title_ci:
            t = str(r[title_ci] or '').strip()[:60]
            if t and t not in ('nan', 'None', ''):
                title_set.add(t)

    return url_set, title_set


def load_articles(excel_path: str) -> list:
    """
    Excel News Database 시트에서 기사 데이터 로드.

    ★ v7.4 Matched 판정:
      Matched_Plan 시트의 URL 세트 + Title 세트와 교차 매칭
      → News DB URL이 Matched_Plan.Link에 있으면 matched='Y'
      → URL 없으면 Title_EN/VI 앞 60자로 2차 매칭

    ★ v7.3 열 매핑 (0-based, 17컬럼):
      row[0] =Area        row[1] =Sector      row[2] =No
      row[3] =Date        row[4] =Title_EN    row[5] =Title_VI
      row[6] =Tit_ko      row[7] =Source      row[8] =Src_Type
      row[9] =Province    row[10]=Plan_ID     row[11]=Grade
      row[12]=URL         row[13]=sum_ko      row[14]=sum_en
      row[15]=sum_vi      row[16]=QC
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        os.system("pip install openpyxl --break-system-packages")
        from openpyxl import load_workbook

    if not os.path.exists(excel_path):
        logger.error(f"Excel file not found: {excel_path}")
        return []

    # ★ v7.4: Matched_Plan 키 세트 미리 로드
    mp_url_set, mp_title_set = _load_matched_plan_keys(excel_path)
    logger.info(f"  Matched_Plan 키: URL {len(mp_url_set)}건 / Title {len(mp_title_set)}건")

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb['News Database']
        articles     = []
        matched_count = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if not row or not any(row):
                continue

            def rv(i):
                """row[i] 안전 추출"""
                return _clean(row[i] if len(row) > i else None)

            # ── ★ v7.3 올바른 열 매핑 ────────────────────────────────
            area_raw   = rv(0)   # Area
            sector_raw = rv(1)   # Sector
            # row[2] = No (행번호) — 사용 안 함
            date_val   = rv(3)   # Date
            title_en   = rv(4)   # Title_EN
            title_vi   = rv(5)   # Title_VI
            title_ko   = rv(6)   # Tit_ko
            source     = rv(7)   # Source
            src_type   = rv(8)   # Src_Type  (예: 'News', 'Specialist')
            province   = rv(9)   # Province
            plan_id    = rv(10)  # Plan_ID
            grade      = rv(11)  # Grade
            url        = rv(12)  # URL
            sum_ko     = rv(13)  # sum_ko
            sum_en     = rv(14)  # sum_en
            sum_vi     = rv(15)  # sum_vi

            # ── 제목 유효성 검사 ──────────────────────────────────────
            # title_en 또는 title_vi 둘 중 하나라도 있어야 수집
            display_title = title_en or title_vi
            if not display_title:
                continue

            # ── Sector 결정 (우선순위) ────────────────────────────────
            # 1순위: Excel Sector 컬럼값 (가장 정확)
            # 2순위: Plan_ID 기반 추론
            # 3순위: 기본값 Environment
            if sector_raw in VALID_SECTORS:
                sector = sector_raw
            elif plan_id:
                sector = _plan_to_sector(plan_id)
            else:
                sector = 'Environment'

            # ── Area 결정 (오염 방지) ─────────────────────────────────
            area = _safe_area(area_raw, sector)

            # ── Grade 정규화 ──────────────────────────────────────────
            grade_upper = grade.upper() if grade else ''
            grade_clean = grade_upper if grade_upper in VALID_GRADES else ''

            # ── ★ v7.4 Matched 판정 (Matched_Plan 시트 기준) ─────────
            # 1순위: URL이 Matched_Plan.Link 세트에 있으면 matched
            # 2순위: Title_EN 또는 Title_VI 앞 60자가 Matched_Plan.Title 세트에 있으면 matched
            url_in_mp   = bool(url and url in mp_url_set)
            title_key   = (title_en or title_vi)[:60]
            title_in_mp = bool(title_key and title_key in mp_title_set)
            is_matched  = url_in_mp or title_in_mp

            if is_matched:
                matched_count += 1

            # plan_id: VN-/HN- 시작만 유효 (출처명 오염 방지)
            plan_valid = bool(plan_id and (
                plan_id.upper().startswith('VN-') or
                plan_id.upper().startswith('HN-')
            ))

            articles.append({
                'id':      idx,
                'title':   {
                    'ko': title_ko,
                    'en': title_en,
                    'vi': title_vi,
                },
                'summary': {
                    'ko': sum_ko,
                    'en': sum_en,
                    'vi': sum_vi,
                },
                'sector':   sector,
                'area':     area,
                'province': province,
                'source':   source,
                'date':     date_val,
                'url':      url,
                'plan_id':  plan_id if plan_valid else '',
                'grade':    grade_clean,
                # ★ matched 플래그: JS 필터에서 사용
                'matched':  'Y' if is_matched else '',
            })

        wb.close()

        # 날짜 역순 정렬 (최신순)
        articles.sort(key=lambda x: x.get('date', '') or '', reverse=True)

        logger.info(f"✓ Loaded {len(articles)} articles (Matched_Plan: {matched_count})")
        return articles

    except Exception as e:
        logger.error(f"Error loading Excel: {e}")
        import traceback; traceback.print_exc()
        return []


def build_dashboard(excel_path: str = EXCEL_PATH,
                    template_path: str = TEMPLATE_PATH,
                    output_path: str = OUTPUT_PATH):
    """
    대시보드 생성.

    워크플로:
    1. Excel 로드 → articles 리스트
    2. JSON 직렬화
    3. 템플릿 HTML 읽기
    4. /*__BACKEND_DATA__*/[] 플레이스홀더에 JSON 주입
    5. index.html 저장
    """
    logger.info("=" * 60)
    logger.info("DASHBOARD BUILD v7.3 (Column Mapping Fix)")
    logger.info("=" * 60)

    logger.info(f"Loading Excel: {excel_path}")
    articles = load_articles(excel_path)

    if not articles:
        logger.error("No articles loaded")
        sys.exit(1)

    # JSON 직렬화
    try:
        json_str = json.dumps(articles, ensure_ascii=False)
        json.loads(json_str)   # 유효성 검증
    except (json.JSONDecodeError, ValueError):
        logger.warning("JSON error — retrying with ASCII mode")
        json_str = json.dumps(articles, ensure_ascii=True)

    # 템플릿 로드
    if not os.path.exists(template_path):
        logger.error(f"Template not found: {template_path}")
        sys.exit(1)

    logger.info(f"Loading template: {template_path}")
    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()

    # BACKEND_DATA 주입
    PLACEHOLDER = '/*__BACKEND_DATA__*/[]'
    if PLACEHOLDER not in html:
        logger.error(f"Placeholder not found in template: {PLACEHOLDER}")
        sys.exit(1)

    html = html.replace(PLACEHOLDER, json_str)

    # 선택적 플레이스홀더
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    html = html.replace('{{UPDATE_TIME}}', now)
    html = html.replace('{{ARTICLE_COUNT}}', str(len(articles)))

    # 저장
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    kb = os.path.getsize(output_path) / 1024
    logger.info(f"✓ Saved: {output_path} ({kb:.1f} KB) | {len(articles)} articles")
    logger.info("=" * 60)


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Build dashboard from Excel')
    parser.add_argument('--excel',    default=EXCEL_PATH)
    parser.add_argument('--template', default=TEMPLATE_PATH)
    parser.add_argument('--output',   default=OUTPUT_PATH)
    args = parser.parse_args()
    build_dashboard(args.excel, args.template, args.output)
