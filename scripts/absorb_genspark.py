#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
absorb_genspark.py
==================
Genspark 수동 실행 후 GitHub shared/에 업로드된 결과물을
Claude 파이프라인에 자동 흡수하는 스크립트.

실행 시점: daily_pipeline.yml에서 quality_context_agent 직후
           (genspark_output.json이 최신화된 후)

흡수 항목:
  1. genspark_output.json → 미번역 기사 요약 보완
  2. genspark_output.json → policy_context 플랜 ID 교차확인
  3. genspark_output.json → 신규 수집 기사 (Claude DB 미포함)
  4. knowledge_index.json → 플랜 ID 등장 빈도 기반 threshold 조정 참고
"""
import json, os, re
from pathlib import Path
from datetime import datetime, timedelta

BASE_DIR       = Path(__file__).parent.parent
SHARED_DOCS    = BASE_DIR / "docs"  / "shared"
AGENT_OUT      = BASE_DIR / "data"  / "agent_output"
DB_PATH        = BASE_DIR / "data"  / "database" / "Vietnam_Infra_News_Database_Final.xlsx"
KI_PATH        = SHARED_DOCS / "knowledge_index.json"

GENSPARK_OUT   = SHARED_DOCS / "genspark_output.json"
ABSORB_LOG     = AGENT_OUT   / "absorb_genspark_log.json"


def load_genspark():
    if not GENSPARK_OUT.exists():
        print("[absorb] genspark_output.json 없음 — 건너뜀")
        return []
    data = json.loads(GENSPARK_OUT.read_text(encoding='utf-8'))
    arts = data if isinstance(data, list) else data.get('articles', [])
    print(f"[absorb] Genspark 기사 로드: {len(arts)}건")
    return arts


def load_existing_urls():
    """Claude DB의 기존 URL 세트 로드"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DB_PATH, read_only=True)
        ws = wb.active
        headers = {str(ws.cell(1, c).value or '').strip(): c
                   for c in range(1, ws.max_column + 1)}
        link_col = headers.get('Link', 7)
        urls = set()
        for row in range(2, ws.max_row + 1):
            url = str(ws.cell(row, link_col).value or '').strip()
            if url:
                urls.add(url)
        print(f"[absorb] 기존 DB URL: {len(urls)}건")
        return urls
    except Exception as e:
        print(f"[absorb] DB 로드 실패: {e}")
        return set()


def analyze_genspark(arts, existing_urls):
    """
    Genspark 결과 분석:
      - 참고 가능: summary_ko (AI 자연어 요약)
      - 참고 가능: policy_context (플랜 맥락)
      - 참고 가능: matched_plans (플랜 ID 목록)
      ※ 매핑 결과는 불완전 — 참고만
    """
    result = {
        "generated_at":     datetime.utcnow().isoformat(),
        "genspark_total":   len(arts),
        "note":             "Genspark 매핑은 참고용 — Claude 키워드 매핑이 기준",
        "plan_frequency":   {},   # 플랜별 등장 빈도
        "new_articles":     [],   # Claude DB 미포함 신규 기사
        "summary_available":[],   # summary_ko 있는 기사 (번역 보완 가능)
        "sector_coverage":  {},   # 섹터별 기사 수
    }

    for art in arts:
        url  = art.get('url', '')
        plans = art.get('matched_plans', [])
        ctx   = art.get('policy_context', {})
        s_ko  = art.get('summary_ko', '')
        sector = art.get('sector', '')

        # 플랜 빈도
        for p in plans:
            result["plan_frequency"][p] = result["plan_frequency"].get(p, 0) + 1

        # 섹터 커버리지
        result["sector_coverage"][sector] = result["sector_coverage"].get(sector, 0) + 1

        # Claude DB 미포함 신규
        if url and url not in existing_urls:
            result["new_articles"].append({
                "title":   art.get('title', '')[:60],
                "url":     url,
                "sector":  sector,
                "plans":   plans,
                "has_summary_ko": bool(s_ko and len(s_ko) > 10),
            })

        # summary_ko 있는 기사 (Claude 미번역 보완 후보)
        if s_ko and len(s_ko) > 10 and url not in existing_urls:
            result["summary_available"].append({
                "url":        url,
                "title":      art.get('title', '')[:60],
                "summary_ko": s_ko[:100],
                "plans":      plans,
            })

    # 플랜 빈도 정렬
    result["plan_frequency"] = dict(
        sorted(result["plan_frequency"].items(), key=lambda x: -x[1])
    )

    return result


def print_report(result):
    print(f"\n{'='*55}")
    print(f"[absorb] Genspark 흡수 분석 결과")
    print(f"{'='*55}")
    print(f"  총 기사:      {result['genspark_total']}건")
    print(f"  신규 기사:    {len(result['new_articles'])}건 (Claude DB 미포함)")
    print(f"  요약 보완가능: {len(result['summary_available'])}건")
    print(f"\n  플랜별 빈도 (참고용 — 매핑 불완전):")
    for p, cnt in list(result['plan_frequency'].items())[:8]:
        bar = '█' * min(cnt, 20)
        print(f"    {p:<28} {cnt:3d}건 {bar}")
    print(f"\n  Genspark 섹터 커버리지:")
    for s, cnt in sorted(result['sector_coverage'].items(), key=lambda x: -x[1]):
        print(f"    {s:<28} {cnt}건")
    print(f"\n  ※ {result['note']}")
    print(f"{'='*55}")


def save_log(result):
    AGENT_OUT.mkdir(parents=True, exist_ok=True)
    ABSORB_LOG.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding='utf-8'
    )
    print(f"[absorb] 로그 저장: {ABSORB_LOG.name}")


def main():
    print(f"[absorb] 시작: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    arts = load_genspark()
    if not arts:
        return

    existing_urls = load_existing_urls()
    result = analyze_genspark(arts, existing_urls)
    print_report(result)
    save_log(result)
    print(f"[absorb] 완료")


if __name__ == '__main__':
    main()
