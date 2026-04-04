"""
backfill_dates.py
News Database 시트의 빈 Date 컬럼을 SQLite published_date로 채운다.

- Link(G열, index 6) URL → articles.url 매칭
- Date(E열, index 4) 가 비어있는 행만 업데이트
- 기존 데이터 절대 삭제/덮어쓰기 금지
"""

import os
import sqlite3

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "database", "Vietnam_Infra_News_Database_Final.xlsx")
DB_PATH    = os.path.join(BASE_DIR, "data", "vietnam_infrastructure_news.db")

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl 미설치 — pip install openpyxl")


def load_url_date_map(db_path):
    """SQLite에서 url → published_date 맵 로드."""
    url_map = {}
    try:
        conn = sqlite3.connect(db_path)
        cur  = conn.cursor()
        cur.execute("SELECT url, published_date FROM articles WHERE url IS NOT NULL AND published_date IS NOT NULL")
        for url, pub_date in cur.fetchall():
            date_str = str(pub_date)[:10]  # YYYY-MM-DD
            if date_str and date_str != "None":
                url_map[url.strip()] = date_str
        conn.close()
        print(f"[DB] {len(url_map)}개 URL-날짜 매핑 로드 완료")
    except Exception as e:
        print(f"[ERROR] SQLite 읽기 실패: {e}")
    return url_map


def backfill(excel_path, url_map):
    """Excel Date 컬럼 백필."""
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        raise SystemExit(f"[ERROR] Excel 열기 실패: {e}")

    if "News Database" not in wb.sheetnames:
        raise SystemExit("[ERROR] 'News Database' 시트 없음")

    ws = wb["News Database"]

    # 컬럼 인덱스 (1-based)
    COL_DATE = 5   # E열
    COL_LINK = 7   # G열

    total    = 0
    filled   = 0
    no_match = 0
    skipped  = 0  # 이미 값 있는 행

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            date_cell = ws.cell(row=row_idx, column=COL_DATE)
            link_cell = ws.cell(row=row_idx, column=COL_LINK)

            # 제목 없는 빈 행 무시
            if not ws.cell(row=row_idx, column=4).value:
                continue

            total += 1

            # 이미 날짜 있으면 건너뜀
            if date_cell.value:
                skipped += 1
                continue

            url = str(link_cell.value or "").strip()
            if not url:
                no_match += 1
                continue

            date_str = url_map.get(url)
            if date_str:
                date_cell.value = date_str
                filled += 1
            else:
                no_match += 1

        except Exception as e:
            print(f"  [WARN] 행 {row_idx} 처리 오류: {e}")
            no_match += 1

    try:
        wb.save(excel_path)
    except Exception as e:
        raise SystemExit(f"[ERROR] Excel 저장 실패: {e}")

    print(f"\n총 {total}행 / 이미채움 {skipped}행 / 채움 {filled}행 / 미매칭 {no_match}행")
    return filled


def main():
    print(f"Excel : {EXCEL_PATH}")
    print(f"DB    : {DB_PATH}")
    print()

    url_map = load_url_date_map(DB_PATH)
    if not url_map:
        print("[SKIP] 매핑 데이터 없음")
        return

    filled = backfill(EXCEL_PATH, url_map)
    if filled > 0:
        print(f"[OK] {filled}개 날짜 채우기 완료 → Excel 저장됨")
    else:
        print("[INFO] 새로 채울 날짜 없음")


if __name__ == "__main__":
    main()
