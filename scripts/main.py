"""
Vietnam Infrastructure News Pipeline - Main Entry Point
수집 → 번역 → Excel저장 → 대시보드 순서로 실행
"""

import argparse
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('pipeline.log', encoding='utf-8')
    ]
)
logger = logging.getLogger('MainPipeline')


def step1_collect_news(hours_back: int = 168) -> list:
    logger.info("=" * 60)
    logger.info(f"STEP 1: 뉴스 수집 시작 (최근 {hours_back}시간 = {hours_back//24}일)")
    logger.info("=" * 60)
    try:
        # news_collector.py는 클래스가 아닌 함수 방식
        # collect_news(hours_back) 함수를 직접 호출
        import importlib, sys
        sys.path.insert(0, '.')
        import scripts.news_collector as nc

        # collect_news() 함수 호출 → (count, articles, stats) 반환
        cnt, articles, stats = nc.collect_news(hours_back=hours_back)

        logger.info(f"[Step1 완료] 수집 기사 수: {len(articles)}")
        _save_raw_backup(articles)
        return articles
    except Exception as e:
        logger.error(f"[Step1 실패] {e}", exc_info=True)
        return []


def _save_raw_backup(articles: list):
    try:
        backup_dir = Path("data/raw")
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        with open(backup_dir / f"raw_{ts}.json", 'w', encoding='utf-8') as f:
            json.dump(articles, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"[Backup] 실패: {e}")


def step2_translate_articles(raw_articles: list) -> list:
    logger.info("=" * 60)
    logger.info("STEP 2: Claude API 번역/요약 시작")
    logger.info("=" * 60)
    if not raw_articles:
        logger.warning("[Step2] 수집된 기사 없음")
        return []
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        logger.error("[Step2] ANTHROPIC_API_KEY 미설정!")
        return _fallback_no_translation(raw_articles)
    try:
        import anthropic  # noqa
    except ImportError:
        logger.error("[Step2] anthropic 패키지 미설치!")
        return _fallback_no_translation(raw_articles)
    try:
        from scripts.ai_summarizer import AISummarizer
        summarizer = AISummarizer()
        processed = summarizer.process_articles(raw_articles)
        translated = sum(
            1 for a in processed
            if a.get('title_en') and a['title_en'] != a.get('title', '')
        )
        logger.info(f"[Step2 완료] 번역 성공: {translated}/{len(processed)}건")
        return processed
    except Exception as e:
        logger.error(f"[Step2 실패] {e}", exc_info=True)
        return _fallback_no_translation(raw_articles)


def _fallback_no_translation(articles: list) -> list:
    logger.warning("[Fallback] 번역 실패 - 원문으로 대체")
    for a in articles:
        title   = a.get('title', '')
        summary = a.get('summary', '') or a.get('content', '')[:200]
        a.setdefault('title_ko',   title)
        a.setdefault('title_en',   title)
        a.setdefault('title_vi',   title)
        a.setdefault('summary_ko', summary)
        a.setdefault('summary_en', summary)
        a.setdefault('summary_vi', summary)
    return articles


def step3_save_to_excel(processed_articles: list) -> bool:
    logger.info("=" * 60)
    logger.info("STEP 3: Excel 데이터베이스 저장")
    logger.info("=" * 60)
    if not processed_articles:
        logger.warning("[Step3] 저장할 기사 없음")
        return False
    try:
        from scripts.excel_updater import ExcelUpdater
        updater = ExcelUpdater()
        updater.update_all(processed_articles)
        logger.info("[Step3 완료] Excel 전체 시트 업데이트")
        return True
    except ImportError:
        logger.warning("[Step3] ExcelUpdater 없음 - 내장 저장 사용")
        return _builtin_excel_save(processed_articles)
    except Exception as e:
        logger.error(f"[Step3 실패] {e}", exc_info=True)
        return _builtin_excel_save(processed_articles)


def _builtin_excel_save(articles: list) -> bool:
    try:
        import openpyxl
        EXCEL_PATH = Path("data/database/Vietnam_Infra_News_Database_Final.xlsx")
        if not EXCEL_PATH.exists():
            logger.error(f"[BuiltinSave] Excel 파일 없음: {EXCEL_PATH}")
            return False
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        url_col = headers.get('Link') or headers.get('URL') or headers.get('url')
        existing_urls = set()
        if url_col:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[url_col - 1]:
                    existing_urls.add(str(row[url_col - 1]).strip())
        added = 0
        for article in articles:
            url = article.get('url', '').strip()
            if url and url in existing_urls:
                continue
            row_data = _map_article_to_row(article, headers, ws.max_column)
            ws.append(row_data)
            existing_urls.add(url)
            added += 1
        wb.save(EXCEL_PATH)
        logger.info(f"[BuiltinSave] {added}건 추가 완료")
        return True
    except Exception as e:
        logger.error(f"[BuiltinSave] 오류: {e}", exc_info=True)
        return False


def _map_article_to_row(article: dict, headers: dict, max_col: int) -> list:
    row = [''] * max_col
    mapping = {
        'Area':            lambda a: a.get('area', ''),
        'Business Sector': lambda a: a.get('sector', ''),
        'Province':        lambda a: a.get('province', ''),
        'News Tittle':     lambda a: a.get('title_en') or a.get('title', ''),
        'Date':            lambda a: a.get('date', ''),
        'Source':          lambda a: a.get('source', ''),
        'Link':            lambda a: a.get('url', ''),
        'Short summary':   lambda a: a.get('summary_en', '')[:200],
        'title_ko':        lambda a: a.get('title_ko', ''),
        'title_en':        lambda a: a.get('title_en') or a.get('title', ''),
        'title_vi':        lambda a: a.get('title_vi') or a.get('title', ''),
        'summary_ko':      lambda a: a.get('summary_ko', ''),
        'summary_en':      lambda a: a.get('summary_en', ''),
        'summary_vi':      lambda a: a.get('summary_vi', ''),
    }
    for col_name, col_idx in headers.items():
        if col_name in mapping and col_idx <= max_col:
            try:
                row[col_idx - 1] = mapping[col_name](article)
            except Exception:
                pass
    return row


def step4_build_dashboard(processed_articles: list) -> bool:
    logger.info("=" * 60)
    logger.info("STEP 4: 대시보드 HTML 생성")
    logger.info("=" * 60)
    try:
        from scripts.dashboard_updater import DashboardUpdater
        updater = DashboardUpdater()
        result = updater.generate(processed_articles)
        logger.info(f"[Step4 완료] {result}")
        return True
    except Exception as e:
        logger.error(f"[Step4 실패] {e}", exc_info=True)
        return False


def main():
    parser = argparse.ArgumentParser(description='Vietnam Infra News Pipeline')
    parser.add_argument(
        '--mode',
        choices=['full', 'collect-only', 'translate-only', 'dashboard-only'],
        default='full',
        help='실행 모드 선택'
    )
    parser.add_argument(
        '--full',
        action='store_true',
        help='전체 파이프라인 실행'
    )
    parser.add_argument(
        '--hours-back',
        type=int,
        default=168,
        dest='hours_back',
        help='수집 기간 시간 단위 (기본값: 168 = 7일)'
    )
    args = parser.parse_args()

    if args.full:
        args.mode = 'full'

    logger.info("╔══════════════════════════════════════════════════════╗")
    logger.info("║   Vietnam Infrastructure News Pipeline               ║")
    logger.info(f"║   실행 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}                   ║")
    logger.info(f"║   모드:     {args.mode:<47}║")
    logger.info(f"║   수집기간: {args.hours_back}시간 ({args.hours_back//24}일)                              ║")
    logger.info("╚══════════════════════════════════════════════════════╝")

    excel_saved        = False
    processed_articles = []

    if args.mode in ('full', 'collect-only'):
        raw_articles = step1_collect_news(hours_back=args.hours_back)
        if args.mode == 'collect-only':
            logger.info(f"[collect-only] 완료: {len(raw_articles)}건")
            return
        processed_articles = step2_translate_articles(raw_articles)
        excel_saved        = step3_save_to_excel(processed_articles)
        step4_build_dashboard(processed_articles)

    elif args.mode == 'translate-only':
        raw_articles       = _load_latest_raw_backup()
        processed_articles = step2_translate_articles(raw_articles)
        excel_saved        = step3_save_to_excel(processed_articles)
        step4_build_dashboard(processed_articles)

    elif args.mode == 'dashboard-only':
        processed_articles = _load_from_excel()
        step4_build_dashboard(processed_articles)

    logger.info("=" * 60)
    logger.info("파이프라인 실행 완료")
    logger.info(f"  처리 기사: {len(processed_articles)}건")
    logger.info(f"  Excel 저장: {'성공' if excel_saved else '실패/생략'}")
    logger.info("=" * 60)


def _load_latest_raw_backup() -> list:
    raw_dir = Path("data/raw")
    if not raw_dir.exists():
        return []
    files = sorted(raw_dir.glob("raw_*.json"), reverse=True)
    if not files:
        return []
    with open(files[0], encoding='utf-8') as f:
        return json.load(f)


def _load_from_excel() -> list:
    try:
        from scripts.excel_updater import ExcelUpdater
        updater = ExcelUpdater()
        return updater._load_excel_articles() if hasattr(updater, '_load_excel_articles') else []
    except Exception:
        return []


if __name__ == "__main__":
    main()
