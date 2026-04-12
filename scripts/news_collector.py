#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vietnam Infrastructure News Collector
Version 5.4 — Specialist Media + Waste Water + Smart City RSS 추가

v5.4 변경사항 (2026-04-05):
  [RSS 추가] 전문미디어 4개: The Investor, VIR, Construction Vietnam, VietnamBiz
  [RSS 추가] Waste Water 전용 4개: MONRE, VEA, Nhadepso, Moitruong Online
  [RSS 추가] Smart City 전용 3개: ICT Vietnam, MIC Vietnam, Smartcity Vietnam
  총 RSS 소스: 기존 33개 + 신규 11개 = 44개

v5.3 변경사항 (2026-03-29):
  [RSS 추가] 환경 전문 P1: VietnamPlus-Moi truong, Nhandan-Moi truong,
             Baotainguyenmoitruong, Kinhtemoitruong, VietnamPlus-Kinh te
  [RSS 추가] 북부 지역 P2: Hanoimoi, Baobacgiang English, Nhandan-Kinh te
  [RSS 추가] 보조 P3: Moitruong Net, Congnghiepmoitruong, VietnamPlus-Giao thong
"""

import os
import sys
import re
import sqlite3
import hashlib
import json
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse, urljoin, quote
import html
import concurrent.futures
import time

import requests
import feedparser
from bs4 import BeautifulSoup

# ============================================================
# CONFIGURATION
# ============================================================

DB_PATH          = os.environ.get('DB_PATH',       'data/vietnam_infrastructure_news.db')
HOURS_BACK       = int(os.environ.get('HOURS_BACK', 24))
EXCEL_PATH       = os.environ.get('EXCEL_PATH',    'data/database/Vietnam_Infra_News_Database_Final.xlsx')

LANGUAGE_FILTER  = os.environ.get('LANGUAGE_FILTER', 'all').lower()

MIN_CLASSIFY_THRESHOLD = 2

GNEWS_API_KEY    = os.environ.get('GNEWS_API_KEY', '')
ENABLE_GNEWS     = os.environ.get('ENABLE_GNEWS', 'false').lower() == 'true'

GNEWS_QUERY       = 'Vietnam infrastructure OR "Vietnam energy" OR "Vietnam transport"'
GNEWS_ENV_QUERY   = 'Vietnam environment OR "Vietnam wastewater" OR "Vietnam solid waste" OR "Vietnam water supply"'
GNEWS_NORTH_QUERY = '"Quang Ninh" infrastructure OR "Bac Giang" industrial OR "Hanoi" infrastructure OR "Hai Phong" port'


# ============================================================
# SECTOR DEFINITIONS
# ============================================================

SECTOR_PRIORITY_ORDER = [
    "Waste Water",
    "Water Supply/Drainage",
    "Oil & Gas",
    "Power",
    "Solid Waste",
    "Transport",
    "Industrial Parks",
    "Smart City",
    "Construction",
]

SECTOR_AREA = {
    "Waste Water":            "Environment",
    "Water Supply/Drainage":  "Environment",
    "Solid Waste":            "Environment",
    "Power":                  "Energy Develop.",
    "Oil & Gas":              "Energy Develop.",
    "Transport":              "Urban Development",
    "Industrial Parks":       "Urban Development",
    "Smart City":             "Urban Development",
    "Construction":           "Urban Development",
}

SECTOR_KEYWORDS = {
    "Waste Water": {
        "primary": [
            "wastewater", "waste water",
            "sewage treatment", "sewage plant", "sewage system",
            "wastewater treatment", "wastewater plant", "wwtp",
            "effluent treatment", "sludge treatment",
            "industrial wastewater", "domestic wastewater",
            "xu ly nuoc thai",
            "nuoc thai",
            "xu\u1eed l\xfd n\u01b0\u1edbc th\u1ea3i",
            "n\u01b0\u1edbc th\u1ea3i",
        ],
        "secondary": [
            "sewage", "sewer network", "sewer line",
            "effluent", "sludge",
            "water pollution control",
            "water quality improvement",
            # VN-ENV-IND-1894 산업폐수 처리
            "industrial wastewater treatment", "industrial effluent treatment",
            "wastewater treatment equipment",
            "xu ly nuoc thai cong nghiep", "nuoc thai khu cong nghiep",
        ],
    },

    "Water Supply/Drainage": {
        "primary": [
            "water supply system", "water supply network",
            "clean water supply", "drinking water supply",
            "tap water", "piped water", "potable water",
            "water treatment plant", "water purification plant",
            "water infrastructure", "water distribution",
            "desalination plant",
            "drainage system", "stormwater management",
            "flood control project", "flood prevention",
            "n\u01b0\u1edbc s\u1ea1ch",
            "c\u1ea5p n\u01b0\u1edbc",
            "tho\u00e1t n\u01b0\u1edbc",
            "ch\u1ed1ng ng\u1eadp",
            "h\u1ed3 ch\u1ee9a n\u01b0\u1edbc",
            "nh\u00e0 m\u00e1y n\u01b0\u1edbc",
            "n\u01b0\u1edbc sinh ho\u1ea1t",
            "l\u0169 l\u1ee5t",
            "h\u1ea1n h\u00e1n",
            "n\u01b0\u1edbc m\u01b0a",
            "ng\u1eadp \u00fang",
            "bi\u1ebfn \u0111\u1ed5i kh\u00ed h\u1eadu",
            "kh\u00ed h\u1eadu",
        ],
        "secondary": [
            "clean water", "drinking water",
            "water project", "water plant", "water reservoir",
            "groundwater management",
            "flood control", "anti-flooding",
            "water pipeline",
            # VN-ENV-IND-1894 수처리 설비/기술
            "water treatment technology", "water treatment equipment",
            "environmental water services",
        ],
    },

    "Oil & Gas": {
        "primary": [
            "oil and gas", "oil & gas",
            "petroleum refinery", "oil refinery",
            "offshore oil", "offshore gas", "offshore drilling",
            "lng terminal", "lng plant", "liquefied natural gas",
            "gas pipeline", "oil pipeline",
            "petrovietnam", "pvn", "pvgas", "pv gas",
            "binh son refinery", "nghi son refinery", "dung quat",
            "block b", "ca voi xanh",
            "l\xf4 b", "c\xe1 voi xanh",
        ],
        "secondary": [
            "petroleum", "petrochemical",
            "natural gas plant", "crude oil",
            "oil exploration", "gas exploration", "drilling",
            "upstream", "downstream", "midstream",
            "gas field", "oil field",
        ],
    },

    "Power": {
        "primary": [
            "power plant", "power station", "power project",
            "wind farm", "offshore wind farm", "solar farm",
            "hydroelectric plant", "hydropower plant",
            "transmission line", "high voltage line",
            "power purchase agreement", "ppa",
            "feed-in tariff",
            "battery storage system", "bess",
            "evn", "vietnam electricity",
            "nh\xe0 m\xe1y \u0111i\u1ec7n",
            "n\u0103ng l\u01b0\u1ee3ng t\xe1i t\u1ea1o",
            "\u0111i\u1ec7n m\u1eb7t tr\u1eddi",
            "\u0111i\u1ec7n gi\xf3",
            "th\u1ee7y \u0111i\u1ec7n",
            "nhi\u1ec7t \u0111i\u1ec7n",
            "l\u01b0\u1edbi \u0111i\u1ec7n",
            "truy\u1ec1n t\u1ea3i \u0111i\u1ec7n",
        ],
        "secondary": [
            "wind power", "solar power", "solar energy", "photovoltaic",
            "renewable energy", "clean energy",
            "thermal power", "coal-fired power",
            "hydropower", "hydroelectric",
            "power generation", "power grid",
            "electricity generation",
            "substation", "transformer station",
            "megawatt", "gigawatt", " mw ", " gw ",
            "energy storage", "energy transition",
            "lng power",
            # VN-PWR-PDP8 (국가전력개발계획 8차 — Decision 500/QĐ-TTg 2023)
            "power development plan", "pdp8",
            "offshore wind power", "offshore wind farm",
            "just energy transition", "jetp",
            "direct power purchase agreement", "dppa",
            "coal power phase-out", "energy transition",
            "long an lng", "ca na lng", "bac lieu lng",
            "transmission line", "power transmission",
            "quy hoach dien", "duong day dien",
            "phong dien ngoai khoi", "dien gio ngoai khoi",
            "thi truong dien canh tranh",
        ],
    },

    "Solid Waste": {
        "primary": [
            "solid waste management",
            "municipal solid waste", "msw",
            "landfill site", "sanitary landfill",
            "waste-to-energy plant", "wte plant", "wte facility",
            "incineration plant", "incinerator",
            "recycling plant", "recycling facility",
            "hazardous waste facility",
            "r\xe1c th\u1ea3i",
            "r\xe1c sinh ho\u1ea1t",
            "l\xf2 \u0111\u1ed1t r\xe1c",
            "b\xe3i r\xe1c",
            "x\u1eed l\xfd r\xe1c",
            "ch\u1ea5t th\u1ea3i r\u1eafn",
            "nh\xe0 m\xe1y x\u1eed l\xfd r\xe1c",
            "thu gom r\xe1c",
            "\u0111\u1ed1t r\xe1c ph\xe1t \u0111i\u1ec7n",
            "\xf4 nhi\u1ec5m m\xf4i tr\u01b0\u1eddng",
            "\xf4 nhi\u1ec5m",
            "ph\xe2n lo\u1ea1i r\xe1c",
            "kh\xf4ng kh\xed",
            "ph\xe1t th\u1ea3i",
            "m\xf4i tr\u01b0\u1eddng",
            "t\xe0i nguy\xean m\xf4i tr\u01b0\u1eddng",
            "kh\xed th\u1ea3i",
            "\xf4 nhi\u1ec5m n\u01b0\u1edbc",
            "ch\u1ea5t th\u1ea3i nguy h\u1ea1i",
        ],
        "secondary": [
            "solid waste", "garbage collection", "garbage disposal",
            "trash collection",
            "landfill",
            "waste-to-energy", "wte",
            "incineration",
            "composting facility",
            "electronic waste", "e-waste",
            "construction waste disposal",
            # VN-ENV-IND-1894 (환경산업 발전 프로그램 — 핵심 섹터)
            "environmental industry", "circular economy",
            "pollution control", "environmental services",
            "hazardous waste", "cong nghiep moi truong",
            "kinh te tuan hoan", "lo dot rac phat dien",
            # 추가: 대기오염/토양복원/모니터링
            "air pollution control", "air quality monitoring",
            "soil remediation", "contaminated land",
            "environmental monitoring system",
            "environmental consulting", "environmental audit",
            "environmental technology", "clean technology",
            "xu ly o nhiem", "quan trac moi truong",
            "cong nghe moi truong", "xu ly khi thai",
            "xu ly dat o nhiem", "thiet bi moi truong",
        ],
    },

    "Transport": {
        "primary": [
            "metro line", "metro project", "metro station",
            "urban rail project", "light rail project",
            "high-speed rail", "high speed rail", "hsr project",
            "railway project", "railroad project",
            "expressway project", "highway project",
            "ring road project", "ring expressway",
            "bridge construction", "cable-stayed bridge",
            "long thanh airport",
            "deep-sea port", "container terminal",
            "tuy\u1ebfn metro",
            "\u0111\u01b0\u1eddng cao t\u1ed1c",
            "s\xe2n bay",
            "c\u1ea3ng bi\u1ec3n",
            "l\u1ea1ch huy\u1ec7n",
            "c\u1ea3ng h\u1ea3i ph\xf2ng",
            "\u0111\u01b0\u1eddng s\u1eaft t\u1ed1c \u0111\u1ed9 cao",
        ],
        "secondary": [
            "metro", "subway",
            "railway", "railroad",
            "expressway", "highway", "motorway",
            "bypass road", "overpass", "flyover", "interchange",
            "road construction", "road upgrade",
            "bridge", "tunnel", "viaduct",
            "airport", "airport terminal", "runway",
            "seaport", "logistics hub", "logistics center",
            "public transport", "bus rapid transit", "brt",
            "inland waterway",
        ],
    },

    "Industrial Parks": {
        "primary": [
            "industrial park", "industrial zone", "industrial complex",
            "special economic zone", "sez",
            "export processing zone", "epz",
            "hi-tech park", "high-tech park", "technology park",
            "industrial estate", "industrial cluster",
            "khu c\xf4ng nghi\u1ec7p",
            "khu kinh t\u1ebf",
            "khu c\xf4ng ngh\u1ec7 cao",
            "vsip b\u1eafc ninh", "vsip b\u1eafc giang", "vsip qu\u1ea3ng ng\xe3i",
            "deep c", "amata",
        ],
        "secondary": [
            "economic zone", "free trade zone",
            "manufacturing hub", "manufacturing zone",
            "factory zone", "industrial area",
            "fdi project",
            # VN-ENV-IND-1894 생태산업단지
            "eco-industrial park", "green industrial zone",
            "environmental industry cluster", "industrial ecology",
            "clean production", "green manufacturing",
            "khu cong nghiep sinh thai", "cong nghiep xanh",
        ],
    },

    "Smart City": {
        "primary": [
            "smart city project", "smart city development",
            "intelligent city", "digital city",
            "smart traffic system", "traffic management system",
            "iot infrastructure", "5g network deployment",
            "e-government system", "digital government",
            "th\xe0nh ph\u1ed1 th\xf4ng minh",
            "\u0111\xf4 th\u1ecb th\xf4ng minh",
        ],
        "secondary": [
            "smart city",
            "smart urban",
            "smart grid", "smart meter",
            "smart building",
            "5g infrastructure",
            "digital transformation",
            "surveillance system", "cctv network", "ai camera",
        ],
    },

    "Construction": {
        "primary": [
            "real estate development", "property development",
            "housing project", "residential complex",
            "new urban area", "new township",
            "satellite city development",
            "commercial building construction",
            "urban development project",
            "khu \u0111\xf4 th\u1ecb",
            "d\u1ef1 \xe1n b\u1ea5t \u0111\u1ed9ng s\u1ea3n",
            "bao xay dung",
        ],
        "secondary": [
            "urban development",
            "city planning", "urban planning",
            "construction project",
            "building construction",
            "urban infrastructure",
            "cement plant", "steel plant",
        ],
    },
}


# ============================================================
# EXCLUDE KEYWORDS
# ============================================================

EXCLUDE_KEYWORDS = [
    "arrest", "jail", "prison", "sentenced", "trafficking", "smuggling",
    "fraud", "murder", "crime", "drug trafficking",
    "killed", "death toll", "crash kills", "fire kills", "collision kills",
    "accident kills", "flood kills",
    "football", "soccer", "tennis", "basketball", "volleyball", "badminton",
    "sports", "world cup", "olympics", "championship", "tournament",
    "golf tournament", "bridge tournament",
    "gold price", "stock market", "forex", "exchange rate",
    "cryptocurrency", "bitcoin",
    "seafood export", "agricultural export", "rice export",
    "tourism promotion", "tourist", "hotel resort", "beach resort",
    "retail sales",
    "university", "school enrollment", "scholarship",
    "beauty pageant", "fashion", "concert",
    "matchmaking", "get married", "marriage club",
    "safety certification", "vinfast vf",
    "night flights", "flight schedule", "airline route",
    "train collision in spain", "earthquake in",
    "party congress", "politburo", "state visit", "diplomatic",
    "multimedia", "social links", "subscribe",
]

NON_VIETNAM_COUNTRIES = [
    "singapore", "malaysia", "thailand", "indonesia", "philippines",
    "cambodia", "laos", "myanmar", "china", "japan", "south korea",
    "taiwan", "hong kong", "australia", "russia", " uk ", " usa ",
    "america", "india", "europe", "africa",
]

VIETNAM_KEYWORDS = [
    "vietnam", "vietnamese", "viet nam",
    "hanoi", "ho chi minh", "hcmc", "saigon",
    "da nang", "hai phong", "can tho",
    "binh duong", "dong nai", "quang ninh",
    "mekong", "evn", "petrovietnam", "pvn",
    "vi\u1ec7t nam", "h\xe0 n\u1ed9i", "tp.hcm", "tp hcm",
    "\u0111\xe0 n\u1eb5ng", "h\u1ea3i ph\xf2ng", "c\u1ea7n th\u01a1",
    "b\xecnh d\u01b0\u01a1ng", "\u0111\u1ed3ng nai", "qu\u1ea3ng ninh",
    "h\xe0 long", "b\u1eafc ninh", "long an",
    "qu\u1ea3ng ng\xe3i", "b\xecnh \u0111\u1ecbnh", "kh\xe1nh h\xf2a",
    "l\xe2m \u0111\u1ed3ng", "\u0111\u1eafk l\u1eafk", "gia lai",
    "ti\u1ec1n giang", "b\u1ebfn tre", "an giang",
    "ki\xean giang", "c\xe0 mau", "s\xf3c tr\u0103ng",
    "thanh ho\xe1", "ngh\u1ec7 an", "h\xe0 t\u0129nh",
    "qu\u1ea3ng b\xecnh", "qu\u1ea3ng tr\u1ecb", "th\u1eeba thi\xean",
    "th\xe1i nguy\xean", "b\u1eafc giang", "h\u01b0ng y\xean",
    "v\u0129nh ph\xfac", "ph\xfa th\u1ecd", "h\xf2a b\xecnh",
    "t\u1eadp \u0111o\xe0n \u0111i\u1ec7n l\u1ef1c", "t\u1eadp \u0111o\xe0n d\u1ea7u kh\xed",
    "b\u1ed9 t\xe0i nguy\xean",
    "t\xe0i nguy\xean v\xe0 m\xf4i tr\u01b0\u1eddng",
    "s\xf4ng c\u1eedu long",
    "\u0111\u1ed3ng b\u1eb1ng s\xf4ng",
    "c\u1eedu long",
    "mi\u1ec1n trung",
    "mi\u1ec1n nam",
    "mi\u1ec1n b\u1eafc",
    "b\u1eafc giang", "b\u1eafc ninh", "h\u1ea3i d\u01b0\u01a1ng",
    "h\u01b0ng y\xean", "th\xe1i b\xecnh", "nam \u0111\u1ecbnh",
    "ninh b\xecnh", "l\u1ea1ng s\u01a1n", "y\xean b\xe1i",
    "l\xe0o cai", "tuy\xean quang", "h\xe0 giang",
    "cao b\u1eb1ng", "b\u1eafc k\u1ea1n", "lai ch\xe2u",
    "\u0111i\u1ec7n bi\xean", "s\u01a1n la",
]


# ============================================================
# PROVINCE KEYWORDS
# ============================================================

PROVINCE_KEYWORDS = {
    "Ho Chi Minh City":  ["ho chi minh", "hcmc", "saigon", "sai gon", "h\u1ed3 ch\xed minh", "tp.hcm", "tp hcm"],
    "Hanoi":             ["hanoi", "ha noi", "h\xe0 n\u1ed9i", "capital hanoi"],
    "Da Nang":           ["da nang", "\u0111\xe0 n\u1eb5ng", "danang"],
    "Hai Phong":         ["hai phong", "h\u1ea3i ph\xf2ng", "haiphong", "l\u1ea1ch huy\u1ec7n", "c\u1ea3ng h\u1ea3i ph\xf2ng"],
    "Can Tho":           ["can tho", "c\u1ea7n th\u01a1"],
    "Binh Duong":        ["binh duong", "b\xecnh d\u01b0\u01a1ng"],
    "Dong Nai":          ["dong nai", "\u0111\u1ed3ng nai"],
    "Ba Ria - Vung Tau": ["ba ria", "vung tau", "v\u0169ng t\xe0u", "b\xe0 r\u1ecba"],
    "Long An":           ["long an"],
    "Quang Ninh":        ["quang ninh", "qu\u1ea3ng ninh", "ha long bay", "h\u1ea1 long",
                          "h\u1ea1 long bay", "v\xe2n \u0111\u1ed3n", "m\xf3ng c\xe1i", "c\u1ea9m ph\u1ea3", "u\xf4ng b\xed"],
    "Bac Ninh":          ["bac ninh", "b\u1eafc ninh", "y\xean phong", "vsip b\u1eafc ninh", "qu\u1ebf v\xf5"],
    "Bac Giang":         ["bac giang", "b\u1eafc giang", "vsip b\u1eafc giang", "viettel b\u1eafc giang",
                          "c\xf4ng h\xf2a", "quang ch\xe2u"],
    "Hai Duong":         ["hai duong", "h\u1ea3i d\u01b0\u01a1ng"],
    "Hung Yen":          ["hung yen", "h\u01b0ng y\xean", "th\u0103ng long ii"],
    "Vinh Phuc":         ["vinh phuc", "v\u0129nh ph\xfac"],
    "Thai Nguyen":       ["thai nguyen", "th\xe1i nguy\xean", "samsung th\xe1i nguy\xean"],
    "Phu Tho":           ["phu tho", "ph\xfa th\u1ecd"],
    "Hoa Binh":          ["hoa binh", "h\xf2a b\xecnh"],
    "Lang Son":          ["lang son", "l\u1ea1ng s\u01a1n", "h\u1eefu ngh\u1ecb"],
    "Yen Bai":           ["yen bai", "y\xean b\xe1i"],
    "Lao Cai":           ["lao cai", "l\xe0o cai", "sa pa", "sapa"],
    "Tuyen Quang":       ["tuyen quang", "tuy\xean quang"],
    "Ha Giang":          ["ha giang", "h\xe0 giang"],
    "Cao Bang":          ["cao bang", "cao b\u1eb1ng"],
    "Bac Kan":           ["bac kan", "b\u1eafc k\u1ea1n"],
    "Thai Binh":         ["thai binh", "th\xe1i b\xecnh"],
    "Nam Dinh":          ["nam dinh", "nam \u0111\u1ecbnh"],
    "Ninh Binh":         ["ninh binh", "ninh b\xecnh", "tr\xe0ng an"],
    "Ha Nam":            ["ha nam", "h\xe0 nam", "\u0111\u1ed3ng v\u0103n"],
    "Son La":            ["son la", "s\u01a1n la", "nho qu\u1ebf"],
    "Dien Bien":         ["dien bien", "\u0111i\u1ec7n bi\xean"],
    "Lai Chau":          ["lai chau", "lai ch\xe2u"],
    "Thanh Hoa":         ["thanh hoa", "thanh ho\xe1", "thanh h\xf3a"],
    "Nghe An":           ["nghe an", "ngh\u1ec7 an"],
    "Ha Tinh":           ["ha tinh", "h\xe0 t\u0129nh"],
    "Quang Binh":        ["quang binh", "qu\u1ea3ng b\xecnh"],
    "Thua Thien Hue":    ["thua thien hue", "hu\u1ebf", " hue ", "thua thien"],
    "Quang Nam":         ["quang nam", "qu\u1ea3ng nam"],
    "Quang Ngai":        ["quang ngai", "qu\u1ea3ng ng\xe3i"],
    "Binh Dinh":         ["binh dinh", "b\xecnh \u0111\u1ecbnh"],
    "Khanh Hoa":         ["khanh hoa", "kh\xe1nh h\xf2a", "nha trang"],
    "Lam Dong":          ["lam dong", "l\xe2m \u0111\u1ed3ng", "da lat", "\u0111\xe0 l\u1ea1t", "dalat"],
    "Dak Lak":           ["dak lak", "\u0111\u1eafk l\u1eafk", "buon ma thuot"],
    "Gia Lai":           ["gia lai"],
    "Kon Tum":           ["kon tum"],
    "Tien Giang":        ["tien giang", "ti\u1ec1n giang"],
    "Ben Tre":           ["ben tre", "b\u1ebfn tre"],
    "Vinh Long":         ["vinh long", "v\u0129nh long"],
    "Dong Thap":         ["dong thap", "\u0111\u1ed3ng th\xe1p"],
    "An Giang":          ["an giang"],
    "Kien Giang":        ["kien giang", "ki\xean giang", "phu quoc", "ph\xfa qu\u1ed1c"],
    "Ca Mau":            ["ca mau", "c\xe0 mau"],
    "Long Thanh":        ["long thanh airport", "long th\xe0nh airport"],
    "Mekong Delta":      ["mekong delta", "mekong region"],
    "Central Highlands": ["central highlands", "tay nguyen"],
}


# ============================================================
# RSS FEEDS — v5.4 (총 44개)
# ============================================================

RSS_FEEDS = {
    # ── 영문 주요 소스 ─────────────────────────────────────────
    "VnExpress English - News":        "https://e.vnexpress.net/rss/news.rss",
    "VnExpress English - Business":    "https://e.vnexpress.net/rss/business.rss",
    "Vietnam News - Economy":          "https://vietnamnews.vn/rss/economy.rss",
    "Tuoi Tre News":                   "https://tuoitre.vn/rss/tin-moi-nhat.rss",
    "SGGP News English":               "https://en.sggp.org.vn/rss/home.rss",
    "Nhan Dan English":                "https://en.nhandan.vn/rss/home.rss",
    # ── 영문 전문 소스 ─────────────────────────────────────────
    "PV-Tech":                         "https://www.pv-tech.org/feed/",
    # ── 베트남어 일반 ──────────────────────────────────────────
    "VnExpress - Kinh doanh":          "https://vnexpress.net/rss/kinh-doanh.rss",
    "VnExpress - Thoi su":             "https://vnexpress.net/rss/thoi-su.rss",
    "Tuoi Tre - Kinh doanh":           "https://tuoitre.vn/rss/kinh-doanh.rss",
    "Thanh Nien - Kinh te":            "https://thanhnien.vn/rss/kinh-te.rss",
    "VietnamNet - Kinh doanh":         "https://vietnamnet.vn/rss/kinh-doanh.rss",
    "Dan Tri - Kinh doanh":            "https://dantri.com.vn/rss/kinh-doanh.rss",
    "CafeBiz":                         "https://cafebiz.vn/rss/home.rss",
    # ── 베트남어 전문 소스 ─────────────────────────────────────
    "Bao Xay Dung":                    "https://baoxaydung.com.vn/rss/home.rss",
    # ── 중부 지역 ─────────────────────────────────────────────
    "Bao Ha Tinh":                     "https://baohatinh.vn/rss/home.rss",
    "Bao Binh Dinh":                   "https://baobinhdinh.vn/rss/home.rss",
    # ── 남부 지역 ─────────────────────────────────────────────
    "SGGP":                            "https://www.sggp.org.vn/rss/home.rss",
    # ── v5.3: 환경 전문 P1 ────────────────────────────────────
    "VietnamPlus - Moi truong":        "https://www.vietnamplus.vn/rss/moitruong-270.rss",
    "VietnamPlus - Kinh te":           "https://www.vietnamplus.vn/rss/kinhte-311.rss",
    "Nhandan - Moi truong":            "https://baotintuc.vn/moi-truong.rss",
    "Kinhtemoitruong":                 "https://kinhtemoitruong.vn/rss",
    "Baotainguyenmoitruong":           "https://baotainguyenmoitruong.vn/rss/tin-tuc.rss",
    # ── v5.3: 북부 지역 P2 ────────────────────────────────────
    "Nhandan - Kinh te":               "https://baotintuc.vn/kinh-te.rss",
    "Hanoimoi - Kinh te":              "https://hanoimoi.vn/rss/kinh-te.rss",
    "Baobacgiang English":             "https://en.baobacgiang.vn/rss",
    "Baoquangninh - Kinh te":          "https://baoquangninh.vn/rss/kinh-te.rss",
    # ── v5.3: 보조 환경/교통 P3 ───────────────────────────────
    "Moitruong Net":                   "https://moitruong.net.vn/rss",
    "Congnghiepmoitruong":             "https://congnghiepmoitruong.vn/rss",
    "VietnamPlus - Giao thong":        "https://www.vietnamplus.vn/rss/xahoi/giaothong-358.rss",
    # ── 기존 에너지 전문 ──────────────────────────────────────
    "Bao Dau Tu - Energy":             "https://baodautu.vn/rss/nang-luong.rss",
    "Vietnam Energy alt":              "https://vietnamenergy.vn/rss/tin-tuc.rss",
    "Tap chi Xay dung":                "https://tapchixaydung.vn/rss/home.rss",
    # ── v5.4 신규: 전문미디어 (전문미디어 비율 30% 목표) ────────
    "The Investor":                    "https://theinvestor.vn/feed",
    "VIR - Vietnam Investment Review": "https://vir.com.vn/rss/news.aspx",
    "Construction Vietnam":            "https://constructionvietnam.net/feed",
    "VietnamBiz":                      "https://vietnambiz.vn/rss.rss",
    # ── v5.4 신규: Waste Water 전용 ──────────────────────────
    "MONRE Official":                  "https://monre.gov.vn/rss/tintuc.aspx",
    "VEA - Vietnam Environment":       "https://vea.gov.vn/vn/tintuc/tintuchangngay/rss",
    "Nhadepso Environment":            "https://nhadepso.com/feed/",
    "Moitruong Online":                "https://moitruong.com.vn/feed",
    # ── v5.4 신규: Smart City 전용 ───────────────────────────
    "ICT Vietnam":                     "https://ictvietnam.vn/feed",
    "MIC Vietnam":                     "https://mic.gov.vn/rss/tintuc.aspx",
    "Smartcity Vietnam":               "https://smartcity.mobi/feed",
}


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def log(msg):
    print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {msg}")


def clean_html(text):
    if not text:
        return ""
    soup = BeautifulSoup(text, 'html.parser')
    return html.unescape(soup.get_text(separator=' ', strip=True))


def generate_url_hash(url):
    return hashlib.md5(url.encode()).hexdigest()


def is_english_text(title):
    if not title:
        return False
    ascii_cnt = sum(1 for c in title if c.isascii() and c.isalpha())
    non_ascii  = sum(1 for c in title if not c.isascii())
    total = ascii_cnt + non_ascii
    return (ascii_cnt / total) > 0.7 if total else False


def is_vietnamese_text(title):
    if not title:
        return False
    vn_chars = set('àáảãạăằắẳẵặâầấẩẫậèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđ'
                   'ÀÁẢÃẠĂẰẮẲẴẶÂẦẤẨẪẬÈÉẺẼẸÊỀẾỂỄỆÌÍỈĨỊÒÓỎÕỌÔỒỐỔỖỘƠỜỚỞỠỢÙÚỦŨỤƯỪỨỬỮỰỲÝỶỸỴĐ')
    return any(c in vn_chars for c in title)


def passes_language_filter(title, mode=None):
    m = mode or LANGUAGE_FILTER
    if m == 'all':
        return True
    elif m == 'vietnamese':
        return is_vietnamese_text(title) or not is_english_text(title)
    else:
        return is_english_text(title)


def is_vietnam_related(title, summary=""):
    text = f"{title} {summary}".lower()
    if not any(kw in text for kw in VIETNAM_KEYWORDS):
        return False
    for country in NON_VIETNAM_COUNTRIES:
        if country in text and text.count("vietnam") < text.count(country):
            return False
    return True


def should_exclude(title, summary=""):
    if not title or len(title.strip()) < 15:
        return True
    text = f"{title} {summary}".lower()
    return any(kw in text for kw in EXCLUDE_KEYWORDS)


def extract_province(title, summary="", full_text=""):
    combined = f"{title} {summary} {full_text}".lower()
    for province, keywords in PROVINCE_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                return province
    return "Vietnam"


# ============================================================
# TRANSLATE ARTICLES
# ============================================================

def translate_text(text, target_lang='ko'):
    if not text or len(text.strip()) < 3:
        return text or ''
    src = 'en' if is_english_text(text) else 'vi'
    try:
        url = (f"https://api.mymemory.translated.net/get"
               f"?q={requests.utils.quote(text[:500])}"
               f"&langpair={src}|{target_lang}")
        r = requests.get(url, timeout=10)
        data = r.json()
        result = data.get('responseData', {}).get('translatedText', '')
        if result and result != text and 'INVALID' not in result.upper():
            return result
    except Exception:
        pass
    try:
        from deep_translator import GoogleTranslator
        result = GoogleTranslator(source='auto', target=target_lang).translate(text[:500])
        if result:
            return result
    except Exception:
        pass
    return text


def translate_articles(articles):
    if not articles:
        return articles

    log(f"Translating {len(articles)} articles (ko/en/vi)...")

    for i, art in enumerate(articles):
        title   = art.get('title', '') or ''
        summary = art.get('summary', '') or ''
        is_en   = is_english_text(title)
        is_vi   = is_vietnamese_text(title)

        try:
            art['title_ko']   = translate_text(title, 'ko')
            art['title_en']   = title if is_en else translate_text(title, 'en')
            art['title_vi']   = title if is_vi else translate_text(title, 'vi')
            art['summary_ko'] = translate_text(summary[:300], 'ko') if summary else ''
            art['summary_en'] = summary if is_en else (translate_text(summary[:300], 'en') if summary else '')
            art['summary_vi'] = summary if is_vi else (translate_text(summary[:300], 'vi') if summary else '')
        except Exception as e:
            log(f"  Translation error [{art.get('title','')[:40]}]: {e}")
            art.setdefault('title_ko', '')
            art.setdefault('title_en', title)
            art.setdefault('title_vi', '')
            art.setdefault('summary_ko', '')
            art.setdefault('summary_en', summary)
            art.setdefault('summary_vi', '')

        if (i + 1) % 3 == 0:
            time.sleep(0.5)
        if (i + 1) % 10 == 0:
            log(f"  Translated {i+1}/{len(articles)}")

    log(f"Translation complete: {len(articles)} articles")
    return articles


# ============================================================
# CLASSIFY SECTOR
# ============================================================

def classify_sector(title, summary=""):
    if should_exclude(title, summary):
        return None, None, 0

    text_title = title.lower()
    text_full  = f"{title} {summary}".lower()

    scores = {}
    for sector, kw_dict in SECTOR_KEYWORDS.items():
        score = 0
        for kw in kw_dict.get("primary", []):
            if kw in text_title:
                score += 3
            elif kw in text_full:
                score += 1
        for kw in kw_dict.get("secondary", []):
            if kw in text_title:
                score += 2
            elif kw in text_full:
                score += 1
        scores[sector] = score

    if not scores:
        return None, None, 0

    best_sector = max(scores, key=scores.get)
    best_score  = scores[best_sector]

    if best_score < MIN_CLASSIFY_THRESHOLD:
        return None, None, 0

    confidence = min(100, int(best_score / 20 * 100))
    return best_sector, SECTOR_AREA[best_sector], confidence


# ============================================================
# RSS FETCH
# ============================================================

def parse_date(date_str):
    if not date_str:
        return None
    formats = [
        "%a, %d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S %Z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    try:
        from email.utils import parsedate_to_datetime
        return parsedate_to_datetime(date_str)
    except Exception:
        pass
    return None


def fetch_rss(url, timeout=30):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0',
        'Accept': 'application/rss+xml, application/xml, text/xml, */*',
    }
    try:
        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
        content = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', resp.text)
        return feedparser.parse(content)
    except Exception as e:
        log(f"  RSS error [{url}]: {e}")
        return type('Feed', (), {'entries': [], 'bozo': True})()


# ============================================================
# GOOGLE NEWS API
# ============================================================

def fetch_gnews(query, hours_back=24, max_articles=20):
    if not GNEWS_API_KEY:
        return []

    articles = []
    try:
        is_newsdata = GNEWS_API_KEY.startswith('pub_')

        if is_newsdata:
            url = (
                f"https://newsdata.io/api/1/news"
                f"?apikey={GNEWS_API_KEY}"
                f"&q={quote(query)}"
                f"&country=vn"
                f"&language=en,vi"
                f"&category=business,politics,technology"
                f"&size={min(max_articles, 10)}"
            )
            resp = requests.get(url, timeout=15)
            data = resp.json()
            for item in data.get('results', []):
                title = item.get('title', '') or ''
                if not title or len(title.strip()) < 15:
                    continue
                articles.append({
                    'title':          title,
                    'url':            item.get('link', ''),
                    'published_date': (item.get('pubDate', '') or '')[:10],
                    'source_name':    item.get('source_id', 'NewsData'),
                    'raw_summary':    item.get('description', '') or '',
                })
            log(f"NewsData.io: {len(articles)} articles for query '{query[:50]}'")
        else:
            from_dt = (datetime.utcnow() - timedelta(hours=min(hours_back, 720))).strftime('%Y-%m-%dT%H:%M:%SZ')
            url = (
                f"https://gnews.io/api/v4/search"
                f"?q={quote(query)}"
                f"&lang=en&country=vn"
                f"&from={from_dt}"
                f"&max={max_articles}"
                f"&apikey={GNEWS_API_KEY}"
            )
            resp = requests.get(url, timeout=15)
            data = resp.json()
            for item in data.get('articles', []):
                articles.append({
                    'title':          item.get('title', ''),
                    'url':            item.get('url', ''),
                    'published_date': item.get('publishedAt', '')[:10],
                    'source_name':    item.get('source', {}).get('name', 'GNews'),
                    'raw_summary':    item.get('description', ''),
                })
            log(f"GNews: {len(articles)} articles for query '{query[:50]}'")

    except Exception as e:
        log(f"News API error: {e}")
    return articles


# ============================================================
# NEWSDATA.io API
# ============================================================

NEWSDATA_API_KEY = os.environ.get('NEWSDATA_API_KEY', '')

NEWSDATA_PROVINCE_QUERIES = {
    'group_a': [
        {'province': 'Da Nang',    'q': '"Da Nang" AND (infrastructure OR "industrial park" OR wastewater OR "water supply" OR transport)', 'language': 'en'},
        {'province': 'Da Nang',    'q': '"Da Nang" AND ("khu cong nghiep" OR "nuoc thai" OR "cap nuoc" OR "giao thong")', 'language': 'vi'},
        {'province': 'Binh Duong', 'q': '"Binh Duong" AND ("industrial park" OR wastewater OR "power" OR infrastructure)', 'language': 'en'},
        {'province': 'Binh Duong', 'q': '"Binh Duong" AND ("khu cong nghiep" OR "nuoc thai" OR "dien")', 'language': 'vi'},
        {'province': 'Quang Ninh', 'q': '"Quang Ninh" AND ("wind farm" OR "power plant" OR "port" OR "industrial" OR "coal")', 'language': 'en'},
        {'province': 'Dong Nai',   'q': '"Dong Nai" AND ("industrial park" OR wastewater OR transport OR infrastructure)', 'language': 'en'},
        {'province': 'Bac Ninh',   'q': '"Bac Ninh" AND ("industrial park" OR "semiconductor" OR wastewater OR infrastructure)', 'language': 'en'},
    ],
    'group_b': [
        {'province': 'Ba Ria Vung Tau', 'q': '"Vung Tau" AND ("oil" OR "gas" OR "LNG" OR "port" OR "petrochemical")', 'language': 'en'},
        {'province': 'Binh Dinh',       'q': '"Binh Dinh" AND (infrastructure OR "industrial park" OR wastewater OR "renewable")', 'language': 'en'},
        {'province': 'Quang Nam',       'q': '"Quang Nam" AND (infrastructure OR "industrial park" OR transport OR "water")', 'language': 'en'},
        {'province': 'Thai Nguyen',     'q': '"Thai Nguyen" AND ("industrial park" OR "Samsung" OR infrastructure OR wastewater)', 'language': 'en'},
        {'province': 'Ben Tre',         'q': '"Ben Tre" AND (infrastructure OR "water supply" OR "renewable energy" OR transport)', 'language': 'en'},
        {'province': 'Bac Giang',       'q': '"Bac Giang" AND ("industrial park" OR "VSIP" OR wastewater OR infrastructure)', 'language': 'en'},
    ],
    'group_c': [
        {'province': 'Tien Giang', 'q': '"Tien Giang" AND (infrastructure OR "water supply" OR transport OR "industrial")', 'language': 'en'},
        {'province': 'Hai Duong',  'q': '"Hai Duong" AND ("industrial park" OR wastewater OR infrastructure)', 'language': 'en'},
        {'province': 'Ninh Thuan', 'q': '"Ninh Thuan" AND ("wind" OR "solar" OR "renewable energy" OR infrastructure)', 'language': 'en'},
        {'province': 'Quang Binh', 'q': '"Quang Binh" AND (infrastructure OR transport OR "industrial" OR "water")', 'language': 'en'},
    ],
}

NEWSDATA_SPECIALIST_QUERIES = [
    {'source': 'The Investor',          'domain': 'theinvestor.vn',  'q': 'infrastructure OR "industrial park" OR wastewater OR "power plant" OR "oil gas"', 'language': 'en'},
    {'source': 'Vietnam Investment Review', 'domain': 'vir.com.vn',  'q': 'infrastructure OR investment OR "industrial zone" OR energy OR transport',        'language': 'en'},
    {'source': 'Hanoi Times',           'domain': 'hanoitimes.vn',   'q': 'infrastructure OR "industrial park" OR wastewater OR metro OR "urban development"','language': 'en'},
    {'source': 'Vietnam Energy',        'domain': 'vietnamenergy.vn','q': 'power OR energy OR "renewable" OR "solar" OR "wind" OR "LNG"',                    'language': 'en'},
    {'source': 'Da Nang News',          'domain': 'baodanang.vn',    'q': 'infrastructure OR "khu cong nghiep" OR "nuoc thai" OR "giao thong"',              'language': 'vi'},
    {'source': 'Bao Dau Tu',            'domain': 'baodautu.vn',     'q': 'infrastructure OR "khu cong nghiep" OR "nang luong" OR "giao thong"',             'language': 'vi'},
    {'source': 'PetroTimes',            'domain': 'petrotimes.vn',   'q': 'oil OR gas OR LNG OR petroleum OR "dau khi" OR petrovietnam',                     'language': 'vi'},
]


def fetch_newsdata(hours_back=24):
    if not NEWSDATA_API_KEY:
        log("NewsData.io: NEWSDATA_API_KEY 없음 — 스킵")
        return []

    if not NEWSDATA_API_KEY.startswith('pub_'):
        log("NewsData.io: 올바른 API 키 형식 아님 — 스킵")
        return []

    import datetime as dt_module
    today        = dt_module.datetime.utcnow()
    day_of_week  = today.weekday()
    day_of_month = today.day

    articles     = []
    credit_used  = 0
    CREDIT_LIMIT = 195

    def _call_newsdata(q, language='en', domain=None, size=10):
        nonlocal credit_used
        if credit_used + size > CREDIT_LIMIT:
            log(f"  NewsData.io: 크레딧 한도 도달 ({credit_used}/{CREDIT_LIMIT})")
            return []
        params = {
            'apikey': NEWSDATA_API_KEY, 'q': q,
            'country': 'vn', 'language': language,
            'category': 'business,politics,technology,environment',
            'size': size,
            'from_date': (today - timedelta(hours=min(hours_back, 720))).strftime('%Y-%m-%d'),
        }
        if domain:
            params['domain'] = domain
        try:
            resp = requests.get('https://newsdata.io/api/1/news', params=params, timeout=15)
            resp.raise_for_status()
            data = resp.json()
            if data.get('status') != 'success':
                log(f"  NewsData.io 오류: {data.get('message','unknown')}")
                return []
            results = data.get('results', [])
            credit_used += len(results)
            return results
        except Exception as e:
            log(f"  NewsData.io 호출 실패: {e}")
            return []

    def _parse_result(item, source_name, province_hint=None):
        title    = (item.get('title') or '').strip()
        url      = (item.get('link') or '').strip()
        summary  = (item.get('description') or '').strip()
        pub_date = (item.get('pubDate') or '')[:10]
        source   = item.get('source_id') or source_name
        if not title or not url or len(title) < 15:
            return None
        if not url.startswith('http'):
            return None
        return {
            'url_hash':       generate_url_hash(url),
            'url':            url,
            'title':          title,
            'summary':        summary[:1000],
            'source':         f"NewsData/{source_name}",
            'source_name':    f"NewsData/{source_name}",
            'published_date': pub_date,
            'raw_summary':    summary[:500],
            '_province_hint': province_hint,
        }

    log(f"NewsData.io: 수집 시작 (크레딧 한도 {CREDIT_LIMIT}건)")

    log("  [Group A] 핵심 급감 Province...")
    for q_info in NEWSDATA_PROVINCE_QUERIES['group_a']:
        if credit_used >= CREDIT_LIMIT:
            break
        results = _call_newsdata(q_info['q'], q_info['language'], size=10)
        for item in results:
            parsed = _parse_result(item, q_info['province'], q_info['province'])
            if parsed:
                articles.append(parsed)
        if results:
            log(f"    {q_info['province']}: {len(results)}건")
        time.sleep(0.3)

    if day_of_month % 2 == 1:
        log("  [Group B] 완전 단절 Province (홀수일)...")
        for q_info in NEWSDATA_PROVINCE_QUERIES['group_b']:
            if credit_used >= CREDIT_LIMIT:
                break
            results = _call_newsdata(q_info['q'], q_info['language'], size=10)
            for item in results:
                parsed = _parse_result(item, q_info['province'], q_info['province'])
                if parsed:
                    articles.append(parsed)
            if results:
                log(f"    {q_info['province']}: {len(results)}건")
            time.sleep(0.3)

    log("  [전문미디어] domain 타겟팅...")
    for q_info in NEWSDATA_SPECIALIST_QUERIES:
        if credit_used >= CREDIT_LIMIT:
            break
        results = _call_newsdata(q_info['q'], q_info['language'], domain=q_info['domain'], size=10)
        for item in results:
            parsed = _parse_result(item, q_info['source'])
            if parsed:
                articles.append(parsed)
        if results:
            log(f"    {q_info['source']}: {len(results)}건")
        time.sleep(0.3)

    if day_of_week in (0, 3):
        log("  [Group C] 소규모 Province (월·목)...")
        for q_info in NEWSDATA_PROVINCE_QUERIES['group_c']:
            if credit_used >= CREDIT_LIMIT:
                break
            results = _call_newsdata(q_info['q'], q_info['language'], size=5)
            for item in results:
                parsed = _parse_result(item, q_info['province'], q_info['province'])
                if parsed:
                    articles.append(parsed)
            if results:
                log(f"    {q_info['province']}: {len(results)}건")
            time.sleep(0.3)

    log(f"NewsData.io 완료: {len(articles)}건 | 크레딧: {credit_used}/{CREDIT_LIMIT}")
    return articles


# ============================================================
# DATABASE
# ============================================================

def init_database(db_path):
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS articles (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            url_hash       TEXT UNIQUE,
            url            TEXT,
            title          TEXT,
            title_vi       TEXT,
            title_ko       TEXT,
            summary        TEXT,
            summary_vi     TEXT,
            summary_ko     TEXT,
            source         TEXT,
            sector         TEXT,
            area           TEXT,
            province       TEXT,
            confidence     INTEGER DEFAULT 0,
            published_date TEXT,
            collected_date TEXT,
            processed      INTEGER DEFAULT 0
        )
    ''')
    conn.commit()
    return conn


def get_existing_hashes(conn):
    cur    = conn.execute("SELECT url_hash FROM articles")
    hashes = {row[0] for row in cur.fetchall()}
    _excel = os.environ.get('EXCEL_PATH', EXCEL_PATH)
    try:
        if Path(_excel).exists():
            import openpyxl
            _wb = openpyxl.load_workbook(_excel, read_only=True, data_only=True)
            _ws = _wb.active
            _link_col = 7
            for _c in range(1, _ws.max_column + 1):
                _h = str(_ws.cell(1, _c).value or '').lower()
                if _h in ('link', 'url'):
                    _link_col = _c
                    break
            for _row in _ws.iter_rows(min_row=2, values_only=True):
                _url = _row[_link_col - 1] if _link_col - 1 < len(_row) else None
                if _url and str(_url).startswith('http'):
                    hashes.add(hashlib.md5(str(_url).encode()).hexdigest())
            _wb.close()
            log(f"  Loaded {len(hashes)} existing URL hashes (SQLite + Excel)")
    except Exception as _e:
        log(f"  Excel hash load warning: {_e}")
    return hashes


def save_article(conn, article):
    try:
        conn.execute('''
            INSERT INTO articles
              (url_hash, url, title, summary, source,
               sector, area, province, confidence,
               published_date, collected_date)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            article['url_hash'], article['url'], article['title'],
            article.get('summary', ''), article['source'],
            article['sector'], article.get('area', ''),
            article.get('province', 'Vietnam'),
            article.get('confidence', 0),
            article.get('published_date', ''),
            datetime.now().isoformat(),
        ))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False


# ============================================================
# MAIN COLLECTION
# ============================================================

def collect_news(hours_back=24):
    conn            = init_database(DB_PATH)
    existing_hashes = get_existing_hashes(conn)
    cutoff          = datetime.now() - timedelta(hours=hours_back)

    log(f"Cutoff: {cutoff:%Y-%m-%d %H:%M} | Language: {LANGUAGE_FILTER} | Threshold: {MIN_CLASSIFY_THRESHOLD}")
    log(f"RSS feeds: {len(RSS_FEEDS)} (v5.4: +11 specialist+wastewater+smartcity)")

    total_collected    = 0
    collected_articles = []
    collection_stats   = {}

    for source_name, feed_url in RSS_FEEDS.items():
        stats = {
            'url': feed_url, 'status': 'Unknown',
            'last_check': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'entries_found': 0, 'collected': 0, 'error': '',
        }
        collection_stats[source_name] = stats

        feed = fetch_rss(feed_url)
        if feed.bozo and not feed.entries:
            stats['status'] = 'Failed'
            stats['error']  = 'Feed error or empty'
            continue

        stats['entries_found'] = len(feed.entries)
        stats['status']        = 'Success'
        source_cnt = 0

        for entry in feed.entries:
            title   = clean_html(getattr(entry, 'title', ''))
            link    = getattr(entry, 'link', '')
            summary = clean_html(getattr(entry, 'summary', getattr(entry, 'description', '')))
            pubdate = getattr(entry, 'published', getattr(entry, 'pubDate', ''))

            if not title or not passes_language_filter(title):
                continue

            url_hash = generate_url_hash(link)
            if url_hash in existing_hashes:
                continue

            pub_dt = parse_date(pubdate)
            if pub_dt:
                if pub_dt.tzinfo:
                    pub_dt = pub_dt.replace(tzinfo=None)
                if pub_dt < cutoff:
                    continue

            if not is_vietnam_related(title, summary):
                continue

            sector, area, confidence = classify_sector(title, summary)
            if not sector:
                continue

            province = extract_province(title, summary)

            article = {
                'url_hash':       url_hash,
                'url':            link,
                'title':          title,
                'summary':        summary[:1000] if summary else '',
                'source':         source_name,
                'source_name':    source_name,
                'sector':         sector,
                'area':           area,
                'province':       province,
                'confidence':     confidence,
                'published_date': pub_dt.isoformat() if pub_dt else '',
                'raw_summary':    summary[:500] if summary else '',
            }

            if save_article(conn, article):
                existing_hashes.add(url_hash)
                source_cnt      += 1
                total_collected += 1
                collected_articles.append(article)
                log(f"  SAVED [{sector}|{confidence}%] [{province}] {title[:55]}...")

        stats['collected'] = source_cnt

    if ENABLE_GNEWS:
        log("GNews: fetching supplemental articles...")
        gnews_raw  = fetch_gnews(GNEWS_QUERY, hours_back)
        gnews_raw += fetch_gnews(GNEWS_ENV_QUERY, hours_back, max_articles=15)
        gnews_raw += fetch_gnews(GNEWS_NORTH_QUERY, hours_back, max_articles=10)

        for item in gnews_raw:
            title   = item.get('title', '')
            link    = item.get('url', '')
            summary = item.get('raw_summary', '')
            if not title or not is_vietnam_related(title, summary):
                continue
            url_hash = generate_url_hash(link)
            if url_hash in existing_hashes:
                continue
            sector, area, confidence = classify_sector(title, summary)
            if not sector:
                continue
            province = extract_province(title, summary)
            article  = {
                'url_hash': url_hash, 'url': link,
                'title': title, 'summary': summary[:1000],
                'source': item.get('source_name', 'GNews'),
                'source_name': item.get('source_name', 'GNews'),
                'sector': sector, 'area': area,
                'province': province, 'confidence': confidence,
                'published_date': item.get('published_date', ''),
                'raw_summary': summary[:500],
            }
            if save_article(conn, article):
                existing_hashes.add(url_hash)
                total_collected += 1
                collected_articles.append(article)

    if NEWSDATA_API_KEY:
        log("NewsData.io: Province 공백 + 전문미디어 보완 수집...")
        newsdata_raw = fetch_newsdata(hours_back)
        for item in newsdata_raw:
            title   = item.get('title', '') or ''
            link    = item.get('url', '') or ''
            summary = item.get('raw_summary', '') or ''
            if not title or len(title.strip()) < 15:
                continue
            if not is_vietnam_related(title, summary):
                continue
            url_hash = item.get('url_hash') or generate_url_hash(link)
            if url_hash in existing_hashes:
                continue
            sector, area, confidence = classify_sector(title, summary)
            if not sector:
                continue
            province_hint = item.get('_province_hint')
            province      = province_hint or extract_province(title, summary)
            article = {
                'url_hash':       url_hash, 'url': link,
                'title':          title, 'summary': summary[:1000],
                'source':         item.get('source_name', 'NewsData'),
                'source_name':    item.get('source_name', 'NewsData'),
                'sector':         sector, 'area': area,
                'province':       province, 'confidence': confidence,
                'published_date': item.get('published_date', ''),
                'raw_summary':    summary[:500],
            }
            if save_article(conn, article):
                existing_hashes.add(url_hash)
                total_collected += 1
                collected_articles.append(article)
                log(f"  [NewsData][{sector}|{province}] {title[:55]}...")

        nd_count = sum(1 for a in collected_articles if 'NewsData' in a.get('source', ''))
        log(f"  NewsData.io 기여: {nd_count}건")

    conn.close()

    from collections import Counter
    sector_counts   = Counter(a['sector']   for a in collected_articles)
    province_counts = Counter(a['province'] for a in collected_articles)
    low_conf        = sum(1 for a in collected_articles if a.get('confidence', 0) < 50)

    print("\n" + "=" * 60)
    print(f"COLLECTION COMPLETE  |  {total_collected} articles")
    print("-" * 60)
    print("Sector breakdown:")
    for s, c in sector_counts.most_common():
        print(f"  {s:<28} {c:3d}")
    print(f"\nLow-confidence articles (<50%): {low_conf}")
    print(f"Top province: {province_counts.most_common(1)}")
    print("=" * 60)

    return total_collected, collected_articles, collection_stats


# ============================================================
# EXCEL UPDATE
# ============================================================

def update_excel_database(articles, collection_stats=None, excel_path=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from collections import Counter
        import shutil
    except ImportError:
        log("openpyxl not installed")
        return False

    _ep_str = excel_path or os.environ.get('EXCEL_PATH', EXCEL_PATH)
    ep = Path(_ep_str)
    if not ep.exists():
        log(f"Excel not found: {ep}")
        return False

    try:
        wb_c = openpyxl.load_workbook(ep, read_only=True)
        ws_c = wb_c.active
        existing_count = sum(1 for r in ws_c.iter_rows(min_row=2, values_only=True) if any(r))
        wb_c.close()
        if existing_count < 100:
            log(f"Safety check failed: only {existing_count} rows")
            return False
    except Exception as e:
        log(f"Safety check error: {e}")
        return False

    try:
        shutil.copy2(ep, ep.with_suffix('.xlsx.backup'))
    except Exception:
        pass

    try:
        wb       = openpyxl.load_workbook(ep)
        ws       = wb.active
        last_row = ws.max_row

        url_col = 7
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h and "link" in str(h).lower():
                url_col = c
                break

        existing_urls = set()
        for row in range(2, last_row + 1):
            v = ws.cell(row=row, column=url_col).value
            if v:
                existing_urls.add(v)

        NEW_FILL    = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        NEW_FONT    = Font(bold=True, color="1A1A1A", size=10)
        ENV_FILL    = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        ENERGY_FILL = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        URBAN_FILL  = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
        PLAIN_FONT  = Font(color="1A1A1A", size=10)
        HDR_FILL    = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
        thin_side   = Side(style='thin', color='E2E8F0')
        thin_border = Border(bottom=thin_side)

        col_map = {
            'area':1,'sector':2,'province':3,'title':4,'date':5,
            'source':6,'url':7,'summary':8,
            'title_ko':9,'title_en':10,'title_vi':11,
            'summary_ko':12,'summary_en':13,'summary_vi':14,
        }

        def area_fill(area):
            a = str(area).lower()
            if 'environment' in a: return ENV_FILL
            if 'energy' in a:      return ENERGY_FILL
            return URBAN_FILL

        added    = 0
        new_urls = set()
        for art in articles:
            if art.get('url') in existing_urls:
                continue
            nr = last_row + 1 + added
            ws.cell(row=nr, column=col_map['area'],       value=art.get('area', ''))
            ws.cell(row=nr, column=col_map['sector'],     value=art.get('sector', ''))
            ws.cell(row=nr, column=col_map['province'],   value=art.get('province', 'Vietnam'))
            ws.cell(row=nr, column=col_map['title'],      value=art.get('title', ''))
            ws.cell(row=nr, column=col_map['date'],       value=(art.get('published_date','') or '')[:10])
            ws.cell(row=nr, column=col_map['source'],     value=art.get('source', ''))
            ws.cell(row=nr, column=col_map['url'],        value=art.get('url', ''))
            ws.cell(row=nr, column=col_map['summary'],    value=art.get('summary', '')[:500])
            ws.cell(row=nr, column=col_map['title_ko'],   value=art.get('title_ko', ''))
            ws.cell(row=nr, column=col_map['title_en'],   value=art.get('title_en', ''))
            ws.cell(row=nr, column=col_map['title_vi'],   value=art.get('title_vi', ''))
            ws.cell(row=nr, column=col_map['summary_ko'], value=art.get('summary_ko', ''))
            ws.cell(row=nr, column=col_map['summary_en'], value=art.get('summary_en', ''))
            ws.cell(row=nr, column=col_map['summary_vi'], value=art.get('summary_vi', ''))
            for c in range(1, 15):
                ws.cell(row=nr, column=c).fill   = NEW_FILL
                ws.cell(row=nr, column=c).font   = NEW_FONT
                ws.cell(row=nr, column=c).border = thin_border
            added += 1
            new_urls.add(art.get('url'))
            existing_urls.add(art.get('url'))

        log(f"  +{added} new articles added (yellow highlight)")

        max_row = ws.max_row
        if added > 0 and max_row > 2:
            max_col_dyn = max(8, ws.max_column)
            rows_data   = []
            for r in range(2, max_row + 1):
                row_vals  = [ws.cell(row=r, column=c).value for c in range(1, max_col_dyn + 1)]
                title_val = row_vals[col_map['title']-1] if col_map['title']-1 < len(row_vals) else None
                if not title_val or str(title_val).strip() == '':
                    continue
                date_key = str(row_vals[col_map['date']-1] or '0000-00-00')[:10]
                url_key  = str(row_vals[col_map['url']-1]  or '')
                rows_data.append({'vals': row_vals, 'date': date_key, 'is_new': url_key in new_urls})

            rows_data.sort(key=lambda x: x['date'], reverse=True)

            for i, rd in enumerate(rows_data, 2):
                fill = NEW_FILL if rd['is_new'] else area_fill(rd['vals'][0])
                font = NEW_FONT if rd['is_new'] else PLAIN_FONT
                for c in range(1, max_col_dyn + 1):
                    cell        = ws.cell(row=i, column=c)
                    cell.value  = rd['vals'][c-1] if c-1 < len(rd['vals']) else None
                    if c <= 8:
                        cell.fill = fill
                        cell.font = font
                    cell.border = thin_border

            log(f"  Sorted {max_row-1} rows newest-first")

        for col, w in zip('ABCDEFGHIJKLMN', [18,22,20,60,12,22,50,60,40,40,40,50,50,50]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A2'

        if collection_stats:
            for sn in ["RSS_Sources"]:
                if sn in wb.sheetnames:
                    wb.remove(wb[sn])
            ws_rss = wb.create_sheet("RSS_Sources")
            for ci, h in enumerate(["Source","URL","Status","Last Check","Entries","Collected","Error"], 1):
                c = ws_rss.cell(row=1, column=ci, value=h)
                c.fill = HDR_FILL; c.font = HDR_FONT
                c.alignment = Alignment(horizontal='center')
            for ri, (src, st) in enumerate(collection_stats.items(), 2):
                ws_rss.cell(row=ri,column=1,value=src)
                ws_rss.cell(row=ri,column=2,value=st.get('url',''))
                ws_rss.cell(row=ri,column=3,value=st.get('status',''))
                ws_rss.cell(row=ri,column=4,value=st.get('last_check',''))
                ws_rss.cell(row=ri,column=5,value=st.get('entries_found',0))
                ws_rss.cell(row=ri,column=6,value=st.get('collected',0))
                ws_rss.cell(row=ri,column=7,value=st.get('error',''))
                sfill = ("D1FAE5" if st.get('status')=='Success' else
                         "FEE2E2" if st.get('status')=='Failed' else "F9FAFB")
                ws_rss.cell(row=ri,column=3).fill = PatternFill(start_color=sfill,end_color=sfill,fill_type="solid")
            for col,w in zip('ABCDEFG',[28,50,12,20,10,12,45]):
                ws_rss.column_dimensions[col].width = w

        _src_sn = "Source"
        if _src_sn not in wb.sheetnames:
            ws_src = wb.create_sheet(_src_sn)
            for _ci, _h in enumerate(["Domain","URL","Type","Status","Last Checked","Check Result","Articles Found","Note"], 1):
                _c = ws_src.cell(row=1, column=_ci, value=_h)
                _c.fill = HDR_FILL; _c.font = HDR_FONT
                _c.alignment = Alignment(horizontal='center')
        else:
            ws_src = wb[_src_sn]

        _domain_idx = {}
        _url_idx    = {}
        for _r in range(2, ws_src.max_row + 1):
            _d = ws_src.cell(row=_r, column=1).value
            _u = ws_src.cell(row=_r, column=2).value
            if _u and str(_u).startswith('http'):
                _url_idx[str(_u).rstrip('/')] = _r
                if _d:
                    _domain_idx[str(_d).lower().replace('www.','')] = _r

        def _ext_domain(_url):
            try:
                return urlparse(_url).netloc.lower().replace('www.','')
            except Exception:
                return _url

        now      = datetime.now()
        _run_date = now.strftime("%Y-%m-%d %H:%M")

        if collection_stats:
            for _sname, _st in collection_stats.items():
                _feed_url  = _st.get('url', '')
                if not _feed_url:
                    continue
                _domain    = _ext_domain(_feed_url)
                _status    = _st.get('status', 'Unknown')
                _collected = _st.get('collected', 0)
                _entries   = _st.get('entries_found', 0)
                _err       = _st.get('error', '') or ''

                if _status == 'Success' and _collected > 0:
                    _result = f"OK — {_entries} entries scanned, {_collected} collected"
                elif _status == 'Success':
                    _result = f"OK — {_entries} entries, 0 collected (no matching infra news)"
                else:
                    _result = f"FAILED: {_err[:70]}"

                _tr = _url_idx.get(_feed_url.rstrip('/')) or _domain_idx.get(_domain)
                if _tr:
                    ws_src.cell(row=_tr, column=4, value="Accessible" if _status=='Success' else "Inaccessible")
                    ws_src.cell(row=_tr, column=5, value=_run_date)
                    ws_src.cell(row=_tr, column=6, value=_result)
                    ws_src.cell(row=_tr, column=7, value=_collected)
                    if _err and _status == 'Failed':
                        ws_src.cell(row=_tr, column=8, value=_err[:120])
                else:
                    _tr = ws_src.max_row + 1
                    ws_src.cell(row=_tr, column=1, value=_domain)
                    ws_src.cell(row=_tr, column=2, value=_feed_url)
                    ws_src.cell(row=_tr, column=3, value="RSS Feed")
                    ws_src.cell(row=_tr, column=4, value="Accessible" if _status=='Success' else "Inaccessible")
                    ws_src.cell(row=_tr, column=5, value=_run_date)
                    ws_src.cell(row=_tr, column=6, value=_result)
                    ws_src.cell(row=_tr, column=7, value=_collected)
                    if _err:
                        ws_src.cell(row=_tr, column=8, value=_err[:120])
                    _url_idx[_feed_url.rstrip('/')] = _tr
                    _domain_idx[_domain] = _tr

                _sf = "D1FAE5" if _status == 'Success' else "FEE2E2"
                ws_src.cell(row=_tr, column=4).fill = PatternFill(start_color=_sf,end_color=_sf,fill_type="solid")

        _gnews_by_pub = {}
        _gnews_total  = 0
        for _art in articles:
            _asrc  = _art.get('source', '') or ''
            _aurl  = _art.get('url', '') or ''
            _is_gn = ('GNews' in _asrc or 'NewsData' in _asrc or
                      'gnews' in _aurl.lower() or 'newsdata' in _aurl.lower())
            if _is_gn:
                _gnews_total += 1
                _pub = (_asrc if _asrc not in ('GNews','NewsData') else _ext_domain(_aurl))
                _gnews_by_pub[_pub] = _gnews_by_pub.get(_pub, 0) + 1

        _gn_key = "gnews.io (Google News API)"
        _gn_row = None
        for _r in range(2, ws_src.max_row + 1):
            if str(ws_src.cell(row=_r, column=1).value or '').startswith('gnews.io'):
                _gn_row = _r
                break
        if not _gn_row:
            _gn_row = ws_src.max_row + 1

        ws_src.cell(row=_gn_row, column=1, value=_gn_key)
        ws_src.cell(row=_gn_row, column=2, value="https://gnews.io/api/v4/search")
        ws_src.cell(row=_gn_row, column=3, value="News API")
        ws_src.cell(row=_gn_row, column=4, value="Accessible" if _gnews_total > 0 else "Checked")
        ws_src.cell(row=_gn_row, column=5, value=_run_date)
        _pub_list = ', '.join(f"{k}({v})" for k, v in sorted(_gnews_by_pub.items(), key=lambda x: -x[1])[:10])
        ws_src.cell(row=_gn_row, column=6, value=(f"OK — {_gnews_total} articles | {_pub_list}" if _gnews_total > 0 else "Queried — 0 new articles")[:200])
        ws_src.cell(row=_gn_row, column=7, value=_gnews_total)
        ws_src.cell(row=_gn_row, column=8, value=f"Queries: Vietnam infra + environment + north | {_run_date}")
        ws_src.cell(row=_gn_row, column=4).fill = PatternFill(start_color="DBEAFE",end_color="DBEAFE",fill_type="solid")

        for _pub, _cnt in _gnews_by_pub.items():
            if not _pub:
                continue
            _ptr = _domain_idx.get(_pub.lower().replace('www.',''))
            if not _ptr:
                _ptr = ws_src.max_row + 1
                ws_src.cell(row=_ptr, column=1, value=_pub)
                ws_src.cell(row=_ptr, column=2, value=f"(via GNews API: {_pub})")
                ws_src.cell(row=_ptr, column=3, value="Media/News (via API)")
                ws_src.cell(row=_ptr, column=4, value="Accessible")
                _domain_idx[_pub.lower()] = _ptr
            ws_src.cell(row=_ptr, column=5, value=_run_date)
            ws_src.cell(row=_ptr, column=6, value=f"Collected via GNews API: {_cnt} articles")
            ws_src.cell(row=_ptr, column=7, value=_cnt)
            ws_src.cell(row=_ptr, column=8, value="Accessed via Google News API aggregation")

        for _col, _w in zip('ABCDEFGH', [30, 52, 18, 14, 20, 60, 16, 55]):
            ws_src.column_dimensions[_col].width = _w
        ws_src.freeze_panes = 'A2'

        _rss_ok  = sum(1 for _s in (collection_stats or {}).values() if _s.get('status')=='Success')
        _rss_tot = len(collection_stats) if collection_stats else 0
        log(f"Source sheet updated | RSS {_rss_ok}/{_rss_tot} OK | GNews {_gnews_total} articles")

        if "Collection_Log" not in wb.sheetnames:
            ws_log = wb.create_sheet("Collection_Log")
            for ci,h in enumerate(["Date","Time","Hours Back","Sources Checked","Success","Failed","New Articles","Total DB"],1):
                c = ws_log.cell(row=1,column=ci,value=h)
                c.fill=HDR_FILL; c.font=HDR_FONT
        else:
            ws_log = wb["Collection_Log"]

        now      = datetime.now()
        tot_src  = len(collection_stats) if collection_stats else 0
        ok_src   = sum(1 for s in (collection_stats or {}).values() if s.get('status')=='Success')
        log_row  = ws_log.max_row + 1
        cur_total = sum(1 for r in ws.iter_rows(min_row=2,values_only=True) if any(r))

        ws_log.cell(row=log_row,column=1,value=now.strftime("%Y-%m-%d"))
        ws_log.cell(row=log_row,column=2,value=now.strftime("%H:%M:%S"))
        ws_log.cell(row=log_row,column=3,value=HOURS_BACK)
        ws_log.cell(row=log_row,column=4,value=tot_src)
        ws_log.cell(row=log_row,column=5,value=ok_src)
        ws_log.cell(row=log_row,column=6,value=tot_src - ok_src)
        ws_log.cell(row=log_row,column=7,value=added)
        ws_log.cell(row=log_row,column=8,value=cur_total)
        today_hl = PatternFill(start_color="DBEAFE",end_color="DBEAFE",fill_type="solid")
        for c in range(1,9):
            ws_log.cell(row=log_row,column=c).fill = today_hl

        for sn in ["Summary"]:
            if sn in wb.sheetnames:
                wb.remove(wb[sn])
        ws_sum = wb.create_sheet("Summary")

        sectors_all = [str(ws.cell(row=r,column=2).value or '') for r in range(2,ws.max_row+1)
                       if any(ws.cell(row=r,column=c).value for c in range(1,9))]
        areas_all   = [str(ws.cell(row=r,column=1).value or '') for r in range(2,ws.max_row+1)
                       if any(ws.cell(row=r,column=c).value for c in range(1,9))]
        prov_all    = [str(ws.cell(row=r,column=3).value or '') for r in range(2,ws.max_row+1)
                       if any(ws.cell(row=r,column=c).value for c in range(1,9))]
        total_arts  = len(sectors_all)

        ws_sum.merge_cells('A1:D1')
        tc = ws_sum.cell(row=1,column=1,value="Vietnam Infrastructure News — Summary")
        tc.font=Font(bold=True,size=14,color="0F766E"); tc.alignment=Alignment(horizontal='center')
        ws_sum.cell(row=2,column=1,value=f"Updated: {now.strftime('%Y-%m-%d %H:%M')}  |  Total: {total_arts:,} articles")
        ws_sum.cell(row=2,column=1).font=Font(size=10,color="475569")

        r = 4
        ws_sum.cell(row=r,column=1,value="Business Sector").font=Font(bold=True,size=11)
        ws_sum.cell(row=r,column=2,value="Articles").font=Font(bold=True,size=11)
        ws_sum.cell(row=r,column=3,value="Share").font=Font(bold=True,size=11)
        r += 1
        for sect,cnt in Counter(sectors_all).most_common():
            ws_sum.cell(row=r,column=1,value=sect)
            ws_sum.cell(row=r,column=2,value=cnt)
            ws_sum.cell(row=r,column=3,value=f"{cnt/total_arts*100:.1f}%" if total_arts else "0%")
            r += 1

        r += 1
        ws_sum.cell(row=r,column=1,value="Area").font=Font(bold=True,size=11)
        r += 1
        for area,cnt in Counter(areas_all).most_common():
            ws_sum.cell(row=r,column=1,value=area); ws_sum.cell(row=r,column=2,value=cnt); r+=1

        r += 1
        ws_sum.cell(row=r,column=1,value="Top 15 Provinces").font=Font(bold=True,size=11)
        r += 1
        for prov,cnt in Counter(prov_all).most_common(15):
            ws_sum.cell(row=r,column=1,value=prov); ws_sum.cell(row=r,column=2,value=cnt); r+=1

        for col,w in zip('ABCD',[30,12,10,10]):
            ws_sum.column_dimensions[col].width = w

        _KH_SECTOR_ORDER = [
            "Waste Water","Water Supply/Drainage","Solid Waste",
            "Power","Oil & Gas","Transport",
            "Industrial Parks","Smart City","Construction"
        ]
        def _kh_sector_rank(s):
            try: return _KH_SECTOR_ORDER.index(str(s))
            except ValueError: return 99

        _AREA_CAT = {
            'Environment': 'Environment',
            'Energy Develop.': 'Energy Development',
            'Urban Development': 'Urban Development',
            'Urban Develop.': 'Urban Development',
        }

        _kh_sn = 'Keywords History'
        if _kh_sn not in wb.sheetnames:
            _ws_kh = wb.create_sheet(_kh_sn)
        else:
            _ws_kh = wb[_kh_sn]

        _kh_url_col = 8
        for _c in range(1, _ws_kh.max_column + 1):
            if str(_ws_kh.cell(1, _c).value or '').lower() == 'url':
                _kh_url_col = _c
                break
        _existing_kh_urls = set()
        for _r in range(2, _ws_kh.max_row + 1):
            _u = _ws_kh.cell(_r, _kh_url_col).value
            if _u:
                _existing_kh_urls.add(str(_u))

        _kh_new_articles = [
            a for a in articles
            if a.get('url') and str(a['url']) not in _existing_kh_urls
        ]

        if _kh_new_articles:
            if _ws_kh.cell(1, 1).value != 'No':
                _kh_headers = ['No','Category','Sector','Province','Date','Title','Source','URL','Summary (EN/KO)']
                _kh_wids    = [5, 16, 22, 18, 12, 65, 22, 50, 55]
                for _ci, (_h, _w) in enumerate(zip(_kh_headers, _kh_wids), 1):
                    _c = _ws_kh.cell(1, _ci, _h)
                    _c.fill = HDR_FILL; _c.font = HDR_FONT
                    _c.alignment = Alignment(horizontal='center')
                for _ci, _w in enumerate(_kh_wids, 1):
                    from openpyxl.utils import get_column_letter
                    _ws_kh.column_dimensions[get_column_letter(_ci)].width = _w
                _ws_kh.freeze_panes = 'A2'

            _start_r = _ws_kh.max_row + 1
            _cur_no  = _ws_kh.cell(_start_r - 1, 1).value or (_start_r - 2)
            try: _cur_no = int(_cur_no)
            except: _cur_no = _start_r - 2

            _kh_new_sorted = sorted(
                _kh_new_articles,
                key=lambda a: (
                    _kh_sector_rank(a.get('sector', '')),
                    str(a.get('province', 'Vietnam')),
                    str(a.get('published_date', '') or ''),
                ),
                reverse=False
            )

            for _a in _kh_new_sorted:
                _cur_no += 1
                _nr      = _ws_kh.max_row + 1
                _title   = (str(_a.get('title_en','') or '').strip() or
                            str(_a.get('title_ko','') or '').strip() or
                            str(_a.get('title','') or ''))
                _summary = (str(_a.get('summary_en','') or '').strip() or
                            str(_a.get('summary_ko','') or '').strip() or
                            str(_a.get('summary','') or ''))[:300]
                _sector  = str(_a.get('sector',''))
                _area    = str(_a.get('area',''))
                _cat     = _AREA_CAT.get(_area, _area)
                _date    = str(_a.get('published_date','') or '')[:10]

                _kh_vals = [
                    _cur_no, _cat, _sector,
                    str(_a.get('province','Vietnam')),
                    _date, _title,
                    str(_a.get('source','')),
                    str(_a.get('url','')),
                    _summary,
                ]
                for _ci, _val in enumerate(_kh_vals, 1):
                    _cell        = _ws_kh.cell(_nr, _ci, _val)
                    _cell.fill   = NEW_FILL
                    _cell.font   = NEW_FONT
                    _cell.border = thin_border
                    if _ci in (1, 5):
                        _cell.alignment = Alignment(horizontal='center', vertical='top')
                    else:
                        _cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
                _ws_kh.row_dimensions[_nr].height = 15

            log(f"Keywords History: +{len(_kh_new_articles)} new articles added (yellow)")
        else:
            log("Keywords History: no new articles to add")

        wb.save(ep)
        wb.close()
        log(f"Excel saved | +{added} new(yellow) | total {cur_total} | sorted by date desc")
        return True

    except Exception as e:
        log(f"Excel update error: {e}")
        import traceback; traceback.print_exc()
        return False


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description='Vietnam Infra News Collector v5.4')
    p.add_argument('--hours-back', type=int, default=HOURS_BACK)
    p.add_argument('--threshold',  type=int, default=MIN_CLASSIFY_THRESHOLD)
    p.add_argument('--gnews',      action='store_true', help='Enable Google News API')
    p.add_argument('--no-excel',   action='store_true', help='Skip Excel update')
    p.add_argument('--agent-mode', action='store_true', help='Save to collector_output.json')
    args = p.parse_args()

    HOURS_BACK             = args.hours_back
    MIN_CLASSIFY_THRESHOLD = args.threshold
    if args.gnews:
        ENABLE_GNEWS = True

    print("=" * 60)
    print("VIETNAM INFRASTRUCTURE NEWS COLLECTOR  v5.4")
    print(f"Hours back: {HOURS_BACK} | Threshold: {MIN_CLASSIFY_THRESHOLD} | Language: {LANGUAGE_FILTER}")
    print(f"RSS feeds: {len(RSS_FEEDS)} | v5.4 신규: 11개 (전문미디어+WasteWater+SmartCity)")
    print("=" * 60)

    cnt, arts, stats = collect_news(HOURS_BACK)

    if cnt > 0:
        arts = translate_articles(arts)

    if args.agent_mode:
        import json as _json
        from datetime import timezone as _tz
        _total        = len(arts)
        _vietnam_ratio = (
            sum(1 for a in arts if a.get('province', '') == 'Vietnam') / _total
            if _total > 0 else 0.0
        )
        _missing = [a.get('title', '')[:40] for a in arts if not a.get('province')]
        _out = {
            'run_timestamp':   datetime.now(_tz.utc).isoformat(),
            'hours_back':      HOURS_BACK,
            'total_collected': cnt,
            'articles': [
                {k: v for k, v in a.items() if k != 'url_hash'}
                for a in arts
            ],
            'quality_flags': {
                'vietnam_ratio':     round(_vietnam_ratio, 3),
                'missing_provinces': _missing,
            },
        }
        _json_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'data', 'agent_output', 'collector_output.json'
        )
        Path(_json_path).parent.mkdir(parents=True, exist_ok=True)
        with open(_json_path, 'w', encoding='utf-8') as _f:
            _json.dump(_out, _f, ensure_ascii=False, default=str)
        log(f"[agent-mode] Saved {cnt} articles to {_json_path}")

    elif args.no_excel:
        import json
        _out = {
            'count': cnt,
            'articles': [
                {k: v for k, v in a.items() if k != 'url_hash'}
                for a in arts
            ],
            'stats': {
                src: {
                    'url':           st.get('url',''),
                    'status':        st.get('status',''),
                    'entries_found': st.get('entries_found', 0),
                    'collected':     st.get('collected', 0),
                    'error':         st.get('error',''),
                }
                for src, st in stats.items()
            }
        }
        _json_path = os.environ.get('COLLECTOR_OUTPUT', 'data/collector_output.json')
        Path(_json_path).parent.mkdir(parents=True, exist_ok=True)
        with open(_json_path, 'w', encoding='utf-8') as _f:
            json.dump(_out, _f, ensure_ascii=False, default=str)
        log(f"Saved {cnt} articles to {_json_path} (Excel deferred to ExcelUpdater)")

    else:
        update_excel_database(arts, stats)

    print("\nRSS SOURCE STATUS:")
    for src, st in stats.items():
        icon = "✓" if st['status'] == 'Success' else "✗"
        print(f"  {icon} {src}: {st['entries_found']} entries → {st['collected']} collected")

    print(f"\nTOTAL: {cnt} new articles collected")
