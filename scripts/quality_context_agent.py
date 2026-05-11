"""
quality_context_agent.py — v3.6 (2026-05-11)
====================================================
SA-6: 품질검증 + 정책매핑 에이전트

★ v3.6 변경사항 (최소 패치):
  [패치 1] SECTOR_KEYWORDS 베트남어 동의어 확충
    - 각 섹터에 베트남어 키워드 추가
    - 법령 번호 패턴 리스트 추가 (score_article에서 보너스 점수)
    - 효과: 키워드 매핑률 7.2% → 목표 15% 이상

  [패치 2] score_article()에 법령 번호 패턴 보너스 추가
    - Decision/Decree/Resolution 번호 포함 기사에 보너스 점수
    - 정책 기사 매핑률 향상

  ★ 나머지 코드 전체 v3.5 동일 — 함수명/구조/로직 변경 없음

[v3.5 버그 수정 — 2026-05-09]
  ★ BUG FIX 1: KeyError: 'jina_matched' 크래시 수정
  ★ BUG FIX 2: run_jina_enrichment_for_matched 함수 중복 정의 제거
  ★ BUG FIX 3: 반환 키 통일 {'jina_matched': N}

[v3.4] API 키 없을 때 Jina 영문 원문 오염 방지
[v3.3] Jina 전용 보강 함수 분리
[v3.2] Matched_Plan 시트 재작성 버그 완전 제거

영구 제약:
  - Anthropic API: claude-haiku-4-5-20251001 (번역 금지, 분석에만)
  - Matched_Plan 시트: ExcelUpdater 전용, SA-6는 절대 건드리지 않음
  - 이메일 Secrets: EMAIL_USERNAME / EMAIL_PASSWORD
"""

import json
import logging
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger('quality_context_agent')

# ── 경로 ───────────────────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent.parent
DATA_DIR  = BASE_DIR / 'data'
AGENT_OUT = DATA_DIR / 'agent_output'
DOCS_OUT  = BASE_DIR / 'docs' / 'shared'

EXCEL_PATH = DATA_DIR / 'database' / 'Vietnam_Infra_News_Database_Final.xlsx'

KI_PATHS = [
    DOCS_OUT / 'knowledge_index.json',
    DATA_DIR / 'shared' / 'knowledge_index.json',
    DATA_DIR / 'shared' / 'layer1_data.json',
]

# ── Anthropic / Jina 설정 ──────────────────────────────────────────────────
ANTHROPIC_API_URL = 'https://api.anthropic.com/v1/messages'
HAIKU_MODEL       = 'claude-haiku-4-5-20251001'
HAIKU_TIMEOUT     = 30
JINA_BASE         = 'https://r.jina.ai/'

HAIKU_CLASSIFY_LIMIT = 50
HAIKU_ENRICH_LIMIT   = 20

JINA_MATCHED_LIMIT   = 30
JINA_SUMKO_THRESHOLD = 300

# ── 설정 ───────────────────────────────────────────────────────────────────
RECENT_DAYS = 90
MIN_SCORE   = 35
GRADE_HIGH  = 65
GRADE_MED   = 45

# ── Province 정규화 (v3.5 그대로) ─────────────────────────────────────────
PROVINCE_MAP = {
    'ho chi minh':   'Ho Chi Minh City',
    'hcmc':          'Ho Chi Minh City',
    'saigon':        'Ho Chi Minh City',
    'hà nội':        'Hanoi',
    'ha noi':        'Hanoi',
    'đà nẵng':       'Da Nang',
    'da nang':       'Da Nang',
    'hải phòng':     'Hai Phong',
    'hai phong':     'Hai Phong',
    'cần thơ':       'Can Tho',
    'ninh thuận':    'Ninh Thuan',
    'ninh thuan':    'Ninh Thuan',
    'bình dương':    'Binh Duong',
    'bình định':     'Binh Dinh',
    'long an':       'Long An',
    'đồng nai':      'Dong Nai',
    'dong nai':      'Dong Nai',
    'quảng ninh':    'Quang Ninh',
    'quang ninh':    'Quang Ninh',
}

def normalize_province(text: str) -> str:
    if not text:
        return 'Vietnam'
    low = text.lower().strip()
    return PROVINCE_MAP.get(low, text.strip())


# ══════════════════════════════════════════════════════════════════════════
# 1. knowledge_index 로드 (v3.5 그대로)
# ══════════════════════════════════════════════════════════════════════════
def load_ki() -> dict:
    for path in KI_PATHS:
        if path.exists():
            with open(path, encoding='utf-8') as f:
                ki = json.load(f)
            plans = ki.get('masterplans', ki)
            log.info(f"knowledge_index 로드: {len(plans)}개 플랜 [{path.name}]")
            return plans
    log.warning("knowledge_index 없음")
    return {}


# ══════════════════════════════════════════════════════════════════════════
# 2. 정책 키워드 딕셔너리 빌드 (v3.5 그대로)
# ══════════════════════════════════════════════════════════════════════════
def build_keyword_dict(plans: dict) -> list:
    result = []
    for pid, p in plans.items():
        kw_en = p.get('keywords_en', p.get('keywords', []))
        kw_vi = p.get('keywords_vi', [])
        sector = p.get('sector', '')

        # ★ v3.6 패치: SECTOR_SUPPLEMENT 키워드 직접 주입
        # knowledge_index에 keywords_en/vi 없는 플랜도 섹터 키워드로 매핑 가능
        if sector in SECTOR_SUPPLEMENT_KEYWORDS:
            extra_en, extra_vi = SECTOR_SUPPLEMENT_KEYWORDS[sector]
            kw_en = list(kw_en) + extra_en
            kw_vi = list(kw_vi) + extra_vi

        # 키워드가 전혀 없는 플랜도 섹터 키워드가 있으면 포함
        if not kw_en and not kw_vi:
            continue

        result.append({
            'plan_id':     pid,
            'keywords_en': [k.lower() for k in kw_en if k],
            'keywords_vi': [k.lower() for k in kw_vi if k],
            'sector':      sector,
            'area':        p.get('area', ''),
            'threshold':   p.get('threshold', MIN_SCORE),
        })
    log.info(f"정책 키워드 딕셔너리: {len(result)}개 플랜")
    return result


# ══════════════════════════════════════════════════════════════════════════
# ★ v3.6 패치: 섹터별 확충 키워드 (베트남어 동의어 + 추가 영문 동의어)
# build_keyword_dict()가 knowledge_index 키워드를 읽지만,
# knowledge_index에 없는 동의어를 여기서 보완
# ══════════════════════════════════════════════════════════════════════════
SECTOR_SUPPLEMENT_KEYWORDS = {
    # (섹터명): ([영문 추가], [베트남어 추가])
    'Waste Water': (
        ['sewage treatment', 'wwtf', 'sanitation', 'effluent', 'to lich river',
         'nhieu loc', 'tham luong', 'binh hung hoa', 'thu duc wwtp'],
        ['nước thải', 'xử lý nước thải', 'thoát nước', 'cống thoát nước',
         'nhà máy xử lý nước thải', 'hệ thống thoát nước',
         'Yên Xá', 'Bình Hưng', 'Nhiêu Lộc', 'Tham Lương'],
    ),
    'Water Supply/Drainage': (
        ['potable water', 'water works', 'water pipe network', 'water reservoir',
         'dam reservoir', 'irrigation canal', 'bau bang water', 'song da water'],
        ['cấp nước', 'nước sạch', 'nước sinh hoạt', 'nhà máy nước',
         'hồ chứa nước', 'đập thủy lợi', 'tưới tiêu', 'thủy lợi',
         'mạng lưới cấp nước'],
    ),
    'Solid Waste': (
        ['waste-to-energy plant', 'incineration plant', 'extended producer responsibility',
         'plastic recycling', 'packaging waste', 'da phuoc landfill',
         'soc son wte', 'nam son landfill', 'cu chi wte', 'phu son wte'],
        ['rác thải', 'chất thải rắn', 'đốt rác phát điện', 'tái chế rác',
         'xử lý rác', 'bãi chôn lấp', 'thu gom rác thải',
         'Sóc Sơn', 'Đa Phước', 'Nam Sơn', 'EPR', 'trách nhiệm mở rộng'],
    ),
    'Power': (
        ['power development plan', 'electricity grid', 'transmission line',
         'rooftop solar', 'offshore wind farm', 'lng power plant',
         'battery energy storage', 'virtual power purchase agreement',
         'decision 768', 'pdp viii', 'evn', 'vinacomin'],
        ['điện', 'nhà máy điện', 'lưới điện', 'truyền tải điện',
         'điện gió ngoài khơi', 'điện mặt trời mái nhà',
         'lưu trữ năng lượng', 'EVN', 'Quy hoạch điện 8',
         'Quyết định 768', 'năng lượng tái tạo'],
    ),
    'Oil & Gas': (
        ['lng terminal', 'gas pipeline', 'petrovietnam', 'pv gas',
         'thi vai lng', 'bach ho field', 'rang dong field',
         'nam con son pipeline', 'binh son refinery', 'dung quat refinery'],
        ['dầu khí', 'khí đốt', 'khí thiên nhiên hóa lỏng', 'LNG',
         'lọc dầu', 'Petrovietnam', 'PVN', 'PV GAS',
         'đường ống khí', 'bể chứa LNG'],
    ),
    'Industrial Parks': (
        ['special economic zone', 'export processing zone', 'deep c industrial',
         'vsip', 'foxconn vietnam', 'samsung vietnam', 'intel vietnam',
         'supply chain vietnam', 'data center campus', 'semiconductor fab'],
        ['khu công nghiệp', 'khu kinh tế', 'khu chế xuất', 'VSIP',
         'nhà máy sản xuất', 'FDI', 'thu hút đầu tư nước ngoài',
         'chuỗi cung ứng', 'trung tâm dữ liệu', 'bán dẫn'],
    ),
    'Smart City': (
        ['smart city project', 'digital infrastructure', 'urban digital twin',
         'fiber optic network', '5g network', 'iot platform',
         'dong anh smart city', 'hoa lac hi-tech park'],
        ['thành phố thông minh', 'chuyển đổi số', 'hạ tầng số',
         'trung tâm dữ liệu', 'mạng 5G', 'IoT', 'đô thị thông minh',
         'Hòa Lạc', 'Đông Anh'],
    ),
    'Transport': (
        ['north south expressway', 'ring road 4', 'ring road 3',
         'long thanh international airport', 'lach huyen port', 'cai mep port',
         'metro line', 'urban railway', 'brt corridor', 'ppp highway'],
        ['sân bay quốc tế Long Thành', 'cao tốc Bắc-Nam', 'vành đai 4',
         'vành đai 3', 'metro', 'tàu điện', 'cảng Lạch Huyện',
         'cảng Cái Mép', 'đường sắt đô thị', 'BRT'],
    ),
}

# ★ v3.6 패치: 법령 번호 패턴 (매칭 시 보너스 점수)
LEGAL_PATTERNS = [
    r'decision\s+\d+[/-]qd[/-]ttg',
    r'decree\s+\d+[/-]nd[/-]cp',
    r'circular\s+\d+[/-]tt[/-]',
    r'resolution\s+\d+[-/]nq[/-]',
    r'quyết\s+định\s+số\s+\d+',
    r'nghị\s+định\s+số\s+\d+',
    r'thông\s+tư\s+số\s+\d+',
    r'nghị\s+quyết\s+số\s+\d+',
    r'decision\s+768',
    r'decision\s+500',
    r'decision\s+1354',
    r'decision\s+491',
]


# ══════════════════════════════════════════════════════════════════════════
# 3. 단일 기사 ↔ 플랜 매칭 점수 계산
# ★ v3.6 패치: 법령 번호 패턴 보너스 + SECTOR_SUPPLEMENT 키워드 반영
# ══════════════════════════════════════════════════════════════════════════
def score_article(title_en: str, title_ko: str, summary_en: str, summary_ko: str,
                  plan: dict) -> float:
    score = 0.0
    en_text = (title_en + ' ' + summary_en).lower()
    ko_text = (title_ko + ' ' + summary_ko).lower()
    all_text = en_text + ' ' + ko_text

    # 기존 v3.5 키워드 매칭
    for kw in plan['keywords_en']:
        if kw in en_text:
            score += 25 if kw in title_en.lower() else 10

    for kw in plan['keywords_vi']:
        if kw in ko_text:
            score += 20 if kw in title_ko.lower() else 8

    # ★ v3.6 패치: SECTOR_SUPPLEMENT 추가 키워드 매칭
    sector = plan.get('sector', '')
    if sector in SECTOR_SUPPLEMENT_KEYWORDS:
        extra_en, extra_vi = SECTOR_SUPPLEMENT_KEYWORDS[sector]
        for kw in extra_en:
            if kw.lower() in en_text:
                score += 15 if kw.lower() in title_en.lower() else 6
        for kw in extra_vi:
            if kw.lower() in all_text:
                score += 12 if kw.lower() in title_ko.lower() else 5

    # ★ v3.6 패치: 법령 번호 패턴 보너스 (정책 기사 매핑 향상)
    for pattern in LEGAL_PATTERNS:
        if re.search(pattern, all_text, re.IGNORECASE):
            score += 8
            break  # 첫 번째 매칭만 보너스 적용

    return min(score, 100.0)


# ══════════════════════════════════════════════════════════════════════════
# 4. Excel 로드 + 매칭 실행 + News DB Grade/Plan_ID 업데이트 (v3.5 그대로)
# ══════════════════════════════════════════════════════════════════════════
def run_matching(plans: dict, keyword_dict: list) -> dict:
    """
    ★ v3.2 핵심: News Database의 Grade·Plan_ID 컬럼만 업데이트.
                 Matched_Plan 시트는 절대 건드리지 않음.
    """
    if not EXCEL_PATH.exists():
        log.error(f"Excel DB 없음: {EXCEL_PATH}")
        return {'error': 'excel_not_found'}

    log.info(f"Excel 로드: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if 'News Database' not in wb.sheetnames:
        log.error("'News Database' 시트 없음")
        wb.close()
        return {'error': 'sheet_not_found'}

    ws = wb['News Database']

    headers = {}
    for cell in ws[1]:
        if cell.value:
            key = str(cell.value).strip().lower().replace(' ', '_')
            headers[key] = cell.column - 1

    log.info(f"Excel 헤더: {list(headers.keys())}")

    def col(names):
        for n in names:
            if n in headers:
                return headers[n]
        return -1

    C = {
        'date':       col(['date']),
        'title_en':   col(['title_(en/vi)', 'title', 'news_title', 'title_en']),
        'title_ko':   col(['tit_ko', 'title_ko']),
        'summary_en': col(['short_summary', 'summary_en', 'sum_en']),
        'summary_ko': col(['sum_ko', 'summary_ko']),
        'province':   col(['province']),
        'plan_id':    col(['plan_id', 'matched_plan']),
        'grade':      col(['grade', 'ctx_grade']),
        'source':     col(['source']),
    }
    log.info(f"컬럼 매핑: {C}")

    cutoff = (datetime.now() - timedelta(days=RECENT_DAYS)).strftime('%Y-%m-%d')

    stats = {
        'total': 0, 'matched': 0, 'high': 0, 'medium': 0,
        'province_fixed': 0, 'skipped_old': 0,
        'plan_counts': {},
    }

    matched_rows = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not any(row):
            continue
        stats['total'] += 1

        def rv(idx):
            if idx < 0 or idx >= len(row): return ''
            return str(row[idx] or '').strip()

        date_val    = rv(C['date'])[:10]
        title_en    = rv(C['title_en'])
        title_ko    = rv(C['title_ko'])
        summary_en  = rv(C['summary_en'])
        summary_ko  = rv(C['summary_ko'])
        province    = rv(C['province'])

        if date_val and date_val < cutoff:
            stats['skipped_old'] += 1
            continue

        if not title_en and not title_ko:
            continue

        best_plan  = None
        best_score = 0.0
        for plan in keyword_dict:
            sc = score_article(title_en, title_ko, summary_en, summary_ko, plan)
            if sc >= plan['threshold'] and sc > best_score:
                best_score = sc
                best_plan  = plan

        if best_plan:
            pid   = best_plan['plan_id']
            grade = ('HIGH'   if best_score >= GRADE_HIGH else
                     'MEDIUM' if best_score >= GRADE_MED  else 'LOW')
            stats['matched'] += 1
            stats['plan_counts'][pid] = stats['plan_counts'].get(pid, 0) + 1
            if grade == 'HIGH':     stats['high']   += 1
            elif grade == 'MEDIUM': stats['medium'] += 1
            matched_rows.append((row_num, pid, grade, best_score))

        normed = normalize_province(province)
        if normed != province:
            stats['province_fixed'] += 1

    log.info(f"분류 완료: {stats['total']}건 (최근 {RECENT_DAYS}일 대상)")
    log.info(f"  POLICY_MATCH: {stats['matched']}건")
    log.info(f"  HIGH: {stats['high']}건  MEDIUM: {stats['medium']}건")
    log.info(f"  오래된 기사 스킵: {stats['skipped_old']}건")
    log.info(f"  Province 정규화: {stats['province_fixed']}건")

    # ★ v3.2: News Database Grade·Plan_ID 컬럼만 업데이트
    wb2 = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws2 = wb2['News Database']

    hdr_row = list(ws2[1])
    existing_headers = {str(c.value or '').strip().lower(): c.column
                        for c in hdr_row if c.value}

    def ensure_col(col_name, after_col=8):
        lower = col_name.lower()
        for h, c in existing_headers.items():
            if lower in h:
                return c
        max_col = max(existing_headers.values()) if existing_headers else after_col
        new_col = max_col + 1
        ws2.cell(row=1, column=new_col, value=col_name)
        existing_headers[lower] = new_col
        return new_col

    plan_col  = ensure_col('Plan_ID')
    grade_col = ensure_col('Grade')

    updated_count = 0
    for row_num, pid, grade, score in matched_rows:
        existing_pid   = str(ws2.cell(row=row_num, column=plan_col).value  or '').strip()
        existing_grade = str(ws2.cell(row=row_num, column=grade_col).value or '').strip()
        if not existing_pid:
            ws2.cell(row=row_num, column=plan_col,  value=pid)
            updated_count += 1
        if not existing_grade:
            ws2.cell(row=row_num, column=grade_col, value=grade)

    wb.close()
    wb2.save(EXCEL_PATH)
    wb2.close()

    log.info(f"Excel 업데이트 완료: News DB Plan_ID/Grade {updated_count}건 기록")
    log.info(f"  ★ Matched_Plan 시트: 변경 없음 (ExcelUpdater 전용)")

    return stats


# ══════════════════════════════════════════════════════════════════════════
# 5. quality_report.json 저장 (v3.5 그대로)
# ══════════════════════════════════════════════════════════════════════════
def save_report(stats: dict):
    AGENT_OUT.mkdir(parents=True, exist_ok=True)
    DOCS_OUT.mkdir(parents=True, exist_ok=True)

    report = {
        'generated_at':   datetime.now().strftime('%Y-%m-%d %H:%M'),
        'total_articles':  stats.get('total', 0),
        'matched_count':   stats.get('matched', 0),
        'match_rate_pct':  round(stats.get('matched', 0) / max(stats.get('total', 1), 1) * 100, 1),
        'grade_high':      stats.get('high', 0),
        'grade_medium':    stats.get('medium', 0),
        'province_fixed':  stats.get('province_fixed', 0),
        'plan_counts':     stats.get('plan_counts', {}),
        'skipped_old':     stats.get('skipped_old', 0),
    }

    out = DOCS_OUT / 'quality_report.json'
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    log.info(f"quality_report.json 저장: {out}")


# ══════════════════════════════════════════════════════════════════════════
# 6~10. Haiku / Jina 함수 (v3.5 그대로)
# ══════════════════════════════════════════════════════════════════════════
def _call_haiku(system_prompt: str, user_prompt: str, api_key: str) -> str:
    import requests as req
    headers = {
        'Content-Type':      'application/json',
        'x-api-key':         api_key.strip(),
        'anthropic-version': '2023-06-01',
    }
    payload = {
        'model':      HAIKU_MODEL,
        'max_tokens': 300,
        'system':     system_prompt,
        'messages':   [{'role': 'user', 'content': user_prompt}],
    }
    try:
        r = req.post(ANTHROPIC_API_URL, headers=headers,
                     json=payload, timeout=HAIKU_TIMEOUT)
        r.raise_for_status()
        for block in r.json().get('content', []):
            if block.get('type') == 'text':
                return block['text'].strip()
    except Exception as e:
        log.warning(f"  Haiku 오류: {e}")
    return ''


def fetch_jina_text(url: str) -> str:
    import requests as req
    jina_url = JINA_BASE + url.strip()
    try:
        r = req.get(jina_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
        if r.status_code == 200:
            return r.text[:3000]
    except Exception as e:
        log.debug(f"  Jina 오류: {e}")
    return ''


def haiku_classify_article(article: dict, plans: dict, api_key: str):
    title_en  = article.get('title_en', '')
    title_ko  = article.get('title_ko', '')
    summ_en   = article.get('summary_en', '')[:300]

    plan_list = []
    for pid, plan in list(plans.items())[:21]:
        kw = (plan.get('keywords_en', []) or plan.get('keywords', []))[:3]
        plan_list.append(f"  {pid} ({plan.get('sector','')}) — {', '.join(kw)}")
    plans_text = '\n'.join(plan_list)

    system_prompt = (
        "당신은 베트남 인프라 뉴스 분류 전문가입니다.\n"
        "아래 마스터플랜 목록과 기사를 비교하여 관련 plan_id를 판단하세요.\n"
        "관련 없으면 반드시 null을 반환하세요.\n\n"
        f"마스터플랜 목록:\n{plans_text}"
    )
    user_prompt = (
        f"기사 제목(EN): {title_en}\n"
        f"기사 제목(KO): {title_ko}\n"
        f"요약(EN): {summ_en}\n\n"
        "이 기사가 위 마스터플랜 중 하나와 명확히 관련되면 아래 JSON만 반환:"
        '{"plan_id":"VN-XXX","grade":"HIGH","reason":"근거 1문장"}'
        "\n관련 없으면: null"
    )

    raw = _call_haiku(system_prompt, user_prompt, api_key)
    if not raw or raw.strip() == 'null':
        return None

    try:
        m = re.search(r'\{.*?\}', raw, re.DOTALL)
        if m:
            result = json.loads(m.group())
            pid   = result.get('plan_id', '')
            grade = result.get('grade', 'MEDIUM')
            if pid in plans and grade in ('HIGH', 'MEDIUM', 'LOW'):
                return result
    except Exception:
        pass
    return None


def enrich_with_jina(article: dict, plans: dict, api_key: str):
    url = article.get('url', '')
    if not url or not url.startswith('http'):
        return None

    body = fetch_jina_text(url)
    if not body or len(body) < 100:
        return None

    plan_ids = ', '.join(list(plans.keys())[:21])
    system_prompt = (
        "당신은 베트남 인프라 뉴스 요약 전문가입니다.\n"
        "아래 기사 본문을 읽고 두 가지를 수행하세요:\n"
        "1. 한국어 요약 200자 이내로 생성\n"
        "2. 관련 마스터플랜 ID 판단 (없으면 null)\n"
        f"마스터플랜 ID 목록: {plan_ids}"
    )
    user_prompt = (
        f"기사 본문:\n{body[:2000]}\n\n"
        "JSON 형식으로만 답변:"
        '{"summary_ko":"200자 이내 한국어 요약","plan_id":"VN-XXX 또는 null","grade":"HIGH/MEDIUM/LOW"}'
    )

    raw = _call_haiku(system_prompt, user_prompt, api_key)
    if not raw:
        return None

    try:
        m = re.search(r'\{.*?\}', raw, re.DOTALL)
        if m:
            result = json.loads(m.group())
            if result.get('summary_ko'):
                pid = result.get('plan_id', '')
                if pid not in plans:
                    result['plan_id'] = ''
                return result
    except Exception:
        pass
    return None


def run_haiku_enhancement(plans: dict, api_key: str) -> dict:
    """방법 A + B 통합 (v3.5 그대로)."""
    if not api_key:
        log.info("[v3.0] ANTHROPIC_API_KEY 없음 — Haiku 보완 건너뜀")
        return {'haiku_classified': 0, 'jina_enriched': 0}

    if not EXCEL_PATH.exists():
        return {'haiku_classified': 0, 'jina_enriched': 0}

    log.info("[v3.0] Haiku 맥락 보완 시작 (방법 A + B)...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if 'News Database' not in wb.sheetnames:
        wb.close()
        return {'haiku_classified': 0, 'jina_enriched': 0}

    ws = wb['News Database']
    headers = [str(c.value or '').strip().lower().replace(' ', '_')
               for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def ci(keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k.lower() in h:
                    return i
        return None

    C = {
        'date':       ci(['date']),
        'title_en':   ci(['title_en', 'title_(en/vi)']),
        'title_ko':   ci(['title_ko', 'tit_ko']),
        'summary_en': ci(['summary_en', 'sum_en']),
        'summary_ko': ci(['summary_ko', 'sum_ko']),
        'url':        ci(['link', 'url']),
        'plan_id':    ci(['plan_id', 'ctx_plans']),
        'grade':      ci(['grade', 'ctx_grade']),
    }

    cutoff = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

    candidates_a = []
    candidates_b = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        date_v = str(row[C['date']] if C['date'] is not None else '').strip()[:10]
        if not date_v or date_v < cutoff:
            continue

        plan_v   = str(row[C['plan_id']]    if C['plan_id']    is not None else '').strip()
        grade_v  = str(row[C['grade']]      if C['grade']      is not None else '').strip()
        summ_ko  = str(row[C['summary_ko']] if C['summary_ko'] is not None else '').strip()
        url_v    = str(row[C['url']]        if C['url']        is not None else '').strip()
        title_en = str(row[C['title_en']]   if C['title_en']   is not None else '').strip()
        title_ko = str(row[C['title_ko']]   if C['title_ko']   is not None else '').strip()
        summ_en  = str(row[C['summary_en']] if C['summary_en'] is not None else '').strip()

        art = {
            'row': i, 'date': date_v, 'plan_id': plan_v,
            'grade': grade_v, 'title_en': title_en,
            'title_ko': title_ko, 'summary_en': summ_en,
            'summary_ko': summ_ko, 'url': url_v,
        }

        if not plan_v and len(candidates_a) < HAIKU_CLASSIFY_LIMIT:
            candidates_a.append(art)

        if (len(summ_ko) < JINA_SUMKO_THRESHOLD
                and (plan_v or grade_v in ('HIGH', 'MEDIUM'))
                and url_v.startswith('http')
                and len(candidates_b) < HAIKU_ENRICH_LIMIT):
            candidates_b.append(art)

    wb.close()

    classified = 0
    updates_a  = []
    log.info(f"  [방법 A] 맥락분류 대상: {len(candidates_a)}건")
    for art in candidates_a:
        result = haiku_classify_article(art, plans, api_key)
        if result:
            updates_a.append((art['row'], result['plan_id'], result['grade']))
            classified += 1
        time.sleep(0.2)

    enriched  = 0
    updates_b = []
    log.info(f"  [방법 B] 요약보강 대상: {len(candidates_b)}건")
    for art in candidates_b:
        result = enrich_with_jina(art, plans, api_key)
        if result:
            updates_b.append((art['row'], result))
            enriched += 1
        time.sleep(0.3)

    if updates_a or updates_b:
        wb2 = openpyxl.load_workbook(EXCEL_PATH)
        ws2 = wb2['News Database']

        headers2  = [str(c.value or '').strip().lower().replace(' ', '_')
                     for c in next(ws2.iter_rows(min_row=1, max_row=1))]
        plan_col  = next((i+1 for i, h in enumerate(headers2) if 'plan_id' in h), None)
        grade_col = next((i+1 for i, h in enumerate(headers2) if 'grade' in h), None)
        sumko_col = next((i+1 for i, h in enumerate(headers2)
                          if 'sum_ko' in h or 'summary_ko' in h), None)

        for row_num, pid, grade in updates_a:
            if plan_col and not str(ws2.cell(row=row_num, column=plan_col).value or '').strip():
                ws2.cell(row=row_num, column=plan_col, value=pid)
            if grade_col and not str(ws2.cell(row=row_num, column=grade_col).value or '').strip():
                ws2.cell(row=row_num, column=grade_col, value=grade)

        for row_num, result in updates_b:
            if sumko_col and result.get('summary_ko'):
                ws2.cell(row=row_num, column=sumko_col, value=result['summary_ko'])
            if plan_col and result.get('plan_id') and result['plan_id'] in plans:
                if not str(ws2.cell(row=row_num, column=plan_col).value or '').strip():
                    ws2.cell(row=row_num, column=plan_col, value=result['plan_id'])
            if grade_col and result.get('grade'):
                if not str(ws2.cell(row=row_num, column=grade_col).value or '').strip():
                    ws2.cell(row=row_num, column=grade_col, value=result['grade'])

        wb2.save(EXCEL_PATH)
        wb2.close()
        log.info(f"  ★ Matched_Plan 시트: 변경 없음 (ExcelUpdater 전용)")

    log.info(f"  [방법 A] Haiku 분류: {classified}건 추가 매핑")
    log.info(f"  [방법 B] Jina 보강: {enriched}건 요약 갱신")
    return {'haiku_classified': classified, 'jina_enriched': enriched}


# ══════════════════════════════════════════════════════════════════════════
# 11. Jina Matched_Plan 전용 보강 (v3.5 그대로 — KeyError 수정 포함)
# ══════════════════════════════════════════════════════════════════════════
def run_jina_enrichment_for_matched(plans: dict, api_key: str = '') -> dict:
    """
    ★ v3.5: 반환 키를 'jina_matched'로 통일 (KeyError 수정)
    Matched_Plan 기반 Jina 보강 — Anthropic API 키 없어도 동작.
    """
    if not EXCEL_PATH.exists():
        return {'jina_matched': 0}

    log.info("[v3.3] Jina Matched_Plan 보강 시작...")

    try:
        wb_r = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"  Excel 로드 실패: {e}")
        return {'jina_matched': 0}

    mp_url_set = set()
    if 'Matched_Plan' in wb_r.sheetnames:
        ws_mp = wb_r['Matched_Plan']
        try:
            mp_hdr = [str(c or '').strip().lower() for c in
                      next(ws_mp.iter_rows(min_row=2, max_row=2, values_only=True))]
            link_ci = next((i for i, h in enumerate(mp_hdr) if 'link' in h or 'url' in h), None)
            if link_ci is not None:
                for row in ws_mp.iter_rows(min_row=3, values_only=True):
                    v = str(row[link_ci] or '').strip()
                    if v and v not in ('nan', 'None', ''):
                        mp_url_set.add(v)
        except StopIteration:
            pass
    log.info(f"  Matched_Plan URL 세트: {len(mp_url_set)}건")

    if 'News Database' not in wb_r.sheetnames:
        wb_r.close()
        return {'jina_matched': 0}

    ws_nd  = wb_r['News Database']
    nd_hdr = [str(c.value or '').strip().lower().replace(' ', '_')
              for c in next(ws_nd.iter_rows(min_row=1, max_row=1))]

    def ci(keys):
        for k in keys:
            for i, h in enumerate(nd_hdr):
                if k in h: return i
        return None

    C = {
        'date':     ci(['date']),
        'title_en': ci(['title_en']),
        'title_vi': ci(['title_vi']),
        'title_ko': ci(['tit_ko', 'title_ko']),
        'sum_ko':   ci(['sum_ko', 'summary_ko']),
        'sum_en':   ci(['sum_en', 'summary_en']),
        'url':      ci(['url']),
        'plan_id':  ci(['plan_id']),
        'grade':    ci(['grade']),
    }

    cutoff     = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    candidates = []

    for i, row in enumerate(ws_nd.iter_rows(min_row=2, values_only=True), 2):
        if not row or not any(row): continue

        def rv(k):
            idx = C.get(k)
            return str(row[idx] or '').strip() if idx is not None and len(row) > idx else ''

        date_v   = rv('date')[:10]
        url_v    = rv('url')
        sumko_v  = rv('sum_ko')
        title_en = rv('title_en') or rv('title_vi')

        if (date_v >= cutoff
                and url_v in mp_url_set
                and len(sumko_v) < JINA_SUMKO_THRESHOLD
                and url_v.startswith('http')
                and title_en):
            candidates.append({
                'row':      i,
                'date':     date_v,
                'url':      url_v,
                'title_en': title_en,
                'title_ko': rv('title_ko'),
                'sum_ko':   sumko_v,
                'sum_en':   rv('sum_en'),
                'plan_id':  rv('plan_id'),
                'grade':    rv('grade'),
            })
            if len(candidates) >= JINA_MATCHED_LIMIT:
                break

    wb_r.close()
    log.info(f"  Jina 보강 대상: {len(candidates)}건 (최근 30일 Matched_Plan 기사, sum_ko<{JINA_SUMKO_THRESHOLD}자)")

    if not candidates:
        log.info("  Jina 보강 대상 없음 — 종료")
        return {'jina_matched': 0}

    updates = []

    for art in candidates:
        url  = art['url']
        body = fetch_jina_text(url)

        if not body or len(body) < 100:
            log.debug(f"    Jina 본문 취득 실패: {url[:60]}")
            time.sleep(0.3)
            continue

        new_sumko = ''

        if api_key:
            plan_ids = ', '.join(list(plans.keys())[:21])
            system_p = (
                "당신은 베트남 인프라 사업 전문 분석가입니다.\n"
                "아래 기사 전문을 읽고 다음 5가지 항목을 반드시 포함해 200자 이내 한국어로 요약하세요:\n"
                "  ① 사업명 또는 프로젝트명\n"
                "  ② 사업규모 (예산금액·설비용량·연장거리 등 수치)\n"
                "  ③ 현재 진행단계 (계획·승인·입찰·착공·준공·운영 중 하나)\n"
                "  ④ 주요 기관 (발주처·시공사·투자자·정부부처)\n"
                "  ⑤ 핵심 일정 또는 다음 이벤트\n"
                f"관련 마스터플랜 ID 목록: {plan_ids}"
            )
            user_p = (
                f"기사 제목: {art['title_en']}\n"
                f"Plan_ID: {art.get('plan_id', '미지정')}\n"
                f"기사 전문 (최대 2500자):\n{body[:2500]}\n\n"
                "아래 JSON 형식으로만 답변:\n"
                '{"summary_ko":"200자 이내 한국어 인사이트 요약"}'
            )
            raw = _call_haiku(system_p, user_p, api_key)
            if raw:
                try:
                    m = re.search(r'{.*?}', raw, re.DOTALL)
                    if m:
                        result = json.loads(m.group())
                        new_sumko = result.get('summary_ko', '')
                except Exception:
                    pass

        if not new_sumko:
            paragraphs = [p.strip() for p in body.split('\n') if len(p.strip()) > 50]
            raw_body   = ' '.join(paragraphs[:3])[:500]
            if raw_body:
                try:
                    from deep_translator import GoogleTranslator
                    body_ko   = GoogleTranslator(source='auto', target='ko').translate(raw_body[:500])
                    new_sumko = f"[Jina] {body_ko[:280]}"
                except Exception as te:
                    log.debug(f"    번역 실패({te}) — 기존 sum_ko 유지")
                    new_sumko = ''

        if new_sumko:
            updates.append((art['row'], new_sumko))
            log.info(f"    ✓ {art['date']} {art['title_en'][:40]} → {len(new_sumko)}자")

        time.sleep(0.5)

    if updates:
        try:
            wb2  = openpyxl.load_workbook(EXCEL_PATH)
            ws2  = wb2['News Database']
            hdr2 = [str(c.value or '').strip().lower().replace(' ', '_')
                    for c in next(ws2.iter_rows(min_row=1, max_row=1))]
            sumko_col = next((i+1 for i, h in enumerate(hdr2)
                              if 'sum_ko' in h or 'summary_ko' in h), None)

            if sumko_col:
                for row_num, new_sumko in updates:
                    ws2.cell(row=row_num, column=sumko_col, value=new_sumko)
                wb2.save(EXCEL_PATH)
                log.info(f"  [v3.3] Jina Matched 보강 완료: {len(updates)}건 sum_ko 갱신")
            wb2.close()
        except Exception as e:
            log.warning(f"  Excel 저장 오류: {e}")

    # ★ v3.5: 반환 키 'jina_matched'로 통일 (KeyError 수정 핵심)
    return {'jina_matched': len(updates)}


# ══════════════════════════════════════════════════════════════════════════
# 메인 (v3.5 그대로)
# ══════════════════════════════════════════════════════════════════════════
def main():
    log.info("=" * 58)
    log.info(f"quality_context_agent v3.6 — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    log.info("=" * 58)

    plans = load_ki()
    if not plans:
        log.error("플랜 데이터 없음 — 종료")
        return

    keyword_dict = build_keyword_dict(plans)
    stats        = run_matching(plans, keyword_dict)
    save_report(stats)

    api_key     = os.getenv('ANTHROPIC_API_KEY', '').strip()
    haiku_stats = run_haiku_enhancement(plans, api_key)

    # ★ v3.3: Jina Matched_Plan 전용 보강 (API 키 없어도 실행)
    jina_stats  = run_jina_enrichment_for_matched(plans, api_key)

    log.info("━" * 58)
    log.info(f"SA-6 v3.6 완료: {stats.get('matched', 0)}건 키워드매핑 / {stats.get('total', 0)}건 전체")
    log.info(f"  키워드 매핑률: {round(stats.get('matched',0)/max(stats.get('total',1),1)*100,1)}%")
    log.info(f"  Haiku 추가분류: {haiku_stats.get('haiku_classified', 0)}건")
    log.info(f"  Jina 요약보강(방법B): {haiku_stats.get('jina_enriched', 0)}건")
    # ★ v3.5: .get()으로 안전하게 접근 — KeyError 완전 방지
    log.info(f"  Jina Matched 전용보강: {jina_stats.get('jina_matched', 0)}건 ← v3.3 신규")
    log.info("━" * 58)


if __name__ == '__main__':
    main()
