#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fix_bad_translations.py
=======================
4/12~15 기간 불량 번역 기사를 찾아 title_ko/summary_ko를 공란으로 초기화.
다음 batch_translate.py 실행 시 자동 재번역됨.

실행:
  python3 scripts/fix_bad_translations.py
  python3 scripts/fix_bad_translations.py --date-from 2026-04-12 --date-to 2026-04-15
  python3 scripts/fix_bad_translations.py --dry-run   (실제 수정 없이 탐지만)
"""
import os, sys, re, argparse
from pathlib import Path
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill

EXCEL_PATH = os.environ.get(
    'EXCEL_PATH',
    'data/database/Vietnam_Infra_News_Database_Final.xlsx'
)
YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
RESET  = PatternFill(fill_type=None)


def is_bad_translation(v: str) -> bool:
    """오번역 패턴 감지"""
    if not v or not v.strip():
        return False
    u = v.upper()
    # MyMemory API 경고 문자열
    if any(k in u for k in ('MYMEMORY WARNING', 'YOU USED ALL',
                             'QUERY LENGTH', 'PLEASE SELECT', 'INVALID')):
        return True
    # 쉼표 반복: "베트남, 베트남, 베트남..."
    parts = [p.strip() for p in v.split(',')]
    if len(parts) >= 3 and len(set(parts[:3])) == 1 and parts[0]:
        return True
    # 공백 반복: "베트남 베트남 베트남 베트남..."
    words = v.split()
    if len(words) >= 4 and len(set(words[:4])) == 1 and words[0]:
        return True
    # 영문 원문 그대로 저장 (한글 없음)
    ko_chars = sum(1 for c in v if '\uAC00' <= c <= '\uD7A3')
    if len(v) > 10 and ko_chars == 0 and v.isascii():
        return True
    return False


def run(date_from: str, date_to: str, dry_run: bool):
    ep = Path(EXCEL_PATH)
    if not ep.exists():
        print(f"[ERROR] Excel 없음: {ep}")
        sys.exit(1)

    print(f"[fix] 로드: {ep}")
    wb = openpyxl.load_workbook(ep)
    ws = wb.active

    # 헤더 매핑
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or '').strip()
        if h:
            headers[h] = c

    date_col    = headers.get('Date', 5)
    title_col   = headers.get('News Title', 4)
    tko_col     = headers.get('title_ko', 9)
    ten_col     = headers.get('title_en', 10)
    tvi_col     = headers.get('title_vi', 11)
    sko_col     = headers.get('summary_ko', 12)
    sen_col     = headers.get('summary_en', 13)
    svi_col     = headers.get('summary_vi', 14)

    print(f"[fix] 대상 기간: {date_from} ~ {date_to}")
    print(f"[fix] dry_run={dry_run}")

    fixed = 0
    scanned = 0

    for r in range(2, ws.max_row + 1):
        date_val = str(ws.cell(r, date_col).value or '')[:10]
        if not date_val:
            continue
        # 날짜 범위 필터
        if date_val < date_from or date_val > date_to:
            continue

        title = str(ws.cell(r, title_col).value or '').strip()
        if not title:
            continue

        scanned += 1
        tko = str(ws.cell(r, tko_col).value or '').strip()
        sko = str(ws.cell(r, sko_col).value or '').strip()

        bad_title   = is_bad_translation(tko)
        bad_summary = is_bad_translation(sko)

        if bad_title or bad_summary:
            fixed += 1
            tag = "[DRY]" if dry_run else "[FIX]"
            print(f"  {tag} row={r} date={date_val} bad_title={bad_title} bad_sum={bad_summary}")
            print(f"       제목원문: {title[:55]}")
            if tko:
                print(f"       현재번역: {tko[:55]}")

            if not dry_run:
                if bad_title:
                    ws.cell(r, tko_col).value = ''    # 재번역 대상
                    ws.cell(r, ten_col).value = ''
                    ws.cell(r, tvi_col).value = ''
                if bad_summary:
                    ws.cell(r, sko_col).value = ''
                    ws.cell(r, sen_col).value = ''
                    ws.cell(r, svi_col).value = ''
                # 노란색 하이라이트 — 재번역 필요 표시
                for col in [tko_col, sko_col]:
                    ws.cell(r, col).fill = YELLOW

    print(f"\n[fix] 스캔: {scanned}건 | 오번역 발견: {fixed}건 | {'초기화 완료' if not dry_run else 'DRY RUN (실제 수정 없음)'}")

    if not dry_run and fixed > 0:
        wb.save(ep)
        print(f"[fix] 저장 완료: {ep}")
        print(f"[fix] 다음 batch_translate 실행 시 {fixed}건 자동 재번역됩니다")
    elif fixed == 0:
        print("[fix] 오번역 기사 없음 — 수정 불필요")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='불량 번역 초기화')
    parser.add_argument('--date-from', default='2026-04-12',
                        help='시작 날짜 YYYY-MM-DD (기본: 2026-04-12)')
    parser.add_argument('--date-to',   default='2026-04-16',
                        help='종료 날짜 YYYY-MM-DD (기본: 2026-04-16)')
    parser.add_argument('--dry-run', action='store_true',
                        help='탐지만 (실제 수정 없음)')
    args = parser.parse_args()
    run(args.date_from, args.date_to, args.dry_run)
