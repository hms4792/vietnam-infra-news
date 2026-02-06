#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vietnam Infrastructure News - Email Notification
Loads from BOTH Excel (historical) AND SQLite (new collected) databases.
"""

import smtplib
import logging
import sys
import os
import sqlite3
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from pathlib import Path
from collections import Counter

# Setup paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / "data"
OUTPUT_DIR = PROJECT_ROOT / "outputs"

sys.path.insert(0, str(PROJECT_ROOT))

from config.settings import (
    EMAIL_SUBJECT, EMAIL_FROM_NAME, EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT,
    DASHBOARD_URL
)

# Get email settings directly from environment
EMAIL_USERNAME = os.getenv("EMAIL_USERNAME", "")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")
EMAIL_RECIPIENTS_RAW = os.getenv("EMAIL_RECIPIENTS", "")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

EXCEL_DB_PATH = DATA_DIR / "database" / "Vietnam_Infra_News_Database_Final.xlsx"
SQLITE_DB_PATH = DATA_DIR / "vietnam_infrastructure_news.db"


def load_articles_from_sqlite():
    """Load articles from SQLite database (newly collected)"""
    if not SQLITE_DB_PATH.exists():
        print(f"SQLite DB not found: {SQLITE_DB_PATH}")
        return []
    
    print(f"Loading from SQLite: {SQLITE_DB_PATH}")
    
    try:
        conn = sqlite3.connect(SQLITE_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, url, title, summary, source, sector, area, province, 
                   published_date, collected_date
            FROM articles
            ORDER BY published_date DESC
        """)
        
        articles = []
        for row in cursor.fetchall():
            date_str = row['published_date'] or row['collected_date'] or ''
            if date_str:
                date_str = date_str[:10]
            
            articles.append({
                "title": row['title'] or "",
                "sector": row['sector'] or "Infrastructure",
                "province": row['province'] or "Vietnam",
                "source": row['source'] or "",
                "url": row['url'] or "",
                "summary": row['summary'] or "",
                "date": date_str,
                "is_new": True
            })
        
        conn.close()
        print(f"Loaded {len(articles)} articles from SQLite")
        return articles
        
    except Exception as e:
        print(f"Error loading SQLite: {e}")
        return []


def load_articles_from_excel():
    """Load articles from Excel"""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed")
        return []
    
    if not EXCEL_DB_PATH.exists():
        print(f"Excel not found: {EXCEL_DB_PATH}")
        return []
    
    print(f"Loading from Excel: {EXCEL_DB_PATH}")
    
    wb = openpyxl.load_workbook(EXCEL_DB_PATH, read_only=True, data_only=True)
    
    print(f"Available sheets: {wb.sheetnames}")
    
    # Find the News data sheet (not Summary or other sheets)
    ws = None
    
    # Priority 1: Look for sheet named "News"
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == 'news':
            ws = wb[sheet_name]
            print(f"Using sheet: {sheet_name}")
            break
    
    # Priority 2: Look for sheet with "Area" header (main data sheet)
    if ws is None:
        for sheet_name in wb.sheetnames:
            if 'summary' in sheet_name.lower() or 'rss' in sheet_name.lower() or 'keyword' in sheet_name.lower() or 'log' in sheet_name.lower():
                continue
            test_ws = wb[sheet_name]
            try:
                first_row = [cell.value for cell in test_ws[1]]
                # Check for expected headers: Area, Business Sector, Province, News Tittle, Date
                if any(h and str(h).strip() == 'Area' for h in first_row):
                    ws = test_ws
                    print(f"Using sheet with Area header: {sheet_name}")
                    break
            except:
                continue
    
    # Priority 3: Use first sheet that's not Summary/RSS/Keywords/Log
    if ws is None:
        for sheet_name in wb.sheetnames:
            if 'summary' not in sheet_name.lower() and 'rss' not in sheet_name.lower() and 'keyword' not in sheet_name.lower() and 'log' not in sheet_name.lower():
                ws = wb[sheet_name]
                print(f"Using fallback sheet: {sheet_name}")
                break
    
    if ws is None:
        ws = wb.active
        print(f"Using active sheet as last fallback: {ws.title}")
    
    headers = [cell.value for cell in ws[1]]
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
    
    print(f"Excel headers: {list(col_map.keys())}")
    
    # Helper function to safely get column value
    def safe_get(row, col_name, default_idx, default_val=""):
        idx = col_map.get(col_name, default_idx)
        if idx < len(row):
            return row[idx] or default_val
        return default_val
    
    articles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        # Safely get date
        date_idx = col_map.get("Date", 4)
        date_val = row[date_idx] if date_idx < len(row) else None
        date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, 'strftime') else str(date_val)[:10] if date_val else ""
        
        articles.append({
            "title": safe_get(row, "News Tittle", 3),
            "sector": safe_get(row, "Business Sector", 1),
            "province": safe_get(row, "Province", 2, "Vietnam"),
            "source": safe_get(row, "Source", 5),
            "url": safe_get(row, "Link", 6),
            "summary": safe_get(row, "Short summary", 7),
            "date": date_str,
            "is_new": False
        })
    
    wb.close()
    print(f"Loaded {len(articles)} articles from Excel")
    return articles


def merge_articles(excel_articles, sqlite_articles):
    """Merge articles from both sources"""
    seen_urls = set()
    merged = []
    
    # SQLite first (newer)
    for article in sqlite_articles:
        url = article.get('url', '')
        if url and url not in seen_urls:
            seen_urls.add(url)
            merged.append(article)
    
    # Then Excel
    for article in excel_articles:
        url = article.get('url', '')
        if url and url not in seen_urls:
            seen_urls.add(url)
            merged.append(article)
    
    return merged


def generate_email_html(articles, new_articles):
    """Generate email HTML content"""
    today = datetime.now().strftime("%Y-%m-%d")
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    
    # Count articles
    today_articles = [a for a in articles if a.get("date") == today]
    yesterday_articles = [a for a in articles if a.get("date") == yesterday]
    week_articles = [a for a in articles if a.get("date", "") >= week_ago]
    
    # Recent articles = today + yesterday + new from collector
    recent_articles = today_articles + [a for a in yesterday_articles if a not in today_articles]
    
    # Also include all new articles from SQLite
    for a in new_articles:
        if a not in recent_articles:
            recent_articles.append(a)
    
    sector_counts = Counter(a.get("sector", "Unknown") for a in articles)
    
    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; max-width: 700px; margin: 0 auto; padding: 20px; background: #f8fafc; }}
        .header {{ background: linear-gradient(135deg, #0d9488, #0f766e); color: white; padding: 25px; border-radius: 12px; text-align: center; }}
        .header h1 {{ margin: 0; font-size: 24px; }}
        .kpi-row {{ display: flex; gap: 15px; margin: 20px 0; flex-wrap: wrap; }}
        .kpi {{ flex: 1; min-width: 80px; background: white; padding: 15px; border-radius: 10px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.05); }}
        .kpi-value {{ font-size: 28px; font-weight: bold; color: #0d9488; }}
        .kpi-value.highlight {{ color: #ef4444; }}
        .kpi-label {{ font-size: 12px; color: #64748b; margin-top: 5px; }}
        .section {{ background: white; border-radius: 10px; padding: 20px; margin: 15px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.05); }}
        .article {{ padding: 12px; border-left: 3px solid #0d9488; margin: 10px 0; background: #f8fafc; border-radius: 0 6px 6px 0; }}
        .article.new {{ border-left-color: #ef4444; background: #fef2f2; }}
        .new-badge {{ background: #ef4444; color: white; padding: 2px 6px; border-radius: 4px; font-size: 10px; font-weight: bold; margin-right: 5px; }}
        .btn {{ display: inline-block; background: #0d9488; color: white; padding: 12px 30px; border-radius: 8px; text-decoration: none; }}
        .footer {{ text-align: center; margin-top: 30px; color: #64748b; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🇻🇳 Vietnam Infrastructure News</h1>
        <p style="margin: 5px 0 0 0; opacity: 0.9;">Daily Report - {datetime.now().strftime('%B %d, %Y')}</p>
    </div>
    
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-value highlight">{len(new_articles)}</div>
            <div class="kpi-label">🆕 New Collected</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{len(today_articles)}</div>
            <div class="kpi-label">Today</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{len(yesterday_articles)}</div>
            <div class="kpi-label">Yesterday</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{len(week_articles)}</div>
            <div class="kpi-label">This Week</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{len(articles):,}</div>
            <div class="kpi-label">Total DB</div>
        </div>
    </div>
    
    <div class="section">
        <h3 style="margin-top: 0;">📊 Top Sectors</h3>
        <p>{" | ".join([f"{s}: {c}" for s, c in sector_counts.most_common(5)])}</p>
    </div>
'''
    
    # Show new collected articles
    if new_articles:
        html += f'''
    <div class="section">
        <h3 style="margin-top: 0; color: #ef4444;">🆕 Newly Collected ({len(new_articles)})</h3>
'''
        for a in new_articles[:10]:
            html += f'''
        <div class="article new">
            <span class="new-badge">NEW</span>
            <strong>[{a['sector']}]</strong> {a['title'][:100]}{'...' if len(str(a['title'])) > 100 else ''}
            <br><small style="color: #64748b;">{a['date']} | {a['source']} | <a href="{a['url']}" style="color: #0d9488;">Read more →</a></small>
        </div>
'''
        html += '</div>'
    
    # Show today's articles if different from new
    today_not_new = [a for a in today_articles if not a.get('is_new')]
    if today_not_new:
        html += f'''
    <div class="section">
        <h3 style="margin-top: 0;">📰 Today's Articles ({len(today_not_new)})</h3>
'''
        for a in today_not_new[:5]:
            html += f'''
        <div class="article">
            <strong>[{a['sector']}]</strong> {a['title'][:100]}{'...' if len(str(a['title'])) > 100 else ''}
            <br><small style="color: #64748b;">{a['source']} | <a href="{a['url']}" style="color: #0d9488;">Read more →</a></small>
        </div>
'''
        html += '</div>'
    
    html += f'''
    <div style="text-align: center; margin: 30px 0;">
        <a href="{DASHBOARD_URL}" class="btn">📊 View Full Dashboard</a>
    </div>
    
    <div class="footer">
        <p>Vietnam Infrastructure News Pipeline</p>
        <p style="font-size: 10px;">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
</body>
</html>'''
    
    return html


def send_email(html_content):
    """Send email notification"""
    print(f"\n--- Email Configuration ---")
    print(f"SMTP: {EMAIL_SMTP_SERVER}:{EMAIL_SMTP_PORT}")
    print(f"Username: {EMAIL_USERNAME[:3]}***" if EMAIL_USERNAME else "Username: NOT SET")
    print(f"Password: {'*' * 8 if EMAIL_PASSWORD else 'NOT SET'}")
    print(f"Recipients: '{EMAIL_RECIPIENTS_RAW}'")
    
    if not EMAIL_USERNAME or not EMAIL_PASSWORD:
        print("ERROR: Email credentials not configured")
        return False
    
    recipients = []
    if EMAIL_RECIPIENTS_RAW:
        for sep in [',', ';']:
            if sep in EMAIL_RECIPIENTS_RAW:
                recipients = [r.strip() for r in EMAIL_RECIPIENTS_RAW.split(sep) if r.strip()]
                break
        if not recipients:
            recipients = [EMAIL_RECIPIENTS_RAW.strip()] if EMAIL_RECIPIENTS_RAW.strip() else []
    
    if not recipients:
        print("ERROR: No recipients configured")
        return False
    
    try:
        print(f"\nSending to {len(recipients)} recipient(s)...")
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"{EMAIL_SUBJECT} - {datetime.now().strftime('%Y-%m-%d')}"
        msg['From'] = f"{EMAIL_FROM_NAME} <{EMAIL_USERNAME}>"
        msg['To'] = ', '.join(recipients)
        
        msg.attach(MIMEText(html_content, 'html', 'utf-8'))
        
        server = smtplib.SMTP(EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT, timeout=30)
        server.starttls()
        server.login(EMAIL_USERNAME, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        print(f"✓ Email sent to: {', '.join(recipients)}")
        return True
        
    except smtplib.SMTPAuthenticationError as e:
        print(f"ERROR: Authentication failed - {e}")
        return False
    except Exception as e:
        print(f"ERROR: {type(e).__name__}: {e}")
        return False


def main():
    """Main function - loads from BOTH Excel and SQLite"""
    print("=" * 60)
    print("EMAIL NOTIFICATION (Merged Sources)")
    print("=" * 60)
    
    # Load from both sources
    excel_articles = load_articles_from_excel()
    sqlite_articles = load_articles_from_sqlite()
    
    # Merge
    all_articles = merge_articles(excel_articles, sqlite_articles)
    
    print(f"\nExcel articles: {len(excel_articles)}")
    print(f"SQLite articles (new): {len(sqlite_articles)}")
    print(f"Total merged: {len(all_articles)}")
    
    # Stats
    today = datetime.now().strftime("%Y-%m-%d")
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    
    today_count = sum(1 for a in all_articles if a.get("date") == today)
    yesterday_count = sum(1 for a in all_articles if a.get("date") == yesterday)
    
    print(f"Today's articles: {today_count}")
    print(f"Yesterday's articles: {yesterday_count}")
    print(f"New from collector: {len(sqlite_articles)}")
    
    # Generate and send email
    html = generate_email_html(all_articles, sqlite_articles)
    
    if send_email(html):
        print("\n✓ Email notification sent")
    else:
        print("\n✗ Email send failed")


if __name__ == "__main__":
    main()
