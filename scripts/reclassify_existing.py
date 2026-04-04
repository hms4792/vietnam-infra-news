"""
reclassify_existing.py
Excel News Database 기존 기사를 7개 허용 섹터 기준으로 재분류.

처리:
  1. 허용 외 섹터 행 → 키워드 점수로 재분류
  2. Province="Vietnam" 행 중 제목에 성/시명 포함 → Province 보완
"""

import json
import os
from collections import defaultdict

BASE_DIR      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH    = os.path.join(BASE_DIR, "data", "database", "Vietnam_Infra_News_Database_Final.xlsx")
REPORT_PATH   = os.path.join(BASE_DIR, "data", "agent_output", "reclassify_report.json")

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl 미설치 — pip install openpyxl")

# ── 상수 ──────────────────────────────────────────────────────────────────────

ALLOWED_SECTORS = {
    "Waste Water", "Water Supply/Drainage", "Solid Waste",
    "Power", "Oil & Gas", "Industrial Parks", "Smart City",
}

SECTOR_AREA = {
    "Waste Water":           "Environment",
    "Water Supply/Drainage": "Environment",
    "Solid Waste":           "Environment",
    "Power":                 "Energy Develop.",
    "Oil & Gas":             "Energy Develop.",
    "Industrial Parks":      "Urban Develop.",
    "Smart City":            "Urban Develop.",
}

# 섹터별 키워드 (소문자 매칭, 3점)
SECTOR_KEYWORDS = {
    "Waste Water": [
        "wastewater", "waste water", "sewage", "wwtp",
        "nước thải", "thoát nước", "sludge",
    ],
    "Water Supply/Drainage": [
        "water supply", "clean water", "drinking water",
        "cấp nước", "nước sạch", "flood control",
        "chống ngập", "drainage", "stormwater",
    ],
    "Solid Waste": [
        "solid waste", "landfill", "recycling",
        "rác thải", "chất thải rắn",
        "incineration", "waste-to-energy", "đốt rác",
    ],
    "Power": [
        "wind power", "solar", "electricity",
        "điện gió", "điện mặt trời",
        "evn", "hydro", "thủy điện", "lng power", "năng lượng",
    ],
    "Oil & Gas": [
        "petroleum", "refinery", "gas pipeline",
        "dầu khí", "pvn", "petrovietnam",
        "lng", "crude oil", "dầu thô",
    ],
    "Industrial Parks": [
        "industrial park", "khu công nghiệp", "export processing",
        "khu chế xuất", "fdi factory", "logistics park",
    ],
    "Smart City": [
        "smart city", "digital infrastructure", "iot",
        "thành phố thông minh", "data center",
        "e-government", "5g",
    ],
}

# Province 감지용 (영문, 타이틀 포함 여부)
PROVINCE_LIST = [
    "Hanoi", "Ho Chi Minh City", "Da Nang", "Binh Duong", "Dong Nai",
    "Hai Phong", "Can Tho", "Quang Ninh", "Binh Dinh", "Gia Lai",
    "Khanh Hoa", "Nghe An", "Ha Tinh", "Thanh Hoa", "Quang Nam",
    "Quang Ngai", "Ba Ria Vung Tau", "Long An", "Tien Giang",
    "An Giang", "Soc Trang", "Dak Lak", "Lam Dong", "Ninh Thuan",
    "Binh Thuan", "Hue", "Bac Ninh", "Vinh Phuc", "Thai Nguyen",
    "Nam Dinh", "Ninh Binh",
]

# 컬럼 인덱스 (1-based)
COL_AREA    = 1
COL_SECTOR  = 2
COL_PROVINCE= 3
COL_TITLE   = 4
COL_SUMMARY = 8   # Short Summary


# ── 헬퍼 ─────────────────────────────────────────────────────────────────────

def score_text(text, sector):
    """텍스트에서 섹터 키워드 점수(3점/개) 계산."""
    low = text.lower()
    return sum(3 for kw in SECTOR_KEYWORDS[sector] if kw in low)


def best_sector(title, summary, current_sector):
    """최고 점수 섹터 반환. 동점이면 current_sector 유지. 0점이면 None."""
    text = f"{title} {summary}"
    scores = {s: score_text(text, s) for s in ALLOWED_SECTORS}
    max_score = max(scores.values())
    if max_score == 0:
        return None
    # 동점이면 current가 그 점수를 가지면 current 유지
    if current_sector in ALLOWED_SECTORS and scores.get(current_sector, 0) == max_score:
        return current_sector
    # 최고 점수 섹터 (순서 고정)
    for s in ["Waste Water", "Water Supply/Drainage", "Solid Waste",
              "Power", "Oil & Gas", "Industrial Parks", "Smart City"]:
        if scores[s] == max_score:
            return s


def detect_province(title):
    """제목에서 Province명 추출 (첫 번째 매칭)."""
    for prov in PROVINCE_LIST:
        if prov.lower() in title.lower():
            return prov
    return None


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print(f"Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    if "News Database" not in wb.sheetnames:
        raise SystemExit("[ERROR] 'News Database' 시트 없음")
    ws = wb["News Database"]

    total_rows        = 0
    sector_reclassify = 0
    province_updated  = 0
    unclassified      = 0
    by_sector         = defaultdict(int)

    max_row = ws.max_row
    print(f"전체 행수: {max_row - 1}행 처리 시작...\n")

    for row_idx in range(2, max_row + 1):

        # 100행마다 진행상황 출력
        if (row_idx - 1) % 100 == 0:
            print(f"  진행 중... {row_idx - 1}/{max_row - 1}행")

        title   = str(ws.cell(row_idx, COL_TITLE).value   or "")
        summary = str(ws.cell(row_idx, COL_SUMMARY).value or "")
        sector  = str(ws.cell(row_idx, COL_SECTOR).value  or "")
        province= str(ws.cell(row_idx, COL_PROVINCE).value or "")

        # 빈 행 무시
        if not title:
            continue

        total_rows += 1
        changed = False

        # ── 1. 섹터 재분류 ───────────────────────────────────────────────────
        if sector not in ALLOWED_SECTORS:
            new_sector = best_sector(title, summary, sector)
            old_sector = sector

            if new_sector:
                ws.cell(row_idx, COL_SECTOR).value = new_sector
                ws.cell(row_idx, COL_AREA).value   = SECTOR_AREA[new_sector]
                by_sector[f"{old_sector}→{new_sector}"] += 1
                sector_reclassify += 1
                changed = True
            else:
                # 점수 0점 + 허용 외 → Unclassified
                ws.cell(row_idx, COL_SECTOR).value = "Unclassified"
                by_sector[f"{old_sector}→Unclassified"] += 1
                unclassified += 1
                changed = True

        # ── 2. Province 보완 (Province="Vietnam"이고 제목에 성/시명 포함) ────
        if province == "Vietnam":
            detected = detect_province(title)
            if detected:
                ws.cell(row_idx, COL_PROVINCE).value = detected
                province_updated += 1

    print(f"\n처리 완료. Excel 저장 중...")
    wb.save(EXCEL_PATH)
    print("[OK] Excel 저장 완료")

    # ── 리포트 저장 ──────────────────────────────────────────────────────────
    report = {
        "total_rows":         total_rows,
        "sector_reclassified": sector_reclassify,
        "province_updated":   province_updated,
        "unclassified":       unclassified,
        "by_sector":          dict(by_sector),
    }
    os.makedirs(os.path.dirname(REPORT_PATH), exist_ok=True)
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"[OK] reclassify_report.json 저장 완료")

    # ── 최종 요약 ─────────────────────────────────────────────────────────────
    print(f"\n{'='*50}")
    print(f"  총 처리 행수:       {total_rows}")
    print(f"  섹터 재분류:        {sector_reclassify}행")
    print(f"  Province 보완:      {province_updated}행")
    print(f"  Unclassified:       {unclassified}행")
    print(f"\n  섹터별 변경 내역:")
    for k, v in sorted(by_sector.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v}건")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
