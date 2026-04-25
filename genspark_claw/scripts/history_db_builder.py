"""
History DB Builder — Vietnam Infrastructure Pipeline
============================================================
역할:
  1. GitHub Excel (2,537건) + 현재 주간 기사 통합
  2. 새 매칭 로직(MASTERPLANS + exclude_if) 적용
  3. URL 기준 중복 제거 후 history_db.json 저장
  4. 이후 매주 신규 기사만 증분 업데이트

파일 위치:
  /home/work/claw/config/history_db.json   ← 누적 전체 DB
  /home/work/claw/config/history_meta.json ← 마지막 업데이트 정보
"""

import json, hashlib, sys, os, ast
from datetime import datetime
from pathlib import Path
from collections import defaultdict

BASE_DIR   = Path(__file__).resolve().parent.parent
HISTORY_DB = BASE_DIR / "config" / "history_db.json"
HISTORY_META = BASE_DIR / "config" / "history_meta.json"
EXCEL_PATH = Path("/home/work/vietnam-infra-news/data/Vietnam_Infra_News_Database_Final.xlsx")

sys.path.insert(0, str(BASE_DIR))

def _uid(url: str, title: str) -> str:
    return hashlib.md5((url or title or "").encode()).hexdigest()

def _date_key(a: dict) -> str:
    d = (a.get("published_date") or "")[:10]
    return d if d else "1900-01-01"

def _match_article(text: str, masterplans: list) -> list:
    """exclude_if + keyword 매칭"""
    matched = []
    for plan in masterplans:
        excluded = any(ex.lower() in text for ex in plan.get("exclude_if", []))
        if excluded:
            continue
        score = sum(1 for kw in plan["keywords"] if kw.lower() in text)
        if score >= plan["threshold"]:
            matched.append(plan["id"])
    return matched

# ── 섹터 정규화 매핑 ──────────────────────────────────────────────
SECTOR_NORM = {
    "Power":             "Power",
    "Power & Energy":    "Power",
    "Energy Develop.":   "Power",
    "Oil & Gas":         "Oil & Gas",
    "Transport":         "Transport",
    "Urban Development": "Transport",
    "Solid Waste":       "Solid Waste",
    "Waste Water":       "Waste Water",
    "Water Supply/Drainage": "Water Supply",
    "Smart City":        "Smart City",
    "Industrial Parks":  "Industrial Parks",
    "Climate Change":    "Carbon & Climate",
    "Construction":      "Construction",
    "EV":                "EV / Mobility",
    "Industrial Parks":  "Industrial Parks",
}

def _norm_sector(s: str) -> str:
    return SECTOR_NORM.get(s, s or "Other")

# ── Excel 역사 데이터 로드 ────────────────────────────────────────
def load_excel_history(path: Path) -> list:
    """두 시트 통합 → 정규화된 기사 리스트"""
    try:
        import openpyxl
    except ImportError:
        print("  ⚠ openpyxl 없음 — pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(str(path), read_only=True)
    articles = {}

    # ── News Database 시트 ──
    ws = wb["News Database"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        area, sector, province, title, date, source, url, summary = (row + (None,)*8)[:8]
        if not title or not url:
            continue
        uid = _uid(str(url), str(title))
        articles[uid] = {
            "id":             uid,
            "title":          str(title).strip(),
            "url":            str(url).strip(),
            "published_date": str(date)[:10] if date else "",
            "source":         str(source) if source else "",
            "sector":         _norm_sector(str(sector) if sector else ""),
            "area":           str(area) if area else "",
            "province":       str(province) if province else "",
            "content":        str(summary) if summary else "",
            "summary_ko":     "",
            "summary_en":     str(summary) if summary else "",
            "summary_vi":     "",
            "lang":           "vi",
            "matched_plans":  [],
            "policy_context": None,
            "qc_status":      "PASS",
            "source_sheet":   "News Database",
        }

    # ── Keywords History 시트 (보완) ──
    ws2 = wb["Keywords History"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        no, cat, sector, keyword, province, date, title, source, url, summary = (row + (None,)*10)[:10]
        if not title or not url:
            continue
        uid = _uid(str(url), str(title))
        if uid not in articles:
            articles[uid] = {
                "id":             uid,
                "title":          str(title).strip(),
                "url":            str(url).strip(),
                "published_date": str(date)[:10] if date else "",
                "source":         str(source) if source else "",
                "sector":         _norm_sector(str(sector) if sector else ""),
                "area":           str(cat) if cat else "",
                "province":       str(province) if province else "",
                "content":        str(summary) if summary else "",
                "summary_ko":     "",
                "summary_en":     str(summary) if summary else "",
                "summary_vi":     "",
                "lang":           "vi",
                "matched_plans":  [],
                "policy_context": None,
                "qc_status":      "PASS",
                "source_sheet":   "Keywords History",
            }

    wb.close()
    result = list(articles.values())
    result.sort(key=_date_key, reverse=True)
    print(f"  ✓ Excel 역사 로드: {len(result)}건 (중복 제거 후)")
    return result

# ── 마스터플랜 매칭 적용 ─────────────────────────────────────────
def apply_matching(articles: list, masterplans: list) -> list:
    matched_count = 0
    for art in articles:
        text = (art.get("title","") + " " + art.get("content","")[:600]).lower()
        mp = _match_article(text, masterplans)
        art["matched_plans"] = mp
        if mp:
            best = mp[0]
            plan_meta = next((p for p in masterplans if p["id"] == best), None)
            score = sum(1 for kw in plan_meta["keywords"] if kw.lower() in text) if plan_meta else 1
            art["policy_context"] = {
                "plan_id":   best,
                "plan_name": plan_meta["name_ko"] if plan_meta else best,
                "score":     score,
            }
            matched_count += 1
        else:
            art["policy_context"] = None
    print(f"  ✓ 매칭 완료: {matched_count}/{len(articles)}건 매칭")
    return articles

# ── 기존 history_db 로드 ─────────────────────────────────────────
def load_history_db() -> dict:
    """기존 누적 DB 로드. 없으면 빈 구조 반환."""
    if HISTORY_DB.exists():
        with open(HISTORY_DB, encoding="utf-8") as f:
            return json.load(f)
    return {"version": "1.0", "articles": {}, "last_updated": ""}

# ── 증분 병합 ────────────────────────────────────────────────────
def merge_into_db(existing_db: dict, new_articles: list) -> tuple:
    """
    새 기사를 기존 DB에 병합 (URL 기준 중복 제거).
    기존 기사에 요약이 있으면 유지.
    반환: (updated_db, added_count, updated_count)
    """
    db_arts = existing_db.get("articles", {})
    added, updated = 0, 0

    for art in new_articles:
        uid = art["id"]
        if uid not in db_arts:
            db_arts[uid] = art
            added += 1
        else:
            # 기존 요약 유지, matched_plans만 갱신
            existing = db_arts[uid]
            if not existing.get("summary_ko") and art.get("summary_ko"):
                existing["summary_ko"] = art["summary_ko"]
                existing["summary_en"] = art["summary_en"]
                existing["summary_vi"] = art["summary_vi"]
            # matched_plans 업데이트 (새 로직 적용)
            if art.get("matched_plans") != existing.get("matched_plans"):
                existing["matched_plans"] = art["matched_plans"]
                existing["policy_context"] = art["policy_context"]
                updated += 1

    existing_db["articles"] = db_arts
    existing_db["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    existing_db["total"] = len(db_arts)
    return existing_db, added, updated

# ── 플랜별 버킷 추출 ─────────────────────────────────────────────
def get_plan_buckets(db: dict) -> dict:
    """history_db에서 플랜별 기사 리스트 추출, 최신순 정렬"""
    buckets = defaultdict(list)
    for art in db["articles"].values():
        for pid in art.get("matched_plans", []):
            buckets[pid].append(art)
    # 최신순 정렬
    for pid in buckets:
        buckets[pid].sort(key=_date_key, reverse=True)
    return dict(buckets)

# ── 저장 ─────────────────────────────────────────────────────────
def save_history_db(db: dict):
    HISTORY_DB.parent.mkdir(parents=True, exist_ok=True)
    with open(HISTORY_DB, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    # 메타
    buckets = get_plan_buckets(db)
    meta = {
        "last_updated": db["last_updated"],
        "total_articles": db["total"],
        "per_plan": {pid: len(arts) for pid, arts in buckets.items()},
    }
    with open(HISTORY_META, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


# ════════════════════════════════════════════════════════════════
# CLI 실행
# ════════════════════════════════════════════════════════════════
def build_full(force_rebuild=False):
    """최초 전체 빌드 (또는 force_rebuild=True 시 재구축)"""
    from scripts.agent_pipeline import MASTERPLANS

    print("="*60)
    print("History DB 전체 구축 시작")
    print("="*60)

    # 기존 DB
    db = {} if force_rebuild else load_history_db()
    if force_rebuild or not db.get("articles"):
        db = {"version": "1.0", "articles": {}, "last_updated": ""}

    # 1. Excel 역사 데이터 로드
    print("\n[1] Excel 역사 데이터 로드...")
    history_arts = load_excel_history(EXCEL_PATH)

    # 2. 매칭 적용
    print("[2] 마스터플랜 매칭 적용...")
    history_arts = apply_matching(history_arts, MASTERPLANS)

    # 3. 현재 주간 기사 병합
    print("[3] 현재 주간 기사 병합...")
    weekly_path = BASE_DIR / "outputs" / "genspark_output.json"
    if weekly_path.exists():
        weekly = json.load(open(weekly_path, encoding="utf-8"))
        if not isinstance(weekly, list):
            weekly = weekly.get("articles", [])
        for a in weekly:
            mp = a.get("matched_plans", [])
            if isinstance(mp, str):
                try: a["matched_plans"] = ast.literal_eval(mp)
                except: a["matched_plans"] = []
        # 주간 기사도 재매칭
        weekly = apply_matching(weekly, MASTERPLANS)
        print(f"     주간 기사: {len(weekly)}건")
        all_arts = history_arts + weekly
    else:
        all_arts = history_arts

    # 4. DB 병합
    print("[4] DB 병합 중...")
    db, added, updated = merge_into_db(db, all_arts)
    print(f"     신규: {added}건, 업데이트: {updated}건, 전체: {db['total']}건")

    # 5. 저장
    save_history_db(db)
    print(f"\n✅ history_db.json 저장: {db['total']}건")

    # 6. 통계 출력
    buckets = get_plan_buckets(db)
    print("\n📊 플랜별 역사 기사 현황:")
    for pid, arts in sorted(buckets.items(), key=lambda x:-len(x[1])):
        print(f"  {pid:25s}: {len(arts):4d}건")

    return db


def update_weekly(weekly_articles: list):
    """매주 실행 — 신규 기사만 증분 추가"""
    from scripts.agent_pipeline import MASTERPLANS

    print("[History] 주간 증분 업데이트 중...")
    db = load_history_db()
    if not db.get("articles"):
        print("  ⚠ history_db 없음 — 전체 빌드 먼저 실행 필요")
        return db

    # 재매칭
    weekly_articles = apply_matching(weekly_articles, MASTERPLANS)
    db, added, updated = merge_into_db(db, weekly_articles)
    save_history_db(db)
    print(f"  ✓ 신규 {added}건 추가, {updated}건 업데이트 → 누적 {db['total']}건")
    return db


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--rebuild", action="store_true", help="강제 전체 재구축")
    args = parser.parse_args()
    build_full(force_rebuild=args.rebuild)
