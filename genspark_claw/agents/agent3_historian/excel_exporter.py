"""
excel_history_exporter.py — 역사 DB 기반 Excel 완전 재구성 v1.0
=================================================================
시트 구조:
  1. [All_Articles]   — 전체 2,668건 (매핑기사: 노란색)
  2. [Matched_Only]   — 매핑된 1,535건 (노란색) 최신순
  3. [플랜ID×24]      — 플랜별 관련기사 시트 (최신순 정렬)
  4. [Stats]          — 플랜별 통계 요약

컬러 규칙:
  - 매핑 기사 행: 노란색 (FFFF99) 하이라이트
  - 매핑 플랜 셀: 각 플랜 테마색
  - 헤더: 네이비 (0F172A) + 흰 글씨
"""
import os, sys, json, re
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2

BASE_DIR = Path("/home/work/claw")

# ─── 색상 상수 ──────────────────────────────────────────────
FILL_HEADER  = PatternFill("solid", fgColor="0F172A")   # 네이비 헤더
FILL_MATCHED = PatternFill("solid", fgColor="FFFF99")   # 노란색 — 매핑 기사
FILL_RECENT  = PatternFill("solid", fgColor="FFF3CD")   # 연노랑 — 2026년 기사
FILL_SUBHDR  = PatternFill("solid", fgColor="1E3A5F")   # 섹션 소헤더
FILL_ALT     = PatternFill("solid", fgColor="F8FAFC")   # 짝수행 밝은색
FILL_STATS   = PatternFill("solid", fgColor="EFF6FF")   # 통계 배경

FONT_HEADER  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_SUBHDR  = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
FONT_BODY    = Font(name="Calibri", size=9)
FONT_BOLD    = Font(name="Calibri", bold=True, size=9)
FONT_LINK    = Font(name="Calibri", size=9, color="1E40AF", underline="single")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
ALIGN_LEFT_NW= Alignment(horizontal="left",   vertical="center", wrap_text=False)

THIN = Side(style="thin",   color="CBD5E1")
MED  = Side(style="medium", color="94A3B8")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 플랜별 테마색 (헤더 색상)
PLAN_COLORS = {
    "VN-PDP8-RENEWABLE": "F59E0B", "VN-PDP8-LNG":       "D97706",
    "VN-PDP8-NUCLEAR":   "B45309", "VN-PDP8-COAL":       "92400E",
    "VN-PDP8-GRID":      "78350F", "VN-PDP8-HYDROGEN":   "FBBF24",
    "VN-WAT-RESOURCES":  "0EA5E9", "VN-WAT-URBAN":       "0284C7",
    "VN-WAT-RURAL":      "0369A1", "HN-URBAN-INFRA":     "DC2626",
    "HN-URBAN-NORTH":    "B91C1C", "HN-URBAN-WEST":      "991B1B",
    "VN-TRAN-2055":      "EA580C", "VN-URB-METRO-2030":  "C2410C",
    "VN-MEKONG-DELTA-2030":"0D9488","VN-RED-RIVER-2030": "BE123C",
    "VN-IP-NORTH-2030":  "7C3AED", "VN-ENV-IND-1894":    "15803D",
    "VN-WW-2030":        "0891B2", "VN-SWM-NATIONAL-2030":"16A34A",
    "VN-SC-2030":        "6D28D9", "VN-OG-2030":         "D97706",
    "VN-EV-2030":        "0D9488", "VN-CARBON-2050":     "166534",
}

PLAN_NAMES = {
    "VN-PDP8-RENEWABLE":    "PDP8-재생에너지",
    "VN-PDP8-LNG":          "PDP8-LNG",
    "VN-PDP8-NUCLEAR":      "PDP8-원자력",
    "VN-PDP8-COAL":         "PDP8-석탄",
    "VN-PDP8-GRID":         "PDP8-전력망",
    "VN-PDP8-HYDROGEN":     "PDP8-수소",
    "VN-WAT-RESOURCES":     "수자원관리",
    "VN-WAT-URBAN":         "도시상수도",
    "VN-WAT-RURAL":         "농촌급수",
    "HN-URBAN-INFRA":       "하노이인프라",
    "HN-URBAN-NORTH":       "하노이북부",
    "HN-URBAN-WEST":        "하노이서부",
    "VN-TRAN-2055":         "교통2055",
    "VN-URB-METRO-2030":    "도시메트로",
    "VN-MEKONG-DELTA-2030": "메콩델타",
    "VN-RED-RIVER-2030":    "홍강델타",
    "VN-IP-NORTH-2030":     "북부산업단지",
    "VN-ENV-IND-1894":      "환경산업1894",
    "VN-WW-2030":           "폐수처리",
    "VN-SWM-NATIONAL-2030": "고형폐기물",
    "VN-SC-2030":           "스마트시티",
    "VN-OG-2030":           "석유가스",
    "VN-EV-2030":           "전기차",
    "VN-CARBON-2050":       "탄소중립",
}

# ─── 헬퍼 함수 ──────────────────────────────────────────────
def _is_vi(text):
    vi_chars = set('ăâêôơưđáàảãạắằẳẵặấầẩẫậéèẻẽẹếềểễệíìỉĩịóòỏõọốồổỗộớờởỡợúùủũụứừửữựýỳỷỹỵ')
    return any(c.lower() in vi_chars for c in (text or ''))

def _date_str(art):
    d = str(art.get('published_date','') or '')
    try:
        if re.match(r'\d{4}-\d{2}-\d{2}', d): return d[:10]
        return datetime.strptime(d.strip(), '%b %d, %Y').strftime('%Y-%m-%d')
    except: return ''

def _best_title(art):
    t = art.get('title','') or ''
    if not _is_vi(t): return t
    en = art.get('title_en','') or art.get('summary_en','') or ''
    if en and not _is_vi(en): return f"[VI→EN] {en[:120]}"
    return t

def _best_summary(art, lang='ko'):
    ko = art.get('summary_ko','') or ''
    en = art.get('summary_en','') or ''
    if lang == 'ko':
        if ko and not _is_vi(ko): return ko[:300]
        if en and not _is_vi(en): return f"[EN] {en[:280]}"
    else:
        if en and not _is_vi(en): return en[:300]
        if ko and not _is_vi(ko): return f"[KO] {ko[:280]}"
    return ''

def _set_header_row(ws, headers, col_widths, row=1, fill=FILL_HEADER, font=FONT_HEADER):
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.fill = fill; cell.font = font
        cell.alignment = ALIGN_CENTER; cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _apply_row_style(ws, row_idx, is_matched, is_2026, n_cols):
    fill = FILL_MATCHED if is_matched else (FILL_RECENT if is_2026 else
           (FILL_ALT if row_idx % 2 == 0 else None))
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        if fill: cell.fill = fill
        cell.border = BORDER_THIN
        cell.font = FONT_BODY
        if col == 1: cell.alignment = ALIGN_CENTER
        else: cell.alignment = ALIGN_LEFT_NW

def _freeze_and_filter(ws, freeze_row=2):
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    ws.auto_filter.ref = ws.dimensions


# ─── 시트 생성 함수 ─────────────────────────────────────────

def build_all_articles_sheet(ws, articles):
    """전체 기사 시트 — 매핑 기사 노란색"""
    headers   = ["No","제목(EN/VI→EN)","출처","날짜","Province","섹터","매핑 플랜","요약(KO)","URL","QC"]
    col_widths = [5,  55,              18,   12,   18,        15,   35,         60,         40,  8]
    _set_header_row(ws, headers, col_widths)

    sorted_arts = sorted(articles, key=lambda a: _date_str(a), reverse=True)
    for i, art in enumerate(sorted_arts, 1):
        plans = art.get('matched_plans') or []
        if isinstance(plans, str):
            try: plans = json.loads(plans)
            except: plans = [plans] if plans else []
        is_matched = len(plans) > 0
        date_s = _date_str(art)
        is_2026 = date_s.startswith('2026')

        row = i + 1
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = _best_title(art)
        ws.cell(row=row, column=3).value = (art.get('source_name') or art.get('source',''))[:25]
        ws.cell(row=row, column=4).value = date_s
        ws.cell(row=row, column=5).value = (art.get('province') or '')[:20]
        ws.cell(row=row, column=6).value = (art.get('sector') or '')[:18]
        ws.cell(row=row, column=7).value = ', '.join(plans[:3]) if plans else ''
        ws.cell(row=row, column=8).value = _best_summary(art, 'ko')[:250]
        url = art.get('url','')
        if url:
            ws.cell(row=row, column=9).hyperlink = url
            ws.cell(row=row, column=9).value = url[:50]
            ws.cell(row=row, column=9).font = FONT_LINK
        ws.cell(row=row, column=10).value = art.get('qc_status','')

        _apply_row_style(ws, row, is_matched, is_2026, 10)

        # 매핑 플랜 셀 강조
        if is_matched:
            plan_color = PLAN_COLORS.get(plans[0], "FFFF99")
            ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=plan_color)
            ws.cell(row=row, column=7).font = Font(name="Calibri", size=9, color="FFFFFF", bold=True)

    ws.row_dimensions[1].height = 22
    _freeze_and_filter(ws)
    return len(sorted_arts)


def build_matched_only_sheet(ws, articles):
    """매핑 기사 전용 시트 — 노란색, 최신순"""
    headers   = ["No","제목(EN/VI→EN)","출처","날짜","Province","매핑 플랜(전체)","요약(KO)","요약(EN)","URL"]
    col_widths = [5,  55,              18,   12,   18,         45,               55,        55,        40]
    _set_header_row(ws, headers, col_widths, fill=PatternFill("solid", fgColor="1E3A5F"))

    matched = [a for a in articles if a.get('matched_plans')]
    sorted_arts = sorted(matched, key=lambda a: _date_str(a), reverse=True)

    for i, art in enumerate(sorted_arts, 1):
        plans = art.get('matched_plans') or []
        if isinstance(plans, str):
            try: plans = json.loads(plans)
            except: plans = [plans] if plans else []
        date_s = _date_str(art)
        is_2026 = date_s.startswith('2026')

        row = i + 1
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = _best_title(art)
        ws.cell(row=row, column=3).value = (art.get('source_name') or art.get('source',''))[:25]
        ws.cell(row=row, column=4).value = date_s
        ws.cell(row=row, column=5).value = (art.get('province') or '')[:20]
        ws.cell(row=row, column=6).value = ' | '.join(plans)
        ws.cell(row=row, column=7).value = _best_summary(art, 'ko')[:250]
        ws.cell(row=row, column=8).value = _best_summary(art, 'en')[:250]
        url = art.get('url','')
        if url:
            ws.cell(row=row, column=9).hyperlink = url
            ws.cell(row=row, column=9).value = url[:50]
            ws.cell(row=row, column=9).font = FONT_LINK

        # 매핑 기사는 항상 노란색
        fill = FILL_RECENT if is_2026 else FILL_MATCHED
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = BORDER_THIN
            ws.cell(row=row, column=col).font = FONT_BODY
            if col == 1: ws.cell(row=row, column=col).alignment = ALIGN_CENTER
            else: ws.cell(row=row, column=col).alignment = ALIGN_LEFT_NW

        # 플랜 셀 색상
        if plans:
            plan_color = PLAN_COLORS.get(plans[0], "FFFF99")
            ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor=plan_color)
            ws.cell(row=row, column=6).font = Font(name="Calibri", size=9, color="FFFFFF", bold=True)

    ws.row_dimensions[1].height = 22
    _freeze_and_filter(ws)
    return len(sorted_arts)


def build_plan_sheet(ws, plan_id, articles, plan_name):
    """플랜별 기사 시트 — 최신순, 노란색"""
    theme_color = PLAN_COLORS.get(plan_id, "1E40AF")
    fill_header = PatternFill("solid", fgColor=theme_color)

    headers   = ["No","제목(EN/VI→EN)","출처","날짜","Province","요약(KO)","요약(EN)","URL","히스토리여부"]
    col_widths = [5,  55,              18,   12,   18,         55,        55,        40,   10]
    _set_header_row(ws, headers, col_widths, fill=fill_header)

    sorted_arts = sorted(articles, key=lambda a: _date_str(a), reverse=True)

    for i, art in enumerate(sorted_arts, 1):
        date_s = _date_str(art)
        is_2026 = date_s.startswith('2026')
        row = i + 1

        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = _best_title(art)
        ws.cell(row=row, column=3).value = (art.get('source_name') or art.get('source',''))[:25]
        ws.cell(row=row, column=4).value = date_s
        ws.cell(row=row, column=5).value = (art.get('province') or '')[:20]
        ws.cell(row=row, column=6).value = _best_summary(art, 'ko')[:250]
        ws.cell(row=row, column=7).value = _best_summary(art, 'en')[:250]
        url = art.get('url','')
        if url:
            ws.cell(row=row, column=8).hyperlink = url
            ws.cell(row=row, column=8).value = url[:50]
            ws.cell(row=row, column=8).font = FONT_LINK
        ws.cell(row=row, column=9).value = "히스토리" if not is_2026 else "최신(2026)"

        # 모두 노란색 (매핑 기사이므로), 2026년은 연노랑
        fill = FILL_RECENT if is_2026 else FILL_MATCHED
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = BORDER_THIN
            ws.cell(row=row, column=col).font = FONT_BODY
            if col == 1: ws.cell(row=row, column=col).alignment = ALIGN_CENTER
            else: ws.cell(row=row, column=col).alignment = ALIGN_LEFT_NW

        # 히스토리 여부 색상
        if is_2026:
            ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor="D1FAE5")
            ws.cell(row=row, column=9).font = Font(name="Calibri", size=9, color="166534", bold=True)
        else:
            ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor="FEF3C7")
            ws.cell(row=row, column=9).font = Font(name="Calibri", size=9, color="92400E")

    ws.row_dimensions[1].height = 22
    _freeze_and_filter(ws)
    return len(sorted_arts)


def build_stats_sheet(ws, articles, plan_counts):
    """통계 요약 시트"""
    # 헤더
    now = datetime.now()
    week = now.isocalendar()[1]
    ws.cell(row=1, column=1).value = "VIETNAM INFRASTRUCTURE NEWS — 마스터플랜별 기사 통계 (누적 DB)"
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14, color="0F172A")
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1).alignment = ALIGN_CENTER

    matched_cnt = sum(1 for a in articles if a.get('matched_plans'))
    ws.cell(row=2, column=1).value = (
        f"최종 업데이트: {now.strftime('%Y-%m-%d %H:%M UTC')}  |  W{week:02d}  |  "
        f"총 기사: {len(articles):,}건  |  매핑: {matched_cnt:,}건 ({matched_cnt/max(len(articles),1)*100:.1f}%)  |  "
        f"※ 매주 파이프라인 실행 시 자동 갱신 (누적)"
    )
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=10, color="64748B")
    ws.merge_cells('A2:H2')

    # 컬럼 헤더
    stat_headers = ["플랜 ID","플랜명(KO)","전체기사","히스토리","2026년","매핑율(%)","최신기사날짜","테마색"]
    stat_widths  = [28,       20,          10,        10,        10,      10,          15,           12]
    _set_header_row(ws, stat_headers, stat_widths, row=4)

    total_arts = len(articles)
    matched_arts = [a for a in articles if a.get('matched_plans')]

    # 연도별 전체 기사
    yr_cnt = Counter(_date_str(a)[:4] for a in articles)

    row = 5
    for plan_id in sorted(plan_counts.keys()):
        plan_arts = plan_counts[plan_id]
        hist_cnt = sum(1 for a in plan_arts if not _date_str(a).startswith('2026'))
        new_cnt  = sum(1 for a in plan_arts if _date_str(a).startswith('2026'))
        dates = sorted([_date_str(a) for a in plan_arts if _date_str(a)], reverse=True)
        last_date = dates[0] if dates else ''
        match_rate = len(plan_arts) / total_arts * 100 if total_arts else 0
        theme = PLAN_COLORS.get(plan_id, "1E40AF")

        ws.cell(row=row, column=1).value = plan_id
        ws.cell(row=row, column=2).value = PLAN_NAMES.get(plan_id, plan_id)
        ws.cell(row=row, column=3).value = len(plan_arts)
        ws.cell(row=row, column=4).value = hist_cnt
        ws.cell(row=row, column=5).value = new_cnt
        ws.cell(row=row, column=6).value = round(match_rate, 1)
        ws.cell(row=row, column=7).value = last_date
        ws.cell(row=row, column=8).value = f"#{theme}"
        ws.cell(row=row, column=8).fill = PatternFill("solid", fgColor=theme)
        ws.cell(row=row, column=8).font = Font(name="Calibri", size=9, color="FFFFFF")

        fill_row = FILL_ALT if row % 2 == 0 else None
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BORDER_THIN
            ws.cell(row=row, column=col).font = FONT_BODY
            ws.cell(row=row, column=col).alignment = ALIGN_LEFT_NW if col > 1 else ALIGN_CENTER
            if col < 8 and fill_row:
                ws.cell(row=row, column=col).fill = fill_row
        row += 1

    # 합계 행
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=2).value = f"24개 플랜"
    ws.cell(row=row, column=3).value = len(matched_arts)
    ws.cell(row=row, column=4).value = sum(1 for a in matched_arts if not _date_str(a).startswith('2026'))
    ws.cell(row=row, column=5).value = sum(1 for a in matched_arts if _date_str(a).startswith('2026'))
    ws.cell(row=row, column=6).value = round(len(matched_arts)/total_arts*100 if total_arts else 0, 1)
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="0F172A")
        ws.cell(row=row, column=col).font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        ws.cell(row=row, column=col).border = BORDER_THIN
        ws.cell(row=row, column=col).alignment = ALIGN_CENTER

    # 연도별 분포
    row += 2
    ws.cell(row=row, column=1).value = "연도별 기사 분포"
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color="0F172A")
    row += 1
    for yr in sorted(yr_cnt.keys()):
        ws.cell(row=row, column=1).value = yr
        ws.cell(row=row, column=2).value = f"{yr_cnt[yr]:,}건"
        ws.cell(row=row, column=1).font = FONT_BOLD
        ws.cell(row=row, column=2).font = FONT_BODY
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.freeze_panes = 'A5'


# ─── 메인 생성 함수 ─────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════
# 참조 시트 4개 — 변경이력 추적
# ═══════════════════════════════════════════════════════════════════════

CHANGE_LOG_PATH = BASE_DIR / "config" / "ref_change_log.json"

def _load_change_log() -> dict:
    """변경이력 JSON 로드 (없으면 초기화)"""
    if CHANGE_LOG_PATH.exists():
        return json.loads(CHANGE_LOG_PATH.read_text(encoding="utf-8"))
    return {"masterplan_keywords": [], "sources": [], "provinces": [], "search_queries": []}

def _save_change_log(log: dict):
    CHANGE_LOG_PATH.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")

def _detect_changes(log_key: str, current_set: set, log: dict) -> list:
    """이전 snapshot과 비교하여 추가/삭제 항목 검출"""
    entries = log.get(log_key, [])
    if not entries:
        # 첫 실행: 현재를 baseline으로 저장
        log[log_key] = [{"date": datetime.now().strftime("%Y-%m-%d"), "action": "init",
                         "items": sorted(current_set)}]
        return []
    # 마지막 snapshot의 items
    last_items = set(entries[-1].get("items", []))
    added = current_set - last_items
    removed = last_items - current_set
    changes = []
    ts = datetime.now().strftime("%Y-%m-%d")
    if added:
        changes.append({"date": ts, "action": "added", "items": sorted(added)})
    if removed:
        changes.append({"date": ts, "action": "removed", "items": sorted(removed)})
    if added or removed:
        entries.extend(changes)
        # 최신 스냅샷 저장
        entries.append({"date": ts, "action": "snapshot", "items": sorted(current_set)})
        log[log_key] = entries
    return changes


def _header_row(ws, headers, row=1):
    """헤더 행 스타일링"""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center")


def build_ref_masterplan_keywords(wb, articles):
    """시트: REF_Plan_Keywords — 마스터플랜별 키워드 + 변경이력"""
    ws = wb.create_sheet("REF_Plan_Keywords")

    # agent_pipeline.py에서 MASTERPLANS 로드
    sys.path.insert(0, str(BASE_DIR / "scripts"))
    try:
        from agent_pipeline import MASTERPLANS
    except ImportError:
        ws.cell(1, 1, value="MASTERPLANS 로드 실패")
        return

    _header_row(ws, ["Plan ID", "Plan Name (KO)", "Parent", "Keywords", "Exclude_If", "Threshold", "기사 수"])
    r = 2
    for plan in MASTERPLANS:
        if not isinstance(plan, dict) or "id" not in plan:
            continue
        pid = plan["id"]
        kws = plan.get("keywords", [])
        exc = plan.get("exclude_if", [])
        # 해당 플랜 매핑 기사 수
        count = sum(1 for a in articles if pid in (a.get("matched_plans") or []))
        ws.cell(r, 1, value=pid)
        ws.cell(r, 2, value=plan.get("name_ko", ""))
        ws.cell(r, 3, value=plan.get("parent", "") or "")
        ws.cell(r, 4, value=", ".join(kws))
        ws.cell(r, 5, value=", ".join(exc))
        ws.cell(r, 6, value=plan.get("threshold", 2))
        ws.cell(r, 7, value=count)
        if count > 0:
            ws.cell(r, 7).fill = PatternFill("solid", fgColor="D1FAE5")
        r += 1

    # 변경이력 섹션
    log = _load_change_log()
    current_ids = {p["id"] for p in MASTERPLANS if isinstance(p, dict) and "id" in p}
    changes = _detect_changes("masterplan_keywords", current_ids, log)
    _save_change_log(log)

    r += 1
    ws.cell(r, 1, value="── 변경이력 ──").font = Font(bold=True, size=11, color="B45309")
    r += 1
    _header_row(ws, ["날짜", "변경 유형", "항목"], row=r)
    r += 1
    for entry in log.get("masterplan_keywords", []):
        if entry["action"] == "snapshot":
            continue
        ws.cell(r, 1, value=entry["date"])
        ws.cell(r, 2, value=entry["action"])
        ws.cell(r, 3, value=", ".join(entry.get("items", [])))
        if entry["action"] == "added":
            ws.cell(r, 2).fill = PatternFill("solid", fgColor="D1FAE5")
        elif entry["action"] == "removed":
            ws.cell(r, 2).fill = PatternFill("solid", fgColor="FEE2E2")
        r += 1

    # 컬럼 너비
    for c, w in enumerate([25, 30, 15, 80, 50, 10, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_ref_sources(wb, articles):
    """시트: REF_Sources — 자료출처 목록 (도메인 기준) + 변경이력"""
    ws = wb.create_sheet("REF_Sources")

    # 도메인별 집계
    from urllib.parse import urlparse as _up
    domain_counts = Counter()
    domain_examples = {}
    for a in articles:
        url = a.get("url", "")
        if not url:
            continue
        domain = _up(url).netloc.replace("www.", "") if url.startswith("http") else (a.get("source", "") or "unknown")
        domain_counts[domain] += 1
        if domain not in domain_examples:
            domain_examples[domain] = url

    _header_row(ws, ["Source Domain", "기사 수", "비율(%)", "예시 URL", "유형"])
    total = sum(domain_counts.values()) or 1

    # 전문미디어 도메인
    PRO_MEDIA = {"vir.com.vn", "theinvestor.vn", "vnexpress.net", "vietnamnet.vn",
                 "tuoitre.vn", "thanhnien.vn", "english.thesaigontimes.vn",
                 "pv-magazine.com", "renewableenergyworld.com", "power-eng.com",
                 "bloomberg.com", "reuters.com", "nikkeiasia.com"}

    for r, (domain, cnt) in enumerate(domain_counts.most_common(), 2):
        ws.cell(r, 1, value=domain)
        ws.cell(r, 2, value=cnt)
        ws.cell(r, 3, value=round(cnt/total*100, 1))
        ws.cell(r, 4, value=domain_examples.get(domain, ""))
        media_type = "전문미디어" if domain in PRO_MEDIA else "정부" if ".gov" in domain else "기타"
        ws.cell(r, 5, value=media_type)
        if media_type == "전문미디어":
            ws.cell(r, 5).fill = PatternFill("solid", fgColor="DBEAFE")
        elif media_type == "정부":
            ws.cell(r, 5).fill = PatternFill("solid", fgColor="FEF3C7")

    # 변경이력
    log = _load_change_log()
    current_domains = set(domain_counts.keys())
    changes = _detect_changes("sources", current_domains, log)
    _save_change_log(log)

    r = len(domain_counts) + 3
    ws.cell(r, 1, value="── 변경이력 ──").font = Font(bold=True, size=11, color="B45309")
    r += 1
    _header_row(ws, ["날짜", "변경 유형", "도메인"], row=r)
    r += 1
    for entry in log.get("sources", []):
        if entry["action"] == "snapshot":
            continue
        ws.cell(r, 1, value=entry["date"])
        ws.cell(r, 2, value=entry["action"])
        ws.cell(r, 3, value=", ".join(entry.get("items", [])[:20]))
        r += 1

    for c, w in enumerate([35, 10, 10, 60, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_ref_provinces(wb, articles):
    """시트: REF_Provinces — Province 매핑 현황 + 키워드 + 변경이력"""
    ws = wb.create_sheet("REF_Provinces")

    # config/province_keywords.json 로드
    prov_kw_path = BASE_DIR / "config" / "province_keywords.json"
    prov_kw = {}
    if prov_kw_path.exists():
        data = json.loads(prov_kw_path.read_text(encoding="utf-8"))
        prov_kw = {k: v for k, v in data.items() if not k.startswith("_")}

    # 기사별 province 집계
    prov_counts = Counter()
    for a in articles:
        prov = a.get("province", "Vietnam")
        prov_counts[prov] += 1

    _header_row(ws, ["Province", "기사 수", "비율(%)", "검색 키워드", "키워드 수"])
    total = sum(prov_counts.values()) or 1

    r = 2
    for prov in sorted(prov_kw.keys()):
        cnt = prov_counts.get(prov, 0)
        kws = prov_kw.get(prov, [])
        ws.cell(r, 1, value=prov)
        ws.cell(r, 2, value=cnt)
        ws.cell(r, 3, value=round(cnt/total*100, 1))
        ws.cell(r, 4, value=", ".join(kws))
        ws.cell(r, 5, value=len(kws))
        if cnt > 0:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor="D1FAE5")
        r += 1
    # 미분류 추가
    unspec_cnt = prov_counts.get("Vietnam", 0)
    ws.cell(r, 1, value="Vietnam")
    ws.cell(r, 2, value=unspec_cnt)
    ws.cell(r, 3, value=round(unspec_cnt/total*100, 1))
    ws.cell(r, 4, value="(default)")
    ws.cell(r, 1).fill = PatternFill("solid", fgColor="FEF3C7")
    r += 1

    # 변경이력
    log = _load_change_log()
    current_provs = set(prov_kw.keys())
    changes = _detect_changes("provinces", current_provs, log)
    _save_change_log(log)

    r += 1
    ws.cell(r, 1, value="── 변경이력 ──").font = Font(bold=True, size=11, color="B45309")
    r += 1
    _header_row(ws, ["날짜", "변경 유형", "Province"], row=r)
    r += 1
    for entry in log.get("provinces", []):
        if entry["action"] == "snapshot":
            continue
        ws.cell(r, 1, value=entry["date"])
        ws.cell(r, 2, value=entry["action"])
        ws.cell(r, 3, value=", ".join(entry.get("items", [])[:20]))
        r += 1

    for c, w in enumerate([25, 10, 10, 60, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_ref_search_queries(wb):
    """시트: REF_Search_Queries — 뉴스 검색 키워드 목록 + 변경이력"""
    ws = wb.create_sheet("REF_Search_Queries")

    sys.path.insert(0, str(BASE_DIR / "scripts"))
    try:
        from agent_pipeline import SEARCH_QUERIES
    except ImportError:
        ws.cell(1, 1, value="SEARCH_QUERIES 로드 실패")
        return

    _header_row(ws, ["#", "검색 쿼리", "비고"])
    for i, q in enumerate(SEARCH_QUERIES, 1):
        ws.cell(i+1, 1, value=i)
        ws.cell(i+1, 2, value=q)

    # 변경이력
    log = _load_change_log()
    current_queries = set(SEARCH_QUERIES)
    changes = _detect_changes("search_queries", current_queries, log)
    _save_change_log(log)

    r = len(SEARCH_QUERIES) + 3
    ws.cell(r, 1, value="── 변경이력 ──").font = Font(bold=True, size=11, color="B45309")
    r += 1
    _header_row(ws, ["날짜", "변경 유형", "쿼리"], row=r)
    r += 1
    for entry in log.get("search_queries", []):
        if entry["action"] == "snapshot":
            continue
        ws.cell(r, 1, value=entry["date"])
        ws.cell(r, 2, value=entry["action"])
        ws.cell(r, 3, value=", ".join(entry.get("items", [])[:10]))
        if entry["action"] == "added":
            ws.cell(r, 2).fill = PatternFill("solid", fgColor="D1FAE5")
        elif entry["action"] == "removed":
            ws.cell(r, 2).fill = PatternFill("solid", fgColor="FEE2E2")
        r += 1

    for c, w in enumerate([5, 65, 30], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def generate_history_excel(output_path: str = None) -> str:
    print("  ▶ history_db.json 로드...")
    with open(BASE_DIR / "config/history_db.json", encoding='utf-8') as f:
        hdb = json.load(f)

    articles = list(hdb.get('articles', {}).values())
    print(f"    총 {len(articles):,}건 로드")

    # 플랜별 기사 버킷
    plan_buckets = defaultdict(list)
    for art in articles:
        plans = art.get('matched_plans') or []
        if isinstance(plans, str):
            try: plans = json.loads(plans)
            except: plans = [plans] if plans else []
        for p in plans:
            plan_buckets[p].append(art)

    matched_count = sum(1 for a in articles if a.get('matched_plans'))
    print(f"    매핑된 기사: {matched_count:,}건 ({matched_count/len(articles)*100:.1f}%)")
    print(f"    매핑된 플랜: {len(plan_buckets)}개")

    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws_all = wb.active; ws_all.title = "All_Articles"

    print("  ▶ [1/3] All_Articles 시트...")
    n1 = build_all_articles_sheet(ws_all, articles)
    print(f"    {n1:,}건 작성")

    print("  ▶ [2/3] Matched_Only 시트...")
    ws_matched = wb.create_sheet("Matched_Only")
    n2 = build_matched_only_sheet(ws_matched, articles)
    print(f"    {n2:,}건 작성")

    print("  ▶ [3/3] 플랜별 시트 24개...")
    PLAN_ORDER = [
        "VN-PWR-PDP8","VN-PWR-PDP8-RENEWABLE","VN-PWR-PDP8-LNG","VN-PWR-PDP8-NUCLEAR",
        "VN-PWR-PDP8-COAL","VN-PWR-PDP8-GRID",
        # Legacy IDs (history에 남아있는 기사)
        "VN-PDP8-RENEWABLE","VN-PDP8-LNG","VN-PDP8-NUCLEAR","VN-PDP8-COAL",
        "VN-PDP8-GRID","VN-PDP8-HYDROGEN",
        "VN-WAT-RESOURCES","VN-WAT-URBAN","VN-WAT-RURAL",
        "HN-URBAN-INFRA","HN-URBAN-NORTH","HN-URBAN-WEST",
        "VN-TRAN-2055","VN-URB-METRO-2030","VN-MEKONG-DELTA-2030","VN-RED-RIVER-2030",
        "VN-IP-NORTH-2030","VN-ENV-IND-1894","VN-WW-2030","VN-SWM-NATIONAL-2030",
        "VN-SC-2030","VN-OG-2030","VN-EV-2030","VN-CARBON-2050",
    ]
    for plan_id in PLAN_ORDER:
        arts = plan_buckets.get(plan_id, [])
        short_name = PLAN_NAMES.get(plan_id, plan_id)[:28]  # 시트명 31자 제한
        ws_plan = wb.create_sheet(short_name)
        n = build_plan_sheet(ws_plan, plan_id, arts, short_name)
        print(f"    {plan_id:30s}: {n:3d}건", flush=True)

    print("  ▶ Stats 시트...")
    ws_stats = wb.create_sheet("Stats")
    build_stats_sheet(ws_stats, articles, plan_buckets)

    # ── 참조 시트 4개 생성 (변경이력 포함) ──
    print("  ▶ 참조 시트 4개 생성 (Keywords, Sources, Provinces, Search Queries)...")
    build_ref_masterplan_keywords(wb, articles)
    build_ref_sources(wb, articles)
    build_ref_provinces(wb, articles)
    build_ref_search_queries(wb)

    # 저장 — 파일명 고정 (누적 DB이므로 덮어쓰기)
    now = datetime.now()
    week = now.isocalendar()[1]
    if not output_path:
        # 고정 파일명: AI Drive에서 항상 같은 경로로 갱신됨
        output_path = str(BASE_DIR / "outputs/reports" / "Vietnam_Infra_History_DB.xlsx")

    # 주차별 사본도 별도 저장 (히스토리 보관용)
    snapshot_path = str(
        BASE_DIR / "outputs/reports" /
        f"Vietnam_Infra_History_DB_W{week:02d}_{now.strftime('%Y%m%d')}.xlsx"
    )

    wb.save(output_path)
    wb.save(snapshot_path)
    size_kb = os.path.getsize(output_path) // 1024
    print(f"\n  ✅ 저장 완료 (고정명): {output_path}")
    print(f"  ✅ 주차 스냅샷:        {snapshot_path}")
    print(f"     크기: {size_kb:,}KB | 시트: {len(wb.sheetnames)}개 (All_Articles + Matched_Only + 24플랜 + Stats)")
    return output_path


if __name__ == "__main__":
    sys.path.insert(0, str(BASE_DIR/"scripts"))
    path = generate_history_excel()
    print(f"\n출력: {path}")
