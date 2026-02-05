#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vietnam Infrastructure News - Email Notification
Can be run directly: python send_notification.py
"""

import smtplib
import logging
import sys
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from pathlib import Path
from collections import Counter

# Setup paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / "data"
OUTPUT_DIR = PROJECT_ROOT / "outputs"

sys.path.insert(0, str(PROJECT_ROOT))

from config.settings import (
    EMAIL_SUBJECT, EMAIL_FROM_NAME, EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT,
    DASHBOARD_URL
)

# Get email settings directly from environment
EMAIL_USERNAME = os.getenv("EMAIL_USERNAME", "")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")
EMAIL_RECIPIENTS_RAW = os.getenv("EMAIL_RECIPIENTS", "")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

EXCEL_DB_PATH = DATA_DIR / "database" / "Vietnam_Infra_News_Database_Final.xlsx"


def load_articles():
    """Load articles from Excel"""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed")
        return []
    
    if not EXCEL_DB_PATH.exists():
        logger.warning(f"Excel not found: {EXCEL_DB_PATH}")
        return []
    
    wb = openpyxl.load_workbook(EXCEL_DB_PATH, read_only=True, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
    
    articles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        date_val = row[col_map.get("Date", 4)] if "Date" in col_map else None
        date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, 'strftime') else str(date_val)[:10] if date_val else ""
        
        articles.append({
            "title": row[col_map.get("News Tittle", 3)] or "",
            "sector": row[col_map.get("Business Sector", 1)] or "",
            "province": row[col_map.get("Province", 2)] or "Vietnam",
            "source": row[col_map.get("Source", 5)] or "",
            "url": row[col_map.get("Link", 6)] or "",
            "summary": row[col_map.get("Short summary", 7)] or "",
            "date": date_str
        })
    
    wb.close()
    return articles


def generate_email_html(articles):
    """Generate email HTML content"""
    today = datetime.now().strftime("%Y-%m-%d")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    
    today_articles = [a for a in articles if a.get("date") == today]
    week_articles = [a for a in articles if a.get("date", "") >= week_ago]
    
    sector_counts = Counter(a.get("sector", "Unknown") for a in articles)
    
    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; max-width: 700px; margin: 0 auto; padding: 20px; background: #f8fafc; }}
        .header {{ background: linear-gradient(135deg, #0d9488, #0f766e); color: white; padding: 25px; border-radius: 12px; text-align: center; }}
        .header h1 {{ margin: 0; font-size: 24px; }}
        .kpi-row {{ display: flex; gap: 15px; margin: 20px 0; }}
        .kpi {{ flex: 1; background: white; padding: 15px; border-radius: 10px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.05); }}
        .kpi-value {{ font-size: 28px; font-weight: bold; color: #0d9488; }}
        .kpi-label {{ font-size: 12px; color: #64748b; margin-top: 5px; }}
        .section {{ background: white; border-radius: 10px; padding: 20px; margin: 15px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.05); }}
        .article {{ padding: 12px; border-left: 3px solid #0d9488; margin: 10px 0; background: #f8fafc; border-radius: 0 6px 6px 0; }}
        .article.new {{ border-left-color: #f59e0b; background: #fef3c7; }}
        .btn {{ display: inline-block; background: #0d9488; color: white; padding: 12px 30px; border-radius: 8px; text-decoration: none; }}
        .footer {{ text-align: center; margin-top: 30px; color: #64748b; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🇻🇳 Vietnam Infrastructure News</h1>
        <p style="margin: 5px 0 0 0; opacity: 0.9;">Daily Report - {datetime.now().strftime('%B %d, %Y')}</p>
    </div>
    
    <div class="kpi-row">
        <div class="kpi">
            <div class="kpi-value">{len(today_articles)}</div>
            <div class="kpi-label">Today</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{len(week_articles)}</div>
            <div class="kpi-label">This Week</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{len(articles):,}</div>
            <div class="kpi-label">Total</div>
        </div>
    </div>
    
    <div class="section">
        <h3 style="margin-top: 0;">📊 Top Sectors</h3>
        <p>{" | ".join([f"{s}: {c}" for s, c in sector_counts.most_common(5)])}</p>
    </div>
'''
    
    if today_articles:
        html += f'''
    <div class="section">
        <h3 style="margin-top: 0;">🆕 Today's Articles ({len(today_articles)})</h3>
'''
        for a in today_articles[:10]:
            html += f'''
        <div class="article new">
            <strong>[{a['sector']}]</strong> {a['title'][:100]}{'...' if len(a['title']) > 100 else ''}
            <br><small style="color: #64748b;">{a['source']} | <a href="{a['url']}" style="color: #0d9488;">Read more →</a></small>
        </div>
'''
        html += '</div>'
    
    html += f'''
    <div style="text-align: center; margin: 30px 0;">
        <a href="{DASHBOARD_URL}" class="btn">📊 View Full Dashboard</a>
    </div>
    
    <div class="footer">
        <p>Vietnam Infrastructure News Pipeline</p>
    </div>
</body>
</html>'''
    
    return html


def send_email(html_content):
    """Send email notification"""
    print(f"\n--- Email Configuration ---")
    print(f"SMTP: {EMAIL_SMTP_SERVER}:{EMAIL_SMTP_PORT}")
    print(f"Username: {EMAIL_USERNAME[:3]}***" if EMAIL_USERNAME else "Username: NOT SET")
    print(f"Password: {'*' * 8 if EMAIL_PASSWORD else 'NOT SET'}")
    print(f"Recipients: '{EMAIL_RECIPIENTS_RAW}'")
    
    if not EMAIL_USERNAME or not EMAIL_PASSWORD:
        print("ERROR: Email credentials not configured")
        return False
    
    # Parse recipients
    recipients = []
    if EMAIL_RECIPIENTS_RAW:
        for sep in [',', ';']:
            if sep in EMAIL_RECIPIENTS_RAW:
                recipients = [r.strip() for r in EMAIL_RECIPIENTS_RAW.split(sep) if r.strip()]
                break
        if not recipients:
            recipients = [EMAIL_RECIPIENTS_RAW.strip()] if EMAIL_RECIPIENTS_RAW.strip() else []
    
    if not recipients:
        print("ERROR: No recipients configured")
        return False
    
    try:
        print(f"\nSending to {len(recipients)} recipient(s)...")
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"{EMAIL_SUBJECT} - {datetime.now().strftime('%Y-%m-%d')}"
        msg['From'] = f"{EMAIL_FROM_NAME} <{EMAIL_USERNAME}>"
        msg['To'] = ', '.join(recipients)
        
        msg.attach(MIMEText(html_content, 'html', 'utf-8'))
        
        server = smtplib.SMTP(EMAIL_SMTP_SERVER, EMAIL_SMTP_PORT, timeout=30)
        server.starttls()
        server.login(EMAIL_USERNAME, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        print(f"✓ Email sent to: {', '.join(recipients)}")
        return True
        
    except smtplib.SMTPAuthenticationError as e:
        print(f"ERROR: Authentication failed - {e}")
        print("TIP: Use Gmail App Password, not regular password")
        return False
    except Exception as e:
        print(f"ERROR: {type(e).__name__}: {e}")
        return False


def main():
    """Main function"""
    print("=" * 60)
    print("EMAIL NOTIFICATION")
    print("=" * 60)
    
    articles = load_articles()
    print(f"Loaded {len(articles)} articles")
    
    today = datetime.now().strftime("%Y-%m-%d")
    today_count = sum(1 for a in articles if a.get("date") == today)
    print(f"Today's articles: {today_count}")
    
    html = generate_email_html(articles)
    
    if send_email(html):
        print("\n✓ Email notification sent")
    else:
        print("\n✗ Email send failed")


if __name__ == "__main__":
    main()
